#!/usr/bin/env python3
"""
pz_dual_export.py — PDF + calculation workbook from one process_batch() result
===============================================================================
Creates both deliverables from a single validated process_batch() result:
  1. Final PZ PDF via save_pz_pdf()
  2. Three-sheet audit workbook (.xlsx) via export_pz_calculation_xlsx()

Both outputs are produced from the same validated result — no second calculation
path is introduced.

Dependencies:
    pip install reportlab openpyxl

Usage:
    from pz_import_processor import process_batch
    from pz_dual_export import save_pz_outputs

    result = process_batch(inv_paths, zc429_path, rate=3.6506, batch_meta=batch_meta)
    outputs = save_pz_outputs(
        result      = result,
        pdf_path    = "PZ_039_044.pdf",
        xlsx_path   = "PZ_039_044_calc.xlsx",
        document_no = "PZ 12/3/2026",
    )
    print(outputs)   # {"pdf": Path(...), "xlsx": Path(...)}

Workbook sheets
---------------
  Summary  — batch-level metadata and PLN totals
  Rows     — per-line landed-cost detail (all key fields from process_batch rows)
  Notes    — exact UWAGI lines from process_batch()["notes"]
"""

from __future__ import annotations

from pathlib import Path
from typing import Any, Dict, List, Optional

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from pz_pdf_export import save_pz_pdf


# ── Number formatting ─────────────────────────────────────────────────────────

# Excel format code — displayed as '1 360,18' in Polish locale
PLN_FMT = "#,##0.00"


# ── Style constants ───────────────────────────────────────────────────────────

BOLD      = Font(bold=True)
GRAY_FILL = PatternFill("solid", fgColor="EDEDED")
WRAP_TOP  = Alignment(wrap_text=True, vertical="top")
CENTER    = Alignment(horizontal="center", vertical="top")


def _thin_border() -> Border:
    s = Side(style="thin")
    return Border(left=s, right=s, top=s, bottom=s)


def _apply_border(ws, min_row: int, max_row: int, min_col: int, max_col: int):
    b = _thin_border()
    for row in ws.iter_rows(min_row=min_row, max_row=max_row,
                             min_col=min_col, max_col=max_col):
        for cell in row:
            cell.border = b


# ── Row field accessors ───────────────────────────────────────────────────────
# Canonical keys (post-rename) come first; legacy aliases accepted for compat.

def _r(row: Dict, *keys, default=0.0):
    """Return the first truthy value found among keys."""
    for k in keys:
        v = row.get(k)
        if v is not None and v != 0 and v != "":
            try:
                return float(v)
            except (TypeError, ValueError):
                return v
    return default

def _rs(row: Dict, *keys, default="") -> str:
    for k in keys:
        v = row.get(k)
        if v:
            return str(v)
    return default


# ── XLSX export ───────────────────────────────────────────────────────────────

def export_pz_calculation_xlsx(
    result: Dict[str, Any],
    output_path: str | Path,
    document_no: str = "",
    warehouse:   str = "Główny",
) -> Path:
    """
    Build a three-sheet audit workbook from a validated process_batch() result.

    Sheet layout
    ────────────
    Summary  — document metadata + all batch-level totals and derived rates
    Rows     — 18-column per-line cost chain (all audit fields from engine)
    Notes    — UWAGI lines verbatim from process_batch()

    No recalculation — every value comes directly from the result dict.
    """
    output_path = Path(output_path)
    rows         = result.get("rows",         [])
    notes        = result.get("notes",        [])
    zc           = result.get("zc429",        {})
    totals       = result.get("totals",       {})
    verification = result.get("verification", {})

    doc_no     = document_no or result.get("document_no") or "PZ"
    issue_date = (zc.get("clearance_date") or zc.get("release_date")
                  or zc.get("acceptance_date") or "")

    total_net            = float(result.get("total_net")   or sum(_r(r, "line_netto_pln",  "total_netto")  for r in rows))
    total_gross          = float(result.get("total_gross") or sum(_r(r, "line_brutto_pln", "total_brutto") for r in rows))
    duty_pln             = float(result.get("duty_pln")    or zc.get("duty_pln") or 0)
    total_cif_usd        = float(totals.get("total_cif_usd")         or result.get("total_cif_usd")        or 0)
    zc_cif_usd           = float(zc.get("total_cif_usd") or 0)
    total_before_duty    = float(totals.get("total_before_duty_pln") or 0)
    total_freight_usd    = float(totals.get("total_freight_usd") or 0)
    total_fob_usd        = float(totals.get("total_fob_usd") or 0)
    duty_rate_pct        = float(totals.get("duty_rate_pct") or 0)
    vat_pln              = float(zc.get("vat_pln") or 0)
    usd_pln              = float(totals.get("usd_pln") or result.get("nbp", {}).get("usd_rate") or 0)

    # Effective freight % across all invoices
    eff_freight_pct = (total_freight_usd / total_fob_usd * 100) if total_fob_usd else 0

    # ── Verification values ───────────────────────────────────────────────────
    def _vc(key) -> str:
        v = verification.get(key)
        if v is True:  return "YES"
        if v is False: return "NO"
        return "(not parsed)"

    amendment_flags = verification.get("amendment_flags", [])
    flags_str       = "; ".join(amendment_flags) if amendment_flags else "none"
    nbp_rate_used   = verification.get("nbp_rate_used") or usd_pln or ""
    sad_customs_rate= verification.get("sad_customs_rate") or zc.get("customs_rate_usd") or ""

    wb = openpyxl.Workbook()
    wb.remove(wb.active)   # drop the default empty sheet

    # ── Sheet 1: Summary ──────────────────────────────────────────────────────
    ws_s = wb.create_sheet("Summary")

    PCT_FMT = "0.0000%"
    summary_rows: List[tuple] = [
        ("Field",                          "Value",                      None),
        ("Document No",                    doc_no,                       None),
        ("Issue Date",                     issue_date,                   None),
        ("Warehouse",                      warehouse,                    None),
        ("Line Count",                     len(rows),                    None),
        ("",                               "",                           None),
        ("Total CIF USD (invoices)",       total_cif_usd,                "#,##0.00"),
        ("SAD/ZC429 CIF USD (declared)",   zc_cif_usd,                  "#,##0.00"),
        ("CIF Difference USD",             round(total_cif_usd - zc_cif_usd, 2), "#,##0.00"),
        ("",                               "",                           None),
        ("A00 Duty PLN",                   duty_pln,                    PLN_FMT),
        ("B00 VAT PLN (ref only)",         vat_pln,                     PLN_FMT),
        ("Total Before Duty PLN",          total_before_duty,           PLN_FMT),
        ("",                               "",                           None),
        ("Effective Freight % (batch)",    eff_freight_pct / 100,       PCT_FMT),
        ("Effective Duty % (batch)",       duty_rate_pct / 100,         PCT_FMT),
        ("NBP Rate USD/PLN",               nbp_rate_used,               "0.0000"),
        ("SAD Customs Rate USD/PLN",       sad_customs_rate,            "0.0000"),
        ("",                               "",                           None),
        ("Total Netto PLN",                total_net,                   PLN_FMT),
        ("Total Brutto PLN",               total_gross,                 PLN_FMT),
        ("",                               "",                           None),
        ("Invoice Refs Match",             _vc("invoice_refs_match"),   None),
        ("CIF Total Match",                _vc("cif_match"),            None),
        ("Qty by Type Match",              _vc("qty_match_by_type"),    None),
        ("Importer Match",                 _vc("importer_match"),       None),
        ("Exporter Match",                 _vc("exporter_match"),       None),
        ("Amendment Flags",                flags_str,                   None),
    ]

    for r_idx, (field, value, fmt) in enumerate(summary_rows, 1):
        cf = ws_s.cell(r_idx, 1, field)
        cv = ws_s.cell(r_idx, 2, value)
        if r_idx == 1:
            cf.font = BOLD; cf.fill = GRAY_FILL
            cv.font = BOLD; cv.fill = GRAY_FILL
        # Highlight NO checks in light red
        if value == "NO":
            cv.fill = PatternFill("solid", fgColor="FFD0D0")
        if fmt and value not in ("", None):
            try:
                cv.number_format = fmt
            except Exception:
                pass

    _apply_border(ws_s, 1, len(summary_rows), 1, 2)
    ws_s.column_dimensions["A"].width = 34
    ws_s.column_dimensions["B"].width = 55

    # ── Sheet 2: Rows ─────────────────────────────────────────────────────────
    # 22 columns — full engine output + item type + audit per-unit columns.
    # Allocation logic is in the engine (calculate_landed), not here.
    ws_r = wb.create_sheet("Rows")

    ROW_HDRS = [
        "Lp",               # 1
        "Invoice No",       # 2
        "English Name",     # 3
        "Polish Name",      # 4
        "Qty",              # 5
        "Unit USD",         # 6
        "Line USD",         # 7
        "Freight %",        # 8
        "Alloc. F+I USD",   # 9
        "Alloc. F+I PLN",   # 10
        "Rate PLN/USD",     # 11
        "Before Duty PLN",  # 12
        "Duty %",           # 13
        "Alloc. Duty PLN",  # 14
        "Unit Netto PLN",   # 15
        "Line Netto PLN",   # 16
        "Line Brutto PLN",  # 17
        "VAT",              # 18
        "Item Type",        # 19
        "Supplier Ref",     # 20  (optional — seller_name from invoice)
        "F+I per pcs PLN",  # 21  display-only audit column
        "Duty per pcs PLN", # 22  display-only audit column
    ]
    # Columns that should display as currency / number (1-based)
    USD_FMT_COLS  = {6, 7, 9}                       # USD amounts
    PLN_FMT_COLS  = {10, 12, 14, 15, 16, 17, 21, 22}  # PLN amounts
    PCT_FMT_COLS  = {8, 13}                          # percentage columns
    RATE_FMT_COLS = {11}                             # exchange rate

    for c_idx, hdr in enumerate(ROW_HDRS, 1):
        cell = ws_r.cell(1, c_idx, hdr)
        cell.font      = BOLD
        cell.fill      = GRAY_FILL
        cell.alignment = WRAP_TOP

    for r_idx, row in enumerate(rows, 2):
        qty     = float(row.get("quantity") or 1)
        qty_val = int(qty) if qty == int(qty) else qty
        frt_pct = float(row.get("freight_rate_pct") or 0)
        dty_pct = duty_rate_pct / 100   # same for all rows (batch-level)

        alloc_ship_pln = _r(row, "allocated_ship_pln")
        alloc_duty_pln = _r(row, "allocated_duty_pln")
        fpi_per_pcs    = alloc_ship_pln / qty if qty else 0.0
        duty_per_pcs   = alloc_duty_pln / qty if qty else 0.0

        row_vals = [
            r_idx - 1,                                           # 1  Lp
            _rs(row, "invoice_no"),                              # 2  Invoice No
            _rs(row, "description_en"),                          # 3  English Name
            _rs(row, "pl_desc"),                                 # 4  Polish Name
            qty_val,                                             # 5  Qty
            _r(row, "unit_price_usd"),                           # 6  Unit USD
            _r(row, "total_usd"),                                # 7  Line USD
            frt_pct,                                             # 8  Freight %
            _r(row, "allocated_ship_usd"),                       # 9  Alloc. F+I USD
            alloc_ship_pln,                                      # 10 Alloc. F+I PLN
            _r(row, "usd_pln"),                                  # 11 Rate PLN/USD
            _r(row, "before_duty_pln"),                          # 12 Before Duty PLN
            dty_pct,                                             # 13 Duty %
            alloc_duty_pln,                                      # 14 Alloc. Duty PLN
            _r(row, "unit_netto_pln",  "landed_per_unit"),       # 15 Unit Netto PLN
            _r(row, "line_netto_pln",  "total_netto"),           # 16 Line Netto PLN
            _r(row, "line_brutto_pln", "total_brutto"),          # 17 Line Brutto PLN
            _rs(row, "vat_rate", default="23%"),                 # 18 VAT
            _rs(row, "item_type"),                               # 19 Item Type
            _rs(row, "seller_name", "supplier_ref"),             # 20 Supplier Ref
            fpi_per_pcs,                                         # 21 F+I per pcs PLN (display)
            duty_per_pcs,                                        # 22 Duty per pcs PLN (display)
        ]

        for c_idx, val in enumerate(row_vals, 1):
            cell = ws_r.cell(r_idx, c_idx, val)
            if c_idx in {3, 4}:
                cell.alignment = WRAP_TOP
            if c_idx in USD_FMT_COLS:
                cell.number_format = "#,##0.00"
            elif c_idx in PLN_FMT_COLS:
                cell.number_format = PLN_FMT
            elif c_idx in PCT_FMT_COLS:
                cell.number_format = "0.00%"
            elif c_idx in RATE_FMT_COLS:
                cell.number_format = "0.0000"

    data_end = len(rows) + 1
    _apply_border(ws_r, 1, data_end, 1, len(ROW_HDRS))
    ws_r.freeze_panes = "E2"   # freeze Lp + Invoice + both name columns

    # Column widths for 22 columns (A=1, V=22)
    col_widths = {
        "A": 5,   "B": 18,  "C": 44,  "D": 44,  "E": 6,
        "F": 10,  "G": 10,  "H": 10,  "I": 14,  "J": 14,
        "K": 12,  "L": 16,  "M": 10,  "N": 16,  "O": 15,
        "P": 15,  "Q": 16,  "R": 6,   "S": 12,  "T": 28,
        "U": 15,  "V": 16,
    }
    for col_letter, width in col_widths.items():
        ws_r.column_dimensions[col_letter].width = width

    # ── Sheet 3: Notes ────────────────────────────────────────────────────────
    ws_n = wb.create_sheet("Notes")

    # Section 1: UWAGI lines
    ws_n.cell(1, 1, "Section").font   = BOLD
    ws_n.cell(1, 1).fill              = GRAY_FILL
    ws_n.cell(1, 2, "Content").font   = BOLD
    ws_n.cell(1, 2).fill              = GRAY_FILL

    r_idx = 2
    for i, line in enumerate(notes, 1):
        ws_n.cell(r_idx, 1, f"UWAGI {i}").alignment = CENTER
        ws_n.cell(r_idx, 2, line).alignment          = WRAP_TOP
        r_idx += 1

    # Section 2: Verification checklist
    ws_n.cell(r_idx, 1, "").value = ""
    r_idx += 1
    ws_n.cell(r_idx, 1, "VERIFICATION CHECKLIST").font = BOLD
    ws_n.cell(r_idx, 1).fill = GRAY_FILL
    r_idx += 1

    def _vc_label(key) -> str:
        v = verification.get(key)
        if v is True:  return "YES ✓"
        if v is False: return "NO ✗"
        return "(not parsed)"

    check_rows = [
        ("Invoice refs match",  "invoice_refs_match"),
        ("CIF total match",     "cif_match"),
        ("Qty by type match",   "qty_match_by_type"),
        ("Importer match",      "importer_match"),
        ("Exporter match",      "exporter_match"),
        ("Blocked phrases clean", "blocked_phrases_clean"),
        ("Duty rate plausible", "duty_rate_ok"),
    ]
    for label, key in check_rows:
        ws_n.cell(r_idx, 1, label).alignment = WRAP_TOP
        ws_n.cell(r_idx, 2, _vc_label(key)).alignment = WRAP_TOP
        r_idx += 1

    # Section 3: Amendment flags
    ws_n.cell(r_idx, 1, "").value = ""
    r_idx += 1
    ws_n.cell(r_idx, 1, "AMENDMENT FLAGS").font = BOLD
    ws_n.cell(r_idx, 1).fill = GRAY_FILL
    r_idx += 1

    if amendment_flags:
        for flag in amendment_flags:
            ws_n.cell(r_idx, 1, "⚑").alignment = CENTER
            ws_n.cell(r_idx, 2, flag).alignment = WRAP_TOP
            r_idx += 1
    else:
        ws_n.cell(r_idx, 2, "None — all checks passed or not applicable").alignment = WRAP_TOP
        r_idx += 1

    _apply_border(ws_n, 1, r_idx - 1, 1, 2)
    ws_n.column_dimensions["A"].width = 28
    ws_n.column_dimensions["B"].width = 90

    wb.save(str(output_path))
    return output_path


# ── Combined one-call save ────────────────────────────────────────────────────

def save_pz_outputs(
    result:      Dict[str, Any],
    pdf_path:    str | Path,
    xlsx_path:   str | Path,
    document_no: str = "",
    warehouse:   str = "Główny",
    recipient:   Optional[Dict[str, str]] = None,
    supplier:    Optional[Dict[str, str]] = None,
) -> Dict[str, Path]:
    """
    Save the final PZ PDF and calculation workbook in one call.

    Both outputs come from the same validated process_batch() result —
    no second calculation path is introduced.
    """
    pdf_file = save_pz_pdf(
        result      = result,
        output_path = pdf_path,
        document_no = document_no,
        warehouse   = warehouse,
        recipient   = recipient,
        supplier    = supplier,
    )
    xlsx_file = export_pz_calculation_xlsx(
        result      = result,
        output_path = xlsx_path,
        document_no = document_no,
        warehouse   = warehouse,
    )
    return {"pdf": pdf_file, "xlsx": xlsx_file}
