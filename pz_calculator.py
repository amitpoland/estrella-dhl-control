#!/usr/bin/env python3
"""
PZ Calculator — Landed Cost Calculator
Reads invoice XLSX files, fetches live USD/PLN rate from NBP,
calculates landed costs, and outputs a formatted Excel report.

Usage:
    python pz_calculator.py --invoices ./invoices --dhl 1181.00
    python pz_calculator.py --invoices ./invoices --dhl 1181.00 --rate 3.7206
    python pz_calculator.py --invoices ./invoices --dhl 1181.00 --output March2026_PZ.xlsx
"""

import argparse
import glob
import os
import sys
from datetime import date, datetime
from pathlib import Path

try:
    import requests
except ImportError:
    sys.exit("Missing dependency: pip install requests openpyxl")

try:
    import openpyxl
    from openpyxl import Workbook
    from openpyxl.styles import (
        Alignment, Border, Font, PatternFill, Side
    )
    from openpyxl.utils import get_column_letter
except ImportError:
    sys.exit("Missing dependency: pip install openpyxl requests")


# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------
DUTY_RATE = 0.12        # 12% import duty on CIF value
VAT_RATE  = 0.23        # 23% VAT
DHL_SHARE = 0.50        # DHL cost split: 50% to freight (rest to insurance)
INSURANCE_PCT = 0.005   # 0.5% of FOB for insurance (CIF calc)

# Column header variants found in supplier invoices (case-insensitive)
POSSIBLE_HEADERS = {
    "item_no":      ["item no", "item no.", "item#", "item number", "no", "lp", "pos"],
    "description":  ["description", "product description", "goods description",
                     "item description", "opis", "nazwa"],
    "qty":          ["qty", "quantity", "pcs", "ilość", "ilosc", "units"],
    "unit_price":   ["unit price", "price", "unit price usd", "price usd",
                     "cena", "cena jednostkowa"],
    "total_usd":    ["total", "total usd", "amount", "amount usd", "wartość",
                     "wartosc", "total amount", "line total"],
}

# Colour palette
CLR_HEADER_DARK  = "1F3864"   # dark navy
CLR_HEADER_MID   = "2E75B6"   # medium blue
CLR_ACCENT       = "BDD7EE"   # light blue
CLR_HIGHLIGHT    = "FFFF00"   # yellow for totals
CLR_GREEN        = "E2EFDA"   # light green for VAT row
CLR_WHITE        = "FFFFFF"
CLR_LIGHT_GREY   = "F2F2F2"


# ---------------------------------------------------------------------------
# NBP rate fetch
# ---------------------------------------------------------------------------

def fetch_nbp_rate(currency: str = "USD") -> float:
    """Fetch today's mid rate from NBP Table A. Falls back to yesterday on weekends."""
    base = "https://api.nbp.pl/api/exchangerates/rates/A"
    for offset in range(5):          # try up to 5 days back (long weekends)
        day = date.today()
        if offset:
            from datetime import timedelta
            day = date.today() - timedelta(days=offset)
        url = f"{base}/{currency}/{day.isoformat()}/?format=json"
        try:
            r = requests.get(url, timeout=10)
            if r.status_code == 200:
                data = r.json()
                rate = data["rates"][0]["mid"]
                print(f"  NBP rate {currency}/PLN on {day}: {rate:.4f}")
                return float(rate)
        except Exception:
            pass
    raise RuntimeError(
        f"Could not fetch {currency}/PLN rate from NBP. "
        "Use --rate to specify it manually."
    )


# ---------------------------------------------------------------------------
# Invoice parsing
# ---------------------------------------------------------------------------

def _find_header_row(ws):
    """Return (row_index, col_map) where col_map maps field→col_index (1-based)."""
    for row_idx in range(1, 30):
        row_vals = [
            str(ws.cell(row=row_idx, column=c).value or "").strip().lower()
            for c in range(1, ws.max_column + 1)
        ]
        found = {}
        for field, variants in POSSIBLE_HEADERS.items():
            for col_i, cell_val in enumerate(row_vals, start=1):
                if cell_val in variants:
                    found[field] = col_i
                    break
        # Need at least qty + (unit_price or total_usd) + description
        if "qty" in found and ("unit_price" in found or "total_usd" in found):
            return row_idx, found
    return None, {}


def parse_invoice(filepath: str) -> dict:
    """
    Parse a single invoice XLSX.
    Returns dict with keys: filename, invoice_no, lines[]
    Each line: {item_no, description, qty, unit_price_usd, total_usd}
    """
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb.active

    header_row, col_map = _find_header_row(ws)
    if header_row is None:
        print(f"  WARNING: Could not find header row in {os.path.basename(filepath)} — skipping.")
        return None

    # Try to extract invoice number from the first ~10 rows
    invoice_no = ""
    for r in range(1, header_row):
        for c in range(1, ws.max_column + 1):
            val = str(ws.cell(row=r, column=c).value or "")
            if any(k in val.upper() for k in ["INVOICE NO", "INVOICE NUMBER", "INV NO", "INV#"]):
                # value might be in same cell or next column
                candidate = str(ws.cell(row=r, column=c + 1).value or "").strip()
                if not candidate:
                    # maybe "Invoice No: EJL-123" in one cell
                    parts = val.split(":")
                    if len(parts) > 1:
                        candidate = parts[-1].strip()
                if candidate:
                    invoice_no = candidate
                    break
        if invoice_no:
            break
    if not invoice_no:
        invoice_no = Path(filepath).stem[:30]

    lines = []
    for row_idx in range(header_row + 1, ws.max_row + 1):
        def cv(field):
            if field not in col_map:
                return None
            return ws.cell(row=row_idx, column=col_map[field]).value

        qty_raw = cv("qty")
        if qty_raw is None or str(qty_raw).strip() == "":
            continue
        try:
            qty = float(qty_raw)
        except (ValueError, TypeError):
            continue
        if qty <= 0:
            continue

        desc = str(cv("description") or "").strip()
        item_no = str(cv("item_no") or row_idx - header_row).strip()

        unit_price = cv("unit_price")
        total_usd  = cv("total_usd")

        # Derive missing values
        try:
            unit_price = float(unit_price) if unit_price is not None else None
        except (ValueError, TypeError):
            unit_price = None
        try:
            total_usd = float(total_usd) if total_usd is not None else None
        except (ValueError, TypeError):
            total_usd = None

        if unit_price is None and total_usd is not None:
            unit_price = total_usd / qty if qty else 0
        if total_usd is None and unit_price is not None:
            total_usd = unit_price * qty

        if unit_price is None or total_usd is None:
            continue

        lines.append({
            "item_no":       item_no,
            "description":   desc,
            "qty":           qty,
            "unit_price_usd": unit_price,
            "total_usd":     total_usd,
        })

    return {
        "filename":   os.path.basename(filepath),
        "invoice_no": invoice_no,
        "lines":      lines,
    }


# ---------------------------------------------------------------------------
# Landed cost calculation
# ---------------------------------------------------------------------------

def calculate_landed(invoices: list, dhl_pln: float, usd_pln: float) -> list:
    """
    Add landed-cost fields to every line across all invoices.

    CIF  = FOB * (1 + INSURANCE_PCT) + freight_share_per_unit
    Duty = CIF * DUTY_RATE
    Landed_USD = CIF + Duty
    Landed_PLN = Landed_USD * usd_pln

    DHL is allocated proportionally by FOB value.
    """
    # Total FOB in USD across all lines
    all_lines = []
    total_fob_usd = 0.0
    for inv in invoices:
        for line in inv["lines"]:
            all_lines.append((inv, line))
            total_fob_usd += line["total_usd"]

    enriched = []
    for inv, line in all_lines:
        fob_usd   = line["total_usd"]
        qty       = line["qty"]

        # DHL share proportional to FOB value
        dhl_line_pln  = (fob_usd / total_fob_usd) * dhl_pln if total_fob_usd else 0
        dhl_line_usd  = dhl_line_pln / usd_pln if usd_pln else 0

        # Freight + insurance → CIF
        insurance_usd = fob_usd * INSURANCE_PCT
        freight_usd   = dhl_line_usd * DHL_SHARE
        cif_usd       = fob_usd + insurance_usd + freight_usd

        # Duty
        duty_usd      = cif_usd * DUTY_RATE

        # Landed
        landed_usd    = cif_usd + duty_usd
        landed_pln    = landed_usd * usd_pln

        # VAT base = landed PLN
        vat_pln       = landed_pln * VAT_RATE
        brutto_pln    = landed_pln + vat_pln

        # Per-piece values
        landed_per_pc_pln = landed_pln / qty if qty else 0
        brutto_per_pc_pln = brutto_pln / qty if qty else 0

        enriched.append({
            **line,
            "invoice_no":          inv["invoice_no"],
            "filename":            inv["filename"],
            "fob_usd":             fob_usd,
            "dhl_share_pln":       dhl_line_pln,
            "dhl_share_usd":       dhl_line_usd,
            "insurance_usd":       insurance_usd,
            "freight_usd":         freight_usd,
            "cif_usd":             cif_usd,
            "duty_usd":            duty_usd,
            "landed_usd":          landed_usd,
            "landed_pln":          landed_pln,
            "vat_pln":             vat_pln,
            "brutto_pln":          brutto_pln,
            "landed_per_pc_pln":   landed_per_pc_pln,
            "brutto_per_pc_pln":   brutto_per_pc_pln,
            "usd_pln_rate":        usd_pln,
        })

    return enriched


# ---------------------------------------------------------------------------
# Excel output helpers
# ---------------------------------------------------------------------------

def _font(bold=False, size=11, color="000000", name="Calibri"):
    return Font(bold=bold, size=size, color=color, name=name)

def _fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def _border(style="thin"):
    s = Side(style=style)
    return Border(left=s, right=s, top=s, bottom=s)

def _align(h="center", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def _money(ws, row, col, value, fmt="#,##0.00"):
    c = ws.cell(row=row, column=col, value=round(value, 2))
    c.number_format = fmt
    return c

def _set_col_width(ws, col, width):
    ws.column_dimensions[get_column_letter(col)].width = width


# ---------------------------------------------------------------------------
# Sheet: INPUT
# ---------------------------------------------------------------------------

def write_input_sheet(ws, invoices, usd_pln, dhl_pln):
    ws.title = "INPUT"
    ws.sheet_view.showGridLines = False

    # Title
    ws.merge_cells("A1:G1")
    t = ws["A1"]
    t.value = "INPUT DATA — Invoice Lines"
    t.font  = _font(bold=True, size=14, color=CLR_WHITE)
    t.fill  = _fill(CLR_HEADER_DARK)
    t.alignment = _align()
    ws.row_dimensions[1].height = 28

    # Parameters row
    ws.merge_cells("A2:G2")
    p = ws["A2"]
    p.value = (f"USD/PLN: {usd_pln:.4f}    |    DHL (PLN): {dhl_pln:,.2f}    |    "
               f"Duty: {DUTY_RATE*100:.0f}%    |    VAT: {VAT_RATE*100:.0f}%    |    "
               f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}")
    p.font  = _font(size=10, color="FFFFFF")
    p.fill  = _fill(CLR_HEADER_MID)
    p.alignment = _align(h="left")

    headers = ["Invoice No", "File", "Item No", "Description", "Qty",
               "Unit Price USD", "Total FOB USD"]
    col_widths = [20, 35, 10, 45, 8, 16, 16]

    for ci, (h, w) in enumerate(zip(headers, col_widths), start=1):
        c = ws.cell(row=3, column=ci, value=h)
        c.font      = _font(bold=True, size=10, color=CLR_WHITE)
        c.fill      = _fill(CLR_HEADER_MID)
        c.alignment = _align()
        c.border    = _border()
        _set_col_width(ws, ci, w)

    row = 4
    for inv in invoices:
        for line in inv["lines"]:
            ws.cell(row=row, column=1, value=inv["invoice_no"]).alignment = _align(h="left")
            ws.cell(row=row, column=2, value=inv["filename"]).alignment   = _align(h="left")
            ws.cell(row=row, column=3, value=line["item_no"]).alignment   = _align()
            ws.cell(row=row, column=4, value=line["description"]).alignment = _align(h="left", wrap=True)
            ws.cell(row=row, column=5, value=line["qty"]).alignment       = _align()
            _money(ws, row, 6, line["unit_price_usd"])
            _money(ws, row, 7, line["total_usd"])
            fill = _fill(CLR_LIGHT_GREY) if row % 2 == 0 else _fill(CLR_WHITE)
            for ci in range(1, 8):
                ws.cell(row=row, column=ci).fill   = fill
                ws.cell(row=row, column=ci).border = _border()
                ws.cell(row=row, column=ci).font   = _font(size=10)
            row += 1

    ws.freeze_panes = "A4"


# ---------------------------------------------------------------------------
# Sheet: PZ_CALC
# ---------------------------------------------------------------------------

def write_pzcalc_sheet(ws, lines):
    ws.title = "PZ_CALC"
    ws.sheet_view.showGridLines = False

    # Title
    ws.merge_cells("A1:P1")
    t = ws["A1"]
    t.value     = "PZ CALCULATION — Landed Cost Breakdown"
    t.font      = _font(bold=True, size=14, color=CLR_WHITE)
    t.fill      = _fill(CLR_HEADER_DARK)
    t.alignment = _align()
    ws.row_dimensions[1].height = 28

    headers = [
        "Invoice No", "Item No", "Description", "Qty",
        "Unit Price\nUSD", "FOB Total\nUSD",
        "Insurance\nUSD", "Freight\nUSD", "CIF\nUSD",
        "Duty 12%\nUSD", "Landed\nUSD",
        "USD/PLN", "Landed\nPLN",
        "Landed/pc\nPLN",
        "VAT 23%\nPLN", "Brutto\nPLN",
    ]
    col_widths = [20, 10, 38, 7, 12, 14, 12, 12, 14, 12, 14, 9, 14, 14, 12, 14]

    for ci, (h, w) in enumerate(zip(headers, col_widths), start=1):
        c = ws.cell(row=2, column=ci, value=h)
        c.font      = _font(bold=True, size=9, color=CLR_WHITE)
        c.fill      = _fill(CLR_HEADER_MID)
        c.alignment = _align(wrap=True)
        c.border    = _border()
        _set_col_width(ws, ci, w)
        ws.row_dimensions[2].height = 32

    row = 3
    for i, line in enumerate(lines):
        alt = (i % 2 == 0)
        bg  = _fill(CLR_ACCENT if alt else CLR_WHITE)

        def cell(col, value, fmt=None, align_h="center"):
            c = ws.cell(row=row, column=col, value=value)
            c.fill   = bg
            c.border = _border()
            c.font   = _font(size=9)
            c.alignment = _align(h=align_h)
            if fmt:
                c.number_format = fmt
            return c

        cell(1,  line["invoice_no"],     align_h="left")
        cell(2,  line["item_no"])
        cell(3,  line["description"],    align_h="left")
        cell(4,  line["qty"])
        cell(5,  line["unit_price_usd"], fmt="#,##0.0000")
        cell(6,  line["fob_usd"],        fmt="#,##0.00")
        cell(7,  line["insurance_usd"],  fmt="#,##0.00")
        cell(8,  line["freight_usd"],    fmt="#,##0.00")
        cell(9,  line["cif_usd"],        fmt="#,##0.00")
        cell(10, line["duty_usd"],       fmt="#,##0.00")
        cell(11, line["landed_usd"],     fmt="#,##0.00")
        cell(12, line["usd_pln_rate"],   fmt="0.0000")
        cell(13, line["landed_pln"],     fmt="#,##0.00")
        cell(14, line["landed_per_pc_pln"], fmt="#,##0.00")
        cell(15, line["vat_pln"],        fmt="#,##0.00")
        cell(16, line["brutto_pln"],     fmt="#,##0.00")
        row += 1

    # Totals row
    ws.row_dimensions[row].height = 18
    total_labels = {1: "TOTAL", 4: sum(l["qty"] for l in lines)}
    total_cols   = {6: "fob_usd", 7: "insurance_usd", 8: "freight_usd",
                    9: "cif_usd", 10: "duty_usd", 11: "landed_usd",
                    13: "landed_pln", 15: "vat_pln", 16: "brutto_pln"}

    for ci in range(1, 17):
        c = ws.cell(row=row, column=ci)
        c.fill   = _fill(CLR_HIGHLIGHT)
        c.border = _border("medium")
        c.font   = _font(bold=True, size=10)
        c.alignment = _align()
        if ci in total_labels:
            c.value = total_labels[ci]
        elif ci in total_cols:
            c.value = sum(l[total_cols[ci]] for l in lines)
            c.number_format = "#,##0.00"

    ws.freeze_panes = "A3"
    return row  # totals row index


# ---------------------------------------------------------------------------
# Sheet: DASHBOARD
# ---------------------------------------------------------------------------

def write_dashboard_sheet(ws, lines, dhl_pln, usd_pln, totals_row_pzcalc):
    ws.title = "DASHBOARD"
    ws.sheet_view.showGridLines = False

    total_fob_usd   = sum(l["fob_usd"]    for l in lines)
    total_cif_usd   = sum(l["cif_usd"]    for l in lines)
    total_duty_usd  = sum(l["duty_usd"]   for l in lines)
    total_landed_usd = sum(l["landed_usd"] for l in lines)
    total_landed_pln = sum(l["landed_pln"] for l in lines)
    total_vat_pln   = sum(l["vat_pln"]    for l in lines)
    total_brutto_pln = sum(l["brutto_pln"] for l in lines)
    total_qty       = sum(l["qty"]         for l in lines)

    # Header
    ws.merge_cells("B2:F2")
    t = ws["B2"]
    t.value     = "PZ LANDED COST — SUMMARY DASHBOARD"
    t.font      = _font(bold=True, size=16, color=CLR_WHITE)
    t.fill      = _fill(CLR_HEADER_DARK)
    t.alignment = _align()
    ws.row_dimensions[2].height = 36

    ws.merge_cells("B3:F3")
    sub = ws["B3"]
    sub.value     = (f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}    |    "
                     f"USD/PLN: {usd_pln:.4f}    |    Lines: {len(lines)}    |    Qty: {total_qty:,.0f}")
    sub.font      = _font(size=10, color=CLR_WHITE)
    sub.fill      = _fill(CLR_HEADER_MID)
    sub.alignment = _align()

    def kv_row(row, label, value, fmt=None, bold_val=True, bg=CLR_LIGHT_GREY):
        ws.merge_cells(f"B{row}:C{row}")
        lc = ws[f"B{row}"]
        lc.value     = label
        lc.font      = _font(bold=True, size=11)
        lc.fill      = _fill(bg)
        lc.alignment = _align(h="left")
        lc.border    = _border()

        ws.merge_cells(f"D{row}:F{row}")
        vc = ws[f"D{row}"]
        vc.value     = value
        vc.font      = _font(bold=bold_val, size=12)
        vc.fill      = _fill(bg)
        vc.alignment = _align(h="right")
        vc.border    = _border()
        if fmt:
            vc.number_format = fmt
        ws.row_dimensions[row].height = 22

    r = 5
    ws.row_dimensions[r].height = 10   # spacer

    kv_row(r+1, "Total Lines",         len(lines),          fmt="#,##0")
    kv_row(r+2, "Total Qty (pcs)",     total_qty,           fmt="#,##0")
    kv_row(r+3, "NBP Rate USD/PLN",    usd_pln,             fmt="0.0000")
    kv_row(r+4, "DHL Cost (PLN)",      dhl_pln,             fmt="#,##0.00")

    ws.row_dimensions[r+5].height = 10  # spacer

    kv_row(r+6,  "FOB Value (USD)",    total_fob_usd,       fmt="#,##0.00", bg=CLR_ACCENT)
    kv_row(r+7,  "CIF Value (USD)",    total_cif_usd,       fmt="#,##0.00", bg=CLR_ACCENT)
    kv_row(r+8,  "Duty 12% (USD)",     total_duty_usd,      fmt="#,##0.00", bg=CLR_ACCENT)
    kv_row(r+9,  "Landed Total (USD)", total_landed_usd,    fmt="#,##0.00", bg=CLR_ACCENT)

    ws.row_dimensions[r+10].height = 10  # spacer

    kv_row(r+11, "Landed Netto (PLN)", total_landed_pln,    fmt="#,##0.00", bg=CLR_GREEN)
    kv_row(r+12, "VAT 23% (PLN)",      total_vat_pln,       fmt="#,##0.00", bg=CLR_GREEN)
    kv_row(r+13, "Landed BRUTTO (PLN)", total_brutto_pln,   fmt="#,##0.00", bg=CLR_HIGHLIGHT)

    _set_col_width(ws, 2, 4)   # A narrow margin col
    _set_col_width(ws, 3, 28)  # label
    _set_col_width(ws, 4, 4)
    _set_col_width(ws, 5, 18)  # value
    _set_col_width(ws, 6, 18)


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser(
        description="PZ Calculator — Landed Cost from Invoice XLSX files"
    )
    parser.add_argument("--invoices", required=True,
                        help="Folder containing invoice XLSX files")
    parser.add_argument("--dhl",  required=True, type=float,
                        help="Total DHL cost in PLN")
    parser.add_argument("--rate", type=float, default=None,
                        help="USD/PLN rate (optional; fetched from NBP if omitted)")
    parser.add_argument("--output", default="PZ_Calc_Output.xlsx",
                        help="Output filename (default: PZ_Calc_Output.xlsx)")
    args = parser.parse_args()

    # 1. Validate inputs
    inv_folder = Path(args.invoices)
    if not inv_folder.is_dir():
        sys.exit(f"ERROR: Invoices folder not found: {inv_folder}")

    xlsx_files = sorted(glob.glob(str(inv_folder / "*.xlsx")) +
                        glob.glob(str(inv_folder / "*.XLSX")))
    if not xlsx_files:
        sys.exit(f"ERROR: No XLSX files found in {inv_folder}")

    print(f"\nPZ Calculator")
    print(f"{'='*55}")
    print(f"  Invoices folder : {inv_folder.resolve()}")
    print(f"  Files found     : {len(xlsx_files)}")
    print(f"  DHL cost (PLN)  : {args.dhl:,.2f}")

    # 2. Fetch or use manual rate
    if args.rate:
        usd_pln = args.rate
        print(f"  USD/PLN rate    : {usd_pln:.4f}  (manual)")
    else:
        print(f"  Fetching USD/PLN from NBP...")
        usd_pln = fetch_nbp_rate("USD")

    # 3. Parse invoices
    print(f"\nParsing invoices...")
    invoices = []
    for fp in xlsx_files:
        print(f"  → {os.path.basename(fp)}")
        inv = parse_invoice(fp)
        if inv and inv["lines"]:
            invoices.append(inv)
            print(f"     Invoice: {inv['invoice_no']}  |  {len(inv['lines'])} lines")

    if not invoices:
        sys.exit("ERROR: No valid invoice lines found.")

    total_lines = sum(len(i["lines"]) for i in invoices)
    total_fob   = sum(l["total_usd"] for i in invoices for l in i["lines"])
    print(f"\n  Total invoice lines : {total_lines}")
    print(f"  Total FOB (USD)     : {total_fob:,.2f}")

    # 4. Calculate landed costs
    lines = calculate_landed(invoices, args.dhl, usd_pln)

    # 5. Build Excel output
    print(f"\nBuilding Excel report...")
    wb = Workbook()

    # Sheet order: DASHBOARD first (active by default)
    ws_dash = wb.active
    ws_pz   = wb.create_sheet()
    ws_in   = wb.create_sheet()

    write_input_sheet(ws_in, invoices, usd_pln, args.dhl)
    totals_row = write_pzcalc_sheet(ws_pz, lines)
    write_dashboard_sheet(ws_dash, lines, args.dhl, usd_pln, totals_row)

    output_path = Path(args.output)
    wb.save(output_path)

    # 6. Summary
    total_landed_pln = sum(l["landed_pln"] for l in lines)
    total_brutto_pln = sum(l["brutto_pln"] for l in lines)
    print(f"\n{'='*55}")
    print(f"  Output file         : {output_path.resolve()}")
    print(f"  Landed NETTO (PLN)  : {total_landed_pln:>14,.2f}")
    print(f"  VAT 23%   (PLN)     : {sum(l['vat_pln'] for l in lines):>14,.2f}")
    print(f"  Landed BRUTTO (PLN) : {total_brutto_pln:>14,.2f}")
    print(f"{'='*55}\n")


if __name__ == "__main__":
    main()
