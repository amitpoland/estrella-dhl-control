"""
routes_wfirma.py — wFirma PZ Export
=====================================
Exports generated PZ data into formats ready for wFirma.pl import.

Modes
-----
  Mode 1  — Clipboard (tab-separated)
    POST /api/v1/upload/shipment/{batch_id}/wfirma/clipboard
    Returns tab-delimited rows ready for paste into wFirma PZ table.

  Mode 1B — PZ_READY.json
    GET  /api/v1/upload/shipment/{batch_id}/wfirma/json
    Returns structured JSON saved as outputs/{batch_id}/PZ_READY.json.

Guards (both modes)
-------------------
  - SAD/ZC429 must be present
  - PZ must be in status "success" or "partial"
  - PZ rows must exist (read from pz_rows.json or XLSX Rows sheet)

Rules
-----
  - Duty A00 is already allocated into cost price (line_netto_pln includes duty)
  - Do NOT add VAT — PZ is stock receipt cost basis
  - Polish numeric format: 1 234,56
  - Uwagi includes: Invoice, AWB, MRN, NBP rate, A00 note
"""

from __future__ import annotations

import json
import os
import time
import xml.etree.ElementTree as ET
from contextlib import contextmanager
from pathlib import Path
from typing import Any, Dict, List, Optional

from fastapi import APIRouter, Depends, Header, HTTPException
from fastapi.responses import FileResponse, JSONResponse
from pydantic import BaseModel

from ..core.config import settings
from ..core.logging import get_logger
from ..core.security import require_api_key
from ..core import timeline as tl
from ..services.batch_service import get_output_dir
from ..services.import_pz_builder import BatchRow, build_pz_request_from_batch
from ..services.wfirma_pz_notes import build_wfirma_pz_notes
from ..services import description_engine as deng
from ..services import wfirma_client
from ..services import wfirma_db
from ..utils.io import write_json_atomic

log    = get_logger(__name__)
router = APIRouter(prefix="/api/v1/upload", tags=["wfirma"])
_auth  = Depends(require_api_key)


def _build_product_code(invoice_no: str, position: int) -> str:
    """Delegate to the engine helper; guarantees ``invoice_no-N`` with no space."""
    try:
        from pz_import_processor import build_product_code  # noqa: PLC0415
        return build_product_code(invoice_no, position)
    except ImportError:
        return f"{invoice_no}-{position}"


# ── Timeline events ───────────────────────────────────────────────────────────
EV_WFIRMA_CLIPBOARD  = "wfirma_clipboard_generated"
EV_WFIRMA_JSON       = "wfirma_json_generated"
EV_WFIRMA_PZ_CREATED = "wfirma_pz_created"
EV_WFIRMA_PZ_ADOPTED = "wfirma_pz_adopted"

# PZ statuses that indicate a completed run
_PZ_DONE = {"success", "partial"}


# ── Helpers ───────────────────────────────────────────────────────────────────

def _read_audit(output_dir: Path) -> dict:
    p = output_dir / "audit.json"
    if not p.exists():
        raise HTTPException(status_code=404, detail="Batch not found.")
    try:
        return json.loads(p.read_text(encoding="utf-8"))
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"audit.json unreadable: {e}")


def _compute_effective_pz_status(audit: dict) -> tuple:
    """Return (effective_status, normalized_flag).

    Normalizes a stale ``audit.status`` back to ``'partial'`` when the
    shipment is in fact PZ-complete but the persisted status string
    lags behind operator decisions. The check below is conservative:

      ✓ ``failed_checks`` is empty
      ✓ ``customs_declaration.mrn`` is populated
      ✓ ``verification.cn_match`` is True
        OR ``cn_decision.approved`` is True
        (operator already accepted SAD CN — see /cn-decision/accept-sad)

    Hard blocks remain (returns the stored status unchanged) when:
      • ``failed_checks`` non-empty (real engine failures)
      • MRN missing
      • CN still unresolved

    The function NEVER mutates the audit; it only computes a value
    for the guard's decision and the pz_preview response payload.
    """
    stored = (audit.get("status") or "").strip()
    if stored in _PZ_DONE:
        return stored, False

    failed = audit.get("failed_checks") or []
    if failed:
        return stored, False

    cd  = audit.get("customs_declaration") or {}
    mrn = (cd.get("mrn") or "").strip()
    if not mrn:
        return stored, False

    ver    = audit.get("verification") or {}
    cn_dec = audit.get("cn_decision")  or {}
    cn_ok  = bool(ver.get("cn_match")) or bool(cn_dec.get("approved"))
    if not cn_ok:
        return stored, False

    # Operator-effective complete state. We use "partial" rather than
    # "success" because the engine did not assert success — this is a
    # post-override normalisation.
    return "partial", True


# ── PZ preview structured blockers ────────────────────────────────────────────
#
# Lesson G + PZ Preview Authority Audit (2026-05-21) — `pz_preview` historically
# raised HTTPException(422) on every prerequisite gap, which forced the frontend
# to invent its own optimistic "Ready for PZ" badge in parallel. That duplicate
# authority lets the badge contradict the engine outcome.
#
# The fix: a single read-only authority. `_collect_pz_preview_blockers` returns
# a list of `{code, message, severity, source}` rows describing every reason
# the batch is not ready. When the list is empty, the batch is ready. The
# legacy 422 path is preserved behind feature flag `PZ_PREVIEW_STRUCTURED_BLOCKERS`
# defaulting ON; OFF restores the prior HTTPException behaviour for one release.
#
# Mutating endpoints (`pz_create`, `pz_adopt`, clipboard, json export) keep
# `_guard_wfirma_export` unchanged — they must hard-fail on prerequisites.

# Stable blocker code vocabulary. Treat as part of the public API.
BLOCKER_WFIRMA_NO_SAD              = "WFIRMA_NO_SAD"
BLOCKER_WFIRMA_PZ_NOT_GENERATED    = "WFIRMA_PZ_NOT_GENERATED"
BLOCKER_WFIRMA_NO_ROWS             = "WFIRMA_NO_ROWS"
BLOCKER_ENGINE_ERROR               = "ENGINE_ERROR"
BLOCKER_WFIRMA_CREATE_DISABLED     = "WFIRMA_CREATE_DISABLED"
BLOCKER_SUPPLIER_NOT_CONFIGURED    = "SUPPLIER_NOT_CONFIGURED"
BLOCKER_WAREHOUSE_NOT_CONFIGURED   = "WAREHOUSE_NOT_CONFIGURED"


def _pz_preview_structured_enabled() -> bool:
    """Feature flag — defaults ON. Set PZ_PREVIEW_STRUCTURED_BLOCKERS=false
    in env to restore the legacy 422 raising path for one release."""
    raw = (os.getenv("PZ_PREVIEW_STRUCTURED_BLOCKERS") or "").strip().lower()
    if raw in ("0", "false", "no", "off"):
        return False
    return True


def _collect_pz_preview_blockers(audit: dict, output_dir: Path) -> List[Dict[str, Any]]:
    """Read-only authority for "is this batch ready for PZ".

    Returns a list of blocker rows. Empty list = the SAD + engine + rows
    prerequisites are all satisfied (independent of wFirma supplier/warehouse
    configuration — those are added by the caller using the existing lock_status
    derivation so we don't duplicate that logic here).

    NEVER raises. NEVER mutates audit or disk.
    """
    blockers: List[Dict[str, Any]] = []

    # 1. SAD/ZC429 present?
    inputs = audit.get("inputs") or {}
    if not inputs.get("zc429"):
        blockers.append({
            "code":     BLOCKER_WFIRMA_NO_SAD,
            "message":  "wFirma export requires SAD (ZC429). Upload SAD before exporting.",
            "severity": "error",
            "source":   "audit.inputs.zc429",
        })

    # 2. Engine outcome — engine_error wins over stored status. A persisted
    # engine failure is the strongest signal that PZ is not ready, and it
    # carries operator-actionable text from pz_import_processor.
    eng_err = (audit.get("engine_error") or "").strip()
    if eng_err:
        blockers.append({
            "code":     BLOCKER_ENGINE_ERROR,
            "message":  eng_err,
            "severity": "error",
            "source":   "audit.engine_error",
        })

    # 3. Effective status — uses the existing reconciler so a stale persisted
    # "failed" status doesn't keep the operator locked after they cleared the
    # underlying failed_checks.
    effective, normalized = _compute_effective_pz_status(audit)
    stored = audit.get("status", "")
    if effective not in _PZ_DONE and not eng_err:
        # Skip emitting a second status blocker when ENGINE_ERROR already
        # surfaced the real reason — they're the same fault expressed twice.
        blockers.append({
            "code":     BLOCKER_WFIRMA_PZ_NOT_GENERATED,
            "message":  (
                f"wFirma export requires a completed PZ. "
                f"Stored status: {stored!r}; effective status: {effective!r}."
            ),
            "severity": "error",
            "source":   "audit.status",
        })

    # 4. PZ rows file present?
    rows_present = (
        (output_dir / "pz_rows.json").exists()
        or any(output_dir.glob("*_calc.xlsx"))
    )
    if not rows_present and not eng_err:
        # When ENGINE_ERROR already surfaced, suppress the duplicate "no rows"
        # blocker — they share the same root cause.
        blockers.append({
            "code":     BLOCKER_WFIRMA_NO_ROWS,
            "message":  "PZ rows not found. pz_rows.json and XLSX Rows sheet both missing.",
            "severity": "error",
            "source":   "outputs.disk",
        })

    return blockers


def _guard_wfirma_export(audit: dict) -> None:
    """Block wFirma export if SAD is missing or PZ not yet generated.

    Uses ``_compute_effective_pz_status`` so a stale persisted
    ``audit.status="failed"`` doesn't keep the operator locked out
    after they cleared all real failed_checks (e.g. via
    /cn-decision/accept-sad). Hard blocks (real failed_checks, missing
    MRN) still raise as before.
    """
    inputs = audit.get("inputs") or {}
    if not inputs.get("zc429"):
        raise HTTPException(
            status_code=422,
            detail={
                "guard": "wfirma",
                "error": "wFirma export requires SAD (ZC429). Upload SAD before exporting.",
                "code": "WFIRMA_NO_SAD",
            },
        )
    effective, normalized = _compute_effective_pz_status(audit)
    stored = audit.get("status", "")
    if effective not in _PZ_DONE:
        raise HTTPException(
            status_code=422,
            detail={
                "guard": "wfirma",
                "error": (
                    f"wFirma export requires a completed PZ. "
                    f"Stored status: {stored!r}; effective status: "
                    f"{effective!r}."
                ),
                "code":              "WFIRMA_PZ_NOT_GENERATED",
                "stored_status":     stored,
                "effective_status":  effective,
                "status_normalized": normalized,
            },
        )


def _fmt_pln(v: float) -> str:
    """Polish numeric format: 1 234,56 (space thousands, comma decimal)."""
    try:
        f = float(v)
    except (TypeError, ValueError):
        return "0,00"
    # Format with comma thousands first, then swap separators
    formatted = f"{f:,.2f}"  # e.g. "1,234.56"
    return formatted.replace(",", "X").replace(".", ",").replace("X", chr(32))


def _load_rows_from_pz_rows_json(output_dir: Path) -> Optional[List[dict]]:
    """Try to load rows from pz_rows.json (written by export_service post-processing)."""
    p = output_dir / "pz_rows.json"
    if not p.exists():
        return None
    try:
        data = json.loads(p.read_text(encoding="utf-8"))
        rows = data if isinstance(data, list) else data.get("rows")
        if rows and isinstance(rows, list):
            return rows
    except Exception as e:
        log.warning("pz_rows.json parse error: %s", e)
    return None


def _load_rows_from_xlsx(output_dir: Path, audit: dict) -> Optional[List[dict]]:
    """
    Fallback: read the Rows sheet of the calc XLSX.
    Extracts columns by header name — robust to column reorder.
    """
    xlsx_name = (audit.get("files") or {}).get("xlsx", {}).get("name")
    if not xlsx_name:
        # Try glob for any *_calc.xlsx
        candidates = list(output_dir.glob("*_calc.xlsx"))
        if not candidates:
            return None
        xlsx_name = candidates[0].name

    xlsx_path = output_dir / xlsx_name
    if not xlsx_path.exists():
        return None

    try:
        import openpyxl
        wb = openpyxl.load_workbook(xlsx_path, data_only=True, read_only=True)
        if "Rows" not in wb.sheetnames:
            wb.close()
            return None

        ws = wb["Rows"]
        # Read header row (row 1) to build column-name → index map
        header_map: Dict[str, int] = {}
        rows_raw = list(ws.iter_rows(values_only=True))
        if not rows_raw:
            wb.close()
            return None

        for c_idx, cell_val in enumerate(rows_raw[0]):
            if cell_val:
                header_map[str(cell_val).strip()] = c_idx

        def _col(row_vals, name, default=None):
            idx = header_map.get(name)
            if idx is None:
                return default
            v = row_vals[idx] if idx < len(row_vals) else None
            return v if v is not None else default

        rows_out = []
        for row_vals in rows_raw[1:]:
            if all(v is None for v in row_vals):
                continue  # skip empty rows
            lp = _col(row_vals, "Lp")
            if lp is None:
                continue

            # Extract a clean unit from Item Type or default "szt."
            item_type = str(_col(row_vals, "Item Type", "")).strip()
            unit = _col(row_vals, "Unit", "szt.") or "szt."

            rows_out.append({
                "lp":               int(lp) if lp else 0,
                "invoice_no":       str(_col(row_vals, "Invoice No", "") or ""),
                "description_en":   str(_col(row_vals, "English Name", "") or ""),
                "pl_desc":          str(_col(row_vals, "Polish Name", "") or ""),
                "quantity":         float(_col(row_vals, "Qty", 1) or 1),
                "unit":             unit,
                "unit_netto_pln":   float(_col(row_vals, "Unit Netto PLN", 0) or 0),
                "line_netto_pln":   float(_col(row_vals, "Line Netto PLN", 0) or 0),
                "line_brutto_pln":  float(_col(row_vals, "Line Brutto PLN", 0) or 0),
                "allocated_duty_pln": float(_col(row_vals, "Alloc. Duty PLN", 0) or 0),
                "usd_pln":          float(_col(row_vals, "Rate PLN/USD", 0) or 0),
                "item_type":        item_type,
            })
        wb.close()
        return rows_out or None

    except Exception as e:
        log.warning("XLSX Rows read error: %s", e)
        return None


def _build_rows(output_dir: Path, audit: dict) -> List[dict]:
    """Load PZ rows from best available source, or raise 422."""
    rows = _load_rows_from_pz_rows_json(output_dir)
    if not rows:
        rows = _load_rows_from_xlsx(output_dir, audit)
    if not rows:
        raise HTTPException(
            status_code=422,
            detail={
                "guard": "wfirma",
                "error": "PZ rows not found. pz_rows.json and XLSX Rows sheet both missing.",
                "code": "WFIRMA_NO_ROWS",
            },
        )
    return rows


def _resolve_supplier(audit: dict) -> tuple[str, str, list[str]]:
    """
    Resolve the supplier name for wFirma export using a priority chain.

    Returns (supplier_name, source_label, risk_flags).
    Never returns blank — always falls back to UNKNOWN_SUPPLIER with a risk flag.

    Priority (first non-empty wins):
      1. customs_declaration.exporter_name  (canonical, from SAD)
      2. verification.invoice_exporter_name (parsed from invoice PDFs)
      3. exporter_check.invoice_exporter    (alternate invoice check)
      4. zc429.exporter_name / .exporter    (raw ZC429 fields)
      5. learning_traces[0].supplier_key    (last-known supplier from learning system)
      6. fallback                           "UNKNOWN_SUPPLIER" + risk flag
    """
    risks: list[str] = []
    cd       = audit.get("customs_declaration") or {}
    ver      = audit.get("verification") or {}
    expchk   = audit.get("exporter_check") or {}
    zc429    = audit.get("zc429") or {}
    traces   = audit.get("learning_traces") or []

    candidates = [
        ("customs_declaration.exporter_name", cd.get("exporter_name")),
        ("verification.invoice_exporter_name", ver.get("invoice_exporter_name")),
        ("exporter_check.invoice_exporter",   expchk.get("invoice_exporter")),
        ("zc429.exporter_name",               zc429.get("exporter_name")),
        ("zc429.exporter",                    zc429.get("exporter")),
    ]
    for source, value in candidates:
        if value and str(value).strip():
            return str(value).strip(), source, risks

    # Soft fallback: learning system supplier key (humanise estrella_jewels_llp → "Estrella Jewels LLP")
    if traces:
        key = (traces[0] or {}).get("supplier_key") or ""
        if key:
            humanised = " ".join(w.capitalize() if w not in ("llp","ltd","pvt") else w.upper()
                                 for w in key.replace("_"," ").split())
            risks.append("supplier_from_learning_only")
            return humanised, "learning_traces[0].supplier_key", risks

    risks.append("supplier_missing_for_wfirma")
    return "UNKNOWN_SUPPLIER", "fallback", risks


def _build_uwagi(row: dict, awb: str, mrn: str, nbp_rate: float, settlement_mode: str = "standard") -> str:
    """Build Uwagi cell content for one PZ row."""
    invoice_no = row.get("invoice_no", "")
    duty_pln   = row.get("allocated_duty_pln", 0.0)

    parts = []
    if invoice_no:
        parts.append(f"Invoice {invoice_no}")
    if awb:
        parts.append(f"AWB {awb}")
    if mrn:
        parts.append(f"MRN {mrn}")
    parts.append("A00 allocated in cost")
    if nbp_rate:
        parts.append(f"NBP {_fmt_pln(nbp_rate)}")
    if settlement_mode == "art33a":
        parts.append("Art.33a")

    return "; ".join(parts)


def _build_wfirma_rows(rows: List[dict], audit: dict) -> List[dict]:
    """
    Transform PZ engine rows into wFirma-ready row dicts.
    Returns list of dicts with keys matching wFirma PZ columns.
    """
    cd   = audit.get("customs_declaration") or {}
    awb  = audit.get("tracking_no", "") or ""
    mrn  = cd.get("mrn", "") or audit.get("inputs", {}).get("zc429_mrn", "") or ""
    nbp  = float(cd.get("nbp_rate", 0) or audit.get("inputs", {}).get("nbp_rate_usd", 0) or 0)
    mode = audit.get("settlement_mode", "standard")

    out = []
    for row in rows:
        # Nazwa towaru: prefer Polish description, fall back to English
        nazwa = (row.get("pl_desc") or row.get("description_en") or
                 row.get("item_type") or "Towar").strip()

        qty     = float(row.get("quantity", 1) or 1)
        unit    = str(row.get("unit", "szt.") or "szt.")
        # Normalise: wFirma uses "szt." for PCS/PIECE
        if unit.upper() in ("PCS", "PIECE", "PIECES", "PC", "ITEM", "ITEMS"):
            unit = "szt."
        elif unit.upper() in ("PAIR", "PAIRS", "PR"):
            unit = "para"
        elif unit.upper() == "SET":
            unit = "zest."

        unit_netto  = float(row.get("unit_netto_pln", 0) or 0)
        line_netto  = float(row.get("line_netto_pln", 0) or 0)
        line_brutto = float(row.get("line_brutto_pln", 0) or 0)

        out.append({
            "nazwa_towaru":   nazwa,
            "ilosc":          qty,
            "jm":             unit,
            "cena_netto":     unit_netto,
            "wartosc_netto":  line_netto,
            "wartosc_brutto": line_brutto,
            "uwagi":          _build_uwagi(row, awb, mrn, nbp, mode),
            # extended fields (not in clipboard, but in JSON)
            "_invoice_no":    row.get("invoice_no", ""),
            "_description_en": row.get("description_en", ""),
            "_pl_desc":       row.get("pl_desc", "") or row.get("description_en", ""),
            "_allocated_duty_pln": row.get("allocated_duty_pln", 0),
            "_product_code":  row.get("product_code", ""),
            "_item_type":     row.get("item_type", ""),
            "_unit_netto_pln": row.get("unit_netto_pln", 0),
        })
    return out


def _build_clipboard_text(wfirma_rows: List[dict]) -> str:
    """
    Build tab-separated clipboard string.
    Columns: Nazwa towaru | Ilość | J.m. | Cena netto | Wartość netto | Wartość brutto | Uwagi
    Numeric values use Polish format (space thousands, comma decimal).
    """
    lines = []
    header = "\t".join([
        "Nazwa towaru", "Ilość", "J.m.",
        "Cena netto", "Wartość netto", "Wartość brutto", "Uwagi",
    ])
    lines.append(header)

    for r in wfirma_rows:
        qty_str = (
            str(int(r["ilosc"])) if r["ilosc"] == int(r["ilosc"])
            else _fmt_pln(r["ilosc"]).replace(" ", "")       # remove thousands sep from qty
        )
        line = "\t".join([
            r["nazwa_towaru"],
            qty_str,
            r["jm"],
            _fmt_pln(r["cena_netto"]),
            _fmt_pln(r["wartosc_netto"]),
            _fmt_pln(r["wartosc_brutto"]),
            r["uwagi"],
        ])
        lines.append(line)

    return "\n".join(lines)


def _patch_pz_doc_id(output_dir: Path, wfirma_pz_doc_id: str) -> Optional[str]:
    """
    Write wfirma_pz_doc_id into audit.json wfirma_export block. Atomic.

    Returns:
        None on success.
        A short error string when the audit could not be written so the caller
        can include it in the response (NEVER silently continues — wFirma has
        already created the document at this point).
    """
    audit_path = output_dir / "audit.json"
    if not audit_path.exists():
        return f"audit.json missing at {audit_path}"
    try:
        audit = json.loads(audit_path.read_text(encoding="utf-8"))
        existing = audit.get("wfirma_export") or {}
        audit["wfirma_export"] = {
            **existing,
            "wfirma_pz_doc_id": wfirma_pz_doc_id,
            "pz_source":        existing.get("pz_source") or "created_via_app",
            "pz_created_at":    time.strftime("%Y-%m-%dT%H:%M:%S"),
        }
        write_json_atomic(audit_path, audit)
        return None
    except Exception as e:
        log.error("[pz_create] AUDIT PATCH FAILED for doc_id=%s — %s", wfirma_pz_doc_id, e)
        return str(e)


def _patch_pz_adopted(output_dir: Path, wfirma_pz_doc_id: str) -> Optional[str]:
    """
    Write wfirma_pz_doc_id + pz_source='adopted_existing' into audit.json.
    Atomic. Does not overwrite other wfirma_export fields.

    Returns:
        None on success, error string on failure (see _patch_pz_doc_id).
    """
    audit_path = output_dir / "audit.json"
    if not audit_path.exists():
        return f"audit.json missing at {audit_path}"
    try:
        audit = json.loads(audit_path.read_text(encoding="utf-8"))
        existing = audit.get("wfirma_export") or {}
        audit["wfirma_export"] = {
            **existing,
            "wfirma_pz_doc_id": wfirma_pz_doc_id,
            "pz_source":        "adopted_existing",
            "pz_adopted_at":    time.strftime("%Y-%m-%dT%H:%M:%S"),
        }
        write_json_atomic(audit_path, audit)
        return None
    except Exception as e:
        log.error("[pz_adopt] AUDIT PATCH FAILED for doc_id=%s — %s", wfirma_pz_doc_id, e)
        return str(e)


# ── PZ idempotency guard + lock ───────────────────────────────────────────────

def _has_pz_terminal_event(audit: dict) -> Optional[str]:
    """
    Return the name of any PZ terminal timeline event already recorded
    (EV_WFIRMA_PZ_CREATED or EV_WFIRMA_PZ_ADOPTED), or None.

    The timeline check runs even when wfirma_export.wfirma_pz_doc_id has been
    manually removed — the audit timeline is append-only and authoritative.
    """
    timeline = audit.get("timeline") or []
    for ev in timeline:
        name = (ev or {}).get("event") or ""
        if name in (EV_WFIRMA_PZ_CREATED, EV_WFIRMA_PZ_ADOPTED):
            return name
    return None


def _assert_pz_not_locked(audit: dict, batch_id: str, action: str) -> None:
    """
    Reject a PZ create/adopt attempt with HTTP 409 when the shipment already
    has *any* PZ provenance recorded.  Three sources are checked:

        1. wfirma_export.wfirma_pz_doc_id  (current state)
        2. wfirma_export.pz_source          (current state — distinguishes
                                             created_via_app vs adopted_existing)
        3. Audit timeline EV_WFIRMA_PZ_CREATED / EV_WFIRMA_PZ_ADOPTED
                                            (history; survives field cleanup)

    The timeline check is intentional: even if an operator removes the doc_id
    field manually, the immutable timeline still proves a PZ was created/adopted
    and a second attempt must fail.

    Args:
        audit:    parsed audit.json dict (re-read inside the file lock)
        batch_id: for logging
        action:   "pz_create" | "pz_adopt"  — controls the error message

    Raises:
        HTTPException(409) with structured detail describing which signal
        triggered the lock.
    """
    wfirma_export      = audit.get("wfirma_export") or {}
    existing_pz_doc_id = (wfirma_export.get("wfirma_pz_doc_id") or "").strip()
    existing_pz_source = (wfirma_export.get("pz_source") or "").strip()
    terminal_event     = _has_pz_terminal_event(audit)

    if not existing_pz_doc_id and not existing_pz_source and not terminal_event:
        return

    # PZ_RECONCILED fast-exit: the operator explicitly cleared the mapping via
    # /wfirma/pz/clear-mapping after a previous create/adopt.  The clearing
    # event is the most recent PZ-mapping event in the timeline, and the export
    # fields are already empty.  Allow recreation — this is the intended workflow
    # after a wFirma PZ has been manually deleted and the local mapping cleared.
    # We only bypass when BOTH export fields are empty (no live doc_id) so a
    # real duplicate (doc_id still present) is never bypassed.
    if (not existing_pz_doc_id and not existing_pz_source
            and _has_pz_mapping_cleared_after_create(audit)):
        return

    # Build a precise reason string + machine-readable code
    if terminal_event == EV_WFIRMA_PZ_CREATED:
        reason = "PZ has already been created for this shipment"
        code   = "PZ_ALREADY_CREATED"
    elif terminal_event == EV_WFIRMA_PZ_ADOPTED:
        reason = "PZ has already been adopted for this shipment"
        code   = "PZ_ALREADY_ADOPTED"
    elif existing_pz_source == "adopted_existing":
        reason = "PZ has already been adopted for this shipment"
        code   = "PZ_ALREADY_ADOPTED"
    elif existing_pz_source == "created_via_app":
        reason = "PZ has already been created for this shipment"
        code   = "PZ_ALREADY_CREATED"
    else:
        reason = f"PZ already linked to this shipment (doc_id={existing_pz_doc_id!r})"
        code   = "PZ_ALREADY_LINKED"

    log.warning(
        "[%s] %s blocked: %s (existing_doc_id=%r, source=%r, timeline_event=%r)",
        batch_id, action, reason, existing_pz_doc_id, existing_pz_source, terminal_event,
    )
    raise HTTPException(
        status_code=409,
        detail={
            "guard":            action,
            "error":            reason,
            "code":             code,
            "existing_doc_id":  existing_pz_doc_id or None,
            "existing_source":  existing_pz_source or None,
            "timeline_event":   terminal_event,
        },
    )


@contextmanager
def _pz_write_lock(output_dir: Path, batch_id: str, action: str):
    """
    Process-wide mutex around the PZ create/adopt critical section.

    Uses an O_EXCL lockfile inside the batch output dir so two concurrent
    operator clicks can never both pass the idempotency check before either
    writes audit.  The lock is held only across:

        re-read audit  →  guard check  →  wFirma call  →  audit patch

    Stale lock recovery: the lock file carries the holder's PID and a
    timestamp.  If acquisition finds a lock older than _PZ_LOCK_STALE_SECS
    the lock is force-removed (e.g. process crashed mid-write).

    Yields nothing.  Raises HTTPException(409) if another writer already
    holds the lock.
    """
    # If output_dir isn't a real filesystem Path (e.g. tests passing MagicMock),
    # silently skip locking — production callers always pass a real Path from
    # get_output_dir().  This degrades to a no-op without compromising the
    # idempotency guard, which still runs against the audit dict.
    try:
        output_dir.mkdir(parents=True, exist_ok=True)
    except (TypeError, AttributeError, OSError):
        yield
        return

    lock_path = output_dir / ".pz_write.lock"

    # Stale-lock takeover — broaden except since output_dir may be unusual in tests
    try:
        if lock_path.exists():
            age = time.time() - lock_path.stat().st_mtime
            if age > _PZ_LOCK_STALE_SECS:
                log.warning("[%s] %s: removing stale PZ lock (age=%.1fs)",
                            batch_id, action, age)
                lock_path.unlink(missing_ok=True)
    except Exception:
        # Path doesn't behave like a real filesystem path — skip locking
        yield
        return

    try:
        fd = os.open(str(lock_path), os.O_CREAT | os.O_EXCL | os.O_WRONLY, 0o600)
    except FileExistsError:
        log.warning("[%s] %s: PZ write lock held by another request", batch_id, action)
        raise HTTPException(
            status_code=409,
            detail={
                "guard": action,
                "error": "Another PZ create/adopt request is in progress for this shipment.",
                "code":  "PZ_WRITE_LOCKED",
            },
        )
    except (FileNotFoundError, PermissionError, OSError):
        # Filesystem unusable in this context — degrade to no-lock
        yield
        return
    try:
        os.write(fd, f"{os.getpid()}@{time.time():.0f}\n".encode("utf-8"))
        os.close(fd)
        yield
    finally:
        try:
            lock_path.unlink(missing_ok=True)
        except OSError:
            pass


_PZ_LOCK_STALE_SECS = 120  # locks older than 2 min are force-released


# ── Read-only PZ lock-status snapshot (for dashboard) ────────────────────────

# ── PZ lifecycle state — single authority for the UI ─────────────────────────
#
# Background
# ----------
# The shipment-detail PZ panel historically composed its render from four
# separate authority sources (pz_preview.ready, pz_lock_status,
# capabilities.create_pz_allowed, timeline events). The four did not have a
# documented precedence, so the UI could simultaneously claim:
#   - "PZ audit write recovery required"
#   - "PZ creation disabled"
#   - ready=true / would_create_pz=true
# The operator had to mentally reconcile contradictions the backend already
# knew how to resolve.
#
# `_compute_pz_lifecycle_state` collapses every signal into ONE canonical
# state value. Every UI surface that renders PZ status MUST consume only the
# `pz_lifecycle.state` field — never combine multiple authorities.
#
# Precedence (first match wins; this is the load-bearing contract):
#   1. PZ_DUPLICATE_DETECTED — another batch owns the same wfirma_pz_doc_id
#   2. PZ_RECOVERY_REQUIRED  — timeline says created, audit doc_id empty
#                              (overrides ANY "creation disabled" messaging)
#   3. PZ_LOCKED             — audit.wfirma_locked or accounting period closed
#                              (reserved; not emitted today)
#   4. PZ_CREATED            — doc_id present + pz_source set
#   5. PZ_RECONCILED         — wfirma_pz_mapping_cleared after wfirma_pz_created
#                              with no later wfirma_pz_created; doc_id empty
#   6. PZ_READY_TO_CREATE    — preview ready + supplier + warehouse + flag on
#   7. PZ_NOT_READY          — default; blockers / flag off / engine_error
#
# PZ_CREATING is reserved for a future async-create path; the current
# synchronous create endpoint never emits it (no audit state to derive it
# from). It stays in the enum so consumers can already pattern-match against
# the full state set.

PZ_LIFECYCLE_STATES: tuple = (
    "PZ_DUPLICATE_DETECTED",
    "PZ_RECOVERY_REQUIRED",
    "PZ_LOCKED",
    "PZ_CREATED",
    "PZ_CREATING",          # reserved — see comment above
    "PZ_RECONCILED",
    "PZ_READY_TO_CREATE",
    "PZ_NOT_READY",
)

_PZ_LIFECYCLE_PRIMARY_ACTIONS = {
    "PZ_DUPLICATE_DETECTED": "resolve_cross_batch_conflict",
    "PZ_RECOVERY_REQUIRED":  "confirm_existing_pz",
    "PZ_LOCKED":             "none",
    "PZ_CREATED":            "none",
    "PZ_CREATING":           "wait_for_completion",
    "PZ_RECONCILED":         "recreate_when_ready",
    "PZ_READY_TO_CREATE":    "create_pz",
    "PZ_NOT_READY":          "resolve_blockers",
}


def _has_pz_mapping_cleared_after_create(audit: Dict[str, Any]) -> bool:
    """Return True when the timeline contains a `wfirma_pz_mapping_cleared`
    event that is the most recent PZ-mapping-related event (i.e. no later
    `wfirma_pz_created` or `wfirma_pz_adopted` after it).

    Used to detect the post-clear / pre-recreate intermediate state.
    """
    timeline = audit.get("timeline") or []
    last_mapping_event = ""
    for ev in timeline:
        name = (ev or {}).get("event") or ""
        if name in (
            EV_WFIRMA_PZ_CREATED,
            EV_WFIRMA_PZ_ADOPTED,
            "wfirma_pz_mapping_cleared",
        ):
            last_mapping_event = name
    return last_mapping_event == "wfirma_pz_mapping_cleared"


def _compute_pz_lifecycle_state(
    audit: Dict[str, Any],
    *,
    preview_ready: bool         = False,
    supplier_configured: bool   = False,
    warehouse_configured: bool  = False,
    create_allowed: bool        = False,
    duplicate_owner_batch_id: Optional[str] = None,
    blocker_codes:  Optional[List[str]]      = None,
) -> Dict[str, Any]:
    """Single authority for the UI's PZ status panel.

    Returns a 10-field envelope:
        state, reason, primary_action, secondary_actions,
        hide_create_button, hide_resolve_products,
        override_create_disabled_message,
        wfirma_pz_doc_id, terminal_event, blocker_codes

    Every UI surface MUST read `pz_lifecycle.state` (and optionally the
    action flags) as the sole source for the PZ panel. The legacy
    `pz_lock_status` field remains for backward compat with already-
    deployed dashboards; new code should target `pz_lifecycle`.

    `duplicate_owner_batch_id`: when provided + non-empty AND non-self,
    the state is `PZ_DUPLICATE_DETECTED`. The caller (preview/create)
    invokes `_find_pz_owner_batch` only when a doc_id is set; passing the
    result here keeps this helper a pure function on the audit dict.

    `blocker_codes`: stable list of codes from `pz_preview.blockers[].code`
    (e.g. ENGINE_ERROR). Echoed in the envelope so a UI consuming only
    `pz_lifecycle` still has access.
    """
    wfirma_export = audit.get("wfirma_export") or {}
    doc_id        = (wfirma_export.get("wfirma_pz_doc_id") or "").strip()
    pz_source     = (wfirma_export.get("pz_source") or "").strip()
    terminal_ev   = _has_pz_terminal_event(audit)
    codes_in      = list(blocker_codes or [])

    # Default envelope template — the matched branch overrides fields.
    out: Dict[str, Any] = {
        "state":                            "PZ_NOT_READY",
        "reason":                           "",
        "primary_action":                   "resolve_blockers",
        "secondary_actions":                [],
        "hide_create_button":               False,
        "hide_resolve_products":            False,
        "override_create_disabled_message": False,
        "wfirma_pz_doc_id":                 doc_id or None,
        "terminal_event":                   terminal_ev,
        "blocker_codes":                    codes_in,
    }

    # ── 1. Cross-batch duplicate ─────────────────────────────────────────
    if duplicate_owner_batch_id:
        out.update({
            "state":                            "PZ_DUPLICATE_DETECTED",
            "reason":                           "wfirma_pz_doc_id_claimed_by_other_batch",
            "primary_action":                   "resolve_cross_batch_conflict",
            "hide_create_button":               True,
            "hide_resolve_products":            True,
            "override_create_disabled_message": True,
            "duplicate_owner_batch_id":         duplicate_owner_batch_id,
        })
        return out

    # ── 2a. Reconciled (operator explicitly cleared the mapping post-create) ─
    # MUST come before PZ_RECOVERY_REQUIRED — when the latest tracked
    # PZ-mapping event is `wfirma_pz_mapping_cleared`, the missing doc_id
    # is intentional, not a half-state. The operator wants to recreate.
    if not doc_id and _has_pz_mapping_cleared_after_create(audit):
        out.update({
            "state":             "PZ_RECONCILED",
            "reason":            "pz_mapping_cleared_awaiting_recreate",
            "primary_action":    "recreate_when_ready",
            "hide_create_button":    not (preview_ready and supplier_configured
                                          and warehouse_configured and create_allowed),
            "hide_resolve_products": False,
            "terminal_event":        terminal_ev,
        })
        return out

    # ── 2b. Recovery required (audit write failed after wFirma success) ─
    # Distinct from PZ_RECONCILED: the operator did NOT clear the mapping,
    # but the audit still lost its doc_id. The operator's job is to
    # Confirm Existing PZ, not to flip a flag. Overrides any "creation
    # disabled" message.
    if terminal_ev and not doc_id:
        out.update({
            "state":                            "PZ_RECOVERY_REQUIRED",
            "reason":                           "audit_write_recovery_required",
            "primary_action":                   "confirm_existing_pz",
            "secondary_actions":                ["refresh_mapping"],
            "hide_create_button":               True,
            "hide_resolve_products":            True,
            "override_create_disabled_message": True,
        })
        return out

    # ── 3. Locked (accounting period close, operator hold, etc.) ────────
    # Reserved — no emitter today. Surfaces immediately when audit flag
    # is set so the UI doesn't have to guess.
    if bool(audit.get("wfirma_locked") or audit.get("accounting_period_closed")):
        out.update({
            "state":                            "PZ_LOCKED",
            "reason":                           "wfirma_locked",
            "primary_action":                   "none",
            "hide_create_button":               True,
            "hide_resolve_products":            True,
            "override_create_disabled_message": True,
        })
        return out

    # ── 4. Already created (canonical happy-path terminal state) ────────
    if doc_id and pz_source:
        out.update({
            "state":          "PZ_CREATED",
            "reason":         "pz_created" if pz_source == "created_via_app"
                              else "pz_adopted_existing" if pz_source == "adopted_existing"
                              else f"pz_source_{pz_source}",
            "primary_action": "none",
            "secondary_actions": ["refresh_mapping"],
            "hide_create_button":    True,
            "hide_resolve_products": False,
        })
        return out

    # ── 5. Ready to create ──────────────────────────────────────────────
    # (PZ_RECONCILED branch lives at priority 2a — operators clearing a
    # mapping is a distinct intent from a half-state audit-write recovery.)
    if (preview_ready and supplier_configured and warehouse_configured
            and create_allowed):
        out.update({
            "state":             "PZ_READY_TO_CREATE",
            "reason":            "ready",
            "primary_action":    "create_pz",
            "hide_create_button":    False,
            "hide_resolve_products": False,
        })
        return out

    # ── 7. Not ready (default) ──────────────────────────────────────────
    if not create_allowed:
        reason = "create_disabled"
    elif not preview_ready:
        reason = "preview_not_ready"
    elif not supplier_configured:
        reason = "supplier_not_configured"
    elif not warehouse_configured:
        reason = "warehouse_not_configured"
    else:
        reason = "not_ready"
    out["reason"] = reason
    out["hide_create_button"] = True
    return out


def _compute_pz_lock_status(
    audit: Dict[str, Any],
    *,
    preview_ready: bool = False,
    supplier_configured: bool = False,
    warehouse_configured: bool = False,
) -> Dict[str, Any]:
    """
    Build a read-only summary of why PZ create/adopt is or isn't allowed.

    This is the dashboard-facing companion of ``_assert_pz_not_locked``.  It
    inspects the same three signals (doc_id field, pz_source field, timeline
    terminal events) and returns a structured envelope so the UI can show a
    precise banner instead of a generic "ready" / "not ready" boolean.

    Audit-write recovery
    --------------------
    A rare half-state: timeline records EV_WFIRMA_PZ_CREATED but
    wfirma_export.wfirma_pz_doc_id is empty (audit-patch failed after the
    wFirma call succeeded — see _patch_pz_doc_id error path).  In that case
    we still report ``locked=True`` but flag ``recovery_required=True`` so
    the dashboard can prompt the operator to use Confirm Existing PZ.

    Args:
        audit:                parsed audit.json dict (same one already loaded
                              by the caller — never re-reads disk).
        preview_ready:        result.ready from build_pz_request_from_batch.
        supplier_configured:  bool — supplier contractor id is set.
        warehouse_configured: bool — warehouse id is set.

    Returns:
        dict — see schema in the docstring of wfirma_pz_preview.
    """
    wfirma_export = audit.get("wfirma_export") or {}
    doc_id        = (wfirma_export.get("wfirma_pz_doc_id") or "").strip()
    raw_source    = (wfirma_export.get("pz_source") or "").strip()
    terminal_ev   = _has_pz_terminal_event(audit)

    # Normalize pz_source for the dashboard:
    #   "created_via_app" → "created_by_system"  (UX-friendly label)
    #   "adopted_existing" → "adopted_existing"
    pz_source: Optional[str]
    if raw_source == "created_via_app":
        pz_source = "created_by_system"
    elif raw_source == "adopted_existing":
        pz_source = "adopted_existing"
    elif raw_source:
        pz_source = raw_source
    else:
        pz_source = None

    locked              = bool(doc_id or pz_source or terminal_ev)
    recovery_required   = bool(terminal_ev) and not doc_id
    reason: str
    code: str

    if not locked:
        reason = "no_pz_linked"
        code   = "NO_PZ_LINKED"
    elif recovery_required and terminal_ev == EV_WFIRMA_PZ_CREATED:
        reason = "audit_write_recovery_required"
        code   = "PZ_AUDIT_RECOVERY_NEEDED"
    elif recovery_required and terminal_ev == EV_WFIRMA_PZ_ADOPTED:
        reason = "audit_write_recovery_required"
        code   = "PZ_AUDIT_RECOVERY_NEEDED"
    elif pz_source == "created_by_system" or terminal_ev == EV_WFIRMA_PZ_CREATED:
        reason = "pz_created_by_system"
        code   = "PZ_ALREADY_CREATED"
    elif pz_source == "adopted_existing" or terminal_ev == EV_WFIRMA_PZ_ADOPTED:
        reason = "pz_adopted_existing"
        code   = "PZ_ALREADY_ADOPTED"
    else:
        reason = "pz_doc_id_set"
        code   = "PZ_ALREADY_LINKED"

    # Action gates
    # - can_create: false whenever locked, OR preview not ready, OR config missing
    # - can_adopt:  false when locked (idempotent same-id adopt is handled by
    #               the route itself, not exposed as a separate button); true
    #               in the recovery_required case so the operator can adopt
    #               the live wFirma doc_id manually.
    can_create = (not locked) and preview_ready and supplier_configured and warehouse_configured
    if recovery_required:
        can_adopt = True   # let operator link the live doc_id back into audit
    else:
        can_adopt = not locked

    return {
        "locked":             locked,
        "reason":             reason,
        "code":               code,
        "wfirma_pz_doc_id":   doc_id or None,
        "pz_source":          pz_source,
        "terminal_event":     terminal_ev,
        "recovery_required":  recovery_required,
        "can_create":         can_create,
        "can_adopt":          can_adopt,
    }


def _find_pz_owner_batch(
    wfirma_pz_doc_id: str,
    exclude_batch_id: str,
) -> Optional[str]:
    """
    Scan all outputs/*/audit.json files and return the first batch_id whose
    audit already holds the given wfirma_pz_doc_id, excluding the caller's own
    batch.  Returns None if no other shipment owns it.

    Used as the cross-shipment duplicate guard in pz_adopt.
    """
    outputs_dir = settings.storage_root / "outputs"
    if not outputs_dir.is_dir():
        return None
    target = wfirma_pz_doc_id.strip()
    for audit_path in outputs_dir.glob("*/audit.json"):
        batch = audit_path.parent.name
        if batch == exclude_batch_id:
            continue
        try:
            a       = json.loads(audit_path.read_text(encoding="utf-8"))
            existing = ((a.get("wfirma_export") or {}).get("wfirma_pz_doc_id") or "").strip()
            if existing == target:
                return batch
        except Exception:
            continue
    return None


def _patch_audit_wfirma(output_dir: Path, mode: str, row_count: int) -> None:
    """Write wfirma_export block into audit.json."""
    audit_path = output_dir / "audit.json"
    if not audit_path.exists():
        return
    try:
        audit = json.loads(audit_path.read_text(encoding="utf-8"))
        existing = audit.get("wfirma_export") or {}
        audit["wfirma_export"] = {
            "clipboard_generated": existing.get("clipboard_generated") or (mode == "clipboard"),
            "json_generated":      existing.get("json_generated")      or (mode == "json"),
            "last_generated_at":   time.strftime("%Y-%m-%dT%H:%M:%S"),
            "row_count":           row_count,
            "mode":                mode,
        }
        if mode == "clipboard":
            audit["wfirma_export"]["clipboard_generated"] = True
        elif mode == "json":
            audit["wfirma_export"]["json_generated"] = True
        write_json_atomic(audit_path, audit)
    except Exception as e:
        log.warning("wfirma_export audit patch failed: %s", e)


# ── Routes ────────────────────────────────────────────────────────────────────

@router.post("/shipment/{batch_id}/wfirma/clipboard", dependencies=[_auth])
async def wfirma_clipboard(batch_id: str) -> JSONResponse:
    """
    Mode 1 — Clipboard Export.

    Returns tab-separated PZ rows ready for copy/paste into wFirma PZ table.
    Guard: SAD must exist + PZ must be generated.
    """
    output_dir = get_output_dir(batch_id)
    audit      = _read_audit(output_dir)
    _guard_wfirma_export(audit)

    rows        = _build_rows(output_dir, audit)
    wfirma_rows = _build_wfirma_rows(rows, audit)
    clipboard   = _build_clipboard_text(wfirma_rows)

    # Audit
    _patch_audit_wfirma(output_dir, "clipboard", len(wfirma_rows))
    tl.log_event(
        output_dir / "audit.json",
        EV_WFIRMA_CLIPBOARD,
        "dashboard",
        "user",
        detail={"batch_id": batch_id, "row_count": len(wfirma_rows)},
    )

    # Resolve supplier + doc_no for warnings (preview should also surface these)
    supplier, supplier_source, risk_flags = _resolve_supplier(audit)
    doc_no_curr = (audit.get("doc_no") or "").strip()
    requires_doc_no = not doc_no_curr
    warnings: list[str] = []
    if requires_doc_no:
        warnings.append("PZ document number not set — confirm via dashboard before final wFirma save")
    if supplier == "UNKNOWN_SUPPLIER":
        warnings.append("Supplier name could not be resolved from invoice or SAD — fill manually in wFirma")
    elif "supplier_from_learning_only" in risk_flags:
        warnings.append(f"Supplier '{supplier}' inferred from learning history only — verify against invoice")

    log.info("[%s] wFirma clipboard generated: %d rows", batch_id, len(wfirma_rows))

    return JSONResponse({
        "batch_id":        batch_id,
        "row_count":       len(wfirma_rows),
        "mode":            "clipboard",
        "supplier":        supplier,
        "supplier_source": supplier_source,
        "doc_no":          doc_no_curr,
        "requires_doc_no": requires_doc_no,
        "risk_flags":      risk_flags,
        "warnings":        warnings,
        "clipboard":       clipboard,
        "rows":            [
            {
                "nazwa_towaru":   r["nazwa_towaru"],
                "ilosc":          r["ilosc"],
                "jm":             r["jm"],
                "cena_netto":     round(r["cena_netto"], 2),
                "wartosc_netto":  round(r["wartosc_netto"], 2),
                "wartosc_brutto": round(r["wartosc_brutto"], 2),
                "uwagi":          r["uwagi"],
            }
            for r in wfirma_rows
        ],
    })


@router.get("/shipment/{batch_id}/wfirma/json", dependencies=[_auth])
async def wfirma_json(batch_id: str) -> FileResponse:
    """
    Mode 1B — PZ_READY.json.

    Generates and returns PZ_READY.json, saved to outputs/{batch_id}/PZ_READY.json.
    Guard: SAD must exist + PZ must be generated.
    """
    output_dir = get_output_dir(batch_id)
    audit      = _read_audit(output_dir)
    _guard_wfirma_export(audit)

    rows        = _build_rows(output_dir, audit)
    wfirma_rows = _build_wfirma_rows(rows, audit)

    cd       = audit.get("customs_declaration") or {}
    totals   = audit.get("totals") or {}
    awb      = audit.get("tracking_no", "") or ""
    mrn      = cd.get("mrn", "") or audit.get("inputs", {}).get("zc429_mrn", "") or ""
    doc_no   = (audit.get("doc_no") or "").strip()
    importer = cd.get("importer_name", "") or "ESTRELLA JEWELS SP. Z O.O. SP. K."
    doc_date = cd.get("clearance_date", "") or audit.get("timestamp", "")[:10]

    # ── Supplier resolution (priority chain + fallback) ──────────────────────
    supplier, supplier_source, risk_flags = _resolve_supplier(audit)

    # ── Warnings (do not block preview/copy/JSON) ────────────────────────────
    warnings: list[str] = []
    requires_doc_no = not doc_no
    if requires_doc_no:
        warnings.append("PZ document number not set — confirm via dashboard before final wFirma save")
    if supplier == "UNKNOWN_SUPPLIER":
        warnings.append("Supplier name could not be resolved from invoice or SAD — fill manually in wFirma")
    elif "supplier_from_learning_only" in risk_flags:
        warnings.append(f"Supplier '{supplier}' inferred from learning history only — verify against invoice")

    payload: Dict[str, Any] = {
        "batch_id":      batch_id,
        "awb":           awb,
        "mrn":           mrn,
        "doc_no":        doc_no,
        "supplier":      supplier,
        "supplier_source": supplier_source,
        "importer":      importer,
        "document_date": doc_date,
        "currency":      "PLN",
        "source":        "SAD/PZ engine",
        "requires_doc_no": requires_doc_no,
        "risk_flags":    risk_flags,
        "warnings":      warnings,
        "rows": [
            {
                "name":             r["nazwa_towaru"],
                "quantity":         r["ilosc"],
                "unit":             r["jm"],
                "net_price_pln":    round(r["cena_netto"], 2),
                "net_value_pln":    round(r["wartosc_netto"], 2),
                "gross_value_pln":  round(r["wartosc_brutto"], 2),
                "invoice_no":       r["_invoice_no"],
                "original_description": r["_description_en"],
                "polish_description":   r["_pl_desc"],
                "notes":            r["uwagi"],
                "product_code":     r["_product_code"],
            }
            for r in wfirma_rows
        ],
        "totals": {
            "net":      round(float(totals.get("net") or 0), 2),
            "gross":    round(float(totals.get("gross") or 0), 2),
            "duty_a00": round(float(cd.get("duty_a00_pln") or 0), 2),
        },
    }

    # Save to disk
    out_path = output_dir / "PZ_READY.json"
    write_json_atomic(out_path, payload)

    # Audit
    _patch_audit_wfirma(output_dir, "json", len(wfirma_rows))
    tl.log_event(
        output_dir / "audit.json",
        EV_WFIRMA_JSON,
        "dashboard",
        "user",
        detail={"batch_id": batch_id, "row_count": len(wfirma_rows), "path": str(out_path)},
    )

    log.info("[%s] PZ_READY.json → %s (%d rows)", batch_id, out_path, len(wfirma_rows))

    return FileResponse(
        path=str(out_path),
        media_type="application/json",
        filename=f"PZ_READY_{batch_id}.json",
    )


@router.get("/shipment/{batch_id}/wfirma/pz_preview", dependencies=[_auth])
async def wfirma_pz_preview(batch_id: str) -> JSONResponse:
    """
    Preview — wFirma PZ creation from internal PZ app output.

    Returns what a wFirma PZ would look like if created directly from the PZ
    engine calculation, instead of from a sales proforma.  Read-only: never
    calls create_warehouse_pz.

    Guard: SAD must exist + PZ must be generated (same as clipboard/json).

    Response fields
    ---------------
    already_created        true if audit.wfirma_export.wfirma_pz_doc_id is set
    wfirma_pz_doc_id       existing wFirma PZ doc id (if already_created)
    would_create_pz        true only when ready=true and not already_created
    ready                  true when all product_codes are mapped and no price conflicts
    unresolved_product_codes  product_codes missing from wfirma_products table
    price_conflicts        product_codes with inconsistent unit_netto_pln
    supplier_wfirma_id     import supplier contractor id (from settings or resolved)
    warehouse_id           wFirma warehouse id
    mrn                    customs MRN (dedup key)
    clearance_date         SAD clearance date (PZ document date)
    planned_lines          per-line preview: product_code, good_id, count, price_pln
    description            PZ description that would be written (batch_id + MRN)
    risk_flags             supplier resolution warnings
    """
    output_dir = get_output_dir(batch_id)
    audit      = _read_audit(output_dir)

    # ── Structured blockers (Lesson G — single read-only authority) ──────────
    # When the feature flag is ON we replace the legacy 422-raising guard with
    # a 200 response carrying a `blockers` list. The mutating create/adopt
    # endpoints below keep their 422 guards.
    #
    # The already_created short-circuit (below) takes precedence — a PZ that
    # has already been written to wFirma is not "blocked"; it is done. We
    # only collect prerequisite blockers when no wFirma doc id is linked.
    _existing_pz_doc_id = ((audit.get("wfirma_export") or {}).get("wfirma_pz_doc_id") or "").strip()
    if _pz_preview_structured_enabled() and not _existing_pz_doc_id:
        blockers = _collect_pz_preview_blockers(audit, output_dir)
        if blockers:
            # Add wFirma-config blockers so the UI sees a single complete list.
            supplier_configured  = bool((settings.wfirma_supplier_contractor_id or "").strip())
            warehouse_configured = bool((settings.wfirma_warehouse_id or "").strip())
            if not supplier_configured:
                blockers.append({
                    "code":     BLOCKER_SUPPLIER_NOT_CONFIGURED,
                    "message":  "WFIRMA_SUPPLIER_CONTRACTOR_ID not configured — set in .env",
                    "severity": "warning",
                    "source":   "settings",
                })
            if not warehouse_configured:
                blockers.append({
                    "code":     BLOCKER_WAREHOUSE_NOT_CONFIGURED,
                    "message":  "WFIRMA_WAREHOUSE_ID not configured — set in .env",
                    "severity": "warning",
                    "source":   "settings",
                })
            eff_status, status_normalized = _compute_effective_pz_status(audit)
            pz_lifecycle = _compute_pz_lifecycle_state(
                audit,
                preview_ready=False,
                supplier_configured=supplier_configured,
                warehouse_configured=warehouse_configured,
                create_allowed=bool(getattr(settings, "wfirma_create_pz_allowed", False)),
                duplicate_owner_batch_id=None,
                blocker_codes=[b.get("code") for b in blockers if b.get("code")],
            )
            return JSONResponse({
                "batch_id":             batch_id,
                "already_created":      False,
                "wfirma_pz_doc_id":     None,
                "would_create_pz":      False,
                "ready":                False,
                "blockers":             blockers,
                "engine_error":         (audit.get("engine_error") or "") or None,
                "pz_lifecycle":         pz_lifecycle,
                "unresolved_product_codes": [],
                "price_conflicts":      [],
                "stored_status":        audit.get("status", ""),
                "effective_status":     eff_status,
                "status_normalized":    status_normalized,
            })
    elif not _pz_preview_structured_enabled():
        _guard_wfirma_export(audit)

    # ── Duplicate guard — check if wFirma PZ already exists ──────────────────
    wfirma_export = audit.get("wfirma_export") or {}
    existing_pz_doc_id = wfirma_export.get("wfirma_pz_doc_id") or ""
    if existing_pz_doc_id:
        _supplier_cfg  = bool((settings.wfirma_supplier_contractor_id or "").strip())
        _warehouse_cfg = bool((settings.wfirma_warehouse_id or "").strip())
        _create_flag   = bool(getattr(settings, "wfirma_create_pz_allowed", False))
        lock_status = _compute_pz_lock_status(
            audit,
            preview_ready=False,            # already created — preview is moot
            supplier_configured=_supplier_cfg,
            warehouse_configured=_warehouse_cfg,
        )
        # Cross-batch duplicate detection — only when doc_id is set.
        _dup_owner = _find_pz_owner_batch(existing_pz_doc_id, batch_id)
        pz_lifecycle = _compute_pz_lifecycle_state(
            audit,
            preview_ready=False,
            supplier_configured=_supplier_cfg,
            warehouse_configured=_warehouse_cfg,
            create_allowed=_create_flag,
            duplicate_owner_batch_id=_dup_owner,
            blocker_codes=[],
        )
        eff_status, status_normalized = _compute_effective_pz_status(audit)
        return JSONResponse({
            "batch_id":             batch_id,
            "already_created":      True,
            "wfirma_pz_doc_id":     existing_pz_doc_id,
            "would_create_pz":      False,
            "ready":                False,
            "blockers":             [],
            "engine_error":         None,
            "unresolved_product_codes": [],
            "price_conflicts":      [],
            "pz_lock_status":       lock_status,
            "pz_lifecycle":         pz_lifecycle,
            "stored_status":        audit.get("status", ""),
            "effective_status":     eff_status,
            "status_normalized":    status_normalized,
        })

    # ── Load rows ─────────────────────────────────────────────────────────────
    rows_raw = _build_rows(output_dir, audit)

    cd             = audit.get("customs_declaration") or {}
    mrn            = cd.get("mrn", "") or audit.get("inputs", {}).get("zc429_mrn", "") or ""
    clearance_date = cd.get("clearance_date", "") or audit.get("timestamp", "")[:10]

    # ── Resolve supplier contractor ───────────────────────────────────────────
    supplier_wfirma_id = (settings.wfirma_supplier_contractor_id or "").strip()
    supplier, supplier_source, risk_flags = _resolve_supplier(audit)
    if not supplier_wfirma_id:
        risk_flags.append("WFIRMA_SUPPLIER_CONTRACTOR_ID not configured — set in .env")

    warehouse_id = (settings.wfirma_warehouse_id or "").strip()

    # ── Load product_code → wfirma_good_id mapping ───────────────────────────
    product_map: Dict[str, str] = {}
    all_products = wfirma_db.list_products()
    for p in all_products:
        pid  = (p.get("wfirma_product_id") or "").strip()
        code = (p.get("product_code") or "").strip()
        if pid and code:
            product_map[code] = pid

    # ── Convert raw rows to BatchRow ──────────────────────────────────────────
    batch_rows = [
        BatchRow(
            product_code   = (str(r.get("product_code", "") or "").strip()
                              or _build_product_code(str(r.get("invoice_no", "") or ""), i + 1)),
            quantity       = float(r.get("quantity", 1) or 1),
            unit_netto_pln = float(r.get("unit_netto_pln", 0) or 0),
            invoice_no     = str(r.get("invoice_no", "") or ""),
            description_en = str(r.get("description_en", "") or r.get("_description_en", "") or ""),
            pl_desc        = str(r.get("pl_desc", "") or r.get("_pl_desc", "") or ""),
            item_type      = str(r.get("item_type", "") or ""),
            unit           = str(r.get("unit", "szt.") or "szt."),
        )
        for i, r in enumerate(rows_raw)
    ]

    # ── Build preview ─────────────────────────────────────────────────────────
    # Compact audit-trail notes go into the wFirma PZ <description>
    # field (operator-visible "Uwagi"). Sourced from the current audit
    # only — never from static templates. See `wfirma_pz_notes.py`.
    pz_notes = build_wfirma_pz_notes(audit, batch_id)
    result = build_pz_request_from_batch(
        rows           = batch_rows,
        contractor_id  = supplier_wfirma_id,
        warehouse_id   = warehouse_id,
        product_map    = product_map,
        batch_id       = batch_id,
        clearance_date = clearance_date,
        mrn            = mrn,
        description_override = pz_notes,
    )

    planned = [
        {
            "product_code": pl.product_code,
            "good_id":      pl.good_id,
            "count":        pl.count,
            "price_pln":    round(pl.price_pln, 4),
            "description":  pl.description,
            "resolved":     pl.resolved,
        }
        for pl in result.planned_lines
    ]

    # `description` in the response is the operator-visible "Uwagi"
    # text that will be written to wFirma on PZ create. It mirrors the
    # compact audit notes computed above. Falls back to the legacy
    # batch/MRN format only when no audit fields were available.
    description = pz_notes or (
        f"batch={batch_id}" + (f" | MRN {mrn}" if mrn else "")
    )

    log.info(
        "[%s] wFirma PZ preview: ready=%s unresolved=%d conflicts=%d",
        batch_id, result.ready, len(result.unresolved_codes), len(result.price_conflicts),
    )

    lock_status = _compute_pz_lock_status(
        audit,
        preview_ready=result.ready,
        supplier_configured=bool(supplier_wfirma_id),
        warehouse_configured=bool(warehouse_id),
    )
    # Single PZ lifecycle authority — UI consumes ONLY this field.
    pz_lifecycle = _compute_pz_lifecycle_state(
        audit,
        preview_ready=result.ready,
        supplier_configured=bool(supplier_wfirma_id),
        warehouse_configured=bool(warehouse_id),
        create_allowed=bool(getattr(settings, "wfirma_create_pz_allowed", False)),
        duplicate_owner_batch_id=None,    # no doc_id on this build path
        blocker_codes=[],
    )

    eff_status, status_normalized = _compute_effective_pz_status(audit)
    return JSONResponse({
        "batch_id":                 batch_id,
        "already_created":          False,
        "wfirma_pz_doc_id":         None,
        "would_create_pz":          result.ready and bool(supplier_wfirma_id) and bool(warehouse_id),
        "ready":                    result.ready,
        "blockers":                 [],
        "engine_error":             None,
        "pz_lifecycle":             pz_lifecycle,
        "unresolved_product_codes": result.unresolved_codes,
        "price_conflicts":          result.price_conflicts,
        "supplier_wfirma_id":       supplier_wfirma_id,
        "supplier_name":            supplier,
        "supplier_source":          supplier_source,
        "warehouse_id":             warehouse_id,
        "mrn":                      mrn,
        "clearance_date":           clearance_date,
        "description":              description,
        "planned_lines":            planned,
        "risk_flags":               risk_flags,
        "pz_lock_status":           lock_status,
        "stored_status":            audit.get("status", ""),
        "effective_status":         eff_status,
        "status_normalized":        status_normalized,
    })


@router.post("/shipment/{batch_id}/wfirma/products/resolve", dependencies=[_auth])
async def wfirma_products_resolve(batch_id: str) -> JSONResponse:
    """
    Batch product resolve — map every product_code in this PZ batch to a
    wFirma good_id, so that the batch is ready for PZ creation.

    For each product_code:
      1. Already in wfirma_products with wfirma_product_id → already_mapped
      2. Not in local table → call goods/find by code (read-only)
         a. Found in wFirma → save mapping, count as found_and_mapped
         b. Missing + WFIRMA_CREATE_PRODUCT_ALLOWED=false → missing (no write)
         c. Missing + WFIRMA_CREATE_PRODUCT_ALLOWED=true → create via goods/add,
            save mapping on confirmed success, count as created
      3. goods/find or goods/add error → failed (no fake mapping written)

    Idempotent: re-running after partial resolve skips already_mapped codes.
    Never calls create_warehouse_pz.

    Response
    --------
    batch_id, considered, already_mapped, found_and_mapped,
    created, missing, failed, ready_for_pz, details
    """
    output_dir = get_output_dir(batch_id)
    audit      = _read_audit(output_dir)
    _guard_wfirma_export(audit)

    rows_raw = _build_rows(output_dir, audit)

    # ── Collect unique (product_code, item_type, description_en) per code ────
    # Use first occurrence of each code for description/item_type (they are
    # per-row metadata used only for the create path).
    seen: Dict[str, dict] = {}
    for i, r in enumerate(rows_raw):
        pc = (r.get("product_code") or "").strip()
        if not pc:
            continue
        if pc not in seen:
            seen[pc] = {
                "item_type":      str(r.get("item_type", "") or ""),
                "description_en": str(r.get("description_en", "") or r.get("_description_en", "") or ""),
                "pl_desc":        str(r.get("pl_desc", "") or r.get("_pl_desc", "") or ""),
                "unit_netto_pln": float(r.get("unit_netto_pln", 0) or 0),
                "quantity":       float(r.get("quantity", 1) or 1),
            }

    considered      = len(seen)
    already_mapped  = 0
    found_and_mapped = 0
    created         = 0
    missing_codes: List[str]  = []
    failed_details: List[dict] = []
    drift_codes:    List[str]  = []   # already-mapped codes whose local name has drifted from authority

    # ── Authority pre-flight (Goods Authority Hardening, 2026-05-22) ──────
    # Determine whether the current audit/rows constitute a valid authority
    # source for wFirma goods creation:
    #   - audit._rows_source == "invoice_positions_authority"  (PR #269)
    #   - OR per-row pl_desc populated (Estrella supplier path)
    # The pre-flight is computed ONCE for the batch (cheap), and a refusal
    # is per-code: every row that lacks usable authority refuses with
    # STALE_AUTHORITY_REFUSED rather than triggering the legacy
    # description_engine cache fallback.
    _global_authority = (audit.get("_rows_source") or "") == "invoice_positions_authority"

    operator_for_log = _operator_from_header(None)  # endpoint has no operator header today

    # Performance: batch-fetch all known local products in one SQL round-trip
    # instead of O(N) individual get_product() calls (C6 T6 hardening).
    _local_cache: Dict[str, Any] = wfirma_db.get_products_batch(list(seen.keys()))

    for pc, meta in seen.items():
        # ── 1. Check local table first ────────────────────────────────────────
        local = _local_cache.get(pc)
        if local and (local.get("wfirma_product_id") or "").strip():
            already_mapped += 1
            # Drift surface: when authority name (built from row) differs
            # from cached product_name, expose the code so operator can
            # run /wfirma/products/sync-names. Does NOT auto-rename.
            authority_name = _build_authority_name(meta)
            cached_name = (local.get("product_name") or "").strip()
            if authority_name and cached_name and authority_name != cached_name:
                drift_codes.append(pc)
            continue

        # ── 2. Live goods/find ────────────────────────────────────────────────
        try:
            found = wfirma_client.get_product_by_code(pc)
        except Exception as exc:
            failed_details.append({"product_code": pc, "error": f"goods/find: {exc}"})
            continue

        if found is not None:
            wfirma_db.upsert_product(
                product_code      = pc,
                wfirma_product_id = found.wfirma_id,
                product_name_pl   = found.name or "",
                product_name      = found.name or "",
                unit              = found.unit or "szt.",
                sync_status       = "matched",
            )
            found_and_mapped += 1
            continue

        # ── 3a. Missing + gate off ────────────────────────────────────────────
        if not settings.wfirma_create_product_allowed:
            missing_codes.append(pc)
            continue

        # ── 3b. Missing + gate on → create from PZ ROWS AUTHORITY (Phase E) ──
        # Authority pre-flight: refuse to create when the row lacks a
        # usable authority source. STALE_AUTHORITY_REFUSED prevents the
        # repeat of the Global AWB 4789974092 incident, where stale
        # description_engine cache produced wrong wFirma good names that
        # were then captured by a PZ document snapshot.
        row_pl   = (meta.get("pl_desc")        or "").strip()
        row_en   = (meta.get("description_en") or "").strip()
        if not _global_authority and not row_pl:
            failed_details.append({
                "product_code": pc,
                "error":        "STALE_AUTHORITY_REFUSED",
                "reason":       (
                    "row has no pl_desc and audit._rows_source is not "
                    "'invoice_positions_authority' — refusing to create "
                    "wFirma good from a non-authority description source. "
                    "Re-run the Polish Description regenerate (force=true) "
                    "before retrying /products/resolve."
                ),
                "audit_rows_source": audit.get("_rows_source"),
                "has_pl_desc":       bool(row_pl),
            })
            continue

        # Build the wFirma good name + description block DIRECTLY from
        # the pz_rows row — the description_engine cache is bypassed for
        # the create path so future runs cannot ship stale text.
        wf_name = _build_authority_name(meta)
        if not wf_name:
            # Defensive fallback only when row genuinely has nothing —
            # falls back to product_code as a last-resort identifier.
            wf_name = pc

        if row_pl:
            material_pl = _material_from_pl_desc(row_pl) or row_pl
            description_block = deng.build_description_block(
                description_pl = row_pl,
                material_pl    = material_pl,
                purpose_pl     = "Ozdoba — biżuteria do noszenia.",
                description_en = row_en,
            )
            authority_source = "pz_rows_authority"
        else:
            # No row pl_desc but invoice_positions_authority is set — use
            # description_engine ONCE to seed the description block, but
            # the name still comes from the row's available description_en.
            block = deng.get_description_block(
                product_code   = pc,
                item_type      = meta["item_type"],
                description_en = row_en,
            )
            description_block = block.get("description_block") or ""
            authority_source  = "description_engine_fallback"

        try:
            result_product = wfirma_client.create_product(
                product_code = pc,
                name         = wf_name,
                unit         = "szt.",
                netto        = 0.0,
                vat_code_id  = wfirma_client.find_vat_code_id(23),
                description  = description_block,
            )
        except Exception as exc:
            log.warning("[%s] products/resolve: goods/add failed for %r: %s", batch_id, pc, exc)
            failed_details.append({"product_code": pc, "error": f"goods/add: {exc}"})
            continue

        if not result_product.wfirma_id:
            failed_details.append({"product_code": pc, "error": "goods/add returned no wfirma_id"})
            continue

        # Derive a short PL name (first item type before comma / "ze ")
        # for the wfirma_products.product_name_pl column.
        if row_pl:
            short_pl = row_pl.split(" ze ")[0].split(",")[0].strip() or row_pl
        else:
            short_pl = ""
        wfirma_db.upsert_product(
            product_code      = pc,
            wfirma_product_id = result_product.wfirma_id,
            product_name_pl   = short_pl,
            product_name      = wf_name,
            description_block = description_block,
            unit              = "szt.",
            sync_status       = "matched",
        )
        # Timeline event — pins the authority source the create used so
        # future audits can prove no stale-name creation happened.
        try:
            tl.log_event(
                output_dir / "audit.json",
                EV_WFIRMA_GOOD_CREATED_FROM_AUTHORITY,
                "dashboard",
                operator_for_log,
                detail={
                    "product_code":      pc,
                    "wfirma_id":         result_product.wfirma_id,
                    "name":              wf_name,
                    "authority_source":  authority_source,
                    "audit_rows_source": audit.get("_rows_source"),
                },
            )
        except Exception:  # noqa: BLE001 — never fail the create because of audit-log
            pass
        created += 1

    # ── Compute ready_for_pz via pz_preview builder ───────────────────────────
    product_map: Dict[str, str] = {}
    for p in wfirma_db.list_products():
        pid  = (p.get("wfirma_product_id") or "").strip()
        code = (p.get("product_code") or "").strip()
        if pid and code:
            product_map[code] = pid

    cd             = audit.get("customs_declaration") or {}
    mrn            = cd.get("mrn", "") or audit.get("inputs", {}).get("zc429_mrn", "") or ""
    clearance_date = cd.get("clearance_date", "") or audit.get("timestamp", "")[:10]
    supplier_id    = (settings.wfirma_supplier_contractor_id or "").strip()
    warehouse_id   = (settings.wfirma_warehouse_id or "").strip()

    batch_rows = [
        BatchRow(
            product_code   = (r.get("product_code") or "").strip(),
            quantity       = float(r.get("quantity", 1) or 1),
            unit_netto_pln = float(r.get("unit_netto_pln", 0) or 0),
            invoice_no     = str(r.get("invoice_no", "") or ""),
            description_en = str(r.get("description_en", "") or ""),
            pl_desc        = str(r.get("pl_desc", "") or ""),
            item_type      = str(r.get("item_type", "") or ""),
        )
        for r in rows_raw
        if (r.get("product_code") or "").strip()
    ]

    preview = build_pz_request_from_batch(
        rows           = batch_rows,
        contractor_id  = supplier_id,
        warehouse_id   = warehouse_id,
        product_map    = product_map,
        batch_id       = batch_id,
        clearance_date = clearance_date,
        mrn            = mrn,
        description_override = build_wfirma_pz_notes(audit, batch_id),
    )

    log.info(
        "[%s] products/resolve: considered=%d already_mapped=%d found=%d "
        "created=%d missing=%d failed=%d ready=%s",
        batch_id, considered, already_mapped, found_and_mapped,
        created, len(missing_codes), len(failed_details), preview.ready,
    )

    return JSONResponse({
        "batch_id":          batch_id,
        "considered":        considered,
        "already_mapped":    already_mapped,
        "found_and_mapped":  found_and_mapped,
        "created":           created,
        "missing":           len(missing_codes),
        "failed":            len(failed_details),
        "missing_codes":     missing_codes,
        "failed_details":    failed_details,
        "ready_for_pz":      preview.ready,
        "unresolved_product_codes": preview.unresolved_codes,
        "price_conflicts":   preview.price_conflicts,
        # Drift surface (Goods Authority Hardening, Phase E): codes that
        # are already_mapped locally but whose cached wFirma name has
        # drifted from the current pz_rows authority. Surfacing only —
        # operator must run /wfirma/products/sync-names to fix.
        "drift_codes":       drift_codes,
        "drift_count":       len(drift_codes),
    })


# ── Timeline event constants for sync-names + clear-mapping ────────────────
# Use string literals (no new tl.EV_* constants) to keep the timeline module
# clean of campaign-specific events; the values become the timeline `event`
# field verbatim.
EV_WFIRMA_GOOD_RENAMED                 = "wfirma_good_renamed"
EV_WFIRMA_PZ_MAPPING_CLEARED           = "wfirma_pz_mapping_cleared"
EV_WFIRMA_GOOD_CREATED_FROM_AUTHORITY  = "wfirma_good_created_from_authority"


def _build_authority_name(row_meta: Dict[str, Any]) -> str:
    """Build a wFirma good name from a pz_rows row's authoritative
    fields. Pure function — `row_meta` is the per-code dict assembled
    in `/products/resolve` (carries pl_desc, description_en, item_type).

    Format mirrors PR #276 sync-names output:
      "<pl_desc> / <description_en>"  when both present
      "<pl_desc>"                      when only pl_desc
      "<description_en>"               when only description_en
      ""                               when neither
    """
    pl = (row_meta.get("pl_desc")        or "").strip()
    en = (row_meta.get("description_en") or "").strip()
    if pl and en:
        return f"{pl} / {en}"
    return pl or en


def _material_from_pl_desc(pl_desc: str) -> str:
    """Derive a wFirma-friendly "Z jakiego materiału" line from a PL
    description by stripping the item-type prefix.

    Examples:
      "Bransoletki ze złota próby 375 z diamentami laboratoryjnymi"
        → "złoto próby 375 z diamentami laboratoryjnymi"
      "Kolczyki ze srebra próby 925 z cyrkoniami"
        → "srebro próby 925 z cyrkoniami"
      "Pierścionki ze srebra próby 925"
        → "srebro próby 925"

    The PR #269 PL grammar always uses "ze złota"/"ze srebra"/"z platyny"
    or simply "ze stali". Strip everything up to and including the
    leading "ze "/"z " preposition and any item-type-list prefix. If the
    pattern is unrecognised, return the input unchanged.
    """
    if not pl_desc:
        return ""
    # PR #269 grammar: "<types>, <types> ze {metal} próby {n} z {stones}"
    # Find the first " ze " or " z " followed by metal name.
    import re as _re
    m = _re.search(r"\s+ze\s+(.+)$", pl_desc)
    if m:
        return m.group(1).strip()
    m = _re.search(r"\s+z\s+(\S+\s+próby\s+\S+.*)$", pl_desc)
    if m:
        return m.group(1).strip()
    return pl_desc.strip()


def _operator_from_header(x_operator: Optional[str]) -> str:
    """Extract operator id from the X-Operator header. Falls back to
    'operator' so old clients that don't set the header still produce
    a stable, non-empty operator label in audit/timeline. Backward-
    compatible — never raises, never alters request flow."""
    return (x_operator or "").strip() or "operator"


@router.post("/shipment/{batch_id}/wfirma/products/sync-names", dependencies=[_auth])
async def wfirma_products_sync_names(
    batch_id:   str,
    x_operator: Optional[str] = Header(None, alias="X-Operator"),
) -> JSONResponse:
    """
    Edit the name + description of wFirma goods already mapped to this
    batch's product codes, so they match the current invoice-position
    authority (`pz_rows.json`).

    Use case
    --------
    The `/products/resolve` flow creates wFirma goods using the description
    block available at THAT moment. If the audit/engine state was stale at
    create time (e.g. PR #269 bridge had not yet repopulated rows), the
    resulting wFirma good names are stuck on stale text. This endpoint
    walks the current authoritative `pz_rows.json`, builds the correct
    name + description from each row's `pl_desc` / `description_en`, and
    writes those values to existing wFirma goods via `goods/edit`. Local
    `wfirma_products` cache is updated atomically.

    Guard order
    -----------
    1. `WFIRMA_CREATE_PRODUCT_ALLOWED` must be true (reuses the existing
       product-write flag; editing is semantically the same write class).
    2. Batch/audit exists, SAD present, PZ rows on disk
       (`_guard_wfirma_export`).
    3. Each row must have a non-empty `pl_desc` and a resolved
       `wfirma_product_id` in `wfirma_products`. Codes that are not
       mapped → `unmapped` count, not failures.

    Behaviour
    ---------
    For each row:
      - new_name = "<pl_desc> / <description_en>" (matches the existing
        wFirma name pattern produced by `/products/resolve`).
      - new_description = a freshly-built bilingual block using
        `description_engine.build_description_block` with
        material_pl derived from pl_desc, purpose_pl fixed to
        "Ozdoba — biżuteria do noszenia.".
      - Diff against the cached `wfirma_products.product_name`. If
        unchanged → counted in `unchanged`. If different → call
        `wfirma_client.edit_product` and persist locally.

    Never mutates code / price / unit / vat / warehouse_id. Idempotent —
    re-running after a successful sync reports `renamed=0`.

    Response
    --------
    {
      batch_id, considered, unchanged, renamed, unmapped, failed,
      failed_details, before_after: [{code, wfirma_id, before, after}],
    }
    """
    if not settings.wfirma_create_product_allowed:
        raise HTTPException(
            status_code=422,
            detail={
                "guard": "wfirma_create_product_allowed",
                "error": "wFirma product writes are disabled. Set WFIRMA_CREATE_PRODUCT_ALLOWED=true to enable.",
                "code":  "WFIRMA_PRODUCT_WRITE_DISABLED",
            },
        )

    output_dir = get_output_dir(batch_id)
    audit      = _read_audit(output_dir)
    _guard_wfirma_export(audit)

    rows_raw = _build_rows(output_dir, audit)
    operator = _operator_from_header(x_operator)

    considered      = 0
    unchanged       = 0
    renamed         = 0
    unmapped_codes: List[str] = []
    failed_details: List[Dict[str, Any]] = []
    before_after:   List[Dict[str, Any]] = []

    # Deduplicate by product_code — multiple rows may share one code in
    # principle; only process the first occurrence per code per run.
    seen_codes: set = set()
    for r in rows_raw:
        pc = (r.get("product_code") or "").strip()
        if not pc or pc in seen_codes:
            continue
        seen_codes.add(pc)

        considered += 1
        pl_desc   = (r.get("pl_desc") or r.get("_pl_desc") or "").strip()
        desc_en   = (r.get("description_en") or r.get("_description_en") or "").strip()
        if not pl_desc:
            failed_details.append({
                "product_code": pc,
                "error":        "pz_rows row has empty pl_desc — cannot derive new name",
            })
            continue

        existing = wfirma_db.get_product(pc)
        wfirma_id = (existing or {}).get("wfirma_product_id") or ""
        wfirma_id = wfirma_id.strip()
        if not wfirma_id:
            unmapped_codes.append(pc)
            continue

        # Build new name: PR #269 PL grammar + slash + English source line.
        new_name = pl_desc
        if desc_en:
            new_name = f"{pl_desc} / {desc_en}"

        # Build new description block matching the existing format.
        material_pl = _material_from_pl_desc(pl_desc)
        new_description = deng.build_description_block(
            description_pl = pl_desc,
            material_pl    = material_pl or pl_desc,
            purpose_pl     = "Ozdoba — biżuteria do noszenia.",
            description_en = desc_en,
        )

        before_name = (existing or {}).get("product_name") or ""
        if before_name.strip() == new_name.strip():
            unchanged += 1
            before_after.append({
                "product_code": pc,
                "wfirma_id":    wfirma_id,
                "before":       before_name,
                "after":        new_name,
                "changed":      False,
            })
            continue

        try:
            edit_result = wfirma_client.edit_product(
                wfirma_product_id = wfirma_id,
                name              = new_name,
                description       = new_description,
            )
        except Exception as exc:  # noqa: BLE001
            log.warning(
                "[%s] products/sync-names: goods/edit failed for %r (wfirma_id=%s): %s",
                batch_id, pc, wfirma_id, exc,
            )
            failed_details.append({
                "product_code": pc,
                "wfirma_id":    wfirma_id,
                "error":        f"goods/edit: {exc}",
            })
            continue

        # Persist new name + description block locally. `upsert_product`
        # has never-erase semantics for product_name; passing a non-empty
        # value overwrites correctly.
        try:
            wfirma_db.upsert_product(
                product_code      = pc,
                wfirma_product_id = wfirma_id,
                product_name_pl   = (existing or {}).get("product_name_pl") or pl_desc.split(" ze ")[0].split(",")[0].strip(),
                product_name      = new_name,
                description_block = new_description,
                unit              = (existing or {}).get("unit") or "szt.",
                sync_status       = "matched",
            )
        except Exception as exc:  # noqa: BLE001
            # wFirma was updated but local DB persistence failed. Surface
            # this clearly — the operator must re-run to reconcile the
            # local cache, but wFirma is already correct.
            log.error(
                "[%s] products/sync-names: wFirma updated but local DB upsert "
                "failed for %r: %s — re-run after fix to reconcile local cache",
                batch_id, pc, exc,
            )
            failed_details.append({
                "product_code":     pc,
                "wfirma_id":        wfirma_id,
                "error":            f"local_db_upsert: {exc}",
                "wfirma_updated":   True,
                "local_db_synced":  False,
            })
            continue

        renamed += 1
        before_after.append({
            "product_code": pc,
            "wfirma_id":    wfirma_id,
            "before":       before_name,
            "after":        new_name,
            "changed":      True,
        })
        tl.log_event(
            output_dir / "audit.json",
            EV_WFIRMA_GOOD_RENAMED,
            "dashboard",
            operator,
            detail={
                "product_code": pc,
                "wfirma_id":    wfirma_id,
                "before":       before_name,
                "after":        new_name,
            },
        )

    log.info(
        "[%s] products/sync-names: considered=%d unchanged=%d renamed=%d "
        "unmapped=%d failed=%d",
        batch_id, considered, unchanged, renamed,
        len(unmapped_codes), len(failed_details),
    )

    return JSONResponse({
        "batch_id":       batch_id,
        "considered":     considered,
        "unchanged":      unchanged,
        "renamed":        renamed,
        "unmapped":       len(unmapped_codes),
        "unmapped_codes": unmapped_codes,
        "failed":         len(failed_details),
        "failed_details": failed_details,
        "before_after":   before_after,
    })


@router.post("/shipment/{batch_id}/wfirma/pz/clear-mapping", dependencies=[_auth])
async def wfirma_pz_clear_mapping(
    batch_id:   str,
    x_operator: Optional[str] = Header(None, alias="X-Operator"),
) -> JSONResponse:
    """
    Clear the `wfirma_pz_doc_id` and related fields from this batch's
    audit, so a subsequent `/wfirma/pz_create` call will issue a NEW PZ
    document. The operator must have ALREADY deleted (or otherwise
    cancelled) the original PZ document in the wFirma admin UI — this
    endpoint trusts that operator-side action and only clears the local
    audit linkage.

    Guard
    -----
    Operator-explicit only. The X-Operator header MUST be present and
    non-empty. No write flag controls this — clearing a stale mapping is
    not a wFirma write; it is a local audit edit. The endpoint does NOT
    attempt to delete the PZ document in wFirma (no such wrapper exists,
    by design — accounting documents are not programmatically destroyed
    from this app).

    What is cleared from `audit.wfirma_export`:
      - `wfirma_pz_doc_id`
      - `wfirma_pz_fullnumber`
      - `pz_source`
      - `pz_created_at`

    Preserved:
      - clipboard_generated, json_generated, last_generated_at, mode,
        row_count, generated_count etc.

    A timeline event `wfirma_pz_mapping_cleared` is appended with the
    operator id, the prior `wfirma_pz_doc_id`, and a free-text reason
    passed via the X-Operator header or query string (omitted by default).

    Idempotent: re-calling when no doc id is set returns
    `status=already_cleared`.

    Response
    --------
    { status, batch_id, previous_wfirma_pz_doc_id, previous_pz_source }
    """
    if not (x_operator or "").strip():
        raise HTTPException(
            status_code=422,
            detail={
                "guard": "operator_explicit",
                "error": "X-Operator header is required for pz/clear-mapping.",
                "code":  "OPERATOR_HEADER_MISSING",
            },
        )
    operator = _operator_from_header(x_operator)

    output_dir = get_output_dir(batch_id)
    audit_path = output_dir / "audit.json"
    if not audit_path.exists():
        raise HTTPException(status_code=404, detail=f"Batch {batch_id} not found.")
    audit = _read_audit(output_dir)

    wfirma_export = audit.get("wfirma_export") or {}
    prev_doc_id   = (wfirma_export.get("wfirma_pz_doc_id") or "").strip()
    prev_source   = (wfirma_export.get("pz_source") or "").strip()

    # True idempotency: already cleared when the timeline's latest PZ-mapping
    # event is wfirma_pz_mapping_cleared (the same check the lifecycle engine
    # uses). Checking only prev_doc_id is insufficient — a /process run can
    # wipe wfirma_export without writing the clearing event, leaving the
    # timeline in PZ_RECOVERY_REQUIRED state with no self-service exit.
    if _has_pz_mapping_cleared_after_create(audit):
        return JSONResponse({
            "status":   "already_cleared",
            "batch_id": batch_id,
            "previous_wfirma_pz_doc_id": None,
            "previous_pz_source":        None,
        })

    # Strip the four post-create fields; preserve everything else in
    # wfirma_export (clipboard_generated, json_generated, etc.).
    for k in ("wfirma_pz_doc_id", "wfirma_pz_fullnumber",
              "pz_source", "pz_created_at"):
        wfirma_export.pop(k, None)
    audit["wfirma_export"] = wfirma_export
    write_json_atomic(audit_path, audit)

    tl.log_event(
        audit_path,
        EV_WFIRMA_PZ_MAPPING_CLEARED,
        "dashboard",
        operator,
        detail={
            "previous_wfirma_pz_doc_id": prev_doc_id,
            "previous_pz_source":        prev_source,
        },
    )

    log.info(
        "[%s] pz/clear-mapping: cleared previous wfirma_pz_doc_id=%s "
        "(source=%s) by operator=%s",
        batch_id, prev_doc_id, prev_source, operator,
    )

    return JSONResponse({
        "status":   "cleared",
        "batch_id": batch_id,
        "previous_wfirma_pz_doc_id": prev_doc_id,
        "previous_pz_source":        prev_source,
    })


@router.post("/shipment/{batch_id}/wfirma/pz_create", dependencies=[_auth])
async def wfirma_pz_create(
    batch_id: str,
    x_operator: Optional[str] = Header(None, alias="X-Operator"),
) -> JSONResponse:
    """
    Create a wFirma warehouse PZ directly from the internal PZ app batch output.

    This is the normal import path: PZ app output → resolved goods → wFirma PZ.
    It does NOT call build_pz_request_from_proforma_snapshot (recovery workaround).
    It does NOT create products, trigger proforma creation, or convert to invoice.

    Guard order
    -----------
    1. WFIRMA_CREATE_PZ_ALLOWED must be true
    2. Batch/audit exists, SAD present, PZ generated (_guard_wfirma_export)
    3. MRN present (customs dedup key)
    4. WFIRMA_SUPPLIER_CONTRACTOR_ID and WFIRMA_WAREHOUSE_ID configured
    5. No existing wfirma_pz_doc_id in audit (idempotency)
    6. pz_preview ready=True — all product_codes resolved, no price conflicts

    On success:  writes wfirma_pz_doc_id to audit atomically, returns status=created.
    On failure:  writes nothing, returns status=failed with wFirma error.
    Idempotent:  returns status=already_created when PZ already exists for batch.

    Response
    --------
    status            created | already_created | failed | not_ready
    wfirma_pz_doc_id  wFirma warehouse_document id (on created/already_created)
    planned_lines     per-line preview with good_id and price_pln (on created)
    line_count        number of lines submitted (on created)
    error             wFirma error message (on failed)
    """
    # ── Guard 1: feature gate ─────────────────────────────────────────────────
    if not getattr(settings, "wfirma_create_pz_allowed", False):
        raise HTTPException(
            status_code=403,
            detail={
                "guard": "pz_create",
                "error": "WFIRMA_CREATE_PZ_ALLOWED is not enabled.",
                "code":  "PZ_CREATE_GATE_OFF",
            },
        )

    # ── Guard 2: batch + SAD + PZ status ─────────────────────────────────────
    output_dir = get_output_dir(batch_id)
    audit      = _read_audit(output_dir)
    _guard_wfirma_export(audit)

    # ── Guard 3: MRN ─────────────────────────────────────────────────────────
    cd  = audit.get("customs_declaration") or {}
    mrn = cd.get("mrn", "") or audit.get("inputs", {}).get("zc429_mrn", "") or ""
    if not mrn:
        raise HTTPException(
            status_code=422,
            detail={
                "guard": "pz_create",
                "error": "MRN required for PZ creation.",
                "code":  "PZ_CREATE_NO_MRN",
            },
        )

    # ── Guard 4: supplier + warehouse configured ──────────────────────────────
    supplier_wfirma_id = (getattr(settings, "wfirma_supplier_contractor_id", None) or "").strip()
    warehouse_id       = (getattr(settings, "wfirma_warehouse_id", None) or "").strip()
    if not supplier_wfirma_id:
        raise HTTPException(
            status_code=422,
            detail={
                "guard": "pz_create",
                "error": "WFIRMA_SUPPLIER_CONTRACTOR_ID not configured.",
                "code":  "PZ_CREATE_NO_SUPPLIER",
            },
        )
    if not warehouse_id:
        raise HTTPException(
            status_code=422,
            detail={
                "guard": "pz_create",
                "error": "WFIRMA_WAREHOUSE_ID not configured.",
                "code":  "PZ_CREATE_NO_WAREHOUSE",
            },
        )

    # ── Guard 5: idempotent fast-path BEFORE acquiring lock ───────────────────
    # If the doc_id is already on disk and matches a prior create, return
    # already_created quickly without taking the file lock.  Adoption-style
    # provenance is rejected here too.
    wfirma_export      = audit.get("wfirma_export") or {}
    existing_pz_doc_id = (wfirma_export.get("wfirma_pz_doc_id") or "").strip()
    existing_pz_source = (wfirma_export.get("pz_source") or "").strip()
    if existing_pz_doc_id and existing_pz_source != "adopted_existing":
        return JSONResponse({
            "batch_id":         batch_id,
            "status":           "already_created",
            "wfirma_pz_doc_id": existing_pz_doc_id,
        })

    # ── Build rows + product map ──────────────────────────────────────────────
    clearance_date = cd.get("clearance_date", "") or audit.get("timestamp", "")[:10]
    rows_raw       = _build_rows(output_dir, audit)

    product_map: Dict[str, str] = {}
    for p in wfirma_db.list_products():
        pid  = (p.get("wfirma_product_id") or "").strip()
        code = (p.get("product_code") or "").strip()
        if pid and code:
            product_map[code] = pid

    batch_rows = [
        BatchRow(
            product_code   = (str(r.get("product_code", "") or "").strip()
                              or _build_product_code(str(r.get("invoice_no", "") or ""), i + 1)),
            quantity       = float(r.get("quantity", 1) or 1),
            unit_netto_pln = float(r.get("unit_netto_pln", 0) or 0),
            invoice_no     = str(r.get("invoice_no", "") or ""),
            description_en = str(r.get("description_en", "") or
                                 r.get("_description_en", "") or ""),
            pl_desc        = str(r.get("pl_desc", "") or r.get("_pl_desc", "") or ""),
            item_type      = str(r.get("item_type", "") or ""),
            unit           = str(r.get("unit", "szt.") or "szt."),
        )
        for i, r in enumerate(rows_raw)
    ]

    # ── Guard 6: preview must be ready ───────────────────────────────────────
    preview = build_pz_request_from_batch(
        rows           = batch_rows,
        contractor_id  = supplier_wfirma_id,
        warehouse_id   = warehouse_id,
        product_map    = product_map,
        batch_id       = batch_id,
        clearance_date = clearance_date,
        mrn            = mrn,
        # Compact audit-trail notes for the wFirma <description> field.
        # See `wfirma_pz_notes.build_wfirma_pz_notes`.
        description_override = build_wfirma_pz_notes(audit, batch_id),
    )

    if not preview.ready:
        log.warning(
            "[%s] pz_create blocked: unresolved=%s conflicts=%s",
            batch_id, preview.unresolved_codes, preview.price_conflicts,
        )
        return JSONResponse(
            status_code=422,
            content={
                "batch_id":                 batch_id,
                "status":                   "not_ready",
                "unresolved_product_codes": preview.unresolved_codes,
                "price_conflicts":          preview.price_conflicts,
            },
        )

    # ── Acquire write lock + re-check inside critical section ────────────────
    # Two-stage check: the fast-path above is for the common already_created
    # case.  Inside the lock we re-read audit.json and consult the *full*
    # idempotency guard (doc_id + pz_source + timeline events) so concurrent
    # requests cannot both race past the check.
    planned: List[Dict[str, Any]] = []
    with _pz_write_lock(output_dir, batch_id, "pz_create"):
        audit_locked = _read_audit(output_dir)
        _assert_pz_not_locked(audit_locked, batch_id, "pz_create")

        # ── Call wFirma ──────────────────────────────────────────────────────
        pz_result = wfirma_client.create_warehouse_pz(preview.pz_request)

        planned = [
            {
                "product_code": pl.product_code,
                "good_id":      pl.good_id,
                "count":        pl.count,
                "price_pln":    round(pl.price_pln, 4),
                "resolved":     pl.resolved,
            }
            for pl in preview.planned_lines
        ]

        if not pz_result.ok:
            log.warning("[%s] pz_create: wFirma returned failure: %s", batch_id, pz_result.error)
            tl.log_event(
                output_dir / "audit.json",
                "wfirma_pz_create_failed",
                "system",
                "wfirma",
                detail={"batch_id": batch_id, "error": pz_result.error},
            )
            return JSONResponse(
                status_code=502,
                content={
                    "batch_id": batch_id,
                    "status":   "failed",
                    "error":    pz_result.error,
                },
            )

        # ── Success: persist doc_id, then log timeline ───────────────────────
        # Audit write must succeed BEFORE we declare success.  If it fails the
        # wFirma document has already been created — surface a structured
        # warning so the operator can manually adopt rather than re-create.
        audit_write_error = _patch_pz_doc_id(output_dir, pz_result.wfirma_pz_doc_id)
        operator = _operator_from_header(x_operator)
        tl.log_event(
            output_dir / "audit.json",
            EV_WFIRMA_PZ_CREATED,
            "system",
            "wfirma",
            detail={
                "batch_id":         batch_id,
                "wfirma_pz_doc_id": pz_result.wfirma_pz_doc_id,
                "line_count":       len(planned),
                "operator":         operator,
            },
        )
        log.info(
            "[%s] pz_create: wFirma PZ %s created (%d lines, operator=%s)",
            batch_id, pz_result.wfirma_pz_doc_id, len(planned), operator,
        )
        # Restamp audit.status if operator-effective normalisation now
        # says the run is done. Best-effort: never breaks the create flow.
        try:
            from ..services.audit_persist import restamp_pz_status_if_done
            _r = restamp_pz_status_if_done(output_dir / "audit.json")
            if _r.get("changed"):
                log.info("[%s] pz_create: audit.status restamped %r → %r",
                         batch_id, _r.get("stored_before"), _r.get("stored_after"))
        except Exception as _exc:
            log.warning("[%s] pz_create: status restamp skipped: %s",
                        batch_id, _exc)
        # Auto-map the canonical wFirma PZ fullnumber. Best-effort: a
        # network/parse miss leaves the doc_id-only stamp in place and
        # the operator can run "Refresh Mapping" later.
        try:
            from ..services.audit_persist import record_wfirma_pz_mapping
            _f = wfirma_client.fetch_warehouse_pz(pz_result.wfirma_pz_doc_id)
            if _f.ok and (_f.pz_number or "").strip():
                _m = record_wfirma_pz_mapping(
                    output_dir / "audit.json",
                    wfirma_pz_doc_id     = pz_result.wfirma_pz_doc_id,
                    wfirma_pz_fullnumber = _f.pz_number,
                    source               = "created_via_app",
                    operator             = operator,
                )
                if _m.get("changed"):
                    log.info("[%s] pz_create: mapped %s → %s",
                             batch_id, pz_result.wfirma_pz_doc_id,
                             _f.pz_number)
            else:
                log.warning("[%s] pz_create: fullnumber fetch unavailable "
                            "(%s) — manual confirm remains as fallback",
                            batch_id, getattr(_f, "error", "?"))
        except Exception as _exc:
            log.warning("[%s] pz_create: auto-mapping skipped: %s",
                        batch_id, _exc)

    if audit_write_error:
        # wFirma succeeded but audit didn't reflect it — return 500 with
        # actionable warning, not 200.  The doc_id is included so the
        # operator can adopt it manually.
        return JSONResponse(
            status_code=500,
            content={
                "batch_id":          batch_id,
                "status":            "audit_write_failed",
                "wfirma_pz_doc_id":  pz_result.wfirma_pz_doc_id,
                "warning":           (
                    "wFirma PZ was created but the local audit could not be updated. "
                    "Use Confirm Existing PZ to link this doc_id to the shipment."
                ),
                "audit_error":       audit_write_error,
                "planned_lines":     planned,
                "line_count":        len(planned),
            },
        )

    return JSONResponse({
        "batch_id":         batch_id,
        "status":           "created",
        "wfirma_pz_doc_id": pz_result.wfirma_pz_doc_id,
        "planned_lines":    planned,
        "line_count":       len(planned),
    })


# ── PZ Adopt ──────────────────────────────────────────────────────────────────

class _PZAdoptBody(BaseModel):
    """
    Request body for POST .../wfirma/pz_adopt.

    At least one of pz_doc_id (preferred) or pz_number must be supplied.
    pz_doc_id is the canonical wFirma internal numeric ID.
    pz_number is the human-readable document number (e.g. "PZ 1/5/2026") used
    only when the internal ID is not known.
    """
    pz_doc_id: Optional[str] = None
    pz_number: Optional[str] = None


@router.post("/shipment/{batch_id}/wfirma/pz_adopt", dependencies=[_auth])
async def wfirma_pz_adopt(
    batch_id: str,
    body: _PZAdoptBody,
    x_operator: Optional[str] = Header(None, alias="X-Operator"),
) -> JSONResponse:
    """
    Attach an already-existing wFirma PZ document to this shipment without
    creating a new one (adoption flow for manual / historical PZ documents).

    Guard order
    -----------
    1. At least one of pz_doc_id / pz_number provided
    2. Batch audit exists (batch must exist on disk)
    3. wFirma document resolves and exists (via fetch or search)
    4. Cross-shipment duplicate guard — no other batch may own the same PZ id
    5. This shipment already holds the SAME PZ id → return already_adopted
    6. This shipment already holds a DIFFERENT PZ id → blocked

    On success:  writes wfirma_pz_doc_id + pz_source='adopted_existing' to audit
                 atomically, returns status=adopted.
    Idempotent:  returns status=already_adopted when called again with same id.

    Response
    --------
    status           adopted | already_adopted | blocked
    wfirma_pz_doc_id wFirma PZ internal id confirmed on adoption
    pz_number        human-readable document number (from wFirma response)
    batch_id         echoed
    blocking_reasons list[str] — present only when status=blocked
    """
    pz_doc_id_raw = (body.pz_doc_id or "").strip()
    pz_number_raw = (body.pz_number or "").strip()

    # ── Guard 0: WFIRMA_CREATE_PZ_ALLOWED kill-switch ────────────────────────
    # pz_adopt writes wfirma_pz_doc_id to audit.json (permanent state change).
    # It must respect the same global kill-switch as pz_create so that
    # disabling the flag truly prevents all PZ state mutations.
    if not getattr(settings, "wfirma_create_pz_allowed", False):
        return JSONResponse(
            status_code=403,
            content={
                "ok":              False,
                "status":          "blocked",
                "blocking_reasons": [
                    "WFIRMA_CREATE_PZ_ALLOWED is not enabled. "
                    "Set WFIRMA_CREATE_PZ_ALLOWED=true in .env to allow PZ adoption."
                ],
            },
        )

    # ── Guard 1: at least one identifier ─────────────────────────────────────
    if not pz_doc_id_raw and not pz_number_raw:
        return JSONResponse({
            "ok":              False,
            "status":          "blocked",
            "blocking_reasons": ["pz_doc_id or pz_number is required"],
        })

    # ── Guard 2: batch audit must exist ──────────────────────────────────────
    output_dir = get_output_dir(batch_id)
    audit      = _read_audit(output_dir)   # raises HTTPException 404 if missing

    # ── Guard 3: resolve the PZ document in wFirma ───────────────────────────
    if pz_doc_id_raw:
        fetch_result = wfirma_client.fetch_warehouse_pz(pz_doc_id_raw)
    else:
        fetch_result = wfirma_client.find_warehouse_pz_by_number(pz_number_raw)

    if not fetch_result.ok:
        log.warning(
            "[%s] pz_adopt: wFirma lookup failed: %s", batch_id, fetch_result.error,
        )
        return JSONResponse({
            "ok":              False,
            "status":          "blocked",
            "blocking_reasons": [
                f"wFirma document not found or unreachable: {fetch_result.error}"
            ],
        })

    resolved_doc_id = fetch_result.pz_doc_id
    resolved_number = fetch_result.pz_number

    # ── Guard 4: cross-shipment duplicate guard ───────────────────────────────
    owner_batch = _find_pz_owner_batch(resolved_doc_id, exclude_batch_id=batch_id)
    if owner_batch:
        log.warning(
            "[%s] pz_adopt: PZ %s already owned by batch %s",
            batch_id, resolved_doc_id, owner_batch,
        )
        return JSONResponse({
            "ok":              False,
            "status":          "blocked",
            "blocking_reasons": [
                f"wFirma PZ {resolved_doc_id!r} is already linked to shipment "
                f"{owner_batch!r} — cannot adopt the same PZ for two shipments"
            ],
        })

    # ── Guard 5: idempotent fast-path (same PZ id already adopted) ───────────
    wfirma_export      = audit.get("wfirma_export") or {}
    existing_pz_doc_id = (wfirma_export.get("wfirma_pz_doc_id") or "").strip()
    if existing_pz_doc_id == resolved_doc_id:
        return JSONResponse({
            "ok":               True,
            "status":           "already_adopted",
            "batch_id":         batch_id,
            "wfirma_pz_doc_id": existing_pz_doc_id,
            "pz_number":        resolved_number,
        })

    # ── Acquire write lock + full idempotency check + audit write ────────────
    # Inside the lock: re-read audit to catch any state change, run the full
    # guard (doc_id + pz_source + timeline events), then write atomically.
    audit_write_error: Optional[str] = None
    with _pz_write_lock(output_dir, batch_id, "pz_adopt"):
        audit_locked = _read_audit(output_dir)

        # Re-check the same-id fast-path inside the lock (state may have changed)
        relocked_existing = ((audit_locked.get("wfirma_export") or {})
                             .get("wfirma_pz_doc_id") or "").strip()
        if relocked_existing == resolved_doc_id:
            return JSONResponse({
                "ok":               True,
                "status":           "already_adopted",
                "batch_id":         batch_id,
                "wfirma_pz_doc_id": relocked_existing,
                "pz_number":        resolved_number,
            })

        # Different PZ recorded, OR pz_source set, OR terminal timeline event
        # → block.  This is the unified lock that catches:
        #   adopt-after-create, adopt-after-adopt, overwrite of existing doc_id
        #   even if doc_id field was manually removed (timeline still proves it).
        _assert_pz_not_locked(audit_locked, batch_id, "pz_adopt")

        # ── Adopt: write to audit ───────────────────────────────────────────
        audit_write_error = _patch_pz_adopted(output_dir, resolved_doc_id)
        operator = _operator_from_header(x_operator)
        tl.log_event(
            output_dir / "audit.json",
            EV_WFIRMA_PZ_ADOPTED,
            "dashboard",
            "user",
            detail={
                "batch_id":         batch_id,
                "wfirma_pz_doc_id": resolved_doc_id,
                "pz_number":        resolved_number,
                "source":           "adopted_existing",
                "operator":         operator,
            },
        )
        log.info(
            "[%s] pz_adopt: adopted wFirma PZ %s (%s, operator=%s)",
            batch_id, resolved_doc_id, resolved_number, operator,
        )

    if audit_write_error:
        return JSONResponse(
            status_code=500,
            content={
                "ok":               False,
                "batch_id":         batch_id,
                "status":           "audit_write_failed",
                "wfirma_pz_doc_id": resolved_doc_id,
                "pz_number":        resolved_number,
                "warning":          (
                    "wFirma PZ was located but the local audit could not be updated. "
                    "Retry the Confirm Existing PZ action once the disk issue is resolved."
                ),
                "audit_error":      audit_write_error,
            },
        )

    return JSONResponse({
        "ok":              True,
        "status":          "adopted",
        "batch_id":        batch_id,
        "wfirma_pz_doc_id": resolved_doc_id,
        "pz_number":       resolved_number,
        "pz_source":       "adopted_existing",
    })


# ── PZ Document (read-only view) ──────────────────────────────────────────────

def _parse_pz_doc_from_xml(xml_text: str) -> dict:
    """
    Parse a wFirma warehouse_document_p_z find response into a structured dict.

    Returns fields: date, contractor_id, warehouse_id, description, lines.
    Lines contain: good_id, good_name, count, price_netto.
    Any field that cannot be parsed is returned as "" or [].
    """
    try:
        root = ET.fromstring(xml_text)
    except ET.ParseError:
        return {}

    wd = root.find(".//warehouse_document")
    if wd is None:
        return {}

    def _txt(*path):
        node = wd
        for tag in path:
            if node is None:
                return ""
            node = node.find(tag)
        return (node.text or "").strip() if node is not None else ""

    contractor_id = _txt("contractor", "id") or _txt("contractor_id")
    warehouse_id  = _txt("warehouse", "id") or _txt("warehouse_id")
    date          = _txt("date") or _txt("document_date")
    description   = _txt("description") or _txt("number") or ""
    # wFirma's PZ read response carries the canonical PZ number under
    # ``<fullnumber>`` (no underscore), e.g. "PZ 4/5/2026". Bare
    # ``<number>`` is just the per-month sequence ("4") — only useful
    # as a last-resort fallback when neither full-form is present.
    # Query bodies in find/edit calls use ``<full_number>``; that is a
    # SEPARATE namespace and is left untouched in those builders.
    pz_number     = (
        _txt("fullnumber")
        or _txt("full_number")
        or _txt("number")
        or ""
    )

    lines: List[dict] = []
    for content in root.findall(".//warehouse_document_content"):
        def _ctxt(*path):
            node = content
            for tag in path:
                if node is None:
                    return ""
                node = node.find(tag)
            return (node.text or "").strip() if node is not None else ""

        good_id   = _ctxt("good", "id") or _ctxt("good_id") or ""
        good_name = _ctxt("good", "name") or _ctxt("name") or ""
        try:
            count = float(_ctxt("count") or 0)
        except (TypeError, ValueError):
            count = 0.0
        try:
            price = float(_ctxt("price") or _ctxt("price_netto") or 0)
        except (TypeError, ValueError):
            price = 0.0

        lines.append({
            "good_id":     good_id,
            "name":        good_name,
            "count":       count,
            "price_netto": price,
        })

    return {
        "pz_number":     pz_number,
        "date":          date,
        "contractor_id": contractor_id,
        "warehouse_id":  warehouse_id,
        "description":   description,
        "lines":         lines,
    }


@router.get("/shipment/{batch_id}/wfirma/pz_document", dependencies=[_auth])
async def wfirma_pz_document(batch_id: str) -> JSONResponse:
    """
    Read-only view of the linked wFirma PZ document.

    Reads the wfirma_pz_doc_id from audit.json, fetches the document from
    wFirma, and returns structured JSON with header fields and line items.

    No writes are performed.

    Response fields
    ---------------
    pz_doc_id       wFirma internal numeric ID
    pz_number       human-readable document number (full_number)
    date            document date from wFirma
    contractor_id   wFirma contractor (supplier) ID
    warehouse_id    wFirma warehouse ID
    description     document description / notes
    line_count      number of line items
    lines           list of {good_id, name, count, price_netto}
    pz_source       how the PZ was linked (created / adopted_existing)
    raw_xml         raw wFirma XML response (for diagnostics)
    """
    output_dir = get_output_dir(batch_id)
    audit      = _read_audit(output_dir)

    wfirma_export = audit.get("wfirma_export") or {}
    pz_doc_id     = (wfirma_export.get("wfirma_pz_doc_id") or "").strip()

    if not pz_doc_id:
        raise HTTPException(
            status_code=404,
            detail={
                "error": "No wFirma PZ linked to this shipment.",
                "code":  "PZ_NOT_LINKED",
                "batch_id": batch_id,
            },
        )

    fetch = wfirma_client.fetch_warehouse_pz(pz_doc_id)
    if not fetch.ok:
        raise HTTPException(
            status_code=502,
            detail={
                "error":  f"wFirma fetch failed: {fetch.error}",
                "code":   "PZ_FETCH_FAILED",
                "batch_id": batch_id,
                "pz_doc_id": pz_doc_id,
            },
        )

    parsed = _parse_pz_doc_from_xml(fetch.raw_response or "")
    pz_source = (wfirma_export.get("pz_source") or "created").strip()

    log.info(
        "[%s] pz_document: fetched PZ %s (%d lines)",
        batch_id, pz_doc_id, len(parsed.get("lines", [])),
    )

    return JSONResponse({
        "batch_id":     batch_id,
        "pz_doc_id":    pz_doc_id,
        "pz_number":    parsed.get("pz_number") or fetch.pz_number or "",
        "date":         parsed.get("date", ""),
        "contractor_id": parsed.get("contractor_id", ""),
        "warehouse_id": parsed.get("warehouse_id", ""),
        "description":  parsed.get("description", ""),
        "line_count":   len(parsed.get("lines", [])),
        "lines":        parsed.get("lines", []),
        "pz_source":    pz_source,
        "raw_xml":      fetch.raw_response or "",
    })


# ── Refresh canonical PZ mapping (historical-batch backfill) ──────────────
#
# Operator action: re-fetches the wFirma PZ by stored doc_id and stamps
# `wfirma_pz_doc_id` + `wfirma_pz_fullnumber` + `pz_mapped_at` into
# audit.json via audit_persist.record_wfirma_pz_mapping. Used when:
#   - A historical batch was created before auto-mapping landed.
#   - The PZ create flow's auto-fetch failed (network blip) and the
#     fullnumber field is empty.
# Idempotent. Never creates/modifies a wFirma document.

@router.post("/shipment/{batch_id}/wfirma/pz/refresh-mapping",
             dependencies=[_auth])
def pz_refresh_mapping(
    batch_id:    str,
    x_operator:  Optional[str] = Header(None, alias="X-Operator"),
) -> JSONResponse:
    """
    Refresh the canonical wFirma PZ mapping for *batch_id* by reading
    the live wFirma PZ document and stamping its full_number into
    ``audit.wfirma_export.wfirma_pz_fullnumber``.

    Returns 404 if no doc_id is stored yet (operator must first run
    pz_create or pz_adopt). Returns 502 if wFirma read fails.
    """
    output_dir = settings.storage_root / "outputs" / batch_id
    audit_path = output_dir / "audit.json"
    if not audit_path.exists():
        raise HTTPException(status_code=404, detail="audit.json not found")
    try:
        audit = json.loads(audit_path.read_text(encoding="utf-8"))
    except Exception as exc:
        raise HTTPException(status_code=500,
                             detail=f"audit.json unreadable: {exc}")

    wf = audit.get("wfirma_export") or {}
    doc_id = (wf.get("wfirma_pz_doc_id") or "").strip()
    if not doc_id:
        raise HTTPException(
            status_code=404,
            detail="no wfirma_pz_doc_id stored — run pz_create or pz_adopt first",
        )

    fetch = wfirma_client.fetch_warehouse_pz(doc_id)
    if not fetch.ok:
        raise HTTPException(
            status_code=502,
            detail={
                "error":    "wFirma fetch failed",
                "pz_doc_id": doc_id,
                "wfirma":   fetch.error or "unknown",
            },
        )
    fullnum = (fetch.pz_number or "").strip()
    if not fullnum:
        raise HTTPException(
            status_code=502,
            detail={
                "error": "wFirma PZ has no full_number",
                "pz_doc_id": doc_id,
            },
        )

    operator = (x_operator or "").strip()
    from ..services.audit_persist import record_wfirma_pz_mapping
    result = record_wfirma_pz_mapping(
        audit_path,
        wfirma_pz_doc_id     = doc_id,
        wfirma_pz_fullnumber = fullnum,
        source               = "refresh_mapping",
        operator             = operator,
    )
    log.info("[%s] pz_refresh_mapping: %s → %s (changed=%s)",
             batch_id, doc_id, fullnum, result.get("changed"))
    return JSONResponse({
        "ok":                   True,
        "batch_id":             batch_id,
        "wfirma_pz_doc_id":     doc_id,
        "wfirma_pz_fullnumber": fullnum,
        "changed":              bool(result.get("changed")),
        "operator":             operator,
    })
