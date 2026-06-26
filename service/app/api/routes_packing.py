"""
routes_packing.py — Invoice + packing list DB endpoints.

POST /api/v1/packing/{batch_id}/upload
    Upload a packing list PDF or XLSX.
    Extracts rows, matches to invoice lines, stores in DB.
    Optional query param: force_reextract=true

POST /api/v1/packing/{batch_id}/reprocess-prices
    Re-read saved packing files and backfill unit_price_eur for rows where
    the current DB value is 0.  Used to recover prices after PR 2A migration
    (packing uploaded before the unit_price_eur column was added).

GET  /api/v1/packing/{batch_id}
    Return combined invoice lines + packing rows for a batch.

GET  /api/v1/packing/{batch_id}/lines
    Return only packing lines for a batch.
"""
from __future__ import annotations

import json
import os
import re
import socket
import sqlite3
from pathlib import Path
from typing import Any, Dict, List, Optional

from fastapi import APIRouter, Depends, Header, HTTPException, Query, UploadFile
from fastapi.responses import FileResponse, JSONResponse, PlainTextResponse
from pydantic import BaseModel

from ..auth.dependencies import get_current_user
from ..core.config import settings
from ..core import timeline as tl
from ..core.logging import get_logger
from ..services.batch_service import get_output_dir
from ..services import packing_db as pdb
from ..services import document_db as ddb
from ..services import inventory_state_engine as ise
from ..services.invoice_packing_extractor import process_packing_upload, _safe_float


def _pdf_text_preview(path: Path) -> str:
    """Return first ~800 characters of PDF text. Returns '' on any error."""
    try:
        import pdfplumber
        with pdfplumber.open(str(path)) as pdf:
            if not pdf.pages:
                return ""
            return (pdf.pages[0].extract_text() or "")[:800]
    except Exception:
        return ""


def _is_commercial_invoice_text(text: str) -> bool:
    """Return True if text looks like a commercial invoice, not a packing list."""
    t = text.lower()
    return "commercial invoice" in t and "packing list" not in t


def seed_purchase_transit(batch_id: str, line_records: List[Dict[str, Any]]) -> int:
    """
    Best-effort: ensure every freshly inserted/updated packing line has its
    inventory state initialised to PURCHASE_TRANSIT.

    Idempotent: a scan_code that already has any state is skipped, so re-uploads
    do not duplicate state events. Failures are logged and swallowed — they
    must never break the packing-upload flow.

    Returns the number of lines transitioned (0 on any failure).
    """
    seeded = 0
    try:
        for line in line_records:
            try:
                sc = pdb._compute_scan_code(line) or ""
                if not sc:
                    continue
                if ise.get_state(sc) is not None:
                    continue
                # GAP-17: advisory validation — product_code must exist in product_master.
                # Emits an advisory warning to audit.json (NOT a hard block) when absent.
                # The inventory transition still proceeds; the operator sees the advisory in Inbox.
                _pc = str(line.get("product_code") or "")
                if _pc:
                    try:
                        from ..services.reservation_db import validate_product_code_in_master as _gap17_validate
                        from ..core.config import settings as _gap17_settings
                        _gap17_rq = _gap17_settings.storage_root / "reservation_queue.db"
                        if _gap17_rq.exists() and not _gap17_validate(_gap17_rq, _pc):
                            from ..pipelines.pz import _advisory_to_action_proposal, _write_advisory_proposal
                            _gap17_audit_path = (
                                __import__("pathlib", fromlist=["Path"]).Path(
                                    _gap17_settings.storage_root) / "outputs" / batch_id / "audit.json"
                            )
                            _gap17_adv = _advisory_to_action_proposal(
                                {
                                    "code": "GAP17_PRODUCT_NOT_IN_MASTER",
                                    "message": f"product_code {_pc!r} has no product_master row "
                                               f"(GAP-17). Register the product before final PZ.",
                                    "action": "Run product master backfill or register the product.",
                                },
                                batch_id, "packing_upload",
                            )
                            _write_advisory_proposal(_gap17_audit_path, _gap17_adv)
                    except Exception as _gap17_exc:
                        log.debug("[%s] GAP-17 check failed (non-fatal): %s", batch_id, _gap17_exc)
                ise.transition(
                    scan_code    = sc,
                    to_state     = ise.PURCHASE_TRANSIT,
                    product_code = _pc,
                    design_no    = str(line.get("design_no") or ""),
                    batch_id     = batch_id,
                )
                seeded += 1
            except Exception as _row_exc:
                log.warning("[%s] inventory_state seed skipped for one line: %s",
                            batch_id, _row_exc)
                # Best-effort per-line failure mirror — never raises into the loop.
                # Bounded payload: error str truncated to 200 chars.
                try:
                    from ..services.batch_service import get_output_dir as _get_output_dir
                    _audit_path_fail = _get_output_dir(batch_id) / "audit.json"
                    tl.log_event(
                        _audit_path_fail,
                        tl.EV_INVENTORY_TRANSITION_FAILED,
                        trigger_source = "packing_upload",
                        actor          = "system",
                        detail = {
                            "batch_id":   batch_id,
                            "scan_code":  pdb._compute_scan_code(line) or "",
                            "to_state":   "purchase_transit",
                            "error":      str(_row_exc)[:200],
                        },
                    )
                except Exception as _tl_exc:
                    log.warning(
                        "[%s] inventory transition failure mirror failed (non-fatal): %s",
                        batch_id, _tl_exc,
                    )
    except Exception as _outer:
        log.warning("[%s] inventory_state seed best-effort failure: %s",
                    batch_id, _outer)

    # ── Best-effort timeline mirror — never breaks the upload flow ───────────
    # Emits one per-batch summary event regardless of seeded count, mirroring
    # the existing EV_PZ_GENERATED idiom.  log_event is itself non-fatal when
    # audit.json is missing; the outer try/except catches any unrelated failure
    # (e.g. settings or path resolution) so this can never raise into the route.
    try:
        from ..services.batch_service import get_output_dir as _get_output_dir
        _audit_path = _get_output_dir(batch_id) / "audit.json"
        tl.log_event(
            _audit_path,
            tl.EV_INVENTORY_PURCHASE_TRANSIT_SEEDED,
            trigger_source = "packing_upload",
            actor          = "system",
            detail = {
                "batch_id":    batch_id,
                "seeded":      seeded,
                "total_lines": len(line_records),
            },
        )
    except Exception as _tl_exc:
        log.warning("[%s] inventory_state seed mirror event failed (non-fatal): %s",
                    batch_id, _tl_exc)

    return seeded

log = get_logger(__name__)

router = APIRouter(prefix="/api/v1/packing", tags=["packing"])
_auth  = Depends(get_current_user)

_ALLOWED_EXTENSIONS = {".pdf", ".xlsx", ".xls"}
_MAX_BYTES: int     = settings.max_upload_bytes   # 20 MB

# ── Barcode helpers ───────────────────────────────────────────────────────────

_LABEL_SIZE_MM = (57, 32)    # width × height, 300 DPI
_ZPL_DPI       = 300
# 1 mm = 11.811 dots at 300 DPI → round to 12 for clean arithmetic
_DOT_PER_MM    = 11.811


def _fmt_qty(q: Any) -> str:
    """2.0 → '2'  |  1.5 → '1.5'"""
    try:
        f = float(q)
        return str(int(f)) if f == int(f) else str(f)
    except (TypeError, ValueError):
        return str(q)


def _zpl_safe(text: str) -> str:
    """
    Strip characters that corrupt ZPL II output.

    ^ starts a ZPL command — inside ^FD...^FS it terminates the field early.
    ~ triggers ZPL control sequences at any position.
    \\ can cause unexpected escaping on some firmware builds.
    All three must never appear in data fields.
    """
    if not text:
        return ""
    return text.replace("^", "").replace("~", "").replace("\\", "")


def _label_line(ln: Dict[str, Any]) -> str:
    """
    One human-readable summary string an operator can read aloud.
    Format: product_code · bag_id · design_no · metal karat · qty uom
    Empty tokens are omitted so the line stays clean.
    """
    parts = [
        ln.get("product_code", ""),
        ln.get("bag_id", ""),
        ln.get("design_no", ""),
        " ".join(filter(None, [ln.get("metal", ""), ln.get("karat", "")])),
        " ".join(filter(None, [_fmt_qty(ln.get("quantity", 0)), ln.get("uom", "")])),
    ]
    return " · ".join(p for p in parts if p)


def _barcode_value(ln: Dict[str, Any]) -> str:
    """
    Build the scannable barcode value for one physical packing line.

    Uniqueness contract — each physical piece must scan to a distinct value.
    Priority order:
      1. <product_code>|<bag_id>           — when packing list tracks bags
      2. <product_code>|sr<pack_sr>|<design_no> — when source-row Sr is known
                                              (handles aggregated invoices
                                              where many designs share one
                                              product_code)
      3. <product_code>|<design_no>        — fallback when no Sr/bag
      4. <product_code>                    — last resort
    """
    pc      = ln.get("product_code", "")
    bag     = ln.get("bag_id", "")
    sr      = ln.get("pack_sr")
    design  = ln.get("design_no", "")

    if bag:
        return f"{pc}|{bag}"
    if sr is not None:
        # Format Sr as integer when whole, else as-is
        try:
            sr_str = str(int(sr)) if float(sr).is_integer() else str(sr)
        except (TypeError, ValueError):
            sr_str = str(sr)
        if design:
            return f"{pc}|sr{sr_str}|{design}"
        return f"{pc}|sr{sr_str}"
    if design:
        return f"{pc}|{design}"
    return pc


def _build_barcode_row(ln: Dict[str, Any]) -> Dict[str, Any]:
    bv = _barcode_value(ln)
    return {
        "product_code":           ln["product_code"],
        "invoice_no":             ln.get("invoice_no", ""),
        "design_no":              ln.get("design_no", ""),
        "batch_no":               ln.get("batch_no", ""),
        "bag_id":                 ln.get("bag_id", ""),
        "quantity":               _fmt_qty(ln.get("quantity", 0)),
        "uom":                    ln.get("uom", ""),
        "metal":                  ln.get("metal", ""),
        "karat":                  ln.get("karat", ""),
        "barcode_value":          bv,
        "scan_code":              bv,      # canonical search/entry key for all scan flows
        "pack_sr":                ln.get("pack_sr"),
        "unit_price":             ln.get("unit_price"),
        "label_line":             _label_line(ln),
        "requires_manual_review": bool(ln.get("requires_manual_review")),
    }


# ── ZPL II renderer ───────────────────────────────────────────────────────────
# Label: 57 × 32 mm at 300 DPI = 672 × 378 dots
# Margins: ~20 dots (≈1.7 mm) each side

def _render_zpl(row: Dict[str, Any]) -> str:
    """
    Render one barcode label as ZPL II for a Zebra 300 DPI printer.

    Layout (dots):
      y=15   Zone A: product_code (left)  |  bag_id (right)   — 22pt bold
      y=50   Zone B: design_no (left)     |  metal+karat (right) — 18pt
      y=80   Zone C: QTY: qty uom                              — 18pt
      y=105  Zone D: Code 128 barcode, height=80 dots
      (human-readable text printed by ^BC Y flag)
    """
    pc       = _zpl_safe(row.get("product_code", ""))
    bag      = _zpl_safe(row.get("bag_id", ""))
    design   = _zpl_safe(row.get("design_no", ""))
    metal_k  = _zpl_safe(" ".join(filter(None, [row.get("metal", ""), row.get("karat", "")])))
    qty_uom  = _zpl_safe(f"QTY: {row.get('quantity', '')} {row.get('uom', '')}".strip())
    bv       = _zpl_safe(row.get("barcode_value", row.get("product_code", "")))

    return (
        "^XA\n"
        "^CI28\n"                                            # UTF-8
        f"^FO20,15^A0N,22,22^FD{pc}^FS\n"                  # product_code
        f"^FO450,15^A0N,22,22^FD{bag}^FS\n"                # bag_id (right)
        f"^FO20,50^A0N,18,18^FD{design}^FS\n"              # design_no
        f"^FO450,50^A0N,18,18^FD{metal_k}^FS\n"            # metal karat
        f"^FO20,80^A0N,18,18^FD{qty_uom}^FS\n"             # quantity
        "^FO20,105^BY2,3,80\n"                              # barcode params
        "^BCN,80,Y,N,N\n"                                   # Code128, HR below
        f"^FD{bv}^FS\n"                                     # barcode data
        "^XZ\n"
    )


def _render_zpl_batch(rows: List[Dict[str, Any]]) -> str:
    """Concatenate ZPL for multiple labels — one print job."""
    return "".join(_render_zpl(r) for r in rows)


def _validate_batch(batch_id: str) -> Path:
    if "/" in batch_id or ".." in batch_id:
        raise HTTPException(status_code=400, detail="Invalid batch_id.")
    output_dir = get_output_dir(batch_id)
    if not output_dir.exists():
        raise HTTPException(status_code=404, detail=f"Batch {batch_id!r} not found.")
    return output_dir


def _validate_file(file: UploadFile) -> None:
    suffix = Path(file.filename or "").suffix.lower()
    if suffix not in _ALLOWED_EXTENSIONS:
        raise HTTPException(
            status_code=400,
            detail=f"Unsupported file type {suffix!r}. Accepted: PDF, XLSX, XLS.",
        )


# ── POST /api/v1/packing/{batch_id}/upload ────────────────────────────────────

@router.post("/{batch_id}/upload", dependencies=[_auth])
async def upload_packing_list(
    batch_id: str,
    file:            UploadFile,
    force_reextract: bool = Query(default=False),
) -> Dict[str, Any]:
    """
    Upload a packing list (PDF or XLSX), extract rows, match to invoice lines,
    and store in DB.

    - Does not overwrite verified rows unless force_reextract=true.
    - Preserves original file in source/packing/ directory.
    - Logs PACKING_LIST_EXTRACTED and PACKING_MATCHED_TO_INVOICE events.
    """
    output_dir = _validate_batch(batch_id)
    _validate_file(file)

    content = await file.read()
    if len(content) > _MAX_BYTES:
        raise HTTPException(
            status_code=413,
            detail=f"File too large. Max {_MAX_BYTES // (1024*1024)} MB.",
        )

    # Save original file to source/packing/
    packing_dir = output_dir / "source" / "packing"
    packing_dir.mkdir(parents=True, exist_ok=True)
    safe_name = Path(file.filename or "packing_list").name
    safe_name = "".join(c if c.isalnum() or c in "._- " else "_" for c in safe_name)
    dest_path = packing_dir / safe_name
    dest_path.write_bytes(content)

    # Content guard: reject commercial invoices uploaded as packing lists
    if dest_path.suffix.lower() == ".pdf":
        _preview = _pdf_text_preview(dest_path)
        if _is_commercial_invoice_text(_preview):
            raise HTTPException(
                status_code=400,
                detail=(
                    "Document appears to be a commercial invoice, not a packing list. "
                    "Upload it using the Purchase Invoice upload instead."
                ),
            )

    # Run extraction + matching pipeline
    try:
        result = process_packing_upload(
            batch_id         = batch_id,
            batch_output_dir = output_dir,
            packing_file_path= dest_path,
            force_reextract  = force_reextract,
        )
    except Exception as exc:
        raise HTTPException(
            status_code=422,
            detail=f"Packing list extraction failed: {exc}",
        )

    # C13B — client name resolution: filename pattern → body-cell preamble fallback.
    # Priority: (1) filename suffix e.g. "148 Client SUOKKO.xlsx" → "SUOKKO"
    #           (2) preamble cell  e.g. "Client: SUOKKO" in top-12 rows of Excel
    #           (3) "" (neither found — operator assigns manually)
    # Result is injected into parser_diagnostic so the dashboard and diagnostics
    # can show HOW the name was resolved without re-scanning the file.
    _filename_client = _guess_client_from_filename(safe_name)
    _preamble_client = ""
    if not _filename_client:
        _preamble_client = _guess_client_from_preamble(str(dest_path))
    _resolved_client  = _filename_client or _preamble_client
    _resolution_method = (
        "filename" if _filename_client
        else ("preamble" if _preamble_client else "none")
    )
    _cnr = {
        "method":          _resolution_method,
        "client_name":     _resolved_client,
        "filename_guess":  _filename_client,
        "preamble_guess":  _preamble_client,
    }
    result["parser_diagnostic"]["client_name_resolution"] = _cnr
    result["document"]["parser_diagnostic"]["client_name_resolution"] = _cnr

    # Store packing document record
    doc_id = pdb.upsert_packing_document(**result["document"])

    # Build line records
    packing_rows = result["packing_rows"]
    line_records: List[Dict[str, Any]] = []
    for row in packing_rows:
        line_records.append({
            "packing_document_id":  doc_id,
            "batch_id":             batch_id,
            "invoice_no":           row.get("invoice_no", ""),
            "invoice_line_position":row.get("invoice_line_position"),
            "product_code":         row.get("product_code"),
            "design_no":            str(row.get("design_no", "") or ""),
            "batch_no":             str(row.get("batch_no", "") or ""),
            "bag_id":               str(row.get("bag_id", "") or ""),
            "tray_id":              str(row.get("tray_id", "") or ""),
            "item_type":            str(row.get("item_type", "") or ""),
            "uom":                  str(row.get("uom", "") or ""),
            "quantity":             _safe_float(row.get("quantity", 0)),
            "gross_weight":         _safe_float(row.get("gross_weight", 0)),
            "net_weight":           _safe_float(row.get("net_weight", 0)),
            "metal":                str(row.get("metal", "") or ""),
            "karat":                str(row.get("karat", "") or ""),
            "stone_type":           str(row.get("stone_type", "") or ""),
            "remarks":              str(row.get("remarks", "") or ""),
            "extracted_confidence": _safe_float(row.get("extracted_confidence", 0)),
            "requires_manual_review": bool(row.get("requires_manual_review", False)),
            # PR 2A — product identity enrichment from packing XLSX
            # unit_price_eur: client billing price (packing list Value column, EUR
            #   namespace) — distinct from unit_price which carries the supplier
            #   USD rate set at upsert from unit_price field above.
            "unit_price_eur":       _safe_float(row.get("unit_price", 0)),
            # metal_color: standalone color code (W/Y/RG/R) — preserved from the
            #   "Col" column or parsed from combined "14KT/Y" tokens by extractor.
            "metal_color":          str(row.get("metal_color", "") or ""),
            # quality_string: full quality/grade string as-is from the packing
            #   list (may be compound, e.g. "G-VS LAB,E-VVS LAB"). Captures both
            #   the standard "Quality" column and the "Qualtity" typo variant.
            "quality_string":       str(row.get("quality_string", "") or ""),
            # Display fields — canonical names from invoice_packing_extractor
            # field alias map (dia_wt→diamond_weight, col_wt→color_weight, size→size).
            # Previously extracted but not stored; added 2026-06-09.
            # Re-upload or force_reextract=True will populate these for existing batches.
            "size":             str(row.get("size", "") or ""),
            "diamond_weight":   float(row.get("diamond_weight", 0) or 0),
            "color_weight":     float(row.get("color_weight", 0) or 0),
        })

    inserted = pdb.upsert_packing_lines(line_records, force_reextract=force_reextract)

    # CPA: populate product_master from packing rows (non-blocking).
    # Runs after packing_lines write succeeds; failures must not block the upload.
    try:
        from ..services.cpa_product_service import upsert_product_master_from_packing as _cpa_upsert
        _cpa_db = settings.storage_root / "reservation_queue.db"
        _cpa_result = _cpa_upsert(_cpa_db, batch_id, line_records)
        log.info("[%s] CPA upsert: %s upserted, %s skipped, %s errors",
                 batch_id, _cpa_result["upserted_count"],
                 _cpa_result["skipped_count"], _cpa_result["error_count"])
    except Exception as _cpa_exc:
        log.warning("[%s] CPA product_master upsert failed (non-fatal): %s", batch_id, _cpa_exc)

    # Seed inventory state → PURCHASE_TRANSIT for every line with a scan_code.
    # Idempotent on re-upload; failures must not break this route.
    seed_purchase_transit(batch_id, line_records)

    # Global Jewellery: generate Polish descriptions from packing lines (non-blocking).
    # EJL uses descriptions from invoice_lines; Global uses packing_lines directly.
    if result.get("supplier") == "global_jewellery":
        try:
            from ..services.description_engine import regenerate_descriptions_for_packing_lines
            _desc_result = regenerate_descriptions_for_packing_lines(
                batch_id=batch_id, dry_run=False
            )
            log.info("[%s] Global packing descriptions: %s", batch_id, _desc_result)
        except Exception as _desc_exc:
            log.warning("[%s] Global description regen failed (non-fatal): %s",
                        batch_id, _desc_exc)

    # Auto-create / sync proforma drafts from sales_packing_lines (non-blocking).
    # Uses sales_packing_lines (not packing_lines) as the source of truth for
    # client grouping and pricing. A no-op if no sales packing lines exist yet.
    try:
        from ..services.proforma_draft_sync import sync_draft_from_packing_upload
        _pf_db_path = settings.storage_root / "proforma_links.db"
        _sync_result = sync_draft_from_packing_upload(
            batch_id=batch_id,
            operator="packing_upload",
            db_path=_pf_db_path,
            audit_path=output_dir / "audit.json",
            master_db_path=settings.storage_root / "master_data.sqlite",
        )
        log.info("[%s] proforma draft sync: %s", batch_id, _sync_result)
    except Exception as _pf_exc:
        log.warning("[%s] proforma draft sync failed (non-fatal): %s", batch_id, _pf_exc)

    # Register packing file in unified document registry (non-blocking)
    try:
        import hashlib as _hl
        _h = _hl.sha256(content).hexdigest()
        _inv_no = result["document"].get("invoice_no", "")
        ddb.register_document(
            batch_id=batch_id, document_type="packing",
            file_name=safe_name, file_path=str(dest_path),
            file_hash=_h,
            related_invoice_no=_inv_no,
            extraction_status="extracted", source="upload",
        )
    except Exception as _e:
        log.warning("[%s] document_db packing register failed (non-fatal): %s", batch_id, _e)

    # Timeline events
    audit_path = output_dir / "audit.json"
    tl.log_event(
        audit_path,
        tl.EV_PACKING_LIST_EXTRACTED,
        "packing_upload",
        actor="operator",
        detail={
            "batch_id":     batch_id,
            "file":         safe_name,
            "total_rows":   result["total_rows"],
            "document_id":  doc_id,
        },
    )
    tl.log_event(
        audit_path,
        tl.EV_PACKING_MATCHED_TO_INVOICE,
        "packing_upload",
        actor="operator",
        detail={
            "batch_id":       batch_id,
            "matched":        result["matched_count"],
            "unmatched":      result["unmatched_count"],
            "total":          result["total_rows"],
            "force_reextract":force_reextract,
        },
    )

    _parse_warning = None
    if result["total_rows"] == 0:
        _diag = result.get("parser_diagnostic") or {}
        _audit = _diag.get("column_mapping_audit") or []
        _parse_warning = {
            "failure_reason": _diag.get("failure_reason"),
            "unresolved_columns": [
                m.get("raw_header") for m in _audit if m.get("method") == "unresolved"
            ],
            "llm_suggestions": [
                {
                    "header":          m.get("raw_header"),
                    "suggested_field": m.get("canonical_field"),
                    "confidence":      m.get("score"),
                    "reason":          m.get("reason"),
                }
                for m in _audit if m.get("method") == "llm"
            ],
            "llm_auto_triggered": _diag.get("llm_auto_triggered", False),
        }

    return {
        "ok":                   True,
        "batch_id":             batch_id,
        "document_id":          doc_id,
        "file":                 safe_name,
        "suggested_client_name": _resolved_client,         # C13B
        "client_name_resolution": _resolution_method,      # C13B
        "total_rows":           result["total_rows"],
        "matched_count":        result["matched_count"],
        "unmatched_count":  result["unmatched_count"],
        "inserted_count":   inserted,
        "force_reextract":  force_reextract,
        "parse_warning":    _parse_warning,
    }


# ── POST /api/v1/packing/{batch_id}/reprocess-prices ─────────────────────────

@router.post("/{batch_id}/reprocess-prices", dependencies=[_auth])
async def reprocess_packing_prices(batch_id: str) -> Dict[str, Any]:
    """
    Re-read saved packing XLSX files for this batch and backfill unit_price_eur
    in packing_lines where the current value is 0.

    Does NOT overwrite rows that already have unit_price_eur > 0.
    Does NOT re-run the full upload pipeline; only updates the price field.

    Use-case: recover prices after PR 2A migration when packing was uploaded
    before the unit_price_eur column was added (batch uploaded before 2026-05-14).

    Returns:
      { batch_id, diagnostic: { rows_with_price_before, rows_updated, rows_with_price_after },
        files: [{file, rows_extracted, rows_updated}] }
    """
    output_dir = _validate_batch(batch_id)

    packing_dir = output_dir / "source" / "packing"
    if not packing_dir.exists():
        raise HTTPException(
            status_code=404,
            detail=f"No packing source directory for batch {batch_id!r}",
        )

    xlsx_files = sorted(
        list(packing_dir.glob("*.xlsx")) + list(packing_dir.glob("*.xls")),
    )
    if not xlsx_files:
        raise HTTPException(
            status_code=404,
            detail=f"No packing XLSX files found in source directory for batch {batch_id!r}",
        )

    # Snapshot before
    all_lines_before = pdb.get_packing_lines_for_batch(batch_id)
    rows_with_price_before = sum(
        1 for r in all_lines_before if float(r.get("unit_price_eur", 0) or 0) > 0
    )

    file_results: List[Dict[str, Any]] = []
    total_updated = 0

    for pf in xlsx_files:
        file_entry: Dict[str, Any] = {"file": pf.name, "rows_extracted": 0,
                                      "rows_updated": 0, "error": None}
        try:
            result = process_packing_upload(
                batch_id         = batch_id,
                batch_output_dir = output_dir,
                packing_file_path= pf,
                force_reextract  = False,   # read-only extraction; no DB upsert here
            )
            packing_rows = result["packing_rows"]
            file_entry["rows_extracted"] = len(packing_rows)

            # Build line_records carrying unit_price_eur from the Value column
            line_records: List[Dict[str, Any]] = []
            for row in packing_rows:
                upe = float(row.get("unit_price", 0) or 0)
                if upe <= 0:
                    continue  # no Value column data for this row
                doc_id = ""  # we don't upsert documents here; matching is positional
                line_records.append({
                    "packing_document_id":  doc_id,
                    "batch_id":             batch_id,
                    "invoice_no":           row.get("invoice_no", ""),
                    "invoice_line_position":row.get("invoice_line_position"),
                    "product_code":         row.get("product_code"),
                    "design_no":            str(row.get("design_no", "") or ""),
                    "bag_id":               str(row.get("bag_id", "") or ""),
                    "pack_sr":              row.get("pack_sr"),
                    "unit_price_eur":       upe,
                })

            updated = pdb.backfill_unit_price_eur(batch_id, line_records)
            file_entry["rows_updated"] = updated
            total_updated += updated
        except Exception as exc:
            log.warning("[%s] reprocess-prices failed for %s: %s", batch_id, pf.name, exc)
            file_entry["error"] = str(exc)
        file_results.append(file_entry)

    # Snapshot after
    all_lines_after = pdb.get_packing_lines_for_batch(batch_id)
    rows_with_price_after = sum(
        1 for r in all_lines_after if float(r.get("unit_price_eur", 0) or 0) > 0
    )

    log.info(
        "[%s] reprocess-prices: %d rows updated, price coverage %d→%d/%d",
        batch_id, total_updated,
        rows_with_price_before, rows_with_price_after, len(all_lines_after),
    )
    return {
        "ok":      True,
        "batch_id": batch_id,
        "diagnostic": {
            "rows_with_price_before": rows_with_price_before,
            "rows_updated":           total_updated,
            "rows_with_price_after":  rows_with_price_after,
            "total_packing_rows":     len(all_lines_after),
        },
        "files": file_results,
    }


# ── GET /api/v1/packing/{batch_id} ────────────────────────────────────────────

@router.get("/{batch_id}", dependencies=[_auth])
def get_batch_packing(batch_id: str) -> Dict[str, Any]:
    """
    Return combined invoice lines (from pz_rows.json) and packing lines (from DB)
    for a batch.
    """
    output_dir = _validate_batch(batch_id)

    # Invoice lines from engine output
    try:
        from ..services.invoice_packing_extractor import load_invoice_lines
        invoice_lines = load_invoice_lines(output_dir)
    except Exception:
        invoice_lines = []

    # Packing lines from DB
    packing_lines = pdb.get_packing_lines_for_batch(batch_id)
    documents     = pdb.get_packing_documents_for_batch(batch_id)

    # P1 parser observability: decode parser_diagnostic_json on each row.
    # Older rows have '{}' (idempotent ALTER default), so decode succeeds.
    import json as _json
    for d in documents:
        raw = d.pop("parser_diagnostic_json", None) if isinstance(d, dict) else None
        try:
            d["parser_diagnostic"] = _json.loads(raw) if raw else {}
        except Exception:
            d["parser_diagnostic"] = {}

    # Tag every parsed (purchase) doc with side="purchase" and the
    # explicit document_type so the UI can render a Purchase badge
    # alongside the Sales badge from the merge block below.  Purchase
    # vs sales separation stays in the data, not in fragile filename
    # heuristics.  packing_documents rows have no document_type column
    # (the table is purchase-only); set it here for API consistency.
    for d in documents:
        if isinstance(d, dict):
            d.setdefault("side", "purchase")
            d.setdefault("document_type", "purchase_packing_list")

    # ── Fallback visibility: 2026-05-17 hotfix (purchase fallback + sales) ─
    # Two distinct concerns share this block:
    #
    # 1. PURCHASE fallback (gated on parsed `documents` being empty):
    #    Atlas intake writes shipment_documents rows BEFORE running the
    #    best-effort packing extractor. When purchase extraction fails
    #    (unsupported spreadsheet schema, corrupt PDF, etc.) the parsed
    #    `packing_documents` stays empty and the card used to render
    #    "No packing list uploaded yet" even though the file is on disk.
    #
    # 2. SALES surfacing (ALWAYS runs):
    #    Sales packing lives in shipment_documents + sales_packing_lines
    #    only — there is no parsed packing_documents row for sales files.
    #    The card MUST list sales files independently of whether purchase
    #    extraction succeeded.  Row counts come from sales_packing_lines.
    #
    # Purchase vs sales separation is preserved: the `side` and
    # `document_type` fields drive UI badges, never filename heuristics.
    fallback_docs: List[Dict[str, Any]] = []
    try:
        from ..services import document_db as _ddb

        # Build a hash set of already-parsed purchase docs so we never
        # duplicate a file that the parsed pass already covered.
        parsed_purchase_hashes = {
            d.get("source_file_hash") for d in documents
            if isinstance(d, dict) and d.get("source_file_hash")
        }

        # ── Purchase-side fallback (only when no parsed purchase docs) ──
        if not documents:
            purchase_rows = _ddb.get_documents_for_batch(
                batch_id, document_type="purchase_packing_list") or []
            for r in purchase_rows:
                if r.get("file_hash") and r["file_hash"] in parsed_purchase_hashes:
                    continue
                diag: Dict[str, Any] = {}
                try:
                    for pdoc in pdb.get_packing_documents_for_batch(batch_id) or []:
                        if pdoc.get("source_file_hash") and pdoc["source_file_hash"] == r.get("file_hash"):
                            raw = pdoc.pop("parser_diagnostic_json", None)
                            if raw:
                                import json as _json
                                diag = _json.loads(raw) if isinstance(raw, str) else {}
                            break
                except Exception:
                    diag = {}
                fallback_docs.append({
                    "id":                   r.get("id"),
                    "batch_id":             r.get("batch_id"),
                    "document_type":        "purchase_packing_list",
                    "side":                 "purchase",
                    "file_name":            r.get("file_name") or "",
                    "source_file_path":     r.get("file_path") or "",
                    "file_hash":            r.get("file_hash") or "",
                    "parser_status":        r.get("parser_status") or "pending",
                    "extraction_status":    r.get("extraction_status") or "pending",
                    "fallback_unparsed":    True,
                    "row_count":            0,
                    "parser_diagnostic":    diag,
                    "created_at":           r.get("created_at"),
                    "updated_at":           r.get("updated_at"),
                })

        # ── Sales-side surfacing (ALWAYS runs) ──────────────────────────
        # Build a count map sales_document_id → row count from the
        # sales_packing_lines table so each sales file row shows its
        # extracted line count (or 0 when extraction hasn't run yet).
        sales_line_counts: Dict[str, int] = {}
        try:
            for ln in _ddb.get_sales_packing_lines(batch_id) or []:
                sd_id = ln.get("sales_document_id") or ""
                if not sd_id:
                    continue
                sales_line_counts[sd_id] = sales_line_counts.get(sd_id, 0) + 1
        except Exception as exc:
            log.warning("[%s] sales_packing_lines count failed (non-fatal): %s",
                        batch_id, exc)

        # Map shipment_documents.id → sales_documents row (when present)
        # so we can resolve row counts whether sales_document_id was
        # stored as the shipment_documents.id (reprocess fallback path)
        # or as a real sales_documents.id (normal sales-invoice path).
        sd_by_doc_id: Dict[str, Dict[str, Any]] = {}
        try:
            for sd in _ddb.get_sales_documents(batch_id) or []:
                key = sd.get("document_id") or ""
                if key:
                    sd_by_doc_id[key] = sd
        except Exception:
            sd_by_doc_id = {}

        # Pre-decode sales-side parser_diagnostic_json from
        # sales_documents so the Packing List card Diagnostic toggle
        # renders for sales rows symmetric with purchase.  Key is the
        # sales_documents.id (== sales_document_id on lines).
        sales_diag_by_sd_id: Dict[str, Dict[str, Any]] = {}
        try:
            import json as _json2
            for sd in _ddb.get_sales_documents(batch_id) or []:
                raw = sd.get("parser_diagnostic_json")
                try:
                    sales_diag_by_sd_id[sd.get("id") or ""] = (
                        _json2.loads(raw) if raw else {}
                    )
                except Exception:
                    sales_diag_by_sd_id[sd.get("id") or ""] = {}
        except Exception:
            sales_diag_by_sd_id = {}

        sales_rows = _ddb.get_documents_for_batch(
            batch_id, document_type="sales_packing_list") or []
        for r in sales_rows:
            # Defensive: a sales file should never share a hash with a
            # parsed purchase doc, but if it does, prefer the parsed row.
            if r.get("file_hash") and r["file_hash"] in parsed_purchase_hashes:
                continue
            doc_id = r.get("id") or ""
            sd_row = sd_by_doc_id.get(doc_id)
            sd_id = (sd_row or {}).get("id") or doc_id
            row_count = int(sales_line_counts.get(sd_id, 0))
            # Also try shipment_doc.id directly (reprocess fallback path
            # stores sales_document_id = shipment_documents.id).
            if row_count == 0 and doc_id and doc_id in sales_line_counts:
                row_count = int(sales_line_counts[doc_id])
            # When sales_packing_lines has rows for this file, the parse
            # definitively succeeded — surface "extracted" regardless of
            # the stale shipment_documents.extraction_status carried over
            # from intake (which is "pending" by table default and was
            # never updated by the reprocess endpoint on the sales side).
            if row_count > 0:
                parser_status_out     = "extracted"
                extraction_status_out = "extracted"
            else:
                parser_status_out     = r.get("parser_status") or "pending"
                extraction_status_out = r.get("extraction_status") or "pending"
            fallback_docs.append({
                "id":                   doc_id,
                "batch_id":             r.get("batch_id"),
                "document_type":        "sales_packing_list",
                "side":                 "sales",
                "file_name":            r.get("file_name") or "",
                "source_file_path":     r.get("file_path") or "",
                "file_hash":            r.get("file_hash") or "",
                "parser_status":        parser_status_out,
                "extraction_status":    extraction_status_out,
                # fallback_unparsed=False when sales rows are present
                # (we have real extracted lines); True when only the
                # shipment_documents row exists.
                "fallback_unparsed":    row_count == 0,
                "row_count":            row_count,
                # parser_diagnostic surfaced from
                # sales_documents.parser_diagnostic_json (mirrors the
                # purchase-side packing_documents column).  Empty {} when
                # the sales doc hasn't been parsed yet — UI suppresses
                # the Diagnostic toggle in that case, which is the
                # honest behaviour.
                "parser_diagnostic":    (
                    sales_diag_by_sd_id.get(sd_id)
                    or sales_diag_by_sd_id.get(doc_id)
                    or {}
                ),
                "created_at":           r.get("created_at"),
                "updated_at":           r.get("updated_at"),
            })
    except Exception as exc:
        log.warning("[%s] packing fallback enumeration failed (non-fatal): %s",
                    batch_id, exc)

    return {
        "batch_id":      batch_id,
        "invoice_lines": invoice_lines,
        "packing_lines": packing_lines,
        "documents":     documents + fallback_docs,
    }


# ── POST /api/v1/packing/{batch_id}/reprocess ───────────────────────────────
#
# Re-run the safe packing parser against every shipment_documents row of
# type purchase_packing_list / sales_packing_list whose file still
# exists on disk. Useful for:
#   • Batches that pre-date a parser/dependency fix (e.g. .xls files
#     that failed because xlrd was missing at intake time).
#   • Operator-driven re-parse after a vendor template change.
#
# Hard rules:
#   - Per-batch only. No cross-batch fan-out.
#   - Purchase vs sales separation preserved: parser_document_type
#     comes from the shipment_documents row, never inferred from
#     filename or path heuristics.
#   - purchase_packing_list → packing_lines (purchase-side)
#   - sales_packing_list    → sales_packing_lines (sales-side)
#   - No DHL/SAD/PZ/wFirma/proforma execution. Sales-side draft seed
#     uses the SAME helper intake already calls (idempotent).
#   - Parser failures non-fatal — endpoint returns 200 with per-file
#     status; diagnostic artifact written via the existing writer.
#   - Idempotent: hash-dedup in upsert_packing_document avoids
#     duplicates; sales_packing_lines.replace mode in
#     store_sales_packing_lines keeps the set canonical.

class _ReprocessRequest(BaseModel):
    document_id: Optional[str] = None


@router.post("/{batch_id}/reprocess", dependencies=[_auth])
async def reprocess_packing_documents(
    batch_id: str,
    body:     Optional[_ReprocessRequest] = None,
) -> Dict[str, Any]:
    """Re-run the safe packing parser against on-disk packing files.

    Returns:
        {
          "batch_id":  "...",
          "files":     [
            {
              "file_name":           str,
              "document_id":         str,
              "document_type":       "purchase_packing_list" | "sales_packing_list",
              "rows_extracted":      int,
              "parser_status":       str,
              "failure_reason":      str | None,
              "diagnostic_artifact": str | None,
            }, ...
          ],
          "summary": {
            "files":    int,
            "rows":     int,        # sum of rows_extracted
            "purchase": int,
            "sales":    int,
          }
        }
    """
    output_dir = _validate_batch(batch_id)
    audit_path = output_dir / "audit.json"
    if not audit_path.exists():
        raise HTTPException(status_code=404, detail=f"Batch {batch_id} not found.")

    # Pull AWB for register_document writes (existing pattern).
    awb_canonical = ""
    try:
        audit = json.loads(audit_path.read_text(encoding="utf-8"))
        awb_canonical = str(audit.get("awb") or "")
    except Exception:
        pass

    only_doc_id = (body.document_id or "").strip() if body else ""

    from ..services import document_db as _ddb
    from ..services.invoice_packing_extractor import (
        process_packing_upload, extract_packing,
    )
    from ..services.parser_diagnostic_writer import write_packing_diagnostic_artifact
    from .. import services
    from ..services import packing_resolution_db as _prdb  # noqa: F401 (kept for parity)

    # Collect candidate shipment_documents rows.
    candidates: List[Dict[str, Any]] = []
    for dtype in ("purchase_packing_list", "sales_packing_list", "packing"):
        rows = _ddb.get_documents_for_batch(batch_id, document_type=dtype) or []
        for r in rows:
            if only_doc_id and r.get("id") != only_doc_id:
                continue
            candidates.append(r)

    results: List[Dict[str, Any]] = []
    sum_purchase = 0
    sum_sales    = 0

    for row in candidates:
        doc_id        = row.get("id") or ""
        document_type = row.get("document_type") or ""
        file_name     = row.get("file_name") or ""
        file_path_str = row.get("file_path") or ""
        result_entry: Dict[str, Any] = {
            "file_name":           file_name,
            "document_id":         doc_id,
            "document_type":       document_type,
            "rows_extracted":      0,
            "parser_status":       "skipped",
            "failure_reason":      None,
            "diagnostic_artifact": None,
        }

        try:
            file_path = Path(file_path_str) if file_path_str else None
            if not file_path or not file_path.exists():
                result_entry["parser_status"]  = "file_missing"
                result_entry["failure_reason"] = "file_not_found_on_disk"
                results.append(result_entry)
                continue

            # ── Purchase-side: full process_packing_upload pipeline ─────
            if document_type == "purchase_packing_list":
                result = process_packing_upload(
                    batch_id=batch_id,
                    batch_output_dir=output_dir,
                    packing_file_path=file_path,
                    force_reextract=False,
                )
                diag = result.get("parser_diagnostic") or {}
                rows_parsed = result.get("packing_rows", []) or []
                # Persist packing_documents row + lines via the SAME helpers
                # intake uses. Hash-dedup makes this idempotent.
                doc_id_pdb = pdb.upsert_packing_document(**result["document"])
                line_records = [
                    {
                        "packing_document_id":   doc_id_pdb,
                        "batch_id":              batch_id,
                        "invoice_no":            r.get("invoice_no", ""),
                        "invoice_line_position": r.get("invoice_line_position"),
                        "product_code":          r.get("product_code"),
                        "design_no":             str(r.get("design_no", "") or ""),
                        "batch_no":              str(r.get("batch_no", "") or ""),
                        "bag_id":                str(r.get("bag_id", "") or ""),
                        "tray_id":               str(r.get("tray_id", "") or ""),
                        "item_type":             str(r.get("item_type", "") or ""),
                        "uom":                   str(r.get("uom", "") or ""),
                        "quantity":              float(r.get("quantity", 0) or 0),
                        "gross_weight":          float(r.get("gross_weight", 0) or 0),
                        "net_weight":            float(r.get("net_weight", 0) or 0),
                        "metal":                 str(r.get("metal", "") or ""),
                        "karat":                 str(r.get("karat", "") or ""),
                        "stone_type":            str(r.get("stone_type", "") or ""),
                        "remarks":               str(r.get("remarks", "") or ""),
                        "extracted_confidence":  float(r.get("extracted_confidence", 0) or 0),
                        "requires_manual_review": bool(r.get("requires_manual_review", False)),
                        "pack_sr":               r.get("line_position"),
                        "unit_price":            float(r.get("unit_price", 0) or 0),
                        "total_value":           float(r.get("total_value", 0) or 0),
                    }
                    for r in rows_parsed
                ]
                if line_records:
                    pdb.upsert_packing_lines(line_records)
                    seed_purchase_transit(batch_id, line_records)
                    # CPA: populate product_master from reprocessed packing rows (non-blocking).
                    try:
                        from ..services.cpa_product_service import upsert_product_master_from_packing as _cpa_upsert
                        _cpa_db = settings.storage_root / "reservation_queue.db"
                        _cpa_result = _cpa_upsert(_cpa_db, batch_id, line_records)
                        log.info("[%s] CPA reprocess upsert: %s upserted, %s skipped",
                                 batch_id, _cpa_result["upserted_count"], _cpa_result["skipped_count"])
                    except Exception as _cpa_exc:
                        log.warning("[%s] CPA reprocess product_master failed (non-fatal): %s",
                                    batch_id, _cpa_exc)
                result_entry["rows_extracted"] = len(rows_parsed)
                result_entry["parser_status"]  = "extracted" if rows_parsed else "empty"
                result_entry["failure_reason"] = diag.get("failure_reason")
                # Artifact on failure
                if not rows_parsed or diag.get("failure_reason"):
                    art = write_packing_diagnostic_artifact(
                        storage_root=settings.storage_root,
                        batch_id=batch_id, document_id=doc_id,
                        filename=file_name, document_type=document_type,
                        source_path=file_path, parser_diagnostic=diag,
                    )
                    result_entry["diagnostic_artifact"] = str(art) if art else None
                sum_purchase += result_entry["rows_extracted"]

                # Reprocess-parity (RC-1): mirror the sales-side flip below.
                # Purchase packing extraction writes packing.db; the Document
                # Registry reads the shipment_documents column, which stays
                # 'pending' unless flipped here. Success = any rows stored
                # (matches the intake gate). Read-before-write on the downgrade
                # so a transient re-parse failure cannot overwrite a good row.
                try:
                    if line_records:
                        _ddb.update_document_status(
                            doc_id,
                            extraction_status="extracted",
                            parser_status="complete",
                        )
                    else:
                        _cur = (_ddb.get_document(doc_id) or {})
                        if (_cur.get("extraction_status") or "") not in ("extracted", "complete"):
                            _ddb.update_document_status(
                                doc_id,
                                extraction_status="extraction_failed",
                                parser_status="failed",
                                requires_manual_review=True,
                            )
                except Exception as _exc:
                    log.warning(
                        "[%s] reprocess purchase status flip failed (non-fatal): %s",
                        batch_id, _exc,
                    )

            # ── Sales-side: extract_packing + store_sales_packing_lines ─
            elif document_type == "sales_packing_list":
                sp_rows, _, _, sp_diag = extract_packing(file_path)

                # ── PR-3: Sales Packing Matcher ─────────────────────
                # Copy canonical product_code from same-batch purchase
                # packing_lines into the parsed sales rows BEFORE the
                # reshape loop reads r["product_code"].  Batch-scoped
                # only — cross-batch design collisions cannot leak.
                # Existing non-empty product_code on a row is
                # preserved untouched.  Never invents codes; never
                # uses design_no as a fallback.
                from ..services.sales_packing_matcher import (
                    match_sales_lines_to_packing as _match_sales,
                )
                sp_rows, _matcher_summary = _match_sales(batch_id, sp_rows)
                if _matcher_summary.get("designs_resolved") \
                        or _matcher_summary.get("designs_ambiguous") \
                        or _matcher_summary.get("designs_unresolved"):
                    log.info(
                        "[%s] sales matcher: %s",
                        batch_id, _matcher_summary,
                    )
                # Surface matcher buckets on the per-file response so
                # the dashboard can render resolved/ambiguous/unresolved
                # without re-deriving from logs.
                result_entry["sales_matcher_summary"] = _matcher_summary
                # Ensure a sales_documents row whose PRIMARY KEY id == doc_id
                # (the sales packing list's shipment_documents.id). The lines
                # below are keyed to doc_id; wfirma_reservation and the
                # v_sales_to_wfirma view join sales_packing_lines on
                # sales_documents.id, so the sales_documents row MUST carry that
                # same id or every persisted line is orphaned from those
                # readers. The old store_sales_document path minted a divergent
                # random UUID — the root of that orphaning.
                # ensure_sales_document_id is idempotent (re-reprocess reuses the
                # row) and removes pre-fix phantom rows. NEVER triggers external
                # flows. sales_doc_id stays == doc_id so all downstream linkage
                # (client_name backfill, diagnostic write) is unchanged.
                sales_doc_id = doc_id
                try:
                    _ddb.ensure_sales_document_id(
                        batch_id, doc_id,
                        document_type="sales_packing_list",
                        source_file_path=str(file_path),
                    )
                except Exception as _exc:
                    log.warning(
                        "[%s] ensure_sales_document_id failed (non-fatal): %s",
                        batch_id, _exc,
                    )

                # ── Preserve operator-supplied identity across reprocess ──
                # The parser does not know client_name / client_ref — the
                # operator sets these at intake via slot metadata.  The
                # atomic DELETE+INSERT in replace_sales_packing_lines
                # would otherwise wipe them, and downstream
                # sync_draft_from_packing_upload would skip the empty
                # client group → zero proforma drafts.
                #
                # Resolver order (post-contamination fix):
                #   Pass 1: existing sales_packing_lines row scoped to
                #           the same sales_document_id.
                #   Pass 2: same shipment-document linkage —
                #           sales_documents.document_id == doc_id, then
                #           sd.client_name (non-empty) OR
                #           sales_packing_lines whose sales_document_id
                #           is one of those sales_documents.id.
                #   Pass 3: authoritative wfirma reverse lookup via
                #           shipment_documents.client_contractor_id.
                #   Pass 4: filename hint (conservative regex).
                #
                # Removed: unsafe batch-scope fallback that previously
                # accepted ANY non-empty client_name from any row in
                # the batch — this leaked cross-document contamination
                # (e.g. stray link_as_sales rows polluting reprocessed
                # rows from a different shipment_document).
                preserved_client_name = ""
                preserved_client_ref  = ""
                existing_rows = []
                try:
                    existing_rows = _ddb.get_sales_packing_lines(batch_id) or []
                    # Pass 1: prefer rows scoped to the same sales_doc_id.
                    for er in existing_rows:
                        if er.get("sales_document_id") != sales_doc_id:
                            continue
                        if er.get("client_name") and not preserved_client_name:
                            preserved_client_name = er["client_name"]
                        if er.get("client_ref") and not preserved_client_ref:
                            preserved_client_ref = er["client_ref"]
                        if preserved_client_name and preserved_client_ref:
                            break

                    # Pass 2: same shipment-document linkage. Scope to
                    # sales_documents whose document_id == current
                    # shipment_documents.id (doc_id). Try sd.client_name
                    # first, then sales_packing_lines linked to those sd.id.
                    if not preserved_client_name:
                        try:
                            linked_sds = (
                                _ddb.get_sales_documents_for_shipment_doc(doc_id)
                                or []
                            )
                        except Exception:
                            linked_sds = []
                        for sd in linked_sds:
                            cn = (sd.get("client_name") or "").strip()
                            if cn:
                                preserved_client_name = cn
                                log.info(
                                    "[%s] sales reprocess: client_name "
                                    "recovered via sales_documents."
                                    "document_id==%s -> %r (Pass 2a)",
                                    batch_id, doc_id, preserved_client_name,
                                )
                                break
                        if not preserved_client_name and linked_sds:
                            linked_sd_ids = {sd["id"] for sd in linked_sds
                                             if sd.get("id")}
                            for er in existing_rows:
                                if (er.get("sales_document_id") in linked_sd_ids
                                        and er.get("client_name")):
                                    preserved_client_name = er["client_name"]
                                    log.info(
                                        "[%s] sales reprocess: client_name "
                                        "recovered via sales_packing_lines "
                                        "linked to shipment_doc=%s -> %r "
                                        "(Pass 2b)",
                                        batch_id, doc_id, preserved_client_name,
                                    )
                                    break

                    # client_ref: same scoping as client_name. Pass 1
                    # same sales_doc_id only; Pass 2 same linkage. No
                    # batch-wide fallback.
                    if not preserved_client_ref:
                        try:
                            linked_sd_ids2 = {
                                sd["id"]
                                for sd in (_ddb.get_sales_documents_for_shipment_doc(doc_id) or [])
                                if sd.get("id")
                            }
                        except Exception:
                            linked_sd_ids2 = set()
                        for er in existing_rows:
                            if (er.get("sales_document_id") in linked_sd_ids2
                                    and er.get("client_ref")):
                                preserved_client_ref = er["client_ref"]
                                break
                except Exception as exc:
                    log.warning("[%s] sales reprocess: client preservation "
                                "lookup failed (non-fatal): %s", batch_id, exc)

                # Pass 3 — self-healing resolver via shipment_documents.
                # client_contractor_id → wfirma_customers reverse lookup.
                # Local DB only; NEVER calls the wFirma API.  Fires only
                # when Pass 1+2 produced empty client_name (e.g. batches
                # corrupted by pre-PR-#187 reprocess that wiped client
                # metadata from sales_packing_lines).
                if not preserved_client_name:
                    try:
                        ccid = (row.get("client_contractor_id") or "").strip()
                        if ccid:
                            from ..services import wfirma_db as _wfdb
                            cust = _wfdb.get_customer_by_wfirma_id(ccid)
                            if cust and (cust.get("client_name") or "").strip():
                                preserved_client_name = cust["client_name"]
                                log.info(
                                    "[%s] sales reprocess: client_name "
                                    "recovered via wfirma_customers[%s] "
                                    "-> %r (Pass 3)",
                                    batch_id, ccid, preserved_client_name,
                                )
                    except Exception as exc:
                        log.warning(
                            "[%s] sales reprocess: Pass-3 wfirma reverse "
                            "lookup failed (non-fatal): %s", batch_id, exc,
                        )

                # Pass 4 — conservative filename hint as last resort.
                # Reuses the existing _guess_client_from_filename regex
                # which requires a numeric prefix or dash separator
                # (never bare "Client X.xlsx").  Every hit logged so the
                # operator can audit inferred client names.
                if not preserved_client_name:
                    try:
                        from_file = _guess_client_from_filename(file_name)
                        if from_file:
                            preserved_client_name = from_file
                            log.info(
                                "[%s] sales reprocess: client_name "
                                "recovered from filename %r -> %r "
                                "(Pass 4 — best-effort)",
                                batch_id, file_name, preserved_client_name,
                            )
                    except Exception:
                        pass

                # Pass 5 — body-cell preamble fallback (C13B).
                # Handles the orphan pattern where the filename ends with
                # "-Client.xlsx" (no actual name after the keyword), so
                # _guess_client_from_filename returns "".  Opens the saved
                # Excel file and scans top-12 rows for "Client:" /
                # "Consignee:" / "Buyer:" / "Ship To:" labels.
                # Returns "" on any failure — never blocks the reprocess.
                if not preserved_client_name and file_path and file_path.exists():
                    try:
                        from_preamble = _guess_client_from_preamble(str(file_path))
                        if from_preamble:
                            preserved_client_name = from_preamble
                            log.info(
                                "[%s] sales reprocess: client_name "
                                "recovered from body preamble file=%r -> %r "
                                "(Pass 5 — body-cell fallback)",
                                batch_id, file_name, preserved_client_name,
                            )
                    except Exception:
                        pass

                if not preserved_client_name:
                    log.warning(
                        "[%s] sales reprocess: NO client_name resolvable "
                        "for sales_doc=%s file=%r — drafts will skip "
                        "this group",
                        batch_id, sales_doc_id, file_name,
                    )
                    # PR 1 (visibility) — emit a draft-birth skip event so the
                    # all-fail resolver branch becomes auditable. Observation
                    # only: does not influence the reprocess outcome.
                    try:
                        from ..services import preamble_signals as _ps
                        _signals = (
                            _ps.extract_all_signals(file_path)
                            if (file_path and file_path.exists())
                            else {"vat": None, "heading_candidate": None}
                        )
                        _has_signal = bool(_signals.get("vat") or _signals.get("heading_candidate"))
                        if _signals.get("vat"):
                            _next_action = "vat_resolver_will_auto_bind_post_pr2"
                        elif _signals.get("heading_candidate"):
                            _next_action = "heading_candidate_requires_corroboration"
                        else:
                            _next_action = "operator_bind_client_name_manually"
                        tl.log_event(
                            audit_path,
                            (tl.EV_PROFORMA_DRAFT_CREATION_PENDING_RESOLUTION
                             if _has_signal
                             else tl.EV_PROFORMA_DRAFT_CREATION_SKIPPED),
                            trigger_source="packing_reprocess",
                            actor="system",
                            detail={
                                "batch_id":                 batch_id,
                                "sales_doc_id":             sales_doc_id,
                                "source_file_path":         str(file_path) if file_path else "",
                                "file_name":                file_name,
                                "reason":                   "client_name_unresolved_all_passes",
                                "resolver_signals_seen":    _signals,
                                "resolver_passes_attempted": ["pass1_sales_doc_scope",
                                                              "pass2_shipment_doc_linkage",
                                                              "pass3_wfirma_reverse_lookup",
                                                              "pass4_filename_hint",
                                                              "pass5_preamble_label_scan"],
                                "next_action":              _next_action,
                            },
                        )
                    except Exception as _exc:
                        log.warning(
                            "[%s] sales reprocess: skip-event emission "
                            "failed (non-fatal): %s", batch_id, _exc,
                        )

                # Reshape parser rows → sales_packing_lines schema.
                line_records = []
                for r in sp_rows:
                    # Quantity: replace_sales_packing_lines reads
                    # ln.get("quantity"); some parser variants emit "qty"
                    # instead.  Accept both so the column never
                    # silently zeroes.
                    qty_val = r.get("quantity")
                    if qty_val is None:
                        qty_val = r.get("qty", 0)
                    # Product code: defensive coerce of None → "" so the
                    # column never stores the literal "None" string.
                    pc_val = r.get("product_code")
                    line_records.append({
                        "batch_id":              batch_id,
                        "sales_document_id":     sales_doc_id,
                        # Operator-supplied identity, preserved across reprocess.
                        "client_name":           preserved_client_name,
                        "client_ref":            preserved_client_ref,
                        "invoice_no":            r.get("invoice_no", ""),
                        "design_no":             str(r.get("design_no", "") or ""),
                        "bag_id":                str(r.get("bag_id", "") or ""),
                        "product_code":          pc_val if pc_val is not None else "",
                        "quantity":              float(qty_val or 0),
                        "unit_price":            float(r.get("unit_price", 0) or 0),
                        "currency":              (r.get("currency") or "").upper(),
                        "total_value":           float(r.get("total_value", 0) or 0),
                        "price_source":          r.get("price_source") or r.get("currency_source") or "",
                        "client_po":             r.get("client_po") or "",
                        "remarks":               str(r.get("remarks", "") or ""),
                    })
                if line_records:
                    try:
                        _ddb.replace_sales_packing_lines(sales_doc_id, batch_id, line_records)
                    except AttributeError:
                        # Helper name varies between writers; fall back.
                        _ddb.store_sales_packing_lines(sales_doc_id, batch_id, line_records)
                    # Reprocess-parity: the sales branch persists rows but the
                    # intake-time shipment_documents.extraction_status stays
                    # 'pending' unless flipped here. The packing card infers
                    # 'extracted' from row count, but the Document Registry
                    # reads the raw column — flip it so both agree. Gate on real
                    # content (a product_code OR design_no on at least one row)
                    # so a parse that yielded only blank rows is not mislabelled
                    # 'extracted/complete'. Non-fatal.
                    _has_content = any(
                        (r.get("product_code") or "").strip()
                        or (r.get("design_no") or "").strip()
                        for r in line_records
                    )
                    if _has_content:
                        try:
                            _ddb.update_document_status(
                                doc_id,
                                extraction_status="extracted",
                                parser_status="complete",
                            )
                        except Exception as _exc:
                            log.warning(
                                "[%s] reprocess sales status flip failed (non-fatal): %s",
                                batch_id, _exc,
                            )

                # Canonical description authority: pre-populate product_descriptions
                # using description_engine.get_description_block() so the item type's
                # Polish translation is seeded before the proforma draft is born.
                # Sales-packing Ctg/Kt/Col/Quality codes are NOT used — they produce
                # short item-type-code names that diverge from the customs PDF
                # (Lesson N / single-authority rule).
                #
                # description_en is intentionally NOT populated here.
                # invoice_lines.description for EJL products is supplier shorthand
                # (e.g. "PCS, 14KT Gold, LGD Stud Jewellery RING") — not a
                # customs-grade English sentence. Writing shorthand as description_en
                # would cause the renderer to produce:
                #   "{customs_pl} / PCS, 14KT Gold, ..."
                # which violates the canonical authority rule.
                # description_en is left blank; build_description_line() outputs PL
                # only until a verified customs-grade English sentence is explicitly
                # provided by the operator or a dedicated English-authority pipeline.
                try:
                    from ..services.description_engine import (
                        get_description_block as _get_desc_block,
                    )
                    _seen_desc_pcs: set = set()
                    for _sr in sp_rows:
                        _pc  = (_sr.get("product_code") or "").strip()
                        if not _pc or _pc in _seen_desc_pcs:
                            continue
                        _seen_desc_pcs.add(_pc)
                        _ctg = (_sr.get("item_type") or "").strip().upper()
                        try:
                            _get_desc_block(_pc, _ctg)
                        except Exception as _block_exc:
                            log.warning(
                                "[%s] sales packing: get_description_block failed "
                                "for %s: %s",
                                batch_id, _pc, _block_exc,
                            )
                except Exception as _desc_exc:
                    log.warning(
                        "[%s] sales packing: description pre-population "
                        "failed (non-fatal): %s",
                        batch_id, _desc_exc,
                    )

                # Backfill sales_documents.client_name when empty.
                # Cosmetic/visual consistency with sales_packing_lines
                # after self-healing recovery (Pass 3/4).  Local-DB only.
                # Linkage: reprocess uses shipment_documents.id as
                # sales_doc_id, but sales_documents.id is a fresh uuid
                # generated by store_sales_document — so match by
                # sales_documents.document_id (the back-reference column).
                if preserved_client_name:
                    try:
                        for sd in (_ddb.get_sales_documents(batch_id) or []):
                            if (sd.get("document_id") == sales_doc_id
                                and not (sd.get("client_name") or "").strip()):
                                _ddb.update_sales_document_client_name(
                                    sd["id"], preserved_client_name)
                    except Exception as exc:
                        log.warning(
                            "[%s] sales reprocess: sales_documents "
                            "client_name backfill failed (non-fatal): %s",
                            batch_id, exc,
                        )

                # Persist parser_diagnostic on sales_documents so the
                # Packing List card Diagnostic toggle renders for sales
                # rows symmetric with purchase.  Always written (success
                # OR failure) — observability symmetry with purchase
                # side (packing_documents.parser_diagnostic_json).
                try:
                    _ddb.update_sales_document_parser_diagnostic(
                        sales_doc_id, sp_diag or {})
                except Exception as exc:
                    log.warning("[%s] sales parser_diagnostic persist failed "
                                "(non-fatal): %s", batch_id, exc)
                result_entry["rows_extracted"] = len(sp_rows)
                result_entry["parser_status"]  = "extracted" if sp_rows else "empty"
                result_entry["failure_reason"] = sp_diag.get("failure_reason")
                if not sp_rows or sp_diag.get("failure_reason"):
                    art = write_packing_diagnostic_artifact(
                        storage_root=settings.storage_root,
                        batch_id=batch_id, document_id=doc_id,
                        filename=file_name, document_type=document_type,
                        source_path=file_path, parser_diagnostic=sp_diag or {},
                    )
                    result_entry["diagnostic_artifact"] = str(art) if art else None
                sum_sales += result_entry["rows_extracted"]

            # ── Packing (xlsx/xls Client): diagnostic-only refresh ───────
            elif document_type == "packing":
                # These documents carry parser_diagnostic_json but are NOT
                # purchase or sales packing lines — they are client-facing
                # xlsx/xls files uploaded alongside the purchase Poland xls.
                # Refresh column_mapping_audit and all observability fields
                # so the column-mapping UI shows current data for files that
                # pre-date PR #524 (which added _map_headers_with_audit).
                # Hard safety rules:
                #   - NO packing_lines, sales records, or wFirma records modified.
                #   - NO business outputs changed.
                #   - Writes ONLY parser_diagnostic_json in packing_documents.
                #   - Idempotent: re-running produces the same diagnostic.
                _, _, _, packing_diag = extract_packing(file_path)

                # Locate the packing_documents row by batch + source path.
                # (packing_documents has no document_type column — match via
                # source_file_path, set at intake time to the same on-disk path.)
                pdb_docs = pdb.get_packing_documents_for_batch(batch_id)
                pdb_doc  = next(
                    (d for d in pdb_docs
                     if d.get("source_file_path") == str(file_path)),
                    None,
                )
                if pdb_doc:
                    pdb.update_packing_document_diagnostic(pdb_doc["id"], packing_diag)
                    result_entry["document_id"] = pdb_doc["id"]

                audit_count = len(packing_diag.get("column_mapping_audit") or [])
                result_entry["rows_extracted"]             = 0
                result_entry["parser_status"]              = "diagnostic_refreshed"
                result_entry["failure_reason"]             = packing_diag.get("failure_reason")
                result_entry["column_mapping_audit_count"] = audit_count
                log.info(
                    "[%s] packing diagnostic refresh: file=%s audit_entries=%d",
                    batch_id, file_name, audit_count,
                )

            else:
                result_entry["parser_status"]  = "skipped_unsupported_type"
                result_entry["failure_reason"] = "unsupported_document_type"

        except Exception as exc:
            log.warning("[%s] reprocess failed (non-fatal) for %s: %s",
                        batch_id, file_name, exc)
            result_entry["parser_status"]  = "extraction_failed"
            result_entry["failure_reason"] = type(exc).__name__

        results.append(result_entry)

    # Auto-sync proforma drafts after reprocess.  Fires once per call
    # (not per file).  Only when batch carries sales rows or sales
    # shipment_documents — otherwise no-op.  Non-blocking: reprocess
    # response is never affected by sync failure.
    try:
        has_sales = False
        try:
            has_sales = bool(_ddb.get_sales_packing_lines(batch_id))
        except Exception:
            has_sales = False
        if not has_sales:
            try:
                has_sales = bool(
                    _ddb.get_documents_for_batch(
                        batch_id, document_type="sales_packing_list") or []
                )
            except Exception:
                has_sales = False
        if has_sales:
            from ..services.proforma_draft_sync import sync_draft_from_packing_upload
            _pf_db_path = settings.storage_root / "proforma_links.db"
            _sync_result = sync_draft_from_packing_upload(
                batch_id=batch_id,
                operator="reprocess",
                db_path=_pf_db_path,
                audit_path=output_dir / "audit.json",
                master_db_path=settings.storage_root / "master_data.sqlite",
            )
            log.info("[%s] reprocess proforma draft sync: %s",
                     batch_id, _sync_result)
    except Exception as _exc:
        log.warning("[%s] reprocess proforma draft sync failed (non-fatal): %s",
                    batch_id, _exc)

    return {
        "batch_id": batch_id,
        "files":    results,
        "summary": {
            "files":    len(results),
            "rows":     sum_purchase + sum_sales,
            "purchase": sum_purchase,
            "sales":    sum_sales,
        },
    }


# ── POST /api/v1/packing/{batch_id}/suggest-column-mapping ───────────────────
#
# Re-runs Excel column-header mapping with the LLM advisory tier enabled for
# ONE packing document chosen by the operator.
#
# Hard constraints (safety gates — immutable):
#   1. Writes ONLY packing_documents.parser_diagnostic_json.
#   2. Does NOT write packing_lines, products, customers, PZ, or wFirma records.
#   3. LLM output is advisory.  It does not enter build_col_map and cannot be
#      used to create or mutate any business entity.
#   4. Scoped to one document.  No cross-batch or cross-document side effects.
#   5. Only applicable to Excel files (.xlsx / .xls).
#   6. Requires explicit operator POST — never fires on a regular upload.

class _SuggestColumnMappingRequest(BaseModel):
    document_id: str


@router.post("/{batch_id}/suggest-column-mapping", dependencies=[_auth])
async def suggest_column_mapping(
    batch_id: str,
    body: _SuggestColumnMappingRequest,
) -> Dict[str, Any]:
    """Re-run Excel column mapping with LLM advisory tier for one document.

    Writes ONLY parser_diagnostic_json.  No business records are created or
    mutated.  LLM output is advisory only and is never included in the
    extraction col_map used to write packing rows.
    """
    import datetime as _dt

    output_dir = _validate_batch(batch_id)

    doc_id = (body.document_id or "").strip()
    if not doc_id:
        raise HTTPException(status_code=400, detail="document_id is required.")

    doc = pdb.get_packing_document(doc_id)
    if not doc or doc.get("batch_id") != batch_id:
        raise HTTPException(
            status_code=404,
            detail=f"Packing document {doc_id!r} not found for batch {batch_id!r}.",
        )

    file_path_str = doc.get("source_file_path") or ""
    file_path = Path(file_path_str) if file_path_str else None
    if not file_path or not file_path.exists():
        raise HTTPException(
            status_code=404,
            detail=f"Source file not found on disk for document {doc_id!r}.",
        )

    suffix = file_path.suffix.lower()
    if suffix not in (".xlsx", ".xls"):
        raise HTTPException(
            status_code=400,
            detail=(
                f"LLM column mapping applies to Excel files only (.xlsx / .xls). "
                f"Document suffix: {suffix!r}."
            ),
        )

    from ..services.invoice_packing_extractor import extract_packing

    # Re-run extraction with llm_fallback=True.
    # Extracted rows are DISCARDED — only the diagnostic (column_mapping_audit)
    # is kept and written back to the document record.
    _, _, _, diag = extract_packing(file_path, llm_fallback=True)

    diag["llm_mapping_meta"] = {
        "triggered_by": "operator",
        "triggered_at": _dt.datetime.utcnow().isoformat() + "Z",
        "advisory_only": True,
        "document_id":   doc_id,
        "file_name":     file_path.name,
    }

    # Write ONLY parser_diagnostic_json — no rows, no lines, no business records.
    pdb.update_packing_document_diagnostic(doc_id, diag)

    audit_entries = diag.get("column_mapping_audit") or []
    summary = {
        "total_columns":  len(audit_entries),
        "alias":          sum(1 for m in audit_entries if m.get("method") == "alias"),
        "fuzzy":          sum(1 for m in audit_entries if m.get("method") == "fuzzy"),
        "fuzzy_warning":  sum(1 for m in audit_entries if m.get("method") == "fuzzy_warning"),
        "llm":            sum(1 for m in audit_entries if m.get("method") == "llm"),
        "unresolved":     sum(1 for m in audit_entries if m.get("method") == "unresolved"),
    }

    log.info(
        "[%s] suggest-column-mapping: doc=%s summary=%s",
        batch_id, doc_id, summary,
    )

    tl.log_event(
        output_dir / "audit.json",
        tl.EV_COLUMN_MAPPING_LLM_REQUESTED,
        trigger_source="suggest_column_mapping",
        actor="operator",
        detail={
            "batch_id":    batch_id,
            "document_id": doc_id,
            "file_name":   file_path.name,
            "advisory_only":  True,
            "write_scope":    "parser_diagnostic_json_only",
            "llm_fallback":   True,
            "summary":        summary,
        },
    )

    return {
        "ok":               True,
        "batch_id":         batch_id,
        "document_id":      doc_id,
        "llm_mapping_meta": diag["llm_mapping_meta"],
        "column_mapping_summary": summary,
    }


# ── POST /api/v1/packing/{batch_id}/approve-header-mapping ───────────────────
#
# Operator explicitly approves one or more header→field mappings and saves them
# as supplier-specific templates (Tier 0). Requires the packing document to have
# a supplier_id set (from intake dropdown selection).
#
# Safety contract:
#   1. Read-only except for supplier_header_templates — no packing_lines, no
#      products, no PZ, no wFirma, no invoice writes.
#   2. supplier_id MUST come from the packing document (operator-selected via
#      intake dropdown). Never inferred from file content or LLM output.
#   3. LLM suggestions are NEVER auto-saved. Only explicit operator POST saves.
#   4. Operator may provide a doc_type override (default: purchase_packing_list).

class _HeaderMappingItem(BaseModel):
    raw_header: str
    canonical_field: str
    col_index: Optional[int] = None
    source_method: Optional[str] = None   # alias/fuzzy/fuzzy_warning/llm
    operator_confirmed: bool = False       # must be True to persist llm-sourced rows


class _ApproveHeaderMappingBody(BaseModel):
    document_id: str
    mappings: List[_HeaderMappingItem]
    doc_type: str = "purchase_packing_list"


@router.post("/{batch_id}/approve-header-mapping", dependencies=[_auth])
async def approve_header_mapping(
    batch_id: str,
    body: _ApproveHeaderMappingBody,
    x_operator: Optional[str] = Header(None, alias="X-Operator"),
) -> Dict[str, Any]:
    """Save operator-approved header→field mappings as supplier templates (Tier 0).

    The packing document must have a supplier_id (set at intake when the operator
    selected a supplier from the Supplier Master dropdown). Mappings are stored in
    supplier_header_templates and applied automatically on the next upload from
    the same supplier.

    Returns: approved_count, supplier_id, doc_type, saved mappings.
    """
    from ..services.excel_column_mapper import CANONICAL_FIELDS

    doc_id = (body.document_id or "").strip()
    if not doc_id:
        raise HTTPException(status_code=400, detail="document_id is required.")
    if not body.mappings:
        raise HTTPException(status_code=400, detail="mappings list must not be empty.")

    doc = pdb.get_packing_document(doc_id)
    if not doc or doc.get("batch_id") != batch_id:
        raise HTTPException(
            status_code=404,
            detail=f"Packing document {doc_id!r} not found for batch {batch_id!r}.",
        )

    supplier_id = doc.get("supplier_id")
    if not supplier_id:
        raise HTTPException(
            status_code=422,
            detail=(
                "This packing document has no supplier_id. "
                "Upload via the intake form with a supplier selected from the "
                "Supplier Master dropdown to enable template learning."
            ),
        )

    operator = (x_operator or "operator").strip()
    doc_type  = (body.doc_type or "purchase_packing_list").strip()

    saved: List[Dict[str, Any]] = []
    rejected: List[Dict[str, Any]] = []

    for item in body.mappings:
        raw    = (item.raw_header or "").strip()
        field  = (item.canonical_field or "").strip()

        if not raw:
            rejected.append({"raw_header": raw, "reason": "empty header"})
            continue
        if field not in CANONICAL_FIELDS:
            rejected.append({
                "raw_header": raw,
                "reason": f"unknown canonical field: {field!r}",
            })
            continue
        # LLM-sourced items require explicit per-item operator confirmation.
        # Sending operator_confirmed=false (the default) is a hard block — the
        # endpoint never auto-promotes AI suggestions into templates.
        if item.source_method == "llm" and not item.operator_confirmed:
            rejected.append({
                "raw_header": raw,
                "reason": (
                    "LLM-sourced mapping requires operator_confirmed=true; "
                    "AI suggestions are never saved automatically."
                ),
            })
            continue

        src = (item.source_method or "operator_approved").strip() or "operator_approved"
        try:
            template_id = pdb.upsert_supplier_template(
                supplier_id     = supplier_id,
                doc_type        = doc_type,
                raw_header      = raw,
                canonical_field = field,
                col_index       = item.col_index,
                approved_by     = operator,
                source_method   = src,
            )
            saved.append({
                "template_id":    template_id,
                "raw_header":     raw,
                "canonical_field": field,
                "source_method":  src,
            })
        except Exception as exc:
            log.warning(
                "[%s] approve-header-mapping save failed: %s → %s: %s",
                batch_id, raw, field, exc,
            )
            rejected.append({"raw_header": raw, "reason": str(exc)})

    log.info(
        "[%s] approve-header-mapping: supplier_id=%s doc_type=%s saved=%d rejected=%d",
        batch_id, supplier_id, doc_type, len(saved), len(rejected),
    )

    return {
        "ok":             True,
        "batch_id":       batch_id,
        "document_id":    doc_id,
        "supplier_id":    supplier_id,
        "doc_type":       doc_type,
        "approved_count": len(saved),
        "rejected_count": len(rejected),
        "saved":          saved,
        "rejected":       rejected,
    }


# ── GET /api/v1/packing/{batch_id}/lane-readiness ────────────────────────────
#
# Read-only operator dashboard. Aggregates lane-level readiness from
# existing tables — no new persistence.  Sales lane: proforma_drafts
# counts by draft_state.  Purchase lane: packing_lines × wfirma_products
# cache + audit.json SAD presence.  Reason taxonomy is enumerated and
# closed; no freeform strings.
#
# Hard rules: NO writes, NO outbound HTTP, NO email/SMTP, NO wFirma
# calls.  All reads wrapped — endpoint returns 200 with degraded data
# (zeros + safe defaults) on any internal failure.

# Closed enum — never expand without API contract bump.
_PZ_BLOCKED_NO_PACKING_ROWS         = "no_packing_rows"
_PZ_BLOCKED_PRODUCTS_MISSING        = "products_missing"
_PZ_BLOCKED_SAD_MISSING             = "sad_missing"
_PZ_BLOCKED_PRODUCT_MASTER_MISSING  = "product_master_missing"      # PR-5
_PZ_BLOCKED_PURCHASE_INVOICE_MISSING = "purchase_invoice_missing"   # PR-8

# Sales blocked_by — new closed enum introduced in PR-5.  Operator
# dashboard derives a reason list parallel to pz_blocked_by.
_SALES_BLOCKED_NO_DRAFTS                = "no_drafts"
_SALES_BLOCKED_DRAFTS_HAVE_NO_LINES     = "drafts_have_no_lines"
_SALES_BLOCKED_WFIRMA_PRODUCTS_MISSING  = "wfirma_products_missing"
_SALES_BLOCKED_PRODUCT_MASTER_MISSING   = "product_master_missing"
_SALES_BLOCKED_POST_FAILED              = "post_failed"
_SALES_BLOCKED_PURCHASE_INVOICE_MISSING = "purchase_invoice_missing"  # PR-8


# PR-8: regex used to derive an invoice_no anchor from sales packing
# filenames. Matches the leading `EJL-YY-YY-NNN` prefix that intake
# uses for every sales packing file. Non-matching filenames are
# ignored — no false positives.
_SALES_FILENAME_INVOICE_RE = re.compile(
    r"^EJL-(\d+)-(\d+)-(\d+)", re.IGNORECASE,
)


@router.get("/{batch_id}/lane-readiness", dependencies=[_auth])
def get_lane_readiness(batch_id: str) -> Dict[str, Any]:
    """Aggregate lane readiness (sales + purchase) for the Documents tab.

    Returns the shape documented in service/docs/lane_readiness.md and
    contract-tested in tests/test_lane_readiness_endpoint.py.
    """
    _validate_batch(batch_id)
    output_dir = get_output_dir(batch_id)

    # ── Sales lane ───────────────────────────────────────────────────────
    sales_counts = {
        "drafts_total":        0,
        "drafts_needs_review": 0,
        "drafts_approved":     0,
        "drafts_posted":       0,
        "drafts_post_failed":  0,
        # Drafts whose editable_lines_json holds at least one line. A draft
        # with editable_lines_json='[]' has no actionable content even
        # when draft_state is editable — so it must not count toward
        # sales_ready.
        "drafts_with_lines":   0,
    }
    # PR-5: collect distinct product_codes referenced in any draft so we
    # can later evaluate product_master + wFirma coverage.
    draft_pcs: set = set()
    try:
        from ..services import proforma_invoice_link_db as _pildb
        _pf_db_path = settings.storage_root / "proforma_links.db"
        drafts = _pildb.list_drafts_for_batch(_pf_db_path, batch_id) or []
        sales_counts["drafts_total"] = len(drafts)
        import json as _json
        for d in drafts:
            state = (
                (d.get("draft_state") if isinstance(d, dict) else None)
                or getattr(d, "draft_state", "")
                or ""
            ).strip()
            if state in ("draft", "editing"):
                sales_counts["drafts_needs_review"] += 1
            elif state == "approved":
                sales_counts["drafts_approved"] += 1
            elif state == "posted":
                sales_counts["drafts_posted"] += 1
            elif state == "post_failed":
                sales_counts["drafts_post_failed"] += 1
            # drafts_with_lines — non-empty editable_lines_json
            raw = (
                (d.get("editable_lines_json") if isinstance(d, dict) else None)
                or getattr(d, "editable_lines_json", "")
                or ""
            )
            try:
                parsed = _json.loads(raw) if raw else []
            except Exception:
                parsed = []
            if isinstance(parsed, list) and len(parsed) > 0:
                sales_counts["drafts_with_lines"] += 1
                for ln in parsed:
                    if isinstance(ln, dict):
                        pc = str(ln.get("product_code") or "").strip()
                        if pc:
                            draft_pcs.add(pc)
    except Exception as exc:
        log.warning("[%s] lane-readiness sales counts failed (non-fatal): %s",
                    batch_id, exc)

    # PR-5: unresolved sales-side designs.  sales_packing_lines rows with
    # empty product_code carry an honest design_no — surface them so the
    # operator can repair the upstream parser/mapping.
    unresolved_sales_pcs: List[str] = []
    try:
        from ..services import document_db as _ddb_lr
        for r in (_ddb_lr.get_sales_packing_lines(batch_id) or []):
            pc = str(r.get("product_code") or "").strip()
            dn = str(r.get("design_no") or "").strip()
            if not pc and dn:
                unresolved_sales_pcs.append(dn)
        unresolved_sales_pcs = sorted(set(unresolved_sales_pcs))
    except Exception as exc:
        log.warning("[%s] lane-readiness unresolved sales scan failed "
                    "(non-fatal): %s", batch_id, exc)

    # ── Purchase lane ────────────────────────────────────────────────────
    purchase: Dict[str, Any] = {
        "packing_rows":           0,
        "distinct_product_codes": 0,
        "products_ready":         0,
        "products_missing":       0,
        "sad_present":            False,
        "pz_ready":               False,
        "pz_blocked_by":          [],
        # PR-5: additive coverage lists (all default to []).
        "product_master_missing":            [],
        "wfirma_products_missing":           [],
        "unresolved_purchase_product_codes": [],
        # PR-8: invoice_no values present in packing_lines for this
        # batch but ABSENT from invoice_lines.  Indicates a purchase
        # invoice document was never uploaded/parsed — product_code
        # cannot be minted, so any sales row referencing the same
        # invoice will fail to resolve.
        "missing_purchase_invoices":         [],
    }
    purch_pcs: set = set()
    wfirma_ready_pcs: set = set()
    unresolved_purchase: List[str] = []
    try:
        plines = pdb.get_packing_lines_for_batch(batch_id) or []
        purchase["packing_rows"] = len(plines)
        # Canonical codes present on purchase packing rows.
        purch_pcs = {
            str(ln.get("product_code") or "").strip()
            for ln in plines
            if (ln.get("product_code") or "").strip()
        }
        purchase["distinct_product_codes"] = len(purch_pcs)

        # PR-5: purchase rows whose product_code is empty/NULL — surface
        # their design_no so the operator can repair the upstream
        # invoice→packing match (parser-side).
        for ln in plines:
            pc = str(ln.get("product_code") or "").strip()
            dn = str(ln.get("design_no") or "").strip()
            if not pc and dn:
                unresolved_purchase.append(dn)
        unresolved_purchase = sorted(set(unresolved_purchase))

        ready_count = 0
        # PR-5: union of purchase + draft codes — same single SELECT
        # services both purchase.wfirma_products_missing and sales
        # blocked_by.wfirma_products_missing.
        union_pcs = purch_pcs | draft_pcs
        if union_pcs:
            try:
                from ..services import wfirma_db as _wfdb  # noqa: F401
                wfdb_path = settings.storage_root / "wfirma.db"
                if wfdb_path.exists():
                    with sqlite3.connect(str(wfdb_path)) as _wcon:
                        _wcon.row_factory = sqlite3.Row
                        placeholders = ",".join(["?"] * len(union_pcs))
                        rows = _wcon.execute(
                            f"SELECT product_code, sync_status "
                            f"FROM wfirma_products "
                            f"WHERE product_code IN ({placeholders})",
                            list(union_pcs),
                        ).fetchall()
                        for r in rows:
                            pc = (r["product_code"] or "").strip()
                            status = (r["sync_status"] or "").strip()
                            if status in ("created", "ready"):
                                wfirma_ready_pcs.add(pc)
                        # Existing legacy count is over the purchase-side
                        # only — preserve back-compat semantics.
                        ready_count = sum(
                            1 for pc in purch_pcs if pc in wfirma_ready_pcs
                        )
            except Exception as exc:
                log.warning("[%s] lane-readiness wfirma cache lookup failed "
                            "(non-fatal): %s", batch_id, exc)
        purchase["products_ready"]   = ready_count
        purchase["products_missing"] = max(
            0, purchase["distinct_product_codes"] - ready_count
        )
        purchase["unresolved_purchase_product_codes"] = unresolved_purchase
        # purchase-side missing = packing codes lacking a wfirma ready row
        purchase["wfirma_products_missing"] = sorted(
            purch_pcs - wfirma_ready_pcs
        )
    except Exception as exc:
        log.warning("[%s] lane-readiness purchase counts failed (non-fatal): %s",
                    batch_id, exc)

    # PR-5: product_master coverage — single SELECT IN(union of purchase
    # + draft codes).  reservation_queue.db is the canonical identity
    # registry (PR #193 schema + PR #196 backfill).  Read-only.
    pm_pcs: set = set()
    try:
        union_pcs_for_pm = purch_pcs | draft_pcs
        if union_pcs_for_pm:
            rdb_path = settings.storage_root / "reservation_queue.db"
            if rdb_path.exists():
                with sqlite3.connect(str(rdb_path)) as _rcon:
                    _rcon.row_factory = sqlite3.Row
                    placeholders = ",".join(["?"] * len(union_pcs_for_pm))
                    for r in _rcon.execute(
                        f"SELECT product_code FROM product_master "
                        f"WHERE product_code IN ({placeholders})",
                        list(union_pcs_for_pm),
                    ).fetchall():
                        pm_pcs.add((r["product_code"] or "").strip())
        purchase["product_master_missing"] = sorted(
            (purch_pcs | draft_pcs) - pm_pcs
        )
    except Exception as exc:
        log.warning("[%s] lane-readiness product_master lookup failed "
                    "(non-fatal): %s", batch_id, exc)

    # SAD presence — read audit.json once, look for importer / sad_number / mrn.
    try:
        audit_path = output_dir / "audit.json"
        if audit_path.exists():
            audit = json.loads(audit_path.read_text(encoding="utf-8"))
            for key in ("importer", "sad_number", "mrn"):
                v = audit.get(key)
                if isinstance(v, str) and v.strip():
                    purchase["sad_present"] = True
                    break
                if v not in (None, "", 0):
                    purchase["sad_present"] = True
                    break
    except Exception as exc:
        log.warning("[%s] lane-readiness audit.json read failed (non-fatal): %s",
                    batch_id, exc)

    # ── PR-8: missing purchase invoice detection ─────────────────────────
    # Purchase side — invoice_no values claimed by packing_lines that
    # have no corresponding invoice_lines anchor.
    pack_invs: set = set()
    inv_lines_invs: set = set()
    sales_filename_invs: set = set()
    try:
        for ln in (plines if 'plines' in locals() else
                   (pdb.get_packing_lines_for_batch(batch_id) or [])):
            inv = str(ln.get("invoice_no") or "").strip()
            if inv:
                pack_invs.add(inv)
        from ..services import document_db as _ddb_inv
        for r in (_ddb_inv.get_invoice_lines_for_batch(batch_id) or []):
            inv = str(r.get("invoice_no") or "").strip()
            if inv:
                inv_lines_invs.add(inv)
    except Exception as exc:
        log.warning("[%s] lane-readiness missing-invoice scan failed "
                    "(non-fatal): %s", batch_id, exc)
    purchase["missing_purchase_invoices"] = sorted(
        pack_invs - inv_lines_invs
    )

    # Sales side — derive an invoice_no anchor per sales_packing_list
    # filename and check it against invoice_lines.
    try:
        from ..services import document_db as _ddb_sales_inv
        sales_docs = _ddb_sales_inv.get_documents_for_batch(
            batch_id, document_type="sales_packing_list",
        ) or []
        for sd in sales_docs:
            fn = sd.get("file_name") or ""
            m = _SALES_FILENAME_INVOICE_RE.match(fn)
            if m:
                anchor = f"EJL/{m.group(1)}-{m.group(2)}/{m.group(3)}"
                sales_filename_invs.add(anchor)
    except Exception as exc:
        log.warning("[%s] lane-readiness sales-filename scan failed "
                    "(non-fatal): %s", batch_id, exc)
    missing_sales_invs = sorted(sales_filename_invs - inv_lines_invs)

    # PZ readiness + closed-enum blocked_by list.
    blocked: List[str] = []
    if purchase["packing_rows"] == 0:
        blocked.append(_PZ_BLOCKED_NO_PACKING_ROWS)
    if purchase["products_missing"] > 0:
        blocked.append(_PZ_BLOCKED_PRODUCTS_MISSING)
    if not purchase["sad_present"]:
        blocked.append(_PZ_BLOCKED_SAD_MISSING)
    # PR-8: surface purchase_invoice_missing on both lanes.
    if purchase["missing_purchase_invoices"]:
        blocked.append(_PZ_BLOCKED_PURCHASE_INVOICE_MISSING)
    # PR-5: product_master coverage gate.  product_master_missing
    # surfaces purchase-side codes that are absent from the canonical
    # identity registry — operator must repair before PZ posts.
    if purchase["product_master_missing"]:
        # Restrict the PZ-side gate to the purchase intersection so it
        # is symmetric with the existing products_missing semantics
        # (sales-only missing codes are flagged on the sales side).
        purch_pm_missing = sorted(purch_pcs - pm_pcs)
        if purch_pm_missing:
            blocked.append(_PZ_BLOCKED_PRODUCT_MASTER_MISSING)
    # Stable order, no duplicates (the conditions above are mutually
    # distinct, but enforce defensively).
    purchase["pz_blocked_by"] = sorted(set(blocked))
    purchase["pz_ready"] = not purchase["pz_blocked_by"]

    # ── Sales blocked_by (PR-5) ──────────────────────────────────────────
    # Closed-enum reason list parallel to pz_blocked_by.  Derived from
    # the already-computed counts + draft_pcs/pm_pcs/wfirma_ready_pcs.
    sales_blocked: List[str] = []
    if sales_counts["drafts_total"] == 0:
        sales_blocked.append(_SALES_BLOCKED_NO_DRAFTS)
    if (sales_counts["drafts_total"] > 0
            and sales_counts["drafts_with_lines"] == 0):
        sales_blocked.append(_SALES_BLOCKED_DRAFTS_HAVE_NO_LINES)
    if draft_pcs and (draft_pcs - wfirma_ready_pcs):
        sales_blocked.append(_SALES_BLOCKED_WFIRMA_PRODUCTS_MISSING)
    if draft_pcs and (draft_pcs - pm_pcs):
        sales_blocked.append(_SALES_BLOCKED_PRODUCT_MASTER_MISSING)
    if sales_counts["drafts_post_failed"] > 0:
        sales_blocked.append(_SALES_BLOCKED_POST_FAILED)
    # PR-8: sales filename references an invoice anchor that was never
    # parsed → product_code can never be minted/matched for those rows.
    if missing_sales_invs:
        sales_blocked.append(_SALES_BLOCKED_PURCHASE_INVOICE_MISSING)
    sales_blocked_sorted = sorted(set(sales_blocked))
    sales_ready = (sales_blocked_sorted == [])

    return {
        "batch_id": batch_id,
        "sales": {
            **sales_counts,
            "ready":                          sales_ready,
            "blocked_by":                     sales_blocked_sorted,
            "unresolved_sales_product_codes": unresolved_sales_pcs,
            # PR-8: invoice anchors implied by sales filenames that
            # have no invoice_lines evidence.  Always a list, sorted.
            "missing_purchase_invoices_for_sales": missing_sales_invs,
        },
        "purchase": purchase,
    }


# ── GET /api/v1/packing/{batch_id}/lines ─────────────────────────────────────

@router.get("/{batch_id}/lines", dependencies=[_auth])
def get_packing_lines(batch_id: str) -> Dict[str, Any]:
    """Return only packing lines for a batch."""
    _validate_batch(batch_id)
    lines = pdb.get_packing_lines_for_batch(batch_id)
    return {
        "batch_id": batch_id,
        "count":    len(lines),
        "lines":    lines,
    }


# ── Helpers for link-as-sales ────────────────────────────────────────────────

# Two filename patterns for client name extraction:
#   Short:  "148 Client SUOKKO.xlsx"      → leading number + space + Client + name
#   Long:   "148 EJL-26-27-148-PND-18KT-...-Client SUOKKO.xlsx"
#                                          → dash + Client + name anywhere in stem
# Handles both "Client" and the common "Cilent" typo.
# Requires either a numeric prefix OR a dash separator so bare "Client NAME.xlsx"
# (no invoice number, no dash) does not match — preserving prior behavior.
_CLIENT_NAME_RE = re.compile(
    r"(?:^\d+\s+|-)(?:client|cilent)\s+(.+)",
    re.IGNORECASE,
)

# Preamble-level label pattern for Excel header-row fallback.
_CLIENT_PREAMBLE_RE = re.compile(
    r"^(?:client|consignee|buyer|ship\s*to)\s*[:#\-]?\s*(.+)",
    re.IGNORECASE,
)

# C22-PERMANENT: company-suffix detection for label-LESS header blocks where
# the buyer name sits as a free-standing cell in the preamble (e.g. the
# 2026-05 DiamondGroup GmbH packing lists where R5 contains the company name
# with no "Client:" prefix). The cell text must END with one of these
# suffixes (optionally followed by punctuation).  Authoritative for the
# header-extraction priority chain documented in
# `_extract_client_from_preamble` below.
_COMPANY_SUFFIX_RE = re.compile(
    r"(?<![A-Za-z0-9])(?:"
    r"GmbH"                 # DE / AT
    r"|Sp\.?\s*z\s*o\.?\s*o\.?"   # PL
    r"|s\.?\s*r\.?\s*o\.?"  # SK / CZ
    r"|B\.?V\.?"            # NL
    r"|S\.?L\.?"            # ES
    r"|S\.?p\.?\s*A\.?"     # IT (S.p.A.)  — listed BEFORE S.A. to avoid prefix collision
    r"|S\.?A\.?"            # FR / CH / ES — generic suffix; lower priority match
    r"|Ltd"                 # UK / IE / IN
    r"|LLP|PLC|LLC|Inc"     # UK / US
    r"|AG"                  # DE / CH / AT
    r"|OY"                  # FI
    r")\b\.?\s*$",
    re.IGNORECASE,
)

# C22-PERMANENT: deny-list of cell texts that must NOT be treated as a
# client name.  Excludes table column headers + common preamble labels so a
# free-standing "Client Po" header row or a "Total" footer row never
# becomes a client.  Compared against the full stripped cell text,
# case-insensitive.
_CLIENT_DENYLIST: frozenset = frozenset({
    s.lower() for s in (
        "client po", "client p.o.", "client po.", "po", "po.",
        "purchase order", "purchase order #", "order", "order #",
        "invoice", "invoice #", "invoice no", "invoice no.", "invoice number",
        "shipment", "shipment packing list", "packing list",
        "sr", "ctg", "category", "qty", "quantity", "value", "total", "total value",
        "design", "design no", "designno", "kt", "col", "color", "quality",
        "size", "dated", "date", "remarks", "name", "client", "customer",
        "buyer", "consignee", "ship to", "bill to",
    )
})


def _guess_client_from_filename(filename: str) -> str:
    """
    Parse the client name from filenames like:
      '148 Client SUOKKO.xlsx'                              (short format)
      '148 EJL-26-27-148-PND-18KT-...-Client SUOKKO.xlsx'  (long format)
    Also handles the 'Cilent' typo.  Returns '' if pattern not found.
    """
    stem = Path(filename).stem
    m = _CLIENT_NAME_RE.search(stem.strip())
    return m.group(1).strip() if m else ""


def _build_matched_sales_lines(
    packing_lines: List[Dict[str, Any]],
    client: str,
) -> tuple:
    """Filter *packing_lines* to invoiceable rows and build sales_packing_lines dicts.

    Returns (matched_list, skipped_count).  Excludes rows where product_code
    is absent/blank or requires_manual_review is set.
    """
    matched = [
        ln for ln in packing_lines
        if str(ln.get("product_code") or "").strip()
        and not ln.get("requires_manual_review")
    ]
    skipped = len(packing_lines) - len(matched)
    sales_lines = [
        {
            "client_name":  client,
            "client_ref":   str(ln.get("invoice_no", "") or ""),
            "product_code": str(ln.get("product_code", "") or ""),
            "design_no":    str(ln.get("design_no", "") or ""),
            "bag_id":       str(ln.get("bag_id", "") or ""),
            "quantity":     float(ln.get("quantity", 0) or 0),
            "remarks":      str(ln.get("remarks", "") or ""),
            "unit_price":   float(ln.get("unit_price_eur") or ln.get("unit_price") or 0),
            "currency":     str(ln.get("currency") or "EUR"),
            "total_value":  float(ln.get("quantity") or 0) * float(ln.get("unit_price_eur") or ln.get("unit_price") or 0),
            "price_source": "packing_xlsx_value" if float(ln.get("unit_price_eur") or ln.get("unit_price") or 0) > 0 else "packing_promote",
        }
        for ln in matched
    ]
    return sales_lines, skipped


def _is_table_header_or_data_row(cells_text: List[str]) -> bool:
    """Return True if the row LOOKS like a table header row (multiple
    short column-name-like cells in the same row) — used to bail out of
    company-name detection once we've reached the data table.

    Heuristic: 3+ non-empty cells with average length <= 8 chars and at
    least one cell in the column-header denylist (Sr / Ctg / Qty / etc.).
    """
    nonempty = [c for c in cells_text if c]
    if len(nonempty) < 3:
        return False
    avg_len = sum(len(c) for c in nonempty) / len(nonempty)
    if avg_len > 14:
        return False
    if any(c.lower() in _CLIENT_DENYLIST for c in nonempty):
        return True
    return False


def _looks_like_company_name(text: str) -> bool:
    """C22-PERMANENT: detect a free-standing buyer/company-name cell in
    the preamble.  Returns True only when:
      - text is non-empty, length 3-80
      - text is NOT in _CLIENT_DENYLIST (excludes column headers)
      - text ENDS with a recognised company-form suffix
        (GmbH / Sp z o.o. / s.r.o. / B.V. / Ltd / S.A. / etc.)
    """
    if not text:
        return False
    t = text.strip()
    if len(t) < 3 or len(t) > 80:
        return False
    if t.lower() in _CLIENT_DENYLIST:
        return False
    return bool(_COMPANY_SUFFIX_RE.search(t))


def _guess_client_from_preamble(file_path: str) -> str:
    """
    Extract the client name from the top rows of the Excel packing file.

    C22-PERMANENT authority chain (in order):
      1. Explicit label match — "Client: …" / "Consignee: …" / "Buyer: …" /
         "Ship To: …"  (C13B behaviour, preserved).
      2. Free-standing company-suffix match — a cell whose text ends in
         a recognised legal-form suffix (GmbH / Sp z o.o. / Ltd / B.V. /
         s.r.o. / etc.), with the cell NOT in the denylist.  This catches
         label-LESS header blocks such as the DiamondGroup GmbH layout
         where R5 col 2 is the bare company name.

    NEVER matches:
      - Any cell whose text is in `_CLIENT_DENYLIST` (column headers
        like "Client Po" / "Total Value" / "Sr" / "Qty").
      - Any cell inside or below a detected table-header row (avoids
        picking up data rows like "Order 50260837").

    Returns '' on any failure (missing file, unreadable format, no match).
    """
    if not file_path:
        return ""
    try:
        import openpyxl as _opx  # type: ignore
        wb = _opx.load_workbook(str(file_path), read_only=True, data_only=True)
        ws = wb.active
        # PASS 1 — explicit-label search (highest priority, C13B behaviour).
        for row in ws.iter_rows(min_row=1, max_row=12, values_only=True):
            cells_text = [str(c or "").strip() for c in row]
            # Stop scanning once we hit the table header — anything below
            # is data, not preamble.
            if _is_table_header_or_data_row(cells_text):
                break
            for raw in cells_text:
                if not raw:
                    continue
                m = _CLIENT_PREAMBLE_RE.match(raw)
                if m:
                    val = m.group(1).strip().strip(":")
                    if val and len(val) < 80 and val.lower() not in _CLIENT_DENYLIST:
                        wb.close()
                        return val
        # PASS 2 — free-standing company-suffix search (C22-PERMANENT).
        # Re-iterate, stopping at the same table-header boundary.
        for row in ws.iter_rows(min_row=1, max_row=12, values_only=True):
            cells_text = [str(c or "").strip() for c in row]
            if _is_table_header_or_data_row(cells_text):
                break
            for raw in cells_text:
                if _looks_like_company_name(raw):
                    wb.close()
                    return raw.strip()
        wb.close()
    except Exception:
        pass
    return ""


# ── GET /api/v1/packing/{batch_id}/packing-documents ─────────────────────────

@router.get("/{batch_id}/packing-documents", dependencies=[_auth])
def get_packing_documents(batch_id: str) -> Dict[str, Any]:
    """
    Return packing_documents rows for a batch, each annotated with:

    ``suggested_client_name``
        Client name parsed from the filename (``{N} Client {name}`` pattern).
        Empty string when the pattern does not match.

    ``line_count``
        Actual number of packing_lines rows linked to this document.
        Ghost rows from pre-dedup re-uploads will show 0.

    ``is_duplicate``
        True when another document in this batch has the same source_file_hash
        and this document is NOT the canonical one (i.e. it was created by a
        re-upload before hash-dedup was in place).

    ``canonical_id``
        The id of the preferred document in a duplicate hash group — the one
        with the highest line_count (ties broken by oldest created_at).
        None for non-duplicate documents.

    Used by the dashboard "Link packing files" flow so the operator can see
    which documents are real vs. ghost, edit client names, ignore duplicates,
    and then call ``POST /{batch_id}/link-as-sales``.
    """
    from collections import defaultdict as _dd

    _validate_batch(batch_id)
    docs = pdb.get_packing_documents_for_batch(batch_id)

    # ── Annotate each doc with suggested client name and real line count ──────
    line_counts = pdb.get_line_counts_for_batch(batch_id)
    for d in docs:
        raw_name = Path(d.get("source_file_path", "")).name
        d["suggested_client_name"] = (
            _guess_client_from_filename(raw_name)
            or _guess_client_from_preamble(d.get("source_file_path", ""))
        )
        d["line_count"] = line_counts.get(d["id"], 0)

    # ── Detect duplicates: group by non-empty source_file_hash ───────────────
    # Ghost rows share the same hash but were created before the hash-dedup
    # guard was in place.  Canonical = most lines; ties → oldest created_at.
    hash_groups: Dict[str, List[Dict[str, Any]]] = _dd(list)
    for d in docs:
        h = d.get("source_file_hash", "")
        if h:
            hash_groups[h].append(d)

    for group in hash_groups.values():
        if len(group) == 1:
            group[0]["is_duplicate"] = False
            group[0]["canonical_id"] = None
        else:
            # Sort: highest line_count first, then oldest created_at first
            sorted_grp = sorted(group, key=lambda x: (-x["line_count"], x["created_at"]))
            canonical_id = sorted_grp[0]["id"]
            for d in group:
                d["is_duplicate"] = (d["id"] != canonical_id)
                d["canonical_id"] = canonical_id

    # Docs with empty hash (very old rows with no hash) — never marked duplicate
    for d in docs:
        if "is_duplicate" not in d:
            d["is_duplicate"] = False
            d["canonical_id"] = None

    return {
        "batch_id":  batch_id,
        "count":     len(docs),
        "documents": docs,
    }


# ── POST /api/v1/packing/{batch_id}/link-as-sales ────────────────────────────


class _ClientMapping(BaseModel):
    packing_document_id: str
    client_name: str
    # Optional operator-selected Customer-Master contractor. When supplied it is
    # the customer authority for this client (contractor_id beats the parsed/
    # free-text name — rules 1–2) and is persisted onto the sales chain so the
    # proforma draft resolves by it. When omitted, the existing name-fallback
    # behaviour is unchanged (rule 6).
    client_contractor_id: str = ""


class _LinkAsSalesBody(BaseModel):
    client_mappings: List[_ClientMapping]


@router.post("/{batch_id}/link-as-sales", dependencies=[_auth])
def link_packing_as_sales(
    batch_id: str,
    body: _LinkAsSalesBody,
) -> Dict[str, Any]:
    """
    Promote purchase packing lines into ``sales_packing_lines`` with operator-
    supplied client attribution, then auto-create/sync proforma drafts.

    This is the backfill path for batches where client packing files were
    uploaded via the purchase-packing route (``POST /{batch_id}/upload``) instead
    of the sales-intake route.

    **Idempotent**: uses ``replace_sales_packing_lines`` which atomically
    replaces lines scoped to (sales_document_id, batch_id) only.  Re-calling
    with the same mappings is safe.

    **Non-destructive**: rows in ``packing_lines`` (purchase side) are never
    touched.  Duplicate packing_document rows (ghost records from pre-dedup
    era) are safe to reference — they share lines with the canonical record.
    """
    output_dir = _validate_batch(batch_id)
    if not body.client_mappings:
        raise HTTPException(status_code=400, detail="client_mappings must not be empty.")

    results: List[Dict[str, Any]] = []
    for mapping in body.client_mappings:
        pdoc_id = (mapping.packing_document_id or "").strip()
        client  = (mapping.client_name or "").strip()
        cid     = (mapping.client_contractor_id or "").strip()
        if not pdoc_id or not client:
            results.append({
                "packing_document_id": pdoc_id,
                "client_name":         client,
                "client_contractor_id": cid,
                "ok":                  False,
                "reason":              "missing packing_document_id or client_name",
            })
            continue

        # Load existing purchase packing lines for this document
        packing_lines = pdb.get_packing_lines_for_document(pdoc_id)
        if not packing_lines:
            # Check if this might be a ghost duplicate doc with 0 lines.
            # Ghost docs share a hash with a canonical doc that has the real lines.
            # Tell the operator which doc to use instead.
            ghost_hint = ""
            pdoc_row = pdb.get_packing_document(pdoc_id)
            if pdoc_row:
                h = pdoc_row.get("source_file_hash", "")
                if h:
                    all_docs = pdb.get_packing_documents_for_batch(batch_id)
                    counts   = pdb.get_line_counts_for_batch(batch_id)
                    siblings = [d for d in all_docs
                                if d.get("source_file_hash") == h and d["id"] != pdoc_id]
                    canonical = next((d for d in siblings if counts.get(d["id"], 0) > 0), None)
                    if canonical:
                        ghost_hint = (
                            f" This appears to be a ghost duplicate — "
                            f"use canonical_id={canonical['id']!r} instead "
                            f"(has {counts[canonical['id']]} lines)."
                        )
            results.append({
                "packing_document_id": pdoc_id,
                "client_name":         client,
                "ok":                  False,
                "reason":              f"no packing lines found for document.{ghost_hint}",
            })
            continue

        # Get-or-create a stable sales_document record for this packing doc +
        # client. The operator-selected contractor_id (when supplied) is the
        # customer authority and is written onto sales_documents.client_contractor_id
        # → projected onto the sales lines → the proforma draft, so the draft
        # resolves Customer Master by contractor_id (selected beats parsed name).
        sales_doc_id = ddb.get_or_create_sales_document_for_packing(
            batch_id=batch_id,
            packing_document_id=pdoc_id,
            client_name=client,
            client_contractor_id=cid,
        )

        # Map packing_lines → sales_packing_lines, excluding unmatched rows.
        sales_lines, unmatched_skipped = _build_matched_sales_lines(packing_lines, client)

        # Atomically replace (idempotent for this sales_document scope)
        repl = ddb.replace_sales_packing_lines(
            sales_document_id=sales_doc_id,
            batch_id=batch_id,
            lines=sales_lines,
        )
        results.append({
            "packing_document_id": pdoc_id,
            "client_name":         client,
            "client_contractor_id": cid,
            "ok":                  True,
            "packing_lines_read":  len(packing_lines),
            "sales_lines_written": repl["inserted"],
            "unmatched_skipped":   unmatched_skipped,
        })
        log.info(
            "[%s] link_as_sales: client=%s pdoc=%s lines=%d→%d (skipped=%d unmatched)",
            batch_id, client, pdoc_id, len(packing_lines), repl["inserted"],
            unmatched_skipped,
        )

    # Seed the per-batch packing_contractor_resolution (the fallback authority
    # store read by derive_customer_resolution_via_packing) ONLY when this call
    # resolves the batch to a SINGLE operator-selected contractor. That store is
    # UNIQUE(batch_id, role='client') — a multi-client backfill cannot be
    # represented there without misrouting, so for multiple distinct contractors
    # we rely solely on the per-document client_contractor_id projected onto the
    # sales chain above (which resolves each draft correctly via step 0 of
    # derive_customer_authority_for_draft). Never infer from text (rule 5).
    cid_to_name: Dict[str, str] = {}
    for r in results:
        rc = (r.get("client_contractor_id") or "").strip()
        if r.get("ok") and rc:
            cid_to_name.setdefault(rc, r.get("client_name") or "")
    if len(cid_to_name) == 1:
        sel_cid, sel_name = next(iter(cid_to_name.items()))
        try:
            from ..services import packing_resolution_db as prdb
            prdb.upsert_resolution(
                settings.storage_root / "packing_resolutions.sqlite",
                batch_id=batch_id,
                role="client",
                verdict={
                    "parsed_name":         sel_name,
                    "matched_master_type": "customer_master",
                    "matched_master_id":   sel_cid,
                    "matched_wfirma_id":   sel_cid,
                    "tier":                1,
                    "confidence":          1.0,
                    "reason":              "link_as_sales_operator_selected",
                },
                operator_user="link_as_sales",
                status_override="confirmed",
            )
            log.info("[%s] link_as_sales seeded client resolution (contractor=%s)",
                     batch_id, sel_cid)
        except Exception as exc:
            log.warning("[%s] link_as_sales resolution seed failed (non-fatal): %s",
                        batch_id, exc)
    elif len(cid_to_name) > 1:
        log.info("[%s] link_as_sales: %d distinct operator-selected contractors — "
                 "per-batch resolution not seeded; per-document contractor_id "
                 "authority used per draft instead.", batch_id, len(cid_to_name))

    # Trigger proforma draft auto-sync (non-blocking; any error is logged, not raised)
    sync_summary: Dict[str, Any] = {}
    try:
        from ..services.proforma_draft_sync import sync_draft_from_packing_upload
        _pf_db_path = settings.storage_root / "proforma_links.db"
        sync_summary = sync_draft_from_packing_upload(
            batch_id=batch_id,
            operator="link_as_sales",
            db_path=_pf_db_path,
            audit_path=output_dir / "audit.json",
            master_db_path=settings.storage_root / "master_data.sqlite",
        )
        log.info("[%s] link_as_sales draft sync: %s", batch_id, sync_summary)
    except Exception as exc:
        log.warning("[%s] link_as_sales draft sync failed (non-fatal): %s", batch_id, exc)
        sync_summary = {"error": str(exc)}

    # Audit timeline
    try:
        tl.log_event(
            output_dir / "audit.json",
            "PACKING_LINKED_AS_SALES",
            "link_as_sales",
            actor="operator",
            detail={
                "batch_id": batch_id,
                "mappings": [
                    {"doc": m.packing_document_id, "client": m.client_name,
                     "contractor_id": (m.client_contractor_id or "").strip()}
                    for m in body.client_mappings
                ],
                "results": results,
            },
        )
    except Exception:
        pass

    ok_count = sum(1 for r in results if r.get("ok"))
    return {
        "ok":        ok_count > 0,
        "batch_id":  batch_id,
        "processed": len(results),
        "linked":    ok_count,
        "failed":    len(results) - ok_count,
        "results":   results,
        "draft_sync": sync_summary,
    }


# ── POST /api/v1/packing/{batch_id}/manual-sales-allocation ──────────────────


class _ManualAllocationLine(BaseModel):
    product_code: str = ""
    design_no: str = ""
    quantity: float
    unit_price: float
    currency: str = "EUR"
    bag_id: str = ""
    remarks: str = ""


class _ManualSalesAllocationBody(BaseModel):
    client_name: str
    client_ref: str = ""
    client_contractor_id: str = ""
    lines: List[_ManualAllocationLine]


@router.post("/{batch_id}/manual-sales-allocation", dependencies=[_auth])
def manual_sales_allocation(
    batch_id: str,
    body: _ManualSalesAllocationBody,
) -> Dict[str, Any]:
    """Write sales_packing_lines without a sales packing XLSX upload.

    Validates each requested line against purchase packing authority, checks
    over-allocation, then writes via replace_sales_packing_lines (idempotent:
    re-POST replaces the prior allocation for this batch+client).
    price_source='manual_allocation' on every written row.
    """
    import hashlib as _hl
    from datetime import datetime as _dt

    output_dir = _validate_batch(batch_id)

    client_name = (body.client_name or "").strip()
    if not client_name:
        raise HTTPException(status_code=422, detail="client_name is required")
    if not body.lines:
        raise HTTPException(status_code=422, detail="lines must not be empty")

    # 1. Build purchase packing authority index for this batch.
    packing_lines = pdb.get_packing_lines_for_batch(batch_id) or []
    purchase_qty: Dict[str, float] = {}
    design_to_pc: Dict[str, str] = {}
    pc_to_dn: Dict[str, str] = {}
    for pl in packing_lines:
        pc = (pl.get("product_code") or "").strip()
        dn = (pl.get("design_no") or "").strip()
        qty = float(pl.get("quantity") or 0)
        if pc:
            purchase_qty[pc] = purchase_qty.get(pc, 0.0) + qty
            if dn:
                design_to_pc.setdefault(dn.upper(), pc)
                pc_to_dn.setdefault(pc, dn)

    # 2. Validate and normalise each requested line.
    validated: List[Dict[str, Any]] = []
    allocation_by_pc: Dict[str, float] = {}

    for i, ln in enumerate(body.lines):
        pc = (ln.product_code or "").strip()
        dn = (ln.design_no or "").strip()
        qty = ln.quantity
        unit_price = ln.unit_price
        currency = (ln.currency or "EUR").upper().strip()

        if qty <= 0:
            raise HTTPException(status_code=422,
                detail=f"Line {i}: quantity must be > 0 (got {qty})")
        if unit_price < 0:
            raise HTTPException(status_code=422,
                detail=f"Line {i}: unit_price must be >= 0 (got {unit_price})")

        # Resolve product_code via purchase authority.
        if not pc and dn:
            pc = design_to_pc.get(dn.upper(), "")
        if not pc or pc not in purchase_qty:
            raw = ln.product_code or ln.design_no or "<blank>"
            raise HTTPException(status_code=422,
                detail=f"Line {i}: {raw!r} not found in purchase packing for batch {batch_id!r}")

        effective_dn = dn or pc_to_dn.get(pc, "")
        allocation_by_pc[pc] = allocation_by_pc.get(pc, 0.0) + qty
        validated.append({
            "product_code": pc,
            "design_no":    effective_dn,
            "quantity":     qty,
            "unit_price":   unit_price,
            "currency":     currency,
            "total_value":  round(qty * unit_price, 6),
            "bag_id":       (ln.bag_id or "").strip(),
            "remarks":      (ln.remarks or "").strip(),
        })

    # 3. Over-allocation check: requested qty must not exceed purchase qty per product_code.
    over = [
        {"product_code": pc, "requested": aq, "available": purchase_qty.get(pc, 0.0)}
        for pc, aq in allocation_by_pc.items()
        if aq > purchase_qty.get(pc, 0.0)
    ]
    if over:
        raise HTTPException(status_code=422,
            detail={"error": "over_allocation", "items": over})

    # 4. Ensure a sales_documents row exists (deterministic id per batch+client).
    sd_seed = f"{batch_id}\x00manual_allocation\x00{client_name}"
    sd_id = _hl.sha256(sd_seed.encode()).hexdigest()[:32]
    client_ref = (body.client_ref or "").strip()
    client_cid = (body.client_contractor_id or "").strip()
    now = _dt.utcnow().isoformat()

    docs_db = settings.storage_root / "documents.db"
    with sqlite3.connect(str(docs_db)) as _con:
        _con.execute(
            """INSERT OR IGNORE INTO sales_documents
               (id, batch_id, document_id, client_name, client_ref,
                document_type, sales_doc_no, sales_doc_date,
                source_file_path, extraction_status,
                client_contractor_id, created_at, updated_at)
               VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?)""",
            (sd_id, batch_id, "", client_name, client_ref,
             "manual_sales_allocation", "", "", "", "extracted",
             client_cid, now, now),
        )
        _con.execute(
            """UPDATE sales_documents
               SET client_name=?, client_ref=?, client_contractor_id=?, updated_at=?
               WHERE id=?""",
            (client_name, client_ref, client_cid, now, sd_id),
        )

    # 5. Build line_records matching the sales_packing_lines INSERT column set.
    client_ref_val = client_ref
    line_records = [
        {
            "batch_id":             batch_id,
            "sales_document_id":    sd_id,
            "client_name":          client_name,
            "client_ref":           client_ref_val,
            "product_code":         v["product_code"],
            "design_no":            v["design_no"],
            "bag_id":               v["bag_id"],
            "quantity":             v["quantity"],
            "unit_price":           v["unit_price"],
            "currency":             v["currency"],
            "total_value":          v["total_value"],
            "price_source":         "manual_allocation",
            "remarks":              v["remarks"],
            "client_contractor_id": client_cid,
        }
        for v in validated
    ]

    # 6. Idempotent write: replaces prior manual allocation for same batch+client.
    write_result = ddb.replace_sales_packing_lines(
        sales_document_id=sd_id,
        batch_id=batch_id,
        lines=line_records,
    )

    # 7. Proforma draft sync (non-blocking).
    sync_summary: Dict[str, Any] = {}
    try:
        from ..services.proforma_draft_sync import sync_draft_from_packing_upload
        _pf_db = settings.storage_root / "proforma_links.db"
        sync_summary = sync_draft_from_packing_upload(
            batch_id=batch_id,
            operator="manual_allocation",
            db_path=_pf_db,
            audit_path=output_dir / "audit.json",
            master_db_path=settings.storage_root / "master_data.sqlite",
        ) or {}
    except Exception as _pf_exc:
        log.warning("[%s] manual_allocation: draft sync failed (non-fatal): %s",
                    batch_id, _pf_exc)
        sync_summary = {"error": str(_pf_exc)}

    return {
        "ok":                         True,
        "batch_id":                   batch_id,
        "sales_document_id":          sd_id,
        "client_name":                client_name,
        "lines_written":              write_result.get("inserted", len(line_records)),
        "lines_replaced":             write_result.get("deleted", 0),
        "allocation_by_product_code": allocation_by_pc,
        "price_source":               "manual_allocation",
        "draft_sync":                 sync_summary,
    }


# ── GET /api/v1/packing/{batch_id}/barcode ───────────────────────────────────

@router.get("/{batch_id}/barcode", dependencies=[_auth])
def get_barcode_preview(batch_id: str) -> Dict[str, Any]:
    """
    Return barcode-ready rows for a batch.

    Each row joins the packing DB record with its product_code.
    Only matched rows (product_code IS NOT NULL) are included.
    Unmatched rows are reported in the summary but excluded from barcode rows
    because they have no stable identifier to encode.

    barcode_value = product_code  (e.g. "EJL/26-27/100-1")
    This is the canonical per-line identifier and is safe to encode as a
    1-D or 2-D barcode for warehouse scanning.

    Fields per row:
      product_code, invoice_no, design_no, batch_no, bag_id, barcode_value,
      requires_manual_review
    """
    _validate_batch(batch_id)
    all_lines = pdb.get_packing_lines_for_batch(batch_id)

    matched   = [ln for ln in all_lines if ln.get("product_code")]
    unmatched = [ln for ln in all_lines if not ln.get("product_code")]

    # One invoice line can span multiple bags.
    # Explode: one barcode row per physical bag, not per invoice line.
    rows = [_build_barcode_row(ln) for ln in matched]

    return {
        "batch_id":        batch_id,
        "count":           len(rows),
        "unmatched_count": len(unmatched),
        "rows":            rows,
    }


# ── GET /api/v1/packing/{batch_id}/barcode/zpl ───────────────────────────────

@router.get("/{batch_id}/barcode/zpl", dependencies=[_auth])
def get_barcode_zpl(batch_id: str) -> PlainTextResponse:
    """
    Return ZPL II label data for all matched bags in a batch.

    One ^XA...^XZ block per bag — safe to spool directly to a Zebra printer
    via TCP port 9100 or USB raw print queue.

    Label spec:
      57 × 32 mm at 300 DPI
      Zone A: product_code / bag_id
      Zone B: design_no / metal+karat
      Zone C: QTY: qty uom
      Zone D: Code 128 barcode (barcode_value = product_code|bag_id)
    """
    _validate_batch(batch_id)
    all_lines = pdb.get_packing_lines_for_batch(batch_id)
    matched   = [ln for ln in all_lines if ln.get("product_code")]
    if not matched:
        raise HTTPException(
            status_code=422,
            detail="No matched packing lines found. Upload and match a packing list first.",
        )
    rows = [_build_barcode_row(ln) for ln in matched]
    zpl  = _render_zpl_batch(rows)
    return PlainTextResponse(
        content=zpl,
        media_type="text/plain",
        headers={
            "Content-Disposition": f'attachment; filename="labels_{batch_id}.zpl"',
            "X-Label-Count": str(len(rows)),
        },
    )


# ── POST /api/v1/packing/{batch_id}/barcode/print ────────────────────────────

@router.post("/{batch_id}/barcode/print", dependencies=[_auth])
def print_barcode_labels(
    batch_id:     str,
    printer_host: str   = Query(..., description="Zebra printer IP address"),
    printer_port: int   = Query(default=9100, description="Raw TCP port (default 9100)"),
    timeout_sec:  float = Query(default=5.0, description="TCP connect+send timeout"),
) -> Dict[str, Any]:
    """
    Send ZPL II labels directly to a Zebra printer via raw TCP (port 9100).

    The printer must be reachable from the server running this service.
    For USB-connected printers, use the /barcode/zpl endpoint and spool locally.

    Returns: ok, label_count, printer, port
    """
    _validate_batch(batch_id)
    all_lines = pdb.get_packing_lines_for_batch(batch_id)
    matched   = [ln for ln in all_lines if ln.get("product_code")]
    if not matched:
        raise HTTPException(
            status_code=422,
            detail="No matched packing lines. Upload and match a packing list first.",
        )

    rows    = [_build_barcode_row(ln) for ln in matched]
    zpl     = _render_zpl_batch(rows)
    payload = zpl.encode("utf-8")

    try:
        with socket.create_connection((printer_host, printer_port), timeout=timeout_sec) as sock:
            sock.sendall(payload)
        log.info(
            "[%s] Printed %d labels → %s:%d (%d bytes)",
            batch_id, len(rows), printer_host, printer_port, len(payload),
        )
    except OSError as exc:
        raise HTTPException(
            status_code=502,
            detail=f"Printer unreachable at {printer_host}:{printer_port} — {exc}",
        )

    return {
        "ok":           True,
        "batch_id":     batch_id,
        "label_count":  len(rows),
        "printer":      printer_host,
        "port":         printer_port,
        "bytes_sent":   len(payload),
    }


# ─────────────────────────────────────────────────────────────────────────────
# Dev-only producer trigger
# ─────────────────────────────────────────────────────────────────────────────
# Reads existing packing_lines for *batch_id* and runs the same producer that
# fires after a real upload (seed_purchase_transit). Does NOT mutate
# packing_lines. Returns the count seeded into PURCHASE_TRANSIT.
#
# Gated by settings.environment == "dev" — returns 404 in prod. No auth so
# validation harnesses can hit it without a session. Should be removed (or
# kept disabled) before any non-dev deployment.

# Single dev router carries both packing trigger and inventory-state seeder.
# Mounted by main.py as packing_dev_router. Prefix is /api/v1/dev so
# sub-paths can name their domain explicitly.
dev_router = APIRouter(prefix="/api/v1/dev", tags=["dev"])


class _DevTriggerBody(BaseModel):
    batch_id: str


@dev_router.post("/packing/trigger")
def dev_trigger_packing_seeding(body: _DevTriggerBody) -> Dict[str, Any]:
    """
    Run seed_purchase_transit() against existing packing_lines for *batch_id*.

    Read-only against packing.db; only writes to inventory_state /
    inventory_state_events via the engine's idempotent transition path.
    """
    if settings.environment != "dev":
        raise HTTPException(status_code=404, detail="Not found.")

    batch_id = (body.batch_id or "").strip()
    if not batch_id or "/" in batch_id or ".." in batch_id:
        raise HTTPException(status_code=400, detail="Invalid batch_id.")

    try:
        lines = pdb.get_packing_lines_for_batch(batch_id)
        processed = seed_purchase_transit(batch_id, lines)
        return {
            "ok":        True,
            "batch_id":  batch_id,
            "lines":     len(lines),
            "processed": processed,
        }
    except Exception as exc:
        log.warning("[%s] dev trigger failed: %s", batch_id, exc)
        return {
            "ok":        False,
            "batch_id":  batch_id,
            "error":     str(exc),
        }


# ─────────────────────────────────────────────────────────────────────────────
# Dev-only batch inventory_state seeder for legacy purchase packing lines
# ─────────────────────────────────────────────────────────────────────────────
# Per-batch, operator-triggered. Reads packing_lines + audit.json, walks
# each scan_code through legal transitions to a target state. Idempotent;
# never demotes; never touches sales_packing_lines, audit, warehouse scans,
# or proforma drafts.

_VALID_TARGETS = {"auto", ise.PURCHASE_TRANSIT, ise.WAREHOUSE_STOCK}


class _SeedBatchBody(BaseModel):
    batch_id:     str
    target_state: str  = "auto"
    dry_run:      bool = False


def _load_audit(batch_id: str) -> Dict[str, Any]:
    """Read audit.json from outputs/ or working/. Raises HTTPException(400) on miss."""
    for sub in ("outputs", "working"):
        p = settings.storage_root / sub / batch_id / "audit.json"
        if p.exists():
            try:
                return json.loads(p.read_text(encoding="utf-8"))
            except (OSError, json.JSONDecodeError) as exc:
                raise HTTPException(
                    status_code=400,
                    detail=f"audit.json for {batch_id} unreadable: {exc}",
                )
    raise HTTPException(
        status_code=400,
        detail=f"No audit.json found for {batch_id} — refusing to guess state.",
    )


def _pz_done(audit: Dict[str, Any]) -> bool:
    """
    Mirror of the inferred PZ-done logic used elsewhere in the app, plus the
    shared read-time evidence helper so a stale ``audit.status="failed"``
    doesn't block auto-target resolution when a legitimate PZ exists (e.g.
    only the timeline event ``wfirma_pz_created`` carried the proof). Strict
    rejection is preserved when no signal exists.
    """
    if (
        audit.get("pz_generated") is True
        or bool(audit.get("pz_pdf_filename"))
        or bool(audit.get("pz_generated_at"))
        or audit.get("status") in ("success", "partial")
    ):
        return True
    try:
        from ..services.audit_evidence import effective_pz_evidence
    except Exception:
        return False
    ev = effective_pz_evidence(audit)
    # Auto-target only flips to WAREHOUSE_STOCK when a *PZ-side* signal is
    # present — customs-only signals (DSK / SAD / clearance_status) prove
    # customs cleared but not that the wFirma PZ was issued, so they alone
    # must NOT promote inventory state. Business semantics preserved.
    pz_side = {
        "wfirma_export.wfirma_pz_doc_id",
        "timeline:wfirma_pz_created",
        "effective_pz_status_done",
    }
    return any(s in pz_side for s in ev["signals"])


@dev_router.post("/inventory-state/seed-batch")
def dev_seed_inventory_state(body: _SeedBatchBody) -> Dict[str, Any]:
    """
    Seed inventory_state for one legacy batch.

    No global backfill. Operator-triggered, batch-scoped, idempotent.
    Walks each scan_code through legal transitions only — never demotes,
    never overwrites a state row that's already at or beyond the target.
    """
    if settings.environment != "dev":
        raise HTTPException(status_code=404, detail="Not found.")

    batch_id = (body.batch_id or "").strip()
    if not batch_id or "/" in batch_id or ".." in batch_id:
        raise HTTPException(status_code=400, detail="Invalid batch_id.")

    requested = body.target_state or "auto"
    if requested not in _VALID_TARGETS:
        raise HTTPException(
            status_code=400,
            detail=f"target_state must be one of {sorted(_VALID_TARGETS)}; "
                   f"SALES_TRANSIT/CLOSED are not valid backfill targets.",
        )

    audit = _load_audit(batch_id)
    decided_target = (
        (ise.WAREHOUSE_STOCK if _pz_done(audit) else ise.PURCHASE_TRANSIT)
        if requested == "auto" else requested
    )

    lines = pdb.get_packing_lines_for_batch(batch_id)
    considered  = 0
    planned     = 0
    transitioned = 0
    skipped     = 0
    errors: List[Dict[str, Any]] = []

    for line in lines:
        considered += 1
        try:
            sc = line.get("scan_code") or pdb._compute_scan_code(line)
            if not sc:
                skipped += 1
                continue

            cur = ise.get_state(sc)
            cur_state = cur["state"] if cur else None

            # Plan the chain from cur_state to decided_target (legal hops only).
            chain = []
            if cur_state is None:
                chain.append(ise.PURCHASE_TRANSIT)
            if (decided_target == ise.WAREHOUSE_STOCK
                    and (cur_state == ise.PURCHASE_TRANSIT
                         or ise.PURCHASE_TRANSIT in chain)):
                chain.append(ise.WAREHOUSE_STOCK)

            if not chain:
                # Already at target or beyond — never demote.
                skipped += 1
                continue

            planned += len(chain)
            if body.dry_run:
                continue

            for next_state in chain:
                ise.transition(
                    scan_code    = sc,
                    to_state     = next_state,
                    product_code = str(line.get("product_code") or ""),
                    design_no    = str(line.get("design_no") or ""),
                    batch_id     = batch_id,
                )
                transitioned += 1
        except Exception as exc:
            errors.append({
                "scan_code": (line.get("scan_code") or "").strip(),
                "error":     f"{type(exc).__name__}: {exc}",
            })

    return {
        "ok":             True,
        "batch_id":       batch_id,
        "decided_target": decided_target,
        "requested_target": requested,
        "dry_run":        body.dry_run,
        "considered":     considered,
        "planned":        planned,
        "transitioned":   transitioned,
        "skipped":        skipped,
        "errors":         errors,
    }


# ── GET /api/v1/packing/{batch_id}/document/{document_id}/download ────────────

@router.get("/{batch_id}/document/{document_id}/download", dependencies=[_auth])
async def download_packing_document(
    batch_id: str,
    document_id: str,
) -> FileResponse:
    """Download the original source file for a packing document."""
    doc = pdb.get_packing_document(document_id)
    if not doc:
        raise HTTPException(status_code=404, detail="Packing document not found")
    if doc.get("batch_id") != batch_id:
        raise HTTPException(status_code=404, detail="Document does not belong to this batch")
    source_path = doc.get("source_file_path") or ""
    if not source_path or not Path(source_path).exists():
        raise HTTPException(status_code=404, detail="Source file not found on disk")
    filename = Path(source_path).name
    return FileResponse(
        path=source_path,
        filename=filename,
        headers={"Cache-Control": "no-store, no-cache, must-revalidate, max-age=0"},
    )


# ── DELETE /api/v1/packing/{batch_id}/document/{document_id} ─────────────────

@router.delete("/{batch_id}/document/{document_id}", dependencies=[_auth])
async def delete_packing_document(
    batch_id: str,
    document_id: str,
    operator: str = Header(default="operator", alias="X-Operator"),
) -> Dict[str, Any]:
    """Delete a packing document: disk file + packing_lines + packing_documents row.

    SALES guard: if any non-cancelled proforma draft exists for this batch,
    returns 409 — operator must cancel/delete the proforma first.
    """
    doc = pdb.get_packing_document(document_id)
    if not doc:
        raise HTTPException(status_code=404, detail="Packing document not found")
    if doc.get("batch_id") != batch_id:
        raise HTTPException(status_code=404, detail="Document does not belong to this batch")

    # Determine document side from storage path
    source_path = doc.get("source_file_path") or ""
    is_sales = "/source/sales/" in source_path.replace("\\", "/")

    # Proforma guard: block SALES file delete if any active draft exists
    if is_sales:
        try:
            from ..services import proforma_invoice_link_db as _pildb
            _pf_db_path = settings.storage_root / "proforma_links.db"
            drafts = _pildb.list_drafts_for_batch(_pf_db_path, batch_id) or []
            active_drafts = [
                d for d in drafts
                if (
                    (d.draft_state if hasattr(d, "draft_state") else (d.get("draft_state") or ""))
                    not in ("cancelled", "superseded")
                )
            ]
            if active_drafts:
                draft_labels = [
                    (d.wfirma_proforma_fullnumber or f"draft #{d.id}")
                    if hasattr(d, "wfirma_proforma_fullnumber")
                    else (d.get("wfirma_proforma_fullnumber") or f"draft #{d.get('id', '?')}")
                    for d in active_drafts
                ]
                raise HTTPException(
                    status_code=409,
                    detail={
                        "code": "PACKING_SALES_DELETE_BLOCKED_BY_PROFORMA",
                        "message": (
                            f"Cannot delete SALES packing file: "
                            f"{len(active_drafts)} active proforma draft(s) exist for this batch. "
                            "Delete or cancel the proforma first, then delete the packing file."
                        ),
                        "active_drafts": draft_labels,
                    },
                )
        except HTTPException:
            raise
        except Exception as _pf_exc:
            log.warning("[%s] proforma guard check failed (non-fatal): %s", batch_id, _pf_exc)

    # Delete disk file (non-fatal if file is already gone)
    disk_deleted = False
    if source_path and Path(source_path).exists():
        try:
            os.unlink(source_path)
            disk_deleted = True
        except OSError as exc:
            log.warning(
                "[%s] delete_packing_document: disk unlink failed for %r: %s",
                batch_id, source_path, exc,
            )

    # Delete DB rows (atomic — lines first, then document)
    try:
        result = pdb.delete_packing_document_and_lines(document_id)
    except KeyError:
        raise HTTPException(status_code=404, detail="Packing document not found")

    log.info(
        "[%s] packing document %s deleted by %s — %d lines removed, disk_deleted=%s",
        batch_id, document_id, operator, result["deleted_lines"], disk_deleted,
    )

    return {
        "ok":            True,
        "doc_id":        document_id,
        "batch_id":      batch_id,
        "deleted_lines": result["deleted_lines"],
        "disk_deleted":  disk_deleted,
    }


# ── Scored-pending operator confirmation ──────────────────────────────────────
#
# When resolve_sales_lines_for_batch() cannot auto-assign product_code with
# HIGH confidence (≥ 0.85), it records the design in scored_pending.json
# alongside audit.json.  The operator reads the recommendations here and
# confirms (or overrides) per-row assignments.
#
# Safety:
#   - product_code must be in the design's candidates list (no invention).
#   - row_id must exist in sales_packing_lines for this batch (no cross-batch).
#   - All assignments validated before any DB write (fail-closed).
#   - Confirms trigger a non-blocking draft re-sync so the draft picks up
#     the resolved product_codes immediately.
#   - No wFirma writes, no inventory mutations, no proforma approval.

class _ScoredPendingAssignment(BaseModel):
    row_id: str        # sales_packing_lines.id
    design_no: str
    product_code: str  # must be in candidates for this design_no


class _ConfirmScoredPendingBody(BaseModel):
    assignments: List[_ScoredPendingAssignment]


@router.get("/{batch_id}/scored-pending", dependencies=[_auth])
def get_scored_pending(batch_id: str) -> Dict[str, Any]:
    """Return designs awaiting operator product-code confirmation."""
    output_dir = _validate_batch(batch_id)
    sp_path = output_dir / "scored_pending.json"
    if not sp_path.exists():
        return {"batch_id": batch_id, "designs": {}, "count": 0}
    try:
        data = json.loads(sp_path.read_text(encoding="utf-8"))
    except Exception as exc:
        log.warning("[%s] scored_pending read error: %s", batch_id, exc)
        return {"batch_id": batch_id, "designs": {}, "count": 0, "read_error": str(exc)}
    designs = data.get("designs") or {}
    return {"batch_id": batch_id, "designs": designs, "count": len(designs)}


@router.post("/{batch_id}/scored-pending/confirm", dependencies=[_auth])
def confirm_scored_pending(
    batch_id: str,
    body: _ConfirmScoredPendingBody,
    x_operator: Optional[str] = Header(default="operator", alias="X-Operator"),
) -> Dict[str, Any]:
    """Confirm operator product-code assignments for MEDIUM/LOW-confidence designs.

    Validates all assignments before writing (fail-closed).  Updates
    sales_packing_lines.product_code, logs a timeline event, removes confirmed
    designs from scored_pending.json, then re-triggers draft sync so the draft
    immediately reflects the resolved product_codes.

    Safety rules enforced:
      - product_code must be a known candidate for the design (no invention).
      - row_id must exist in sales_packing_lines for this batch (no cross-batch).
      - No wFirma writes, no inventory mutations, no fiscal actions.
    """
    output_dir = _validate_batch(batch_id)
    sp_path    = output_dir / "scored_pending.json"

    if not sp_path.exists():
        raise HTTPException(status_code=404, detail="No scored_pending data for this batch.")
    try:
        pending_data = json.loads(sp_path.read_text(encoding="utf-8"))
    except Exception as exc:
        raise HTTPException(status_code=500, detail=f"scored_pending read error: {exc}")

    designs = pending_data.get("designs") or {}

    # ── Phase 1: validate all assignments before any write (fail-closed) ───────
    for asgn in body.assignments:
        pending = designs.get(asgn.design_no)
        if pending is None:
            raise HTTPException(
                status_code=422,
                detail=f"design_no {asgn.design_no!r} not found in scored_pending.",
            )
        candidates = pending.get("candidates", [])
        if asgn.product_code not in candidates:
            raise HTTPException(
                status_code=422,
                detail=(
                    f"product_code {asgn.product_code!r} is not a valid candidate "
                    f"for design {asgn.design_no!r}. Valid candidates: {candidates}"
                ),
            )

    rows_by_id = {r["id"]: r for r in (ddb.get_sales_packing_lines(batch_id) or [])}
    for asgn in body.assignments:
        if asgn.row_id not in rows_by_id:
            raise HTTPException(
                status_code=422,
                detail=(
                    f"row_id {asgn.row_id!r} not found in sales_packing_lines "
                    f"for batch {batch_id}."
                ),
            )

    # ── Phase 2: apply updates ─────────────────────────────────────────────────
    applied    = 0
    failed_ids: List[str] = []
    for asgn in body.assignments:
        ok = ddb.update_sales_packing_line_product_code(batch_id, asgn.row_id, asgn.product_code)
        if ok:
            applied += 1
        else:
            failed_ids.append(asgn.row_id)

    if failed_ids:
        log.warning("[%s] scored_pending confirm: %d row(s) failed to update: %s",
                    batch_id, len(failed_ids), failed_ids)

    # ── Phase 3: audit timeline ────────────────────────────────────────────────
    operator   = x_operator or "operator"
    audit_path = output_dir / "audit.json"
    tl.log_event(
        audit_path,
        "SCORED_PENDING_CONFIRMED",
        "packing_scored_pending_confirm",
        actor=operator,
        detail={
            "batch_id":    batch_id,
            "applied":     applied,
            "failed":      len(failed_ids),
            "assignments": [
                {
                    "row_id":       a.row_id,
                    "design_no":    a.design_no,
                    "product_code": a.product_code,
                }
                for a in body.assignments
            ],
        },
    )

    # ── Phase 4: remove confirmed designs from scored_pending.json ─────────────
    confirmed_designs = {a.design_no for a in body.assignments}
    remaining = {dn: dinfo for dn, dinfo in designs.items() if dn not in confirmed_designs}
    pending_data["designs"] = remaining
    try:
        sp_path.write_text(json.dumps(pending_data, ensure_ascii=False, indent=2),
                           encoding="utf-8")
    except Exception as _e:
        log.warning("[%s] scored_pending update after confirm failed (non-fatal): %s",
                    batch_id, _e)

    # ── Phase 5: re-trigger draft sync (non-blocking) ─────────────────────────
    if applied > 0:
        try:
            from ..services.proforma_draft_sync import sync_draft_from_packing_upload
            _pf_db_path = settings.storage_root / "proforma_links.db"
            _sync_result = sync_draft_from_packing_upload(
                batch_id=batch_id,
                operator=operator,
                db_path=_pf_db_path,
                audit_path=audit_path,
                master_db_path=settings.storage_root / "master_data.sqlite",
            )
            log.info("[%s] post-confirm draft sync: %s", batch_id, _sync_result)
        except Exception as _sync_exc:
            log.warning("[%s] post-confirm draft sync failed (non-fatal): %s",
                        batch_id, _sync_exc)

    return {
        "ok":                True,
        "batch_id":          batch_id,
        "applied":           applied,
        "failed":            len(failed_ids),
        "failed_row_ids":    failed_ids,
        "remaining_designs": len(remaining),
    }
