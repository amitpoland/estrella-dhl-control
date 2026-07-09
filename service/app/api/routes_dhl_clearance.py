"""
routes_dhl_clearance.py — DHL Customs Clearance API endpoints.

GET  /api/v1/dhl/scan-inbox                          — scan Zoho Mail for DHL customs emails
POST /api/v1/dhl/match-and-handle                    — match AWB to batch + run clearance handler
GET  /api/v1/dhl/clearance-status/{batch_id}         — get clearance status for a batch
POST /api/v1/dhl/generate-description/{batch_id}     — manually trigger Polish description
GET  /api/v1/dhl/download/{filename}                 — download generated PDF
POST /api/v1/dhl/generate-customs-package/{batch_id} — generate full customs description package (PDF + SAD JSON)
GET  /api/v1/dhl/sad-ready/{batch_id}                — return SAD-ready JSON data for a batch
POST /api/v1/dhl/approve/{batch_id}                  — approve the customs description (stores name + timestamp)

All endpoints require API key auth.
None of these endpoints send email — they only prepare packages.
"""
from __future__ import annotations

import json
import os
import re
import sys
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Dict, List, Optional

from fastapi import APIRouter, Body, Depends, HTTPException
from fastapi.responses import FileResponse, JSONResponse
from pydantic import BaseModel

from ..core.config import settings
from ..core.logging import get_logger
from ..core.security import require_api_key
from ..auth.dependencies import require_role
from ..core.guards import guard_dhl_requires_email
from ..services.clearance_path_alias import (
    is_agency_clearance,
    is_dhl_self_clearance,
)
from ..pipelines.dhl import receive_dhl_email as _pipeline_dhl_email
from ..core import timeline as tl
from ..config.email_routing import (
    DHL_TO,
    DHL_DSK_SOURCE,
    INTERNAL_CC,
    format_to,
    format_cc,
    is_dsk_source,
)

# ── Shared grammar import (Phase 2C) ────────────────────────────────────────
# Import shared grammar authority so the packing renderer's local PL
# dictionaries are verified at import time against the single source of
# truth.  The renderer keeps its composite dict-of-dicts structure (EN + PL
# together) and key format ("14KT GOLD" vs "14KT"), but import-time parity
# assertions catch any PL-side grammar drift.
#
# Path fix (Lesson J): use settings.engine_dir instead of parents[3].
#   dev:  parents[3] = C:\PZ-verify (repo root) ✓ but settings.engine_dir is also correct
#   prod: parents[3] = C:\ (system root) ✗ — settings.engine_dir = C:\PZ\engine ✓
_grammar_engine_dir = str(settings.engine_dir)
if _grammar_engine_dir not in sys.path:
    sys.path.insert(0, _grammar_engine_dir)
from description_grammar import (  # noqa: E402
    ITEM_TYPE_PL,
    METAL_PREPOSITIONAL,
)

log      = get_logger(__name__)
router   = APIRouter(prefix="/api/v1/dhl", tags=["dhl-clearance"])
_auth    = Depends(require_api_key)
_op_auth = Depends(require_role("admin", "logistics"))

# ── Engine path setup ─────────────────────────────────────────────────────────
_engine_dir = str(settings.engine_dir)
if _engine_dir not in sys.path:
    sys.path.insert(0, _engine_dir)

# ── Output directories ────────────────────────────────────────────────────────
_DSK_OUTPUT_DIR = (
    Path(os.environ.get("APPDATA", ""))
    / "estrellajewels" / "storage" / "dsk_outputs"
    if os.name == "nt"
    else Path.home() / "Library" / "Application Support"
    / "estrellajewels" / "storage" / "dsk_outputs"
)
if not _DSK_OUTPUT_DIR.parent.exists():
    _DSK_OUTPUT_DIR = settings.storage_root / "dsk_outputs"
_DSK_OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

_POLISH_DESC_DIR = settings.storage_root / "polish_descriptions"
_POLISH_DESC_DIR.mkdir(parents=True, exist_ok=True)

_SAD_READY_DIR = settings.storage_root / "sad_ready"
_SAD_READY_DIR.mkdir(parents=True, exist_ok=True)

# ── Zoho Mail constants ───────────────────────────────────────────────────────
_ZOHO_ACCOUNT_ID   = "2261204000000002002"
_ZOHO_INBOX_FOLDER = "2261204000000002014"


# ── Schemas ───────────────────────────────────────────────────────────────────

class MatchAndHandleRequest(BaseModel):
    awb:                str
    dhl_ticket:         Optional[str] = None
    message_id:         Optional[str] = None
    thread_id:          Optional[str] = None
    subject:            Optional[str] = None
    value_usd_override: Optional[float] = None


class ClearanceStatusResponse(BaseModel):
    batch_id:         str
    clearance_status: Optional[str] = None
    clearance_action: Optional[str] = None
    dhl_ticket:       Optional[str] = None
    awb:              Optional[str] = None
    updated_at:       Optional[str] = None
    found:            bool


class ScanInboxResult(BaseModel):
    scanned:      int
    matched:      int
    emails:       List[Dict[str, Any]]
    scan_method:  str
    scanned_at:   str


# ── Helpers ───────────────────────────────────────────────────────────────────

# Sentinel: imported from the engine at call time so both sides always agree.
# If the engine is unreachable the fallback string keeps the meaning intact.
def _get_unresolved_sentinel() -> str:
    """Return the consignor-unresolved sentinel — always sourced from the engine."""
    try:
        _svc_engine = str(Path(__file__).parent.parent.parent.parent)
        import sys as _sys
        if _svc_engine not in _sys.path:
            _sys.path.insert(0, _svc_engine)
        import customs_description_engine as _cde  # type: ignore
        return _cde._CONSIGNOR_UNRESOLVED_SENTINEL
    except Exception:
        return "[DOSTAWCA NIEOKRESLONY / SUPPLIER UNRESOLVED]"


def _resolve_customs_identities(batch_id: str) -> "tuple[Optional[str], Optional[str]]":
    """Resolve (consignee_name, consignor_name) from governed masters.

    Only called when ``settings.customs_identity_from_masters`` is True.

    Returns
    -------
    (consignee_name, consignor_name) — both Optional[str].

    consignee_name:
        company_profile.legal_name when the row exists and the field is
        non-empty; empty string otherwise (caller falls back to the
        hardcoded constant).

    consignor_name:
        - Resolved supplier name when shipment_documents.supplier_contractor_id
          is non-empty for this batch and the corresponding suppliers row exists.
        - _CONSIGNOR_UNRESOLVED_SENTINEL when the link is absent (supplier not
          set on intake). This surfaces an explicit flag on the PDF so an
          operator knows to set the supplier — rather than silently printing the
          wrong company name on a legal customs document.
        - Empty string on any read error (caller falls back to batch-parse path).

    Never raises. Any DB / import error produces an empty string for that field
    so the PDF can still be generated (with the current hardcoded fallback).
    """
    consignee_name: Optional[str] = None
    consignor_name: Optional[str] = None

    # ── Consignee from company_profile ───────────────────────────────────────
    try:
        from ..services.master_data_db import get_company_profile as _get_cp
        cp = _get_cp(settings.storage_root / "master_data.sqlite")
        consignee_name = (cp.legal_name or "").strip() if cp else ""
    except Exception as _exc:
        log.warning("[%s] _resolve_customs_identities: company_profile read "
                    "failed (non-fatal): %s", batch_id, _exc)
        consignee_name = ""

    # ── Consignor from supplier master ────────────────────────────────────────
    try:
        import sqlite3 as _sqlite3
        docs_db_path = settings.storage_root / "documents.db"
        sup_db_path  = settings.storage_root / "suppliers.sqlite"

        supplier_cid: Optional[str] = None
        if docs_db_path.exists():
            with _sqlite3.connect(str(docs_db_path)) as _dcon:
                _dcon.row_factory = _sqlite3.Row
                _row = _dcon.execute(
                    "SELECT supplier_contractor_id "
                    "FROM shipment_documents "
                    "WHERE batch_id=? AND supplier_contractor_id != '' "
                    "LIMIT 1",
                    (batch_id,),
                ).fetchone()
                if _row:
                    supplier_cid = str(_row["supplier_contractor_id"]).strip()

        _sentinel = _get_unresolved_sentinel()
        if not supplier_cid:
            # No supplier link on intake — surface as explicit flag, not silent
            # constant (which would be wrong for third-party supplier batches).
            consignor_name = _sentinel
        else:
            # Look up supplier by primary key (supplier_contractor_id is the
            # suppliers.id integer).
            if sup_db_path.exists():
                with _sqlite3.connect(str(sup_db_path)) as _scon:
                    _scon.row_factory = _sqlite3.Row
                    _srow = _scon.execute(
                        "SELECT name FROM suppliers WHERE id=?",
                        (supplier_cid,),
                    ).fetchone()
                    if _srow:
                        consignor_name = (_srow["name"] or "").strip()
                    else:
                        consignor_name = _sentinel
            else:
                consignor_name = _sentinel

    except Exception as _exc:
        log.warning("[%s] _resolve_customs_identities: supplier lookup "
                    "failed (non-fatal): %s", batch_id, _exc)
        consignor_name = ""   # fall back to batch-parse on error

    return consignee_name, consignor_name


def _load_audit(batch_id: str) -> Optional[dict]:
    """Load audit.json for a batch_id from storage."""
    for sub in ("outputs", "working"):
        p = settings.storage_root / sub / batch_id / "audit.json"
        if p.exists():
            try:
                return json.loads(p.read_text(encoding="utf-8"))
            except Exception:
                return None
    return None


def _write_audit(batch_id: str, audit: dict) -> None:
    """Write audit.json back to disk atomically."""
    for sub in ("outputs", "working"):
        p = settings.storage_root / sub / batch_id / "audit.json"
        if p.exists():
            tmp = p.with_suffix(".tmp")
            tmp.write_text(json.dumps(audit, indent=2, ensure_ascii=False), encoding="utf-8")
            tmp.replace(p)
            return


def _find_generated_file(filename: str) -> Optional[Path]:
    """Search DSK output dir, Polish descriptions dir, and SAD-ready dir for a named file."""
    for search_dir in (_DSK_OUTPUT_DIR, _POLISH_DESC_DIR, _SAD_READY_DIR):
        candidate = search_dir / filename
        if candidate.exists():
            return candidate
    return None


def _inject_rows_from_xlsx(batch_id: str, audit: dict) -> dict:
    """
    Read the PZ calculation XLSX for a batch and inject invoice rows into the audit dict.

    This enriches the audit dict with structured line-item data so that
    customs_description_engine can build a proper goods table even when
    audit.json only has aggregate invoice_totals.

    Returns the audit dict (modified in-place, also returned for chaining).
    """
    # Already has rows — nothing to do
    if audit.get("rows") or audit.get("invoices"):
        return audit

    # Find the XLSX in the batch output folder
    batch_dir: Optional[Path] = None
    for sub in ("outputs", "working"):
        candidate = settings.storage_root / sub / batch_id
        if candidate.is_dir():
            batch_dir = candidate
            break

    if batch_dir is None:
        return audit

    # Find any *_calc.xlsx file in the batch dir
    xlsx_files = list(batch_dir.glob("*_calc.xlsx")) or list(batch_dir.glob("*.xlsx"))
    if not xlsx_files:
        return audit

    xlsx_path = xlsx_files[0]
    try:
        import openpyxl
        wb = openpyxl.load_workbook(str(xlsx_path), data_only=True)

        # The "Rows" sheet has columns:
        # Lp, Invoice No, English Name, Polish Name, Qty, Unit USD, Line USD, ..., Item Type, ...
        if "Rows" not in wb.sheetnames:
            return audit

        ws = wb["Rows"]
        rows_iter = ws.iter_rows(values_only=True)
        header = [str(c).strip() if c is not None else "" for c in next(rows_iter)]

        def _col(name: str) -> Optional[int]:
            try:
                return header.index(name)
            except ValueError:
                return None

        c_inv  = _col("Invoice No")
        c_desc = _col("English Name")
        c_qty  = _col("Qty")
        c_usd  = _col("Unit USD")
        c_tot  = _col("Line USD")
        c_type = _col("Item Type")

        if c_inv is None or c_desc is None:
            return audit

        injected: list[dict] = []
        for row in rows_iter:
            if not any(row):
                continue
            inv_no  = str(row[c_inv]).strip()  if c_inv  is not None and row[c_inv]  is not None else ""
            desc    = str(row[c_desc]).strip() if c_desc is not None and row[c_desc] is not None else ""
            qty     = float(row[c_qty])  if c_qty  is not None and row[c_qty]  is not None else 0.0
            unit_p  = float(row[c_usd])  if c_usd  is not None and row[c_usd]  is not None else 0.0
            line_t  = float(row[c_tot])  if c_tot  is not None and row[c_tot]  is not None else 0.0
            itype   = str(row[c_type]).strip().upper() if c_type is not None and row[c_type] is not None else ""
            if not desc or not inv_no:
                continue
            injected.append({
                "invoice_number": inv_no,
                "description":    desc,
                "item_type":      itype,
                "quantity":       qty,
                "unit_price":     unit_p,
                "line_total":     line_t,
            })

        if injected:
            audit["rows"] = injected
            log.debug("Injected %d invoice rows from XLSX into audit for description generation", len(injected))

    except Exception as _xe:
        log.warning("_inject_rows_from_xlsx: could not read XLSX %s — %s", xlsx_path, _xe)

    return audit


# ── PR-206: DB-first row injection ─────────────────────────────────────────
#
# Source priority for batch["rows"] (engine input):
#   1. invoice_lines from documents.db for THIS batch_id        (primary)
#   2. invoice_lines from documents.db for batches sharing the
#      same AWB                                                  (cross-batch
#                                                                 union fallback)
#   3. XLSX Rows sheet (existing _inject_rows_from_xlsx)         (legacy)
#   4. nothing — guard fires below, generation refused
#
# Synthetic per-piece averaging (_build_synthetic_lines_from_totals in the
# engine) is intentionally unreachable in production after this PR: when
# none of the four sources produce rows, the route refuses generation
# with HTTP 422 lines_missing_for_description.
#
# Pure / read-only.  No wFirma / PZ / DHL email / proforma posting writes.


def _row_uom_from_description(desc: str) -> str:
    """Derive UoM from the leading token of an EJL invoice line description.

    Real lines start with "PCS, " or "PRS, " — we honour that prefix.
    Falls back to PCS when the prefix is absent or unrecognised.
    """
    if not desc:
        return "PCS"
    head = desc.strip().split(",", 1)[0].strip().upper()
    if head in ("PCS", "PRS"):
        return head
    return "PCS"


def _project_invoice_line_to_engine_row(line: dict) -> dict:
    """Map a documents.db invoice_lines row to the engine row shape used
    by customs_description_engine._extract_invoices / process_batch_items.

    Permissive aliases on the engine side:
      description: description | desc | name
      quantity:    quantity | qty | line_qty
      unit_price:  unit_price | rate | price
      line_total:  line_total | amount | total
      hsn_code:    hsn_code | hs_code | hsn
      uom:         unit | uom (default 'PCS')
    """
    desc      = str(line.get("description") or "")
    unit_p    = float(line.get("unit_price") or line.get("rate_usd") or 0.0)
    line_tot  = float(line.get("total_value") or line.get("amount_usd") or 0.0)
    hsn       = str(line.get("hsn_code") or line.get("hs_code") or "")
    return {
        "invoice_number": str(line.get("invoice_no") or ""),
        "line_position":  int(line.get("line_position") or 0),
        "product_code":   str(line.get("product_code") or ""),
        "description":    desc,
        "item_type":      "",            # engine derives from description
        "quantity":       float(line.get("quantity") or 0.0),
        "unit_price":     unit_p,
        "line_total":     line_tot,
        "hsn_code":       hsn,
        "currency":       str(line.get("currency") or "USD"),
        "uom":            _row_uom_from_description(desc),
    }


def _is_placeholder_invoice_line(row: dict) -> bool:
    """Detect invoice_intake_parser's fallback placeholder row.

    The intake parser falls back to a single zero-valued placeholder row
    when it can't extract real line items (e.g. for the global_jewellery
    invoice template whose line layout differs from the EJL template the
    extractor was tuned for). These rows have qty=0, total_value=0, and a
    description starting with "(placeholder". Projecting them into the
    reconciler poisons the FOB sum to 0 even when aggregate invoice_totals
    were extracted correctly.

    Returns True only for the unambiguous placeholder signature.
    """
    if not isinstance(row, dict):
        return False
    qty   = float(row.get("quantity")    or 0.0)
    total = float(row.get("total_value") or row.get("amount_usd") or 0.0)
    desc  = str(row.get("description") or "")
    return qty == 0.0 and total == 0.0 and desc.lstrip().startswith("(placeholder")


def _inject_rows_from_db_invoice_lines(batch_id: str, audit: dict) -> dict:
    """Project documents.db invoice_lines into ``audit["rows"]``.

    Primary path:
      - Read invoice_lines for ``batch_id``.
      - If empty AND the audit carries an AWB, union invoice_lines from
        all batches sharing the same AWB (dedup by
        (invoice_no, line_position, product_code)).

    Placeholder filter:
      - Intake-time zero-valued placeholder rows (qty=0, total_value=0,
        description starts with "(placeholder") are dropped here. Their
        presence in the DB is a record-of-existence marker, not real
        per-line data — projecting them would crash the reconciler.

    Idempotent: if ``audit`` already has ``rows`` or ``invoices``, no-op.
    """
    if audit.get("rows") or audit.get("invoices"):
        return audit

    try:
        from app.services import document_db as _ddb
        rows = _ddb.get_invoice_lines_for_batch(batch_id) or []
        rows = [r for r in rows if not _is_placeholder_invoice_line(r)]

        if not rows:
            # Cross-batch fallback: same AWB, different batch_id.
            awb = (
                audit.get("dhl_awb")
                or audit.get("awb")
                or (audit.get("batch_meta") or {}).get("awb")
                or audit.get("tracking_no")
                or ""
            )
            if awb:
                seen: set = set()
                docs = _ddb.get_documents_by_awb(awb, "purchase_invoice")
                # Iterate distinct batch_ids referenced by these documents.
                visited_batches: set = set()
                for d in docs:
                    bid = str(d.get("batch_id") or "")
                    if not bid or bid == batch_id or bid in visited_batches:
                        continue
                    visited_batches.add(bid)
                    for ln in (_ddb.get_invoice_lines_for_batch(bid) or []):
                        # Same placeholder filter as primary path.
                        if _is_placeholder_invoice_line(ln):
                            continue
                        key = (
                            str(ln.get("invoice_no") or ""),
                            int(ln.get("line_position") or 0),
                            str(ln.get("product_code") or ""),
                        )
                        if key in seen:
                            continue
                        seen.add(key)
                        rows.append(ln)

        if not rows:
            return audit

        projected = [_project_invoice_line_to_engine_row(r) for r in rows]
        audit["rows"]            = projected
        audit["_rows_source"]    = "db_invoice_lines"
        audit["_rows_row_count"] = len(projected)
        log.info(
            "[%s] _inject_rows_from_db_invoice_lines: projected %d rows from "
            "invoice_lines (this batch + shared-AWB union)",
            batch_id, len(projected),
        )
    except Exception as exc:
        log.warning(
            "[%s] _inject_rows_from_db_invoice_lines: failed (non-fatal): %s",
            batch_id, exc,
        )
    return audit


def _synthesize_rows_from_invoice_aggregates(batch_id: str, audit: dict) -> dict:
    """Last-resort grouped-row synthesizer.

    When the DB and XLSX paths produce no per-line rows but the engine
    successfully parsed aggregate invoice_totals (fob_usd, freight_usd,
    insurance_usd, product_counts_by_unit) — as happens for the
    `global_jewellery` template whose per-line layout differs from the
    EJL template the intake extractor was tuned for — synthesize one
    grouped row per (invoice, unit_type) so the reconciler can verify
    against the same aggregate the engine produced.

    Properties:
      - Pure read of files in source/invoices/ via engine parse_invoice
      - Re-uses the C27.1 PDF-magic quarantine helper to skip non-PDF
        masqueraders
      - One grouped row per (invoice_file, PCS|PRS) — preserves the
        unit breakdown the operator's customs description needs
      - Row totals sum exactly to the engine's parsed FOB per file
      - Idempotent: no-op if audit already has rows or no aggregates

    NEVER touches CIF formula, customs threshold logic, SAD/ZC429 gate,
    wFirma/PZ write paths.
    """
    if audit.get("rows") or audit.get("invoices"):
        return audit

    inv_totals = audit.get("invoice_totals") or {}
    declared_fob = 0.0
    try:
        declared_fob = float(inv_totals.get("total_fob_usd") or 0.0)
    except Exception:
        declared_fob = 0.0
    if declared_fob <= 0:
        return audit  # no aggregate to synthesize from

    # Resolve batch dir using the same pattern as other helpers in this file
    inv_dir: Optional[Path] = None
    for sub in ("outputs", "working"):
        candidate = settings.storage_root / sub / batch_id / "source" / "invoices"
        if candidate.is_dir():
            inv_dir = candidate
            break
    if inv_dir is None:
        return audit

    inv_pdfs_all = sorted(inv_dir.glob("*.pdf"))
    # Re-use the C27.1 magic-header quarantine — never feed a non-PDF
    # to the engine's parser (it would crash or emit empty results that
    # poison the synthesized rows).
    try:
        # Import locally to avoid circular reference; helper lives in
        # routes_dashboard for the recheck loops.
        from .routes_dashboard import _partition_valid_pdfs as _ppvp
        inv_pdfs, _bad = _ppvp(inv_pdfs_all)
    except Exception:
        inv_pdfs, _bad = inv_pdfs_all, []

    if not inv_pdfs:
        return audit

    # Ensure engine importable
    engine_dir = str(settings.engine_dir)
    if engine_dir not in sys.path:
        sys.path.insert(0, engine_dir)

    synthesized: List[dict] = []
    try:
        from pz_import_processor import parse_invoice as _pi  # noqa: PLC0415
    except Exception as exc:
        log.warning("[%s] synthesize_rows: engine import failed: %s", batch_id, exc)
        return audit

    line_pos = 0
    for pdf in inv_pdfs:
        try:
            inv = _pi(str(pdf), [])
        except Exception:
            inv = None
        if not isinstance(inv, dict):
            continue
        fob = float(inv.get("fob_usd") or 0.0)
        if fob <= 0:
            continue
        invoice_no = (
            str(inv.get("invoice_no") or "").strip()
            or pdf.stem
        )
        counts_by_unit = inv.get("product_counts_by_unit") or {}
        # Sum qty across PCS / PRS groups; if engine returned empty dict,
        # fall back to a single PCS group inferred from total_units or
        # qty=1 sentinel so the row still carries SOMETHING.
        pcs_qty = 0
        prs_qty = 0
        try:
            pcs_qty = int(round(sum(float(v or 0) for v in
                                    (counts_by_unit.get("PCS") or {}).values())))
            prs_qty = int(round(sum(float(v or 0) for v in
                                    (counts_by_unit.get("PRS") or {}).values())))
        except Exception:
            pcs_qty = prs_qty = 0

        # Fallback: scan the engine's raw text for the GLOBAL/IEC summary
        # block layout ``PCS 183.0`` / ``PRS 62.0``. Pure read of engine
        # output — no parser arithmetic, no customs logic.
        if pcs_qty == 0 and prs_qty == 0:
            import re as _re_local
            raw = str(inv.get("_raw_text") or "")
            m_pcs = _re_local.search(r"\bPCS\s+(\d+(?:\.\d+)?)\b", raw)
            m_prs = _re_local.search(r"\bPRS\s+(\d+(?:\.\d+)?)\b", raw)
            if m_pcs:
                try: pcs_qty = int(round(float(m_pcs.group(1))))
                except Exception: pass
            if m_prs:
                try: prs_qty = int(round(float(m_prs.group(1))))
                except Exception: pass

        total_qty = pcs_qty + prs_qty
        if total_qty == 0:
            # Fallback: try engine top-level qty
            try:
                total_qty = int(round(float(
                    inv.get("total_pcs") or inv.get("total_units") or 0)))
            except Exception:
                total_qty = 0
            pcs_qty = total_qty
            prs_qty = 0
            if total_qty == 0:
                # Nothing to split — emit one grouped row with qty=1
                # so the reconciler sees a row sum = fob.
                pcs_qty = 1
                total_qty = 1

        # Allocate FOB proportionally to qty per unit so the sum is
        # exact. Rounding residue placed on the last row.
        groups: list[tuple[str, int]] = []
        if pcs_qty > 0:
            groups.append(("PCS", pcs_qty))
        if prs_qty > 0:
            groups.append(("PRS", prs_qty))

        alloc_remaining = round(fob, 2)
        for i, (unit, qty) in enumerate(groups):
            if i == len(groups) - 1:
                line_total = round(alloc_remaining, 2)
            else:
                # Proportional allocation by qty.
                line_total = round(fob * (qty / total_qty), 2)
                alloc_remaining = round(alloc_remaining - line_total, 2)
            line_pos += 1
            unit_price = round(line_total / qty, 6) if qty > 0 else 0.0
            synthesized.append({
                "invoice_number": invoice_no,
                "line_position":  line_pos,
                "product_code":   f"{invoice_no}-AGG-{unit}",
                "description":    (
                    f"{unit}, grouped invoice aggregate from "
                    f"{inv.get('invoice_format') or 'invoice'} ({pdf.name})"
                ),
                "item_type":      "",
                "quantity":       float(qty),
                "unit_price":     unit_price,
                "line_total":     line_total,
                "hsn_code":       "",
                "currency":       "USD",
                "uom":            unit,
            })

    if not synthesized:
        return audit

    audit["rows"]            = synthesized
    audit["_rows_source"]    = "synthesized_from_invoice_aggregates"
    audit["_rows_row_count"] = len(synthesized)
    log.info(
        "[%s] _synthesize_rows_from_invoice_aggregates: produced %d grouped "
        "row(s) summing to USD %.2f across %d invoice file(s)",
        batch_id, len(synthesized),
        sum(r["line_total"] for r in synthesized),
        len({r["invoice_number"] for r in synthesized}),
    )
    return audit


# ─────────────────────────────────────────────────────────────────────────────
# Supplier-detect + Global Jewellery PL/EN render (locked vocabulary)
# ─────────────────────────────────────────────────────────────────────────────

_GLOBAL_TYPE_TABLE: Dict[str, Dict[str, str]] = {
    "RING":     {"en": "RING",      "pl": "Pierścionek",          "label": "RING"},
    "PENDANT":  {"en": "PENDANT",   "pl": "Wisiorek",             "label": "PENDANT"},
    "EARRING":  {"en": "EARRINGS",  "pl": "Kolczyki",             "label": "EARRINGS"},
    "EARRINGS": {"en": "EARRINGS",  "pl": "Kolczyki",             "label": "EARRINGS"},
    "BRACELET": {"en": "BRACELET",  "pl": "Bransoletka",          "label": "BRACELET"},
    "BANGLE":   {"en": "BANGLE",    "pl": "Bransoletka sztywna",  "label": "BANGLE"},
    "NECKLACE": {"en": "NECKLACE",  "pl": "Naszyjnik",            "label": "NECKLACE"},
    "CHAIN":    {"en": "NECKLACE",  "pl": "Łańcuszek",            "label": "CHAIN"},
    "CUFFLINK": {"en": "CUFFLINKS", "pl": "Spinki do mankietów",  "label": "CUFFLINKS"},
    "CUFFLINKS":{"en": "CUFFLINKS", "pl": "Spinki do mankietów",  "label": "CUFFLINKS"},
}

_GLOBAL_METAL_TABLE: Dict[str, Dict[str, str]] = {
    # canonical key → PL phrase + EN form
    # Keys use the NNkt GOLD / PTNNN / 925 SILVER convention.
    # _normalise_metal_key() maps every parser variant to one of these keys.
    "925 SILVER":   {"pl": "ze srebra próby 925", "en": "925 Silver"},
    "9KT GOLD":     {"pl": "ze złota próby 375",  "en": "09KT Gold"},
    "9 GOLD":       {"pl": "ze złota próby 375",  "en": "09KT Gold"},
    "14KT GOLD":    {"pl": "ze złota próby 585",  "en": "14KT Gold"},
    "18KT GOLD":    {"pl": "ze złota próby 750",  "en": "18KT Gold"},
    "22KT GOLD":    {"pl": "ze złota próby 916",  "en": "22KT Gold"},
    "24KT GOLD":    {"pl": "ze złota próby 999",  "en": "24KT Gold"},   # fine gold
    "PT850":        {"pl": "z platyny próby 850", "en": "PT850 Platinum"},
    "PT900":        {"pl": "z platyny próby 900", "en": "PT900 Platinum"},
    "PT950":        {"pl": "z platyny próby 950", "en": "PT950 Platinum"},
}

# Karat-integer → standard Polish próby value.
# DO NOT derive via arithmetic (14×1000/24=583≠585, 22×1000/24=917≠916).
# These are industry-standard fineness marks, not mathematical results.
_KARAT_FINENESS: Dict[int, int] = {
    9: 375, 10: 417, 14: 585, 18: 750, 21: 875, 22: 916, 24: 999,
}


# ── Import-time grammar parity assertions (Phase 2C) ────────────────────────
# Verify that the packing renderer's local PL values match the shared grammar
# authority.  The renderer uses composite dict-of-dicts with a different key
# format (e.g. "14KT GOLD" vs "14KT"), so we compare VALUES, not keys.
#
# TYPE TABLE: every PL value in _GLOBAL_TYPE_TABLE must exist in ITEM_TYPE_PL.
# Note: CHAIN maps to "Łańcuszek" which IS in ITEM_TYPE_PL. The shared grammar
# has more keys (BROOCH, SET, ANKLET, STUD, HOOP) not present here — that's
# fine, the renderer only handles the subset it sees from packing lines.
_SHARED_TYPE_PL_VALUES = set(ITEM_TYPE_PL.values())
_RENDERER_TYPE_PL_VALUES = {v["pl"] for v in _GLOBAL_TYPE_TABLE.values()}
_TYPE_PL_DRIFT = _RENDERER_TYPE_PL_VALUES - _SHARED_TYPE_PL_VALUES
if _TYPE_PL_DRIFT:
    raise ImportError(
        f"routes_dhl_clearance: _GLOBAL_TYPE_TABLE PL values drifted from "
        f"shared grammar ITEM_TYPE_PL: {_TYPE_PL_DRIFT!r}. "
        f"Update description_grammar.py or fix the local table."
    )

# METAL TABLE: every PL value in _GLOBAL_METAL_TABLE must exist in
# METAL_PREPOSITIONAL.  Key format differs ("14KT GOLD" here vs "14KT" in
# shared), so compare value sets.
# Note: "9 GOLD" is a normalisation alias that maps to the same PL as "9KT GOLD".
_SHARED_METAL_PL_VALUES = set(METAL_PREPOSITIONAL.values())
_RENDERER_METAL_PL_VALUES = {v["pl"] for v in _GLOBAL_METAL_TABLE.values()}
_METAL_PL_DRIFT = _RENDERER_METAL_PL_VALUES - _SHARED_METAL_PL_VALUES
if _METAL_PL_DRIFT:
    raise ImportError(
        f"routes_dhl_clearance: _GLOBAL_METAL_TABLE PL values drifted from "
        f"shared grammar METAL_PREPOSITIONAL: {_METAL_PL_DRIFT!r}. "
        f"Update description_grammar.py or fix the local table."
    )

# Clean up module namespace
del _SHARED_TYPE_PL_VALUES, _RENDERER_TYPE_PL_VALUES, _TYPE_PL_DRIFT
del _SHARED_METAL_PL_VALUES, _RENDERER_METAL_PL_VALUES, _METAL_PL_DRIFT


def _normalise_type_key(item_type: str) -> str:
    """Normalise packing-list item-type codes to _GLOBAL_TYPE_TABLE keys.

    Uses the SAME alias map as the packing parser (_EJL_TOKEN_MAP) so the two
    vocabularies can't drift. Resolves every code the parser can emit, including
    2-letter aliases (ER→EARRING) and plurals (EARS→EARRING).

    Returns the full English word in UPPER CASE (e.g. "PENDANT", "RING") or
    the original uppercased code if unrecognised (caller treats as a miss).
    """
    # Shared with invoice_packing_extractor._EJL_TOKEN_MAP — keep in sync.
    # Map to UPPER CASE values matching _GLOBAL_TYPE_TABLE keys.
    _ALIAS: Dict[str, str] = {
        # Pendant
        "PND": "PENDANT",  "PEND": "PENDANT",  "PENDANT": "PENDANT",
        # Ring
        "RNG": "RING",     "RING":  "RING",
        # Earring — all EJL aliases including 2-letter ER and plural EARS
        "ERG": "EARRING",  "EAR":   "EARRING",  "ER":       "EARRING",
        "EARS": "EARRING", "ERS":   "EARRING",  "EARRING":  "EARRING",
        "EARRINGS": "EARRING",
        "PRS": "EARRING",   # EJL packing: "PRS" (pairs) = earrings
        # Bracelet
        "BRC": "BRACELET", "BR":    "BRACELET", "BRACELET": "BRACELET",
        # Necklace
        "NCK": "NECKLACE", "NK":    "NECKLACE", "NECKLACE": "NECKLACE",
        # Bangle
        "BNG": "BANGLE",   "BANGLE": "BANGLE",
        # Cufflinks
        "CFL": "CUFFLINK", "CUFFLINK": "CUFFLINK", "CUFFLINKS": "CUFFLINK",
        # Chain
        "CHN": "CHAIN",    "CHAIN": "CHAIN",
    }
    return _ALIAS.get((item_type or "").strip().upper(),
                      (item_type or "").strip().upper())


def _normalise_metal_key(metal: str) -> str:
    """Normalise ANY packing-list metal code to a _GLOBAL_METAL_TABLE key.

    Strategy (pattern-class approach — a new code in the same family resolves
    without another patch):

    Step 1 — strip qualifier suffix: split on '/' and keep only the first token.
             Handles 18KT/Y, 18KT/P, 18KT/RG, 18KT/YWPD, 925/-, PT950/-, etc.
    Step 2 — classify the first token by structure:
      GOLD    : matches r'(\\d+)\\s*KT?' (karat number)
                → look up fineness in _KARAT_FINENESS (NOT arithmetic)
                → return "<N>KT GOLD"  e.g. "18KT GOLD", "14KT GOLD"
                  (if karat unknown → legible error: not silent fallback)
      PLATINUM: contains 'PT' or 'PLAT' with optional 3-digit suffix
                → normalise to "PTNNN" e.g. "PT950", "PT900", "PT850"
      SILVER  : contains '925' or '999' or starts with SL/SS/SILVER
                → "925 SILVER" or "999 SILVER"
                  EXCEPTION: bare '999' is ambiguous (fine gold or fine silver)
                  → return sentinel "999_AMBIGUOUS" so the caller raises legibly.
    Step 3 — pass-through if already a canonical key.

    Returns the canonical key string. Unrecognised codes that clear step 1-3
    return the post-strip token; caller sees a miss in the table and raises.

    NOTE: '999' alone is intentionally ambiguous. Do not map it silently.
    The operator or a future mapping decision must resolve it.
    """
    import re as _re_m

    s = (metal or "").strip().upper()
    if not s:
        return s

    # Step 3 fast-path: already a canonical key
    if s in _GLOBAL_METAL_TABLE:
        return s

    # Step 1: strip qualifier suffix — everything after the first '/'
    base = s.split("/")[0].strip()

    # Step 2a — GOLD: NNkt or NN karat
    m_karat = _re_m.match(r'^(\d+)\s*K(?:T)?$', base)
    if m_karat:
        karat_int = int(m_karat.group(1))
        # Bare '24' from '24KT' is unambiguous gold; '999' without KT is ambiguous.
        if karat_int == 24:
            return "24KT GOLD"
        if karat_int not in _KARAT_FINENESS:
            # Unknown karat — return raw so caller raises legibly
            return base
        return f"{karat_int}KT GOLD"

    # Step 2b — PLATINUM: PT followed by 3 digits (any order, optional space)
    m_pt = _re_m.search(r'PT\s*(\d{3})', base) or _re_m.search(r'(\d{3})\s*PT', base)
    if m_pt or 'PLAT' in base:
        if m_pt:
            fineness = m_pt.group(1)
            return f"PT{fineness}"
        # "PLAT" without digits — pass through as-is (caller may miss)
        return base

    # Step 2c — SILVER: 925, SL925, SS925, SILVER 925, 999
    # Bare 999 is ambiguous — return sentinel
    if _re_m.search(r'(?<!\d)999(?!\d)', base):
        return "999_AMBIGUOUS"
    if _re_m.search(r'(?<!\d)925(?!\d)', base) or base.startswith(("SL", "SS")) or "SILVER" in base:
        return "925 SILVER"

    # Pass-through: caller looks up in table; miss = legible failure
    return base


class _UnrecognisedMetalCode(ValueError):
    """Raised by _global_render_pl_en when a metal code cannot be classified.

    Surfaced as a legible 422 (not a silent 'metal szlachetny' fallback).
    """


def _global_render_pl_en(item_type: str, metal: str, stone_text: str,
                         _row_context: str = "") -> Dict[str, str]:
    """Render operator-locked PL/EN description for a Global packing row.

    Inputs come straight from packing_lines columns. Rules:
      - Type token (Ring/Bracelet/...) → Pierścionek/Bransoletka/...
      - Metal canonical: colour-suffix codes (18KT/Y, 18KT/P, …) and all
        silver/platinum variants normalised by _normalise_metal_key().
      - Stone text scanned for vocabulary markers.
      - Unknown TYPE → returns empty strings (caller skips row).
      - Unknown METAL → raises _UnrecognisedMetalCode (legible error surfaced
        to operator; never falls through to 'metal szlachetny' placeholder).
      - '999' alone → raises _UnrecognisedMetalCode (ambiguous: fine gold vs
        fine silver — operator must map it before the Polish description runs).
    """
    import re as _re_g
    t_key = _normalise_type_key(item_type)
    if t_key.endswith("S") and t_key[:-1] in _GLOBAL_TYPE_TABLE:
        t_key = t_key[:-1]
    type_info = _GLOBAL_TYPE_TABLE.get(t_key)
    if not type_info:
        # Unknown type — skip row (same as before; not a legible error)
        return {"pl": "", "en": "", "item_type": "", "item_type_pl": ""}

    metal_key = _normalise_metal_key(metal)
    if metal_key == "999_AMBIGUOUS":
        raise _UnrecognisedMetalCode(
            f"Metal code '999' is ambiguous (fine gold vs fine silver) on "
            f"row {_row_context!r}. Set a specific metal code "
            f"(e.g. '24KT GOLD' or '999 SILVER') before generating the "
            f"Polish description."
        )
    metal_info = _GLOBAL_METAL_TABLE.get(metal_key)
    if not metal_info:
        raise _UnrecognisedMetalCode(
            f"Unrecognised metal code {metal!r} (normalised to {metal_key!r}) "
            f"on row {_row_context!r}. Add it to _GLOBAL_METAL_TABLE or "
            f"correct the packing-list data before generating the Polish description."
        )

    # Stone vocabulary scan
    stone_up = (stone_text or "").upper()
    stone_pl, stone_en = "", "Plain Jewellery"
    if _re_g.search(r"\bLGD\b|\bLAB\s*ROUND\s*DIA\b|\bLAB\s*GROWN\b", stone_up):
        stone_pl, stone_en = ("z diamentami laboratoryjnymi", "Lab Grown Diamond Jewellery")
    elif _re_g.search(r"\bDIA\b.*\bCZ\b|\bCZ\b.*\bDIA\b", stone_up):
        stone_pl, stone_en = (
            "wysadzany diamentami i cyrkoniami",
            "Diamond & CZ Stud Jewellery",
        )
    elif _re_g.search(r"\bCZ\b", stone_up) and _re_g.search(
            r"\bCLS\b|\b(COLOUR|COLOR)\s+STONE\b|\bSAPPHIRE\b|\bRUBY\b|\bEMERALD\b|\bAMETHYST\b|\bTOPAZ\b|\bTANZANITE\b",
            stone_up):
        stone_pl, stone_en = (
            "wysadzany cyrkoniami i kamieniami kolorowymi",
            "CZ & Colour Stone Jewellery",
        )
    elif _re_g.search(r"\bCZ\b", stone_up):
        stone_pl, stone_en = ("wysadzany cyrkoniami", "CZ Stud Jewellery")
    elif _re_g.search(r"\bDIA\b|\bDIAMOND\b", stone_up):
        stone_pl, stone_en = ("z diamentami", "Diamond Jewellery")
    elif _re_g.search(
            r"\bCLS\b|\b(COLOUR|COLOR)\s+STONE\b|\bSAPPHIRE\b|\bRUBY\b|\bEMERALD\b|\bAMETHYST\b|\bTOPAZ\b|\bTANZANITE\b",
            stone_up):
        stone_pl, stone_en = (
            "wysadzany kamieniami kolorowymi",
            "Colour Stone Jewellery",
        )

    pl = (type_info["pl"] + " " + metal_info["pl"]
          + (" " + stone_pl if stone_pl else "")).strip()
    en = (metal_info["en"] + " " + stone_en + " " + type_info["label"]).strip()
    return {
        "pl":           pl,
        "en":           en,
        "item_type":    type_info["en"],
        "item_type_pl": type_info["pl"],
    }


def _detect_global_supplier_for_batch(batch_id: str) -> bool:
    """Detect Global Jewellery supplier from any uploaded source file.

    Scans first 1k chars of any PDF in source/invoices/ or source/packing/
    via supplier_detect.detect_supplier. Pure read; no DB writes.
    """
    try:
        from ..services.supplier_detect import detect_supplier  # noqa: PLC0415
        import pdfplumber  # noqa: PLC0415
    except Exception:
        return False
    for sub in ("outputs", "working"):
        base = settings.storage_root / sub / batch_id / "source"
        if not base.is_dir():
            continue
        for cat in ("invoices", "packing"):
            d = base / cat
            if not d.is_dir():
                continue
            for pdf in d.glob("*.pdf"):
                try:
                    with pdfplumber.open(str(pdf)) as p:
                        if not p.pages:
                            continue
                        head = (p.pages[0].extract_text() or "")[:1000]
                    if detect_supplier(head) == "global_jewellery":
                        return True
                except Exception:
                    continue
        break
    return False


_STALE_AGGREGATE_MARKERS = (
    "synthesized_from_invoice_aggregates",
    "grouped invoice aggregate",
    "-AGG-PCS",
    "-AGG-PRS",
)


def _audit_rows_are_stale_aggregate(audit: dict) -> bool:
    """Detect rows persisted from C27.2's aggregate synthesizer that
    pre-date the packing-first authority. Must be purged for Global
    batches so the chain rebuilds from packing_lines."""
    if audit.get("_rows_source") == "synthesized_from_invoice_aggregates":
        return True
    for r in (audit.get("rows") or []):
        pc = str(r.get("product_code") or "")
        desc = str(r.get("description") or "")
        if any(m in pc or m in desc for m in _STALE_AGGREGATE_MARKERS):
            return True
    return False


def _purge_stale_audit_rows(audit: dict) -> None:
    """Remove rows + source markers in-place. Used when a Global batch
    has cached aggregate rows from C27.2."""
    for k in ("rows", "invoices", "_rows_source", "_rows_row_count",
              "_global_packing_present_but_empty"):
        audit.pop(k, None)


def _force_reparse_global_packing(batch_id: str) -> int:
    """Re-parse Global packing PDFs against the LIVE parser and refresh
    packing.db with the result. Used by the `?force=true` regenerate
    path to heal stale packing_lines rows persisted by an older parser
    version (e.g. rows with empty `metal` extracted before the lenient
    style-metal split landed).

    Returns the number of rows upserted. Returns 0 on any failure
    (logged) — callers treat this as best-effort recovery.

    NEVER touches CIF / customs threshold / wFirma gates / SAD/ZC429.
    Pure refresh of the per-row product authority for Global supplier.
    """
    try:
        from ..services.global_packing_parser import parse_global_packing_pdf  # noqa: PLC0415
        from ..services import packing_db as _pdb  # noqa: PLC0415
    except Exception as exc:
        log.warning("[%s] force_reparse: import failed: %s", batch_id, exc)
        return 0

    # Locate packing PDFs in source/packing/
    pkg_dir: Optional[Path] = None
    for sub in ("outputs", "working"):
        d = settings.storage_root / sub / batch_id / "source" / "packing"
        if d.is_dir():
            pkg_dir = d
            break
    if pkg_dir is None:
        return 0

    pdfs = sorted(pkg_dir.glob("*.pdf"))
    if not pdfs:
        return 0

    # Need the existing packing_document_id to attach reparsed rows to.
    try:
        from ..services import document_db as _ddb  # noqa: PLC0415
        docs = _ddb.get_documents_for_batch(
            batch_id, document_type="purchase_packing_list",
        ) or []
    except Exception as exc:
        log.warning("[%s] force_reparse: doc lookup failed: %s", batch_id, exc)
        docs = []
    if not docs:
        return 0

    # Use the first matching doc as the packing_document_id parent
    pkg_doc_id = str(docs[0].get("id") or "")
    if not pkg_doc_id:
        return 0

    # Parse all PDFs in source/packing/ — combine rows
    all_rows: List[dict] = []
    for pdf in pdfs:
        try:
            rows, _, _, _ = parse_global_packing_pdf(pdf)
        except Exception:
            rows = []
        for r in rows:
            r["batch_id"]            = batch_id
            r["packing_document_id"] = pkg_doc_id
            all_rows.append(r)

    if not all_rows:
        return 0

    # Replace existing packing_lines for this batch atomically. The
    # delete + upsert pattern matches what the operator's "Reparse
    # Packing" endpoint does internally; we just call it from the
    # force-regenerate path so the operator doesn't need two clicks.
    import sqlite3 as _sql
    try:
        # Best-effort delete of stale rows (packing_db doesn't expose a
        # clear_batch helper; use direct SQL with the same connection
        # pattern packing_db uses).
        db_path = getattr(_pdb, "_db_path", None)
        if db_path:
            with _sql.connect(str(db_path)) as con:
                con.execute(
                    "DELETE FROM packing_lines WHERE batch_id = ?",
                    (batch_id,),
                )
                con.commit()
    except Exception as exc:
        log.warning(
            "[%s] force_reparse: stale row delete failed (proceeding "
            "with upsert anyway): %s", batch_id, exc,
        )

    try:
        n = _pdb.upsert_packing_lines(all_rows, force_reextract=True)
    except Exception as exc:
        log.warning("[%s] force_reparse: upsert failed: %s", batch_id, exc)
        return 0

    log.info(
        "[%s] force_reparse: refreshed %d packing_lines rows from "
        "%d PDF(s) using live parser",
        batch_id, n, len(pdfs),
    )
    return n


def _inject_rows_from_packing_lines(batch_id: str, audit: dict) -> dict:
    """Project packing.db packing_lines into ``audit["rows"]`` — first
    authority for Global Jewellery shipments.

    Operator spec:
      1. Packing lines are the FIRST authority for customs rows when the
         supplier is Global Jewellery.
      2. If packing_lines count > 0 → use them; never fall through to the
         aggregate synthesizer.
      3. If a Global packing file exists on disk but produced 0 rows →
         flag ``audit["_global_packing_present_but_empty"]`` so the route
         layer raises 422.
      4. **Stale-aggregate purge.** When audit carries rows persisted by
         C27.2's aggregate synthesizer AND the supplier is Global, those
         rows are evicted before the packing path runs — otherwise the
         idempotency-on-rows guard would prevent rebuilding from
         packing_lines.

    Behaviour for non-Global suppliers: no-op. Estrella EJL path is
    untouched.
    """
    is_global = _detect_global_supplier_for_batch(batch_id)
    if not is_global:
        return audit  # non-Global → fall through to existing chain

    # Global path: evict stale aggregate cache before idempotency check.
    if audit.get("rows") and _audit_rows_are_stale_aggregate(audit):
        log.info(
            "[%s] _inject_rows_from_packing_lines: purging %d stale "
            "aggregate row(s) so packing_lines can become authority",
            batch_id, len(audit.get("rows") or []),
        )
        _purge_stale_audit_rows(audit)

    if audit.get("rows") or audit.get("invoices"):
        return audit

    # Check whether any Global packing file exists on disk (to support the
    # 422 guard at the route layer when parser produced 0 rows).
    packing_files_exist = False
    for sub in ("outputs", "working"):
        d = settings.storage_root / sub / batch_id / "source" / "packing"
        if d.is_dir():
            packing_files_exist = bool(list(d.glob("*.pdf")) or list(d.glob("*.xlsx"))
                                       or list(d.glob("*.xls")))
            break

    try:
        from ..services.packing_db import get_packing_lines_for_batch  # noqa: PLC0415
        pkg_rows = get_packing_lines_for_batch(batch_id) or []
    except Exception as exc:
        log.warning("[%s] packing_lines read failed: %s", batch_id, exc)
        pkg_rows = []

    if not pkg_rows:
        if packing_files_exist:
            # Signal the route layer to block with 422 instead of falling
            # through to the aggregate synthesizer that would emit UNKNOWN.
            audit["_global_packing_present_but_empty"] = True
        return audit

    # Build audit rows from packing_lines with operator-locked PL/EN render.
    out_rows: List[dict] = []
    skipped: int = 0
    for ln in pkg_rows:
        item_type = str(ln.get("item_type") or "").strip()
        metal     = str(ln.get("metal") or "").strip()
        stone     = str(ln.get("stone_type") or ln.get("remarks") or "").strip()
        product_code = str(ln.get("product_code") or "").strip()
        if not product_code or not item_type or not metal:
            skipped += 1
            continue

        row_ctx = f"{product_code} ({item_type}/{metal})"
        try:
            desc = _global_render_pl_en(item_type, metal, stone,
                                        _row_context=row_ctx)
        except _UnrecognisedMetalCode as _exc:
            # Legible failure: propagate immediately so the operator sees
            # the exact code + row — never fall through to "metal szlachetny".
            raise HTTPException(
                status_code=422,
                detail={
                    "guard":  "unrecognised_metal_code",
                    "error":  str(_exc),
                    "code":   "unrecognised_metal_code",
                    "row":    row_ctx,
                    "hint":   "Correct the metal code on the packing list or "
                              "add a mapping before regenerating.",
                },
            ) from _exc
        if not desc.get("pl") or not desc.get("en"):
            # Unknown item type (not metal — that would have raised above).
            skipped += 1
            continue

        qty       = float(ln.get("quantity") or 0)
        unit_p    = float(ln.get("unit_price") or 0)
        line_tot  = float(ln.get("total_value") or unit_p * qty)
        out_rows.append({
            "invoice_number":             str(ln.get("invoice_no") or ""),
            "line_position":              int(ln.get("invoice_line_position") or 0),
            "product_code":               product_code,
            "description":                desc["en"],
            "description_pl":             desc["pl"],
            "description_en":             desc["en"],
            "polish_customs_description": desc["pl"],
            "item_type":                  desc["item_type"],
            "item_type_pl":               desc["item_type_pl"],
            "material":                   "",   # engine derives from desc text
            "quantity":                   qty,
            "unit_price":                 unit_p,
            "line_total":                 line_tot,
            "line_total_usd":             line_tot,
            "hsn_code":                   "",
            "currency":                   "USD",
            "uom":                        str(ln.get("uom") or "PCS"),
            "net_weight":                 float(ln.get("net_weight") or 0),
            "gross_weight":               float(ln.get("gross_weight") or 0),
            "design_no":                  str(ln.get("design_no") or ""),
            "_supplier_profile":          "global_jewellery",
            "_rows_source":               "packing_lines",
        })

    if not out_rows:
        if packing_files_exist:
            audit["_global_packing_present_but_empty"] = True
        return audit

    audit["rows"]            = out_rows
    audit["_rows_source"]    = "packing_lines"
    audit["_rows_row_count"] = len(out_rows)
    log.info(
        "[%s] _inject_rows_from_packing_lines: %d Global packing rows "
        "(skipped=%d, fob_sum=%.2f)",
        batch_id, len(out_rows), skipped,
        sum(r["line_total"] for r in out_rows),
    )
    return audit


def _try_inject_invoice_positions_for_global(batch_id: str, audit: dict) -> bool:
    """For Global Jewellery supplier, parse the commercial invoice PDF
    directly into per-position customs rows.

    This is the authority for the Customs Description Report when
    supplier is Global: ONE row per invoice commercial line. Replaces
    the packing-row aggregation path that was collapsing distinct
    invoice lines into artificial groups (e.g. positions with
    "CZ + Colour Stone" vs "CZ alone" both being merged because both
    text-matched on CZ).

    Returns True on success (audit["rows"] populated with invoice
    positions). Returns False on any failure — caller continues with
    the existing packing-row chain.

    Pure read of files in source/invoices/. Never touches CIF,
    customs threshold, wFirma, PZ, DB schema, or engine arithmetic.
    """
    try:
        from ..services.global_invoice_position_parser import (
            parse_invoice_positions_from_pdf, positions_to_audit_rows,
        )
    except Exception as exc:
        log.warning("[%s] invoice-position parser import failed: %s",
                    batch_id, exc)
        return False

    # Locate the invoice PDF — Global ships ONE commercial invoice per
    # batch in source/invoices/. Use first valid PDF.
    inv_dir: Optional[Path] = None
    for sub in ("outputs", "working"):
        candidate = settings.storage_root / sub / batch_id / "source" / "invoices"
        if candidate.is_dir():
            inv_dir = candidate
            break
    if inv_dir is None:
        return False

    pdfs = sorted(inv_dir.glob("*.pdf"))
    if not pdfs:
        return False

    # Re-use the C27.1 quarantine helper so we don't feed a non-PDF
    # (XLS-renamed-to-PDF) to the parser.
    try:
        from .routes_dashboard import _partition_valid_pdfs as _ppvp
        valid_pdfs, _ = _ppvp(pdfs)
    except Exception:
        valid_pdfs = pdfs

    if not valid_pdfs:
        return False

    # Use the first valid invoice PDF for parsing. Multi-invoice
    # batches are out of scope for this PR (Global ships one
    # commercial invoice per batch per the spec).
    inv_pdf = valid_pdfs[0]
    positions = parse_invoice_positions_from_pdf(inv_pdf)
    if not positions:
        return False

    # Resolve invoice_no from the engine's parse_invoice (it knows the
    # Global format) — fallback to the file stem.
    invoice_no = ""
    try:
        engine_dir = str(settings.engine_dir)
        if engine_dir not in sys.path:
            sys.path.insert(0, engine_dir)
        from pz_import_processor import parse_invoice as _pi  # noqa: PLC0415
        inv = _pi(str(inv_pdf), [])
        if isinstance(inv, dict):
            # Prefer the Global Exporter's Ref pattern (088/2026-2027)
            # from raw_text; fall back to engine's invoice_no field.
            import re as _re_local
            raw = str(inv.get("_raw_text") or "")
            m = _re_local.search(r"\b(\d{1,4}/\d{4}-\d{4})\b", raw)
            if m:
                invoice_no = m.group(1)
            else:
                invoice_no = str(inv.get("invoice_no") or "").strip()
    except Exception:
        pass
    if not invoice_no:
        invoice_no = inv_pdf.stem

    rows = positions_to_audit_rows(positions, invoice_no)
    if not rows:
        return False

    # Reconciliation: row sum must match the engine's parsed FOB
    # (or declared invoice_totals.total_fob_usd) within $1.
    declared_fob = 0.0
    try:
        declared_fob = float((audit.get("invoice_totals") or {})
                             .get("total_fob_usd") or 0.0)
    except Exception:
        declared_fob = 0.0
    row_sum = round(sum(float(r.get("line_total") or 0) for r in rows), 2)
    if declared_fob > 0 and abs(row_sum - declared_fob) > 1.00:
        log.warning(
            "[%s] invoice-position rows reject: sum %.2f differs from "
            "declared FOB %.2f by %.2f — falling back to packing path",
            batch_id, row_sum, declared_fob, row_sum - declared_fob,
        )
        return False

    audit["rows"]            = rows
    audit["_rows_source"]    = "invoice_positions_authority"
    audit["_rows_row_count"] = len(rows)
    audit["_customs_aggregation"] = {
        "source": "commercial_invoice_lines",
        "position_count": len(rows),
        "fob_sum_preserved": row_sum,
    }
    # ── PZ engine authority sidecar (Bridge Persistence, 2026-05-21) ──────
    # `audit.rows` is legitimately overwritten by the PZ engine on every
    # /process run (it carries the engine's per-row pipeline output). The
    # invoice-position authority must therefore live in a dedicated key the
    # engine never touches. The engine bridge (pz_import_processor.
    # _try_invoice_from_authority_rows) prefers this key; audit_merge.
    # PRESERVED_KEYS grants it safe passage across regenerations.
    audit["_pz_engine_authority_rows"] = rows
    audit["_pz_engine_authority_meta"] = {
        "source":            "invoice_positions_authority",
        "captured_at":       datetime.utcnow().strftime("%Y-%m-%dT%H:%M:%SZ"),
        "fob_sum_preserved": row_sum,
        "row_count":         len(rows),
        "invoice_pdf":       inv_pdf.name,
    }
    log.info(
        "[%s] customs authority = invoice positions: %d positions "
        "from %s (sum USD %.2f)",
        batch_id, len(rows), inv_pdf.name, row_sum,
    )
    return True


def _inject_rows_from_sources(
    batch_id: str,
    audit: dict,
    *,
    customs_view: str = "invoice_positions",
) -> dict:
    """Chain row sources in priority order.

    Order:
      -1. Global supplier invoice-position parser (NEW; authoritative
          customs row source — one row per commercial invoice line)
      0. _inject_rows_from_packing_lines           (Global per-row, used for packing_rows view)
      1. _inject_rows_from_db_invoice_lines        (DB primary, placeholder-filtered)
      2. _inject_rows_from_xlsx                    (legacy XLSX Rows sheet)
      3. _synthesize_rows_from_invoice_aggregates  (engine-aggregate grouped fallback)

    Operator spec (invoice-line authority):
      - Customs Description Report uses INVOICE COMMERCIAL LINE
        authority — ONE row per invoice position. The previous packing-
        row-aggregation approach (PR #267) collapsed distinct invoice
        lines into artificial groups; this is corrected here.
      - Packing Description Report still available via
        ``customs_view="packing_rows"`` — falls through to per-row chain.

    Idempotent on subsequent calls. Caller MUST apply the lines-missing
    guard (HTTP 422) when the chain still produces no rows.
    """
    # Step -1: Global supplier invoice-position authority (DEFAULT
    # customs_view). When the supplier is Global Jewellery and the
    # caller wants the customs view (not warehouse-detail), parse the
    # commercial invoice directly. On success, the rest of the chain
    # no-ops because audit["rows"] is populated.
    if customs_view == "invoice_positions":
        try:
            if _detect_global_supplier_for_batch(batch_id):
                # Purge any stale aggregate cache first (PR #260 helper)
                if audit.get("rows") and _audit_rows_are_stale_aggregate(audit):
                    _purge_stale_audit_rows(audit)
                _try_inject_invoice_positions_for_global(batch_id, audit)
        except Exception as exc:
            log.warning(
                "[%s] invoice-position injection failed (non-fatal): %s",
                batch_id, exc,
            )

    audit = _inject_rows_from_packing_lines(batch_id, audit)
    audit = _inject_rows_from_db_invoice_lines(batch_id, audit)
    audit = _inject_rows_from_xlsx(batch_id, audit)
    audit = _synthesize_rows_from_invoice_aggregates(batch_id, audit)

    return audit


def _reconcile_rows_with_audit_totals(audit: dict) -> dict:
    """Validate ``audit["rows"]`` (or projected invoices) against the
    aggregate totals declared in ``audit["invoice_totals"]``.

    Pure / deterministic / side-effect free.  Returns a dict::

        {
          "ok":            bool,   # strict: all checks pass (backward-compat)
          "ok_hard":       bool,   # generation gate: hard failures only
          "hard_warnings": list[str],
          "soft_warnings": list[str],
          "warnings":      list[str],  # = hard + soft
          "details": {
            "row_count":              int,
            "row_invoice_numbers":    list[str],
            "row_fob_sum":            float,
            "row_qty_total":          int,
            "audit_invoice_names":    list[str],
            "audit_fob_total":        float,
            "audit_qty_total":        int,
            "fob_drift_usd":          float,
            "qty_drift_units":        int,
            "missing_in_rows":        list[str],
            "extra_in_rows":          list[str],
          }
        }

    Hard blocks (ok_hard=False):
      - no rows projected
      - missing invoice authority (invoices present in audit but absent from rows)
      - FOB drift above $1 tolerance
      - negative line_total in any row
      - zero or negative quantity in any row with an explicit quantity field

    Advisory only (ok=False, ok_hard=True):
      - quantity drift when FOB is within tolerance
        (parser divergence between invoice_intake_parser and pz_import_processor
        produces ±2–3 unit noise; FOB is the financial authority)
    """
    import re as _re

    def _safe_float(x):
        try: return float(x or 0)
        except Exception: return 0.0

    rows = audit.get("rows") or []
    invoice_totals = audit.get("invoice_totals") or {}
    audit_names    = audit.get("invoice_names") or []

    # Row-side aggregates + per-row validity scan
    row_inv_set: set = set()
    row_fob_sum = 0.0
    row_qty_tot = 0
    neg_value_indices: list[int] = []
    invalid_qty_indices: list[int] = []
    for i, r in enumerate(rows):
        inv = str(r.get("invoice_number") or r.get("invoice_no") or "").strip()
        if inv:
            row_inv_set.add(inv)
        lt = _safe_float(r.get("line_total") or r.get("amount") or r.get("total"))
        row_fob_sum += lt
        if lt < 0:
            neg_value_indices.append(i)
        # Explicit key check: `quantity or qty` treats 0 as missing (falsy).
        if "quantity" in r:
            raw_qty = r["quantity"]
        elif "qty" in r:
            raw_qty = r["qty"]
        else:
            raw_qty = None
        if raw_qty is not None:
            try:
                qty_val = int(round(float(raw_qty)))
                row_qty_tot += qty_val
                if qty_val <= 0:
                    invalid_qty_indices.append(i)
            except Exception:
                pass

    # Audit-side aggregates
    audit_fob = _safe_float(invoice_totals.get("total_fob_usd"))
    audit_units = invoice_totals.get("total_units")
    try:
        audit_qty = int(round(_safe_float(audit_units))) if audit_units is not None else 0
    except Exception:
        audit_qty = 0

    # Extract invoice-number tokens from both sides.
    #
    # invoice_names entries look like one of:
    #     "180 Invoice EJL-26-27-180-16-05-26.pdf"   ← leading numeric token
    #     "Invoice EJL-26-27-180-16-05-26.pdf"       ← no leading number
    # Row-side invoice_number is the canonical "EJL/26-27/180" form.
    #
    # We try (in order): EJL-NN-NN-NNN  pattern → leading "^\d+" →
    # last "\d{3,}" digit run in the stem.  Same rule for both sides
    # so the resulting sets compare apples to apples.
    def _token(s: str) -> Optional[str]:
        if not s:
            return None
        # Normalise separators
        norm = s.replace("\\", "/").replace("_", "-")
        # EJL[-/]NN[-/]NN[-/](NNN…)
        m = _re.search(r"EJL[\-/]\d+[\-/]\d+[\-/](\d+)", norm, _re.IGNORECASE)
        if m:
            return m.group(1)
        stem = Path(norm).stem
        m = _re.match(r"^(\d+)\b", stem)
        if m:
            return m.group(1)
        m = _re.search(r"(\d{3,})", stem)
        return m.group(1) if m else None

    audit_tokens = {t for t in (_token(n) for n in audit_names) if t}
    row_tokens   = {t for t in (_token(i) for i in row_inv_set)  if t}

    missing_in_rows = sorted(audit_tokens - row_tokens) if audit_tokens else []
    extra_in_rows   = sorted(row_tokens - audit_tokens) if audit_tokens else []

    fob_drift = round(row_fob_sum - audit_fob, 2) if audit_fob else 0.0
    qty_drift = (row_qty_tot - audit_qty) if audit_qty else 0

    hard_warnings: list[str] = []
    soft_warnings: list[str] = []

    # Hard failure — no rows: self-sufficient guard (callers also check, but
    # the function must not silently pass an empty row set).
    if not rows:
        hard_warnings.append(
            "no_rows: no projected per-line rows available; "
            "generating from aggregate totals would lose per-line detail"
        )

    # Hard failure — negative line values or zero/negative quantities.
    if neg_value_indices:
        hard_warnings.append(
            f"negative_line_total: rows at index {neg_value_indices} "
            "have negative FOB value — credit notes or data corruption"
        )
    if invalid_qty_indices:
        hard_warnings.append(
            f"zero_or_negative_qty: rows at index {invalid_qty_indices} "
            "have zero or negative quantity — invalid for a customs line"
        )

    # Hard failures — financial integrity or missing invoice coverage.
    if audit_fob and abs(fob_drift) > 1.00:
        hard_warnings.append(
            f"fob_total_drift: row sum USD {row_fob_sum:,.2f} differs from "
            f"audit total USD {audit_fob:,.2f} by USD {fob_drift:+,.2f}"
        )
    if missing_in_rows:
        hard_warnings.append(
            "invoices_missing_in_rows: "
            + ", ".join(missing_in_rows)
            + " present in invoice_names but not in projected rows"
        )

    # Soft advisory — qty drift when FOB is exact.
    # invoice_intake_parser and pz_import_processor parse the same PDFs via
    # different code paths and can produce ±2–3 unit counting differences
    # (e.g. PRS unit-vs-pair convention).  FOB is the financial authority for
    # the customs document; exact FOB with minor qty noise is not a document
    # integrity risk.  Callers that need the strict check should inspect
    # ok_strict or soft_warnings directly.
    if audit_qty and qty_drift != 0:
        soft_warnings.append(
            f"qty_total_drift: row qty {row_qty_tot} differs from "
            f"audit total {audit_qty} by {qty_drift:+d}"
        )

    warnings = hard_warnings + soft_warnings
    return {
        "ok":            not warnings,       # strict: all checks pass (backward-compat)
        "ok_hard":       not hard_warnings,  # generation gate: FOB + missing only
        "hard_warnings": hard_warnings,
        "soft_warnings": soft_warnings,
        "warnings":      warnings,
        "details":       {
            "row_count":           len(rows),
            "row_invoice_numbers": sorted(row_inv_set),
            "row_fob_sum":         round(row_fob_sum, 2),
            "row_qty_total":       row_qty_tot,
            "audit_invoice_names": list(audit_names),
            "audit_fob_total":     round(audit_fob, 2),
            "audit_qty_total":     audit_qty,
            "fob_drift_usd":       fob_drift,
            "qty_drift_units":     qty_drift,
            "missing_in_rows":     missing_in_rows,
            "extra_in_rows":       extra_in_rows,
        },
    }


def _find_sad_json(batch_id: str, awb: Optional[str] = None) -> Optional[Path]:
    """Find the SAD_READY JSON file for a batch."""
    # Search by awb-derived filename pattern or scan the sad_ready dir
    search_dirs = [
        _SAD_READY_DIR,
        _POLISH_DESC_DIR,
    ]
    for d in search_dirs:
        if not d.exists():
            continue
        for f in d.iterdir():
            if f.suffix == ".json" and f.name.startswith("SAD_READY_"):
                # If awb is known, match on it
                if awb:
                    awb_clean = re.sub(r"\s+", "", awb)
                    if awb_clean in f.name:
                        return f
                else:
                    return f  # Return first found
    return None


# ── New request schemas ───────────────────────────────────────────────────────

class GenerateCustomsPackageRequest(BaseModel):
    awb:          str
    dhl_email_id: Optional[str] = None
    date_override: Optional[str] = None


class ApproveDescriptionRequest(BaseModel):
    approved_by: str


class MarkEmailReceivedRequest(BaseModel):
    sender:       Optional[str] = "odprawacelna@dhl.com"
    subject:      Optional[str] = None
    ticket:       Optional[str] = None
    request_type: Optional[str] = "unknown"   # polish_description | dsk_broker | unknown
    received_at:  Optional[str] = None        # ISO datetime string; defaults to now
    note:         Optional[str] = None


# ── Endpoints ─────────────────────────────────────────────────────────────────

@router.get("/scan-inbox", dependencies=[_auth])
async def scan_dhl_inbox(
    limit:      int = 50,
    account_id: str = _ZOHO_ACCOUNT_ID,
    folder_id:  str = _ZOHO_INBOX_FOLDER,
    batch_id:   Optional[str] = None,
    awb:        Optional[str] = None,
    refresh:    bool = False,
) -> Dict[str, Any]:
    """
    Scan Zoho Mail inbox for shipment correspondence.

    When `batch_id` is provided, the AWB is auto-resolved from that batch's
    audit.json and an AWB-targeted search is performed. Otherwise, the most
    recent `limit` messages are scanned.

    Matching is permissive: subject, body, attachments, and forwarded content
    are all checked across DHL/agency/Ganther/internal/FedEx senders.

    Optional: `awb` — explicit AWB to search for (overrides batch_id resolution).
    """
    from dhl_email_monitor import scan_for_dhl_customs_emails
    from ..services.zoho_auth import (
        has_zoho_credentials,
        get_valid_access_token,
        ZohoAuthError,
    )

    # Routing mode — see settings.email_scan_mode for full semantics
    scan_mode = (settings.email_scan_mode or "auto").lower()
    api_allowed = scan_mode != "bridge_only"
    bridge_allowed = scan_mode != "api_only"
    creds_present = has_zoho_credentials() and api_allowed
    # Override account ID from settings if configured
    if settings.zoho_mail_account_id:
        account_id = settings.zoho_mail_account_id

    # ── Resolve AWB from batch_id if not given ────────────────────────────────
    target_awb = (awb or "").strip() or None
    if batch_id and not target_awb:
        for sub in ("outputs", "working"):
            _ap = settings.storage_root / sub / batch_id / "audit.json"
            if _ap.exists():
                try:
                    import json as _json
                    a = _json.loads(_ap.read_text(encoding="utf-8"))
                    target_awb = (
                        a.get("awb")
                        or a.get("tracking_no")
                        or (a.get("batch_meta") or {}).get("awb")
                    )
                except Exception:
                    pass
                break

    # Bridge dispatcher — used by bridge_only mode and as auto-fallback
    def _dispatch_to_bridge(reason: str) -> Dict[str, Any]:
        from ..services.ai_bridge import create_task as _bridge_create_task
        from ..services.email_search_context import build_email_search_context

        # Pull the audit so we can extract invoice numbers / DHL ticket / MRN
        ctx_audit: Dict[str, Any] = {}
        if batch_id:
            for sub in ("outputs", "working"):
                _ap = settings.storage_root / sub / batch_id / "audit.json"
                if _ap.exists():
                    try:
                        import json as _json
                        ctx_audit = _json.loads(_ap.read_text(encoding="utf-8"))
                    except Exception:
                        ctx_audit = {}
                    break
        # If awb wasn't resolved earlier, seed audit minimally so the helper
        # at least knows the AWB
        if target_awb and not ctx_audit.get("awb"):
            ctx_audit.setdefault("awb", target_awb)

        ctx = build_email_search_context(ctx_audit)

        # Connector pins come from the helper (canonical Estrella mailbox)
        PREFERRED_CONN_ID = "mcp__620999a3"

        task = _bridge_create_task(
            batch_id=batch_id or "_global_",
            task_type="email_scan",
            payload={
                "awb":             ctx["awb"] or target_awb or "",
                "invoice_numbers": ctx["invoice_numbers"],
                "dhl_ticket":      ctx["dhl_ticket"],
                "mrn":             ctx["mrn"],
                "search_terms":    ctx["search_terms"],
                "known_senders":   ctx["known_senders"],
                "known_domains":   ctx["known_domains"],
                "batch_id":        batch_id or "",
                "account_id":      account_id,
                "folder_id":       folder_id,
                # ── Mailbox identity (single account; multiple identities) ───
                "target_account_id":            ctx["target_account_id"],
                "target_mailbox":               ctx["target_mailbox"],
                "related_identities":           ctx["related_identities"],
                "preferred_mcp_connector_hint": PREFERRED_CONN_ID,
                "instructions": (
                    f"STEP 0: Verify mailbox binding — call getMailAccounts on "
                    f"connector starting with {PREFERRED_CONN_ID}. Confirm "
                    f"accountId == {ctx['target_account_id']} and "
                    f"primaryEmailAddress == {ctx['target_mailbox']}. "
                    "If mismatch, return connector_mismatch=true and stop. "
                    "related_identities[] are aliases of the same mailbox — "
                    "match To/Cc against any of them as in-scope.\n"
                    "Then: search by AWB first, then invoice numbers, then "
                    "sender domain combos. Inspect attachment filenames and "
                    "forwarded chains. Do NOT stop at the first 0-result step. "
                    "If matched=0 but search_terms had >1 entry, mark "
                    "search_unreliable=true."
                ),
            },
            note=f"Bridge scan for AWB {target_awb or '(none)'} — reason: {reason}.",
        )
        return {
            "scanned":     0,
            "matched":     0,
            "emails":      [],
            "scan_method": "ai_bridge_pending",
            "search_mode": "awb_targeted" if target_awb else "broad_recent",
            "query_used":  "ai_bridge:email_scan",
            "awb_used":    target_awb,
            "search_context": {
                "awb":              ctx["awb"],
                "invoice_numbers":  ctx["invoice_numbers"],
                "dhl_ticket":       ctx["dhl_ticket"],
                "mrn":              ctx["mrn"],
                "search_terms":     ctx["search_terms"],
                "search_terms_count": len(ctx["search_terms"]),
            },
            "scanned_at":  datetime.now(timezone.utc).isoformat(),
            "bridge_task": {
                "task_id":     task["task_id"],
                "task_type":   "email_scan",
                "result_file": task.get("result_file"),
                "message": (
                    "Email scan dispatched to AI Bridge. Open the AI Bridge tab "
                    "to execute the task in Cowork, then import the result."
                ),
                "reason": reason,
            },
        }

    # ── Path 0: stored email intelligence cache (skip Cowork if we have data)─
    # Verified prior scans are reusable. Unreliable cached results still let
    # the operator re-run via the dashboard's explicit button.
    # Skip cache when refresh=true (operator forced re-scan).
    try:
        if refresh:
            raise LookupError("operator-forced refresh")
        from ..services.email_intelligence_store import find_existing_email_context
        # Only check when we have an audit context to look up against
        ctx_audit_for_cache: Dict[str, Any] = {}
        if batch_id:
            for sub in ("outputs", "working"):
                _ap = settings.storage_root / sub / batch_id / "audit.json"
                if _ap.exists():
                    try:
                        import json as _json
                        ctx_audit_for_cache = _json.loads(_ap.read_text(encoding="utf-8"))
                    except Exception:
                        ctx_audit_for_cache = {}
                    break
        if target_awb and not ctx_audit_for_cache.get("awb"):
            ctx_audit_for_cache["awb"] = target_awb

        cached = find_existing_email_context(ctx_audit_for_cache) if ctx_audit_for_cache else None
        # Use cache if it's a verified record (not unreliable, matched > 0 or
        # explicitly verified-with-zero). Operator can force re-run from UI.
        if cached and cached.get("matched", 0) > 0 and not cached.get("search_unreliable"):
            log.info(
                "[scan-inbox] using cached email intelligence for AWB %s (matched=%d, scanned %s)",
                cached.get("awb"), cached.get("matched"), cached.get("last_scanned_at", ""),
            )
            return {
                "scanned":     cached.get("matched", 0),
                "matched":     cached.get("matched", 0),
                "emails":      cached.get("emails") or [],
                "scan_method": "email_intelligence_cache",
                "search_mode": "awb_targeted" if target_awb else "broad_recent",
                "query_used":  f"cache:by_awb:{cached.get('awb')}",
                "awb_used":    cached.get("awb"),
                "scanned_at":  datetime.now(timezone.utc).isoformat(),
                "email_scan_results": cached,
                "search_context": {
                    "awb":               cached.get("awb"),
                    "invoice_numbers":   cached.get("invoice_numbers", []),
                    "dhl_ticket":        cached.get("dhl_ticket"),
                    "mrn":               cached.get("mrn"),
                    "search_terms":      [],
                    "search_terms_count": 0,
                },
                "cached": {
                    "source":          cached.get("source"),
                    "last_scanned_at": cached.get("last_scanned_at"),
                    "connector_used":  cached.get("connector_used"),
                    "linked_batches":  cached.get("linked_batches", []),
                },
            }
    except Exception as exc:
        log.debug("[scan-inbox] email intelligence cache lookup failed (non-fatal): %s", exc)

    # ── Path A: native backend scan (when Zoho creds are configured) ──────────
    if creds_present:
        try:
            result = scan_for_dhl_customs_emails(
                zoho_account_id=account_id,
                zoho_folder_id=folder_id,
                limit=limit,
                target_awb=target_awb,
                api_base=settings.zoho_mail_api_base,
                token_provider=get_valid_access_token,
            )
        except ZohoAuthError as exc:
            # Auth failure — auto-fallback to bridge in 'auto' mode; hard 401 in 'api_only'
            log.warning("Zoho auth error during scan: %s", exc)
            if bridge_allowed:
                log.info("Falling back to AI Bridge after auth_error")
                return _dispatch_to_bridge(reason="zoho_auth_error")
            raise HTTPException(status_code=401, detail=str(exc)) from exc
        except Exception as exc:
            log.error("DHL inbox scan failed: %s", exc)
            if bridge_allowed:
                log.info("Falling back to AI Bridge after API error")
                return _dispatch_to_bridge(reason="zoho_api_error")
            raise HTTPException(status_code=500, detail=f"Inbox scan error: {exc}") from exc

        # Auto-fallback to bridge if scan returned auth_error path silently
        if result.get("scan_method") == "auth_error" and bridge_allowed:
            log.info("Scan returned auth_error — auto-falling back to bridge")
            return _dispatch_to_bridge(reason="zoho_auth_error_silent")
    # ── Path B: AI Bridge (forced or no creds) ────────────────────────────────
    elif bridge_allowed:
        reason = "bridge_only_mode" if scan_mode == "bridge_only" else "no_credentials"
        try:
            return _dispatch_to_bridge(reason=reason)
        except Exception as exc:
            log.error("AI Bridge task creation failed: %s", exc)
            raise HTTPException(status_code=500, detail=f"Bridge task error: {exc}") from exc
    else:
        # api_only mode but no creds — clean error, no bridge fallback
        raise HTTPException(
            status_code=503,
            detail=(
                "Email scan unavailable: EMAIL_SCAN_MODE=api_only but no Zoho "
                "credentials configured. Set ZOHO_CLIENT_ID/SECRET/REFRESH_TOKEN "
                "or change EMAIL_SCAN_MODE to 'auto' or 'bridge_only'."
            ),
        )

    scanned_at = datetime.now(timezone.utc).isoformat()

    emails      = result.get("emails", [])
    scanned     = result.get("scanned", 0)
    matched     = result.get("matched", 0)
    scan_method = result.get("scan_method", "no_credentials")
    search_mode = result.get("search_mode", "broad_recent")
    query_used  = result.get("query_used", "")

    # ── Timeline logging — only when something actually matched ───────────────
    if matched > 0 and batch_id and "/" not in batch_id and ".." not in batch_id:
        for sub in ("outputs", "working"):
            _ap = settings.storage_root / sub / batch_id / "audit.json"
            if _ap.exists():
                emails_preview = [
                    {
                        "subject":        e.get("subject") or e.get("raw_subject", ""),
                        "from":           e.get("from", ""),
                        "received_at":    e.get("received_at", ""),
                        "ticket":         e.get("dhl_ticket", ""),
                        "awb":            e.get("awb", ""),
                        "matched_fields": e.get("matched_fields", []),
                        "detected_type":  e.get("detected_type", ""),
                    }
                    for e in emails[:3]
                ]
                tl.log_event(
                    _ap,
                    tl.EV_DHL_INBOX_SCANNED,
                    trigger_source="dashboard",
                    actor="admin",
                    detail={
                        "scanned":        scanned,
                        "matched":        matched,
                        "scan_method":    scan_method,
                        "search_mode":    search_mode,
                        "query_used":     query_used,
                        "awb_used":       target_awb or "",
                        "emails_preview": emails_preview,
                    },
                )
                log.info(
                    "DHL inbox scan logged to timeline: batch=%s scanned=%d matched=%d "
                    "method=%s mode=%s awb=%s",
                    batch_id, scanned, matched, scan_method, search_mode, target_awb,
                )

                # ── Mark dhl_email.received when DHL customs email found ────────────
                # odprawacelna@dhl.com sends T# emails when a high-value shipment
                # arrives at customs. The active_shipment_monitor's B2 path
                # (_ensure_dhl_dsk_transfer_reply) fires automatically on the next
                # sweep once audit.dhl_email.received is True. Without this write,
                # the scan only logs EV_DHL_INBOX_SCANNED and the B2 path never fires
                # because its entry gate checks dhl_email.received first.
                # Idempotent: skipped if received is already set.
                _DHL_CUSTOMS_SENDERS = frozenset({
                    "odprawacelna@dhl.com",
                    "administracja_centralna@dhl.com",
                })
                _customs_hit = next(
                    (
                        e for e in emails
                        if (e.get("from") or "").lower().strip() in _DHL_CUSTOMS_SENDERS
                        and (e.get("dhl_ticket") or e.get("ticket"))
                    ),
                    None,
                )
                if _customs_hit:
                    try:
                        from ..utils.io import write_json_atomic as _wja_scan  # noqa: PLC0415
                        from ..services.active_shipment_monitor import (  # noqa: PLC0415
                            _is_active as _scan_batch_is_active,
                        )
                        _cur_audit = json.loads(_ap.read_text(encoding="utf-8"))
                        # GAP-1 guard: only write dhl_email.received for active
                        # (non-terminal, non-delivered) batches. A terminal batch
                        # whose AWB happens to appear in a new DHL email must NOT
                        # be re-flagged — doing so would restart B2 automation on
                        # a closed shipment.
                        # Note on GAP-2 (deferred): the native email classifier maps
                        # odprawacelna@dhl.com → type "dhl_arrival" → timeline event
                        # "carrier_arrived" (correct for SLA/intelligence chains).
                        # The AI Bridge cache path uses "dhl_customs_request" →
                        # "dhl_customs_email_received". These are intentionally
                        # separate: carrier_arrived feeds sla_engine.py SLA anchors;
                        # dhl_email.received drives B2 DSK-reply automation. Do not
                        # merge them — see routes_dhl_clearance.py audit note.
                        if not _scan_batch_is_active(_cur_audit):
                            log.info(
                                "[scan-inbox] skipping dhl_email.received — batch=%s "
                                "is terminal/delivered (clearance=%s tracking=%s)",
                                batch_id,
                                _cur_audit.get("clearance_status", ""),
                                (_cur_audit.get("tracking") or {}).get("status", ""),
                            )
                        elif not (_cur_audit.get("dhl_email") or {}).get("received"):
                            _ticket = (
                                _customs_hit.get("dhl_ticket")
                                or _customs_hit.get("ticket")
                                or ""
                            )
                            _cur_audit["dhl_email"] = {
                                "received":     True,
                                "source":       "scan_dhl_inbox",
                                "sender":       _customs_hit.get("from", ""),
                                "subject":      (
                                    _customs_hit.get("subject")
                                    or _customs_hit.get("raw_subject", "")
                                ),
                                "ticket":       _ticket,
                                "request_type": "dhl_customs_request",
                                "received_at":  _customs_hit.get("received_at", ""),
                            }
                            if _ticket:
                                _cur_audit["dhl_ticket"] = _ticket
                            _wja_scan(_ap, _cur_audit)
                            tl.log_event(
                                _ap,
                                tl.EV_DHL_EMAIL_RECEIVED,
                                trigger_source="dashboard",
                                actor="admin",
                                detail={
                                    "ticket":     _ticket,
                                    "written_by": "scan_dhl_inbox",
                                },
                            )
                            log.info(
                                "[scan-inbox] dhl_email.received set for batch=%s ticket=%s",
                                batch_id, _ticket,
                            )
                    except Exception as _set_exc:
                        log.warning(
                            "[scan-inbox] dhl_email.received write failed for "
                            "batch=%s: %s",
                            batch_id, _set_exc,
                        )

                break

    log.info(
        "DHL inbox scan: scanned=%d matched=%d method=%s mode=%s awb=%s",
        scanned, matched, scan_method, search_mode, target_awb,
    )

    return {
        "scanned":     scanned,
        "matched":     matched,
        "emails":      emails,
        "scan_method": scan_method,
        "search_mode": search_mode,
        "query_used":  query_used,
        "awb_used":    target_awb,
        "scanned_at":  scanned_at,
    }


def _scan_status_path() -> "Path":
    """Canonical path for the DHL auto-scan status file."""
    return settings.storage_root / "dhl_auto_scan_status.json"


def _write_scan_status(status: Dict[str, Any]) -> None:
    """Atomically write the scan status to disk. Non-fatal on failure."""
    try:
        # Import was MISSING — every lane-a status write failed with a
        # swallowed NameError (live prod WARNING 2026-07-02 14:06; infra
        # health pass d67d3722 finding #3). Local aliased import per this
        # file's idiom (:2038, :2874).
        from ..utils.io import write_json_atomic  # noqa: PLC0415
        write_json_atomic(_scan_status_path(), status)
    except Exception as exc:
        log.warning("[lane-a] status write failed (non-fatal): %s", exc)


@router.get("/auto-scan-status", dependencies=[_auth])
def get_auto_scan_status() -> Dict[str, Any]:
    """
    GET /api/v1/dhl/auto-scan-status — read-only DHL inbox-scanner status card.

    Returns the last recorded scan outcome from storage_root/dhl_auto_scan_status.json.
    Computes next_run_at as started_at + 10 minutes when status is success/failed.

    Status values: running | success | failed | timed_out | never_run
    Read-only: never triggers a scan, never modifies audit files, never sends email.
    """
    p = _scan_status_path()
    if not p.exists():
        return {
            "status":           "never_run",
            "started_at":       None,
            "completed_at":     None,
            "duration_seconds": None,
            "batches_checked":  None,
            "received_set":     None,
            "b2_triggered":     None,
            "b2_sent":          None,
            "skipped_inactive": None,
            "skipped_excluded": None,
            "errors_count":     None,
            "last_error":       None,
            "next_run_at":      None,
        }
    try:
        st = json.loads(p.read_text(encoding="utf-8"))
    except Exception as exc:
        return {"status": "status_read_error", "error": str(exc)}

    # Compute next_run_at from started_at + 10-minute interval
    next_run: Optional[str] = None
    if st.get("started_at"):
        try:
            from datetime import datetime, timezone, timedelta
            started = datetime.fromisoformat(
                str(st["started_at"]).replace("Z", "+00:00")
            )
            next_run = (started + timedelta(minutes=10)).isoformat()
        except Exception:
            pass

    return {
        "status":           st.get("status", "unknown"),
        "started_at":       st.get("started_at"),
        "completed_at":     st.get("completed_at"),
        "duration_seconds": st.get("duration_seconds"),
        "batches_checked":  st.get("batches_checked"),
        "received_set":     st.get("received_set"),
        "b2_triggered":     st.get("b2_triggered"),
        "b2_sent":          st.get("b2_sent"),
        "skipped_inactive": st.get("skipped_inactive"),
        "skipped_excluded": st.get("skipped_excluded"),
        "errors_count":     st.get("errors_count"),
        "last_error":       st.get("last_error"),
        "next_run_at":      next_run,
    }


@router.post("/scheduled-inbox-check", dependencies=[_auth])
def run_scheduled_inbox_check() -> Dict[str, Any]:
    """
    Lane A — automated DHL customs-email scanner (every 10 minutes).

    Kill switch: DHL_AUTO_SCAN_ENABLED=false returns immediately.
    Writes scan status to storage_root/dhl_auto_scan_status.json for the status card.

    Pipeline (one Zoho scan per call — NOT per batch):
      1. Kill-switch check (dhl_auto_scan_enabled)
      2. run_ingestion_cycle()  — scan Zoho inbox once, cache matched emails
      3. For each active batch (excluding manually-excluded AWBs):
           a. find_existing_email_context() → _apply_cache_to_audit()
              writes dhl_email.received when a T# match is found
           b. _ensure_dhl_reply()     → triggers B2 DSK reply if conditions met
      4. Write final status to dhl_auto_scan_status.json
      5. Return summary {lane, batches_checked, received_set, b2_triggered, errors}

    Guards (all inherited):
      - dhl_auto_scan_enabled kill switch
      - _is_active() per batch — terminal/delivered batches counted in skipped_inactive
      - AWB exclusion list — manually excluded batches (e.g. 5665916826 pending operator)
      - B2 idempotency: build_started_at prevents duplicate sends
      - DSK-present gate: dsk_path must resolve to a real file
    """
    from datetime import datetime, timezone as _tz

    def _now_iso() -> str:
        return datetime.now(_tz.utc).isoformat()

    # ── Kill switch ───────────────────────────────────────────────────────────
    if not settings.dhl_auto_scan_enabled:
        log.info("[lane-a] DHL_AUTO_SCAN_ENABLED=false — skipping")
        return {"ok": False, "lane": "A", "skipped": "DHL_AUTO_SCAN_ENABLED=false"}

    from ..services.email_ingestion_worker import run_ingestion_cycle as _run_ing
    from ..services.email_intelligence_store import (
        find_existing_email_context as _find_cache,
    )
    from ..services.active_shipment_monitor import (
        _all_audit_paths   as _audit_paths,
        _is_active         as _batch_active,
        _apply_cache_to_audit,
        _ensure_dhl_reply,
    )

    _EXCLUDED_AWBS: frozenset = frozenset({"5665916826"})
    _started = _now_iso()

    # ── Write "running" status at scan start ──────────────────────────────────
    _write_scan_status({"status": "running", "started_at": _started})

    out: Dict[str, Any] = {
        "ok":               True,
        "lane":             "A",
        "batches_checked":  0,
        "received_set":     0,
        "b2_triggered":     0,
        "b2_sent":          0,
        "skipped_inactive": 0,
        "skipped_excluded": 0,
        "errors":           [],
        "ingestion":        {},
    }

    try:
        # ── Step 1: one global Zoho scan, results cached per AWB ─────────────
        try:
            ing = _run_ing()
            out["ingestion"] = {
                "ok":             ing.get("ok"),
                "active_batches": ing.get("active_batches"),
                "shipments":      len(ing.get("shipments") or []),
            }
        except Exception as exc:
            out["ingestion"] = {"ok": False, "error": str(exc)}
            log.warning("[lane-a] ingestion cycle failed (non-fatal): %s", exc)

        # ── Step 2: apply cached evidence + trigger B2 per active batch ───────
        for ap in _audit_paths():
            try:
                audit = json.loads(ap.read_text(encoding="utf-8"))
            except Exception:
                continue

            _awb = (audit.get("awb") or audit.get("tracking_no") or "").strip()
            if _awb in _EXCLUDED_AWBS:
                out["skipped_excluded"] += 1
                continue

            if not _batch_active(audit):
                out["skipped_inactive"] += 1
                continue

            out["batches_checked"] += 1
            batch_id = ap.parent.name

            try:
                cached = _find_cache(audit)
                if cached and cached.get("matched", 0) > 0:
                    _apply_cache_to_audit(ap, audit, cached)
                    audit = json.loads(ap.read_text(encoding="utf-8"))
                    if (audit.get("dhl_email") or {}).get("received"):
                        out["received_set"] += 1
            except Exception as exc:
                out["errors"].append(f"{batch_id}:cache:{exc}")
                log.warning("[lane-a] cache apply failed batch=%s: %s", batch_id, exc)

            try:
                reply_result = _ensure_dhl_reply(ap, audit)
                if reply_result.get("built"):
                    out["b2_triggered"] += 1
                if reply_result.get("sent"):
                    out["b2_sent"] += 1
                    log.info("[lane-a] B2 sent batch=%s", batch_id)
            except Exception as exc:
                out["errors"].append(f"{batch_id}:b2:{exc}")
                log.warning("[lane-a] B2 failed batch=%s: %s", batch_id, exc)

        _completed = _now_iso()
        try:
            from datetime import datetime as _dt
            _dur = round(
                (_dt.fromisoformat(_completed.replace("Z", "+00:00")) -
                 _dt.fromisoformat(_started.replace("Z", "+00:00"))).total_seconds(),
                1,
            )
        except Exception:
            _dur = None

        log.info(
            "[lane-a] done: checked=%d received_set=%d b2_triggered=%d "
            "b2_sent=%d skipped_inactive=%d skipped_excluded=%d errors=%d",
            out["batches_checked"], out["received_set"], out["b2_triggered"],
            out["b2_sent"], out["skipped_inactive"], out["skipped_excluded"],
            len(out["errors"]),
        )

        # ── Write "success" status on clean completion ────────────────────────
        _write_scan_status({
            "status":           "success",
            "started_at":       _started,
            "completed_at":     _completed,
            "duration_seconds": _dur,
            "batches_checked":  out["batches_checked"],
            "received_set":     out["received_set"],
            "b2_triggered":     out["b2_triggered"],
            "b2_sent":          out["b2_sent"],
            "skipped_inactive": out["skipped_inactive"],
            "skipped_excluded": out["skipped_excluded"],
            "errors_count":     len(out["errors"]),
            "last_error":       out["errors"][-1] if out["errors"] else None,
        })

    except Exception as _fatal:
        _completed = _now_iso()
        log.error("[lane-a] fatal error in scan: %s", _fatal)
        _write_scan_status({
            "status":     "failed",
            "started_at": _started,
            "completed_at": _completed,
            "errors_count": 1,
            "last_error": str(_fatal),
        })
        out["ok"] = False
        out["errors"].append(str(_fatal))

    return out


@router.get("/daily-summary", dependencies=[_auth])
def get_dhl_daily_summary() -> Dict[str, Any]:
    """
    GET /api/v1/dhl/daily-summary — read-only daily DHL operations report.

    Aggregates across:
      - storage_root/dhl_auto_scan_status.json  (Lane A last run)
      - C:\\PZ\\logs\\dhl-auto-scan.log          (24h run history)
      - storage_root/outputs/*/audit.json        (per-shipment state)
      - storage_root/email_queue.json            (sent replies)

    Returns:
      lane_a_health         — last run, 24h run/fail counts, averages
      active_shipments      — per-shipment dashboard with DHL state
      dhl_waiting_queue     — DSK sent, no reply yet (oldest first)
      lane_b_candidates     — read-only preview of who would qualify
      exceptions            — scanner failures, missing DSK, excluded
      summary               — executive counters

    Read-only: never triggers scan, never sends email, never modifies audits.
    """
    from datetime import datetime, timezone as _tz, timedelta as _td
    import re as _re

    _now = datetime.now(_tz.utc)
    _24h_ago = _now - _td(hours=24)
    _today_start = _now.replace(hour=0, minute=0, second=0, microsecond=0)

    # ── Lane A health (from status file + log) ────────────────────────────────
    lane_a_health: Dict[str, Any] = {
        "last_run_at":       None,
        "last_run_status":   "never_run",
        "last_run_duration_s": None,
        "runs_24h":          0,
        "failed_runs_24h":   0,
        "avg_duration_s":    None,
        "avg_batches_checked": None,
        "avg_matches_found": None,
    }

    # Last run from status file (dhl_auto_scan_status.json written by scan handler)
    _sp = settings.storage_root / "dhl_auto_scan_status.json"
    if _sp.exists():
        try:
            _st = json.loads(_sp.read_text(encoding="utf-8"))
            lane_a_health["last_run_at"]        = _st.get("started_at")
            lane_a_health["last_run_status"]     = _st.get("status", "unknown")
            lane_a_health["last_run_duration_s"] = _st.get("duration_seconds")
        except Exception:
            pass

    # 24h history from log file
    _log_path = Path("C:/PZ/logs/dhl-auto-scan.log")
    if not _log_path.exists():
        _log_path = settings.storage_root.parent.parent / "logs" / "dhl-auto-scan.log"
    _runs: list = []  # list of {ts, status, duration, checked, received}
    if _log_path.exists():
        try:
            _log_lines = _log_path.read_text(encoding="utf-8", errors="replace").splitlines()
            _pending_start: Optional[str] = None
            for _line in _log_lines:
                # Parse start timestamp
                _ts_m = _re.match(r"\[(\d{4}-\d{2}-\d{2}T\d{2}:\d{2}:\d{2})\]", _line)
                if not _ts_m:
                    continue
                try:
                    _ts = datetime.fromisoformat(_ts_m.group(1) + "+00:00")
                except Exception:
                    continue
                if _ts < _24h_ago:
                    continue  # outside 24h window
                if "[Lane-A] Starting" in _line:
                    _pending_start = _ts_m.group(1)
                elif "[Lane-A] done:" in _line and _pending_start:
                    _dur = None
                    try:
                        _s = datetime.fromisoformat(_pending_start + "+00:00")
                        _dur = round((_ts - _s).total_seconds(), 1)
                    except Exception:
                        pass
                    _checked = 0
                    _received = 0
                    _cm = _re.search(r"checked=(\d+)", _line)
                    _rm = _re.search(r"received_set=(\d+)", _line)
                    if _cm:
                        _checked = int(_cm.group(1))
                    if _rm:
                        _received = int(_rm.group(1))
                    _runs.append({
                        "ts": _pending_start, "status": "success",
                        "duration": _dur, "checked": _checked, "received": _received,
                    })
                    _pending_start = None
                elif ("[Lane-A] HTTP" in _line or "[Lane-A] error" in _line.lower()) and _pending_start:
                    _runs.append({"ts": _pending_start, "status": "failed", "duration": None,
                                  "checked": 0, "received": 0})
                    _pending_start = None
        except Exception:
            pass

    lane_a_health["runs_24h"]       = len(_runs)
    lane_a_health["failed_runs_24h"] = sum(1 for r in _runs if r["status"] != "success")
    _durs = [r["duration"] for r in _runs if r.get("duration") is not None]
    _chks = [r["checked"] for r in _runs if r.get("checked", 0) > 0]
    _mats = [r["received"] for r in _runs]
    if _durs:
        lane_a_health["avg_duration_s"] = round(sum(_durs) / len(_durs), 1)
    if _chks:
        lane_a_health["avg_batches_checked"] = round(sum(_chks) / len(_chks), 1)
    if _mats:
        lane_a_health["avg_matches_found"] = round(sum(_mats) / len(_mats), 2)

    # ── Per-batch state ───────────────────────────────────────────────────────
    from ..services.active_shipment_monitor import (
        _all_audit_paths    as _audit_paths,
        _is_active          as _batch_active,
        _is_customs_complete,
    )

    _EXCLUDED_AWBS: frozenset = frozenset({"5665916826"})
    active_shipments:   list = []
    dhl_waiting_queue:  list = []
    lane_b_candidates:  list = []
    exceptions:         list = []

    for ap in sorted(_audit_paths(), key=lambda p: p.stat().st_mtime):
        try:
            audit = json.loads(ap.read_text(encoding="utf-8"))
        except Exception as exc:
            exceptions.append({"type": "read_error", "batch": ap.parent.name, "detail": str(exc)})
            continue

        _awb = (audit.get("awb") or audit.get("tracking_no") or "").strip()
        _is_excluded = _awb in _EXCLUDED_AWBS
        _is_act = _batch_active(audit)

        # Compute days_open from batch directory month or first timeline event
        _days_open: Optional[float] = None
        try:
            _tl = audit.get("timeline") or []
            if _tl:
                _first_ts = _tl[0].get("ts") or _tl[0].get("timestamp") or ""
                if _first_ts:
                    _first_dt = datetime.fromisoformat(
                        str(_first_ts).replace("Z", "+00:00")
                    )
                    _days_open = round((_now - _first_dt).total_seconds() / 86400, 1)
        except Exception:
            pass

        _cd       = audit.get("clearance_decision") or {}
        _dhl_email = audit.get("dhl_email") or {}
        _drp      = audit.get("dhl_reply_package") or {}
        _dhl_recv = bool(_dhl_email.get("received"))
        _dsk_sent = _drp.get("status") in ("queued", "sent")
        _supplier = (
            audit.get("exporter") or audit.get("supplier_name")
            or _cd.get("agency") or "—"
        )

        _row: Dict[str, Any] = {
            "awb":            _awb or ap.parent.name[:16],
            "batch_id":       ap.parent.name,
            "supplier":       _supplier,
            "clearance_path": _cd.get("clearance_path", "—"),
            "cif_usd":        _cd.get("total_value_usd"),
            "status":         audit.get("clearance_status", audit.get("status", "—")),
            "dhl_received":   _dhl_recv,
            "dsk_sent":       _dsk_sent,
            "days_open":      _days_open,
            "excluded":       _is_excluded,
            "active":         _is_act,
        }

        if _is_excluded:
            exceptions.append({
                "type":    "excluded_awb",
                "batch":   ap.parent.name,
                "awb":     _awb,
                "reason":  "manual exclusion — pending operator decision",
            })
            continue

        if not _is_act:
            continue  # only active batches in the dashboard

        active_shipments.append(_row)

        # DHL Waiting Queue: DSK sent, no reply yet
        if _dsk_sent and not _dhl_recv:
            dhl_waiting_queue.append({
                "awb":         _awb,
                "batch_id":    ap.parent.name,
                "supplier":    _supplier,
                "clearance_path": _cd.get("clearance_path", "—"),
                "dsk_sent_at": _drp.get("queued_at") or _drp.get("sent_at"),
                "days_open":   _days_open,
                "ticket":      _dhl_email.get("ticket") or audit.get("dhl_ticket"),
            })

        # Missing DSK exception: received DHL email but no DSK path or reply
        if _dhl_recv and not _dsk_sent:
            _dsk_path = (audit.get("dsk_path") or "").strip()
            exceptions.append({
                "type":    "dsk_not_sent",
                "batch":   ap.parent.name,
                "awb":     _awb,
                "reason":  "DHL email received but DSK reply not sent" +
                           (" — dsk_path missing" if not _dsk_path else ""),
            })

        # Lane B candidates: no DHL reply, active, customs NOT complete,
        # check if follow-up SLA eligible.
        # Customs-complete check is first: if SAD/ZC429/PZC exists, the batch
        # is excluded from Lane B regardless of hours_waiting.
        if not _dhl_recv and not _dsk_sent and not _is_customs_complete(audit):
            # Would follow-up be eligible if Lane B were ON?
            _fu_state = audit.get("dhl_followup") or {}
            _fu_active = bool(_fu_state.get("active"))
            _next_at = _fu_state.get("next_followup_at") or _fu_state.get("first_followup_at")
            _trigger_at = None
            try:
                _tl = audit.get("timeline") or []
                for _ev in reversed(_tl):
                    if "customs" in (_ev.get("event") or "").lower() or \
                       "agency" in (_ev.get("event") or "").lower():
                        _trigger_at = _ev.get("ts") or _ev.get("timestamp")
                        break
            except Exception:
                pass

            _hours_waiting = None
            _eligible = False
            if _trigger_at:
                try:
                    _trig_dt = datetime.fromisoformat(
                        str(_trigger_at).replace("Z", "+00:00")
                    )
                    _hours_waiting = round((_now - _trig_dt).total_seconds() / 3600, 1)
                    _eligible = _hours_waiting >= 4
                except Exception:
                    pass

            lane_b_candidates.append({
                "awb":              _awb,
                "batch_id":         ap.parent.name,
                "supplier":         _supplier,
                "clearance_path":   _cd.get("clearance_path", "—"),
                "hours_waiting":    _hours_waiting,
                "eligible":         _eligible,
                "follow_up_active": _fu_active,
                "next_followup_at": _next_at,
                "reason": (
                    "customs trigger detected, 4h+ elapsed" if _eligible
                    else ("customs trigger detected, <4h elapsed" if _hours_waiting is not None
                          else "no customs trigger detected from tracking events")
                ),
                "lane_b_status": "ON" if settings.dhl_followup_enabled else "OFF",
            })

    # Sort DHL waiting queue oldest first
    dhl_waiting_queue.sort(key=lambda x: x.get("days_open") or 0, reverse=True)

    # ── Email queue counts (replies sent today) ────────────────────────────────
    _replies_today = 0
    try:
        _eq = json.loads(
            (settings.storage_root / "email_queue.json").read_text(encoding="utf-8")
        )
        _today_iso = _today_start.isoformat()
        _replies_today = sum(
            1 for e in _eq
            if e.get("status") in ("sent", "queued")
            and (e.get("queued_at") or "") >= _today_iso
            and "odprawacelna" in (e.get("to") or "")
        )
    except Exception:
        pass

    return {
        "generated_at":    _now.isoformat(),
        "lane_a_health":   lane_a_health,
        "active_shipments": active_shipments,
        "dhl_waiting_queue": dhl_waiting_queue,
        "lane_b_candidates": sorted(
            lane_b_candidates,
            key=lambda x: x.get("hours_waiting") or 0, reverse=True,
        ),
        "exceptions": exceptions,
        "summary": {
            "active_shipments":    len(active_shipments),
            "waiting_for_dhl":     len(dhl_waiting_queue),
            "replies_sent_today":  _replies_today,
            "scanner_runs_24h":    lane_a_health["runs_24h"],
            "scanner_failures_24h": lane_a_health["failed_runs_24h"],
            "lane_b_eligible":     sum(1 for c in lane_b_candidates if c.get("eligible")),
            "excluded_batches":    len([e for e in exceptions if e.get("type") == "excluded_awb"]),
            "errors_count":        len([e for e in exceptions if e.get("type") not in
                                        ("excluded_awb", "dsk_not_sent")]),
        },
    }


@router.post("/scheduled-followup-check", dependencies=[_auth])
def run_scheduled_followup_check() -> Dict[str, Any]:
    """
    Lane B — DHL follow-up SLA check (every 60 minutes, working hours only).

    Calls the existing _process_dhl_followup for each eligible active batch.
    That function implements the full SLA contract:
      - 4h initial wait after customs trigger detected from tracking events
      - Max once per hour during 08:00–16:00 Warsaw time
      - Idempotent: last_followup_sent_at + followup_count stored in audit
      - Stops immediately on any stop condition

    Outer kill switch: DHL_FOLLOWUP_ENABLED=false → endpoint returns immediately.
    Inner kill switch: DHL_ORCH_AUTO_SEND_DHL_FOLLOWUP controls whether emails
      are actually sent (via validate_followup_send_preconditions in dhl_followup_guard).
    Both must be True for follow-up emails to send.

    Eligibility per batch (all required — enforced inside _process_dhl_followup):
      - _is_active(audit) = True           not delivered / terminal
      - dhl_email.received != True         DHL hasn't replied to the DSK
      - customs_docs.received != True      customs docs not yet received
      - customs_trigger present             shipment at customs stage in Poland
      - working hours (08:00–16:00 Warsaw) enforced by dhl_followup_sla.is_due()
      - 4h initial wait from trigger       enforced by calculate_first_followup_at()
      - 1h repeat interval                 enforced by calculate_next_followup_at()
      - idempotency key in audit           no duplicate send for same time slot

    AWB 5665916826 and any other excluded AWBs are skipped.
    Never touches financial fields, wFirma, or product-description logic.
    """
    # ── Outer kill switch ─────────────────────────────────────────────────────
    if not settings.dhl_followup_enabled:
        log.info("[lane-b] DHL_FOLLOWUP_ENABLED=false — skipping")
        return {
            "ok":    False,
            "lane":  "B",
            "skipped": "DHL_FOLLOWUP_ENABLED=false",
        }

    from ..services.active_shipment_monitor import (
        _all_audit_paths        as _audit_paths,
        _is_active              as _batch_active,
        _is_customs_complete,
        _process_dhl_followup,
    )
    from ..services.tracking_intelligence import detect_tracking_triggers

    _EXCLUDED_AWBS: frozenset = frozenset({"5665916826"})

    out: Dict[str, Any] = {
        "ok":                       True,
        "lane":                     "B",
        "batches_checked":          0,
        "followup_started":         0,
        "followup_sent":            0,
        "followup_stopped":         0,
        "followup_suppressed":      0,
        "skipped_inactive":         0,
        "skipped_excluded":         0,
        "skipped_received":         0,
        "skipped_customs_complete": 0,
        "errors":                   [],
    }

    for ap in _audit_paths():
        try:
            audit = json.loads(ap.read_text(encoding="utf-8"))
        except Exception:
            continue

        # Exclusion list — skip manually excluded AWBs
        _awb = (audit.get("awb") or audit.get("tracking_no") or "").strip()
        if _awb in _EXCLUDED_AWBS:
            out["skipped_excluded"] += 1
            continue

        # Active-batch guard — skip terminal/delivered batches
        if not _batch_active(audit):
            out["skipped_inactive"] += 1
            continue

        # Customs-complete gate — SAD/ZC429/PZC received means customs is done.
        # No follow-up email should ever be sent for a customs-cleared shipment.
        if _is_customs_complete(audit):
            out["skipped_customs_complete"] += 1
            continue

        # Skip if DHL already replied (Lane A will have set this flag)
        if (audit.get("dhl_email") or {}).get("received"):
            out["skipped_received"] += 1
            continue

        out["batches_checked"] += 1
        batch_id = ap.parent.name

        # Compute customs trigger from tracking events
        # No trigger → _process_dhl_followup returns without starting SLA
        tr_events = (
            audit.get("tracking_events")
            or (audit.get("tracking") or {}).get("events")
            or []
        )
        customs_trigger = None
        if tr_events:
            triggers = detect_tracking_triggers(tr_events, audit)
            customs_trigger = next(
                (t for t in triggers
                 if t.get("trigger") == "DHL_CUSTOMS_EMAIL_CHECK_REQUIRED"),
                None,
            )

        try:
            result = _process_dhl_followup(ap, audit, customs_trigger)
            if result.get("started"):
                out["followup_started"] += 1
            if result.get("sent"):
                out["followup_sent"] += 1
                log.info("[lane-b] follow-up sent batch=%s", batch_id)
            if result.get("stopped"):
                out["followup_stopped"] += 1
            # Normal non-event: SLA not due / working hours / flag off
            if not any(result.get(k) for k in ("started", "sent", "stopped")):
                out["followup_suppressed"] += 1
        except Exception as exc:
            out["errors"].append(f"{batch_id}:{exc}")
            log.warning("[lane-b] follow-up failed batch=%s: %s", batch_id, exc)

    log.info(
        "[lane-b] done: checked=%d started=%d sent=%d stopped=%d "
        "skipped_inactive=%d skipped_received=%d errors=%d",
        out["batches_checked"], out["followup_started"], out["followup_sent"],
        out["followup_stopped"], out["skipped_inactive"],
        out["skipped_received"], len(out["errors"]),
    )
    return out


@router.post("/match-and-handle", dependencies=[_auth, _op_auth])
async def match_and_handle(body: MatchAndHandleRequest) -> Dict[str, Any]:
    """
    Match an AWB number to a batch and run the DHL clearance handler.

    Steps:
    1. Search all audit.json files for the AWB
    2. Determine clearance route (DHL self-clear vs broker)
    3. Generate DSK (broker) or Polish description (DHL direct)
    4. Return reply package (does NOT send email)
    """
    from dhl_email_monitor import match_awb_to_batch
    from dhl_clearance_handler import handle_dhl_customs_email

    # Build synthetic dhl_email dict from request
    dhl_email = {
        "message_id":  body.message_id or "",
        "thread_id":   body.thread_id or "",
        "subject":     body.subject or f"[{body.dhl_ticket or ''}] - Agencja Celna DHL - przesyłka numer: {body.awb}",
        "from":        "odprawacelna@dhl.com",
        "received_at": datetime.now(timezone.utc).isoformat(),
        "dhl_ticket":  body.dhl_ticket or "",
        "awb":         body.awb,
    }

    # 1. Find matching batch
    batch = match_awb_to_batch(
        awb=body.awb,
        storage_root=str(settings.storage_root),
    )
    if batch is None:
        raise HTTPException(
            status_code=404,
            detail=f"No batch found matching AWB {body.awb}. "
                   "Ensure the batch has been processed and audit.json contains the AWB.",
        )

    # 2. Apply value override if provided
    if body.value_usd_override is not None:
        # Inject into batch for handler routing
        batch.setdefault("result", {}).setdefault("verification", {})[
            "invoice_cif_total_usd"
        ] = body.value_usd_override

    log.info("DHL match-and-handle: awb=%s batch_path=%s", body.awb, batch.get("_audit_path"))

    # 3. Run handler
    try:
        result = handle_dhl_customs_email(
            dhl_email=dhl_email,
            batch=batch,
            storage_root=str(settings.storage_root),
            dsk_output_dir=str(_DSK_OUTPUT_DIR),
        )
    except Exception as exc:
        log.error("DHL clearance handler error: %s", exc, exc_info=True)
        raise HTTPException(status_code=500, detail=f"Clearance handler error: {exc}") from exc

    # ── DSK source tracking (automated path) ─────────────────────────────────
    # If the incoming email originated from DHL_DSK_SOURCE, record dsk_received.
    _dhl_sender = dhl_email.get("from", "")
    _batch_audit_path = Path(batch.get("_audit_path", ""))
    if _dhl_sender and is_dsk_source(_dhl_sender) and _batch_audit_path.exists():
        try:
            from ..utils.io import write_json_atomic as _wja
            _ba = json.loads(_batch_audit_path.read_text(encoding="utf-8"))
            _ba["dsk_received"]    = True
            _ba["dsk_source"]      = _dhl_sender
            _ba["dsk_received_at"] = datetime.now(timezone.utc).isoformat()
            _wja(_batch_audit_path, _ba)
            log.info("DSK source email auto-detected in match-and-handle: awb=%s sender=%s",
                     body.awb, _dhl_sender)
        except Exception as _dsk_exc:
            log.warning("DSK source tracking (non-fatal): %s", _dsk_exc)

    # Timeline: DHL email received
    if _batch_audit_path.exists():
        try:
            await _pipeline_dhl_email(
                batch,
                _batch_audit_path,
                ticket=body.dhl_ticket or "",
                awb=body.awb,
                actor="dhl_monitor",
            )
        except Exception as _tl_exc:
            log.warning("DHL timeline event failed (non-fatal): %s", _tl_exc)

    # 4. Enrich reply_package attachments with download URLs (strip local paths)
    reply_pkg = result.get("reply_package", {})
    attachments_safe = []
    for att in (reply_pkg.get("attachments") or []):
        fn = Path(att.get("path", "")).name
        attachments_safe.append({
            "label":        att.get("label", ""),
            "filename":     fn,
            "download_url": f"{settings.fastapi_public_url}/api/v1/dhl/download/{fn}" if fn else None,
        })

    reply_validation = result.get("reply_validation") or {}
    decision_reason  = result.get("decision_reason") or {}

    return {
        "action":            result["action"],
        "clearance_status":  result["clearance_status"],
        "awb":               body.awb,
        "batch_found":       True,
        "batch_path":        batch.get("_audit_path", ""),
        "dsk":               _safe_dsk(result.get("dsk")),
        "polish_description": _safe_polish(result.get("polish_description")),
        "decision_reason":   decision_reason,
        "reply_validation":  reply_validation,
        "reply_package": {
            "to":          format_to(DHL_TO),   # always use centralized config
            "cc":          format_cc(INTERNAL_CC),
            "subject":     reply_pkg.get("subject", ""),
            "thread_id":   reply_pkg.get("thread_id", ""),
            "message_id":  reply_pkg.get("message_id", ""),
            "body_pl":     reply_pkg.get("body_pl", ""),
            "body_en":     reply_pkg.get("body_en", ""),
            "attachments": attachments_safe,
            "blocked":     reply_validation.get("blocked", False),
        },
    }


@router.get("/clearance-status/{batch_id}", response_model=ClearanceStatusResponse, dependencies=[_auth])
async def get_clearance_status(batch_id: str) -> ClearanceStatusResponse:
    """Get the current clearance status for a batch."""
    audit = _load_audit(batch_id)
    if audit is None:
        return ClearanceStatusResponse(
            batch_id=batch_id,
            found=False,
        )

    return ClearanceStatusResponse(
        batch_id=batch_id,
        clearance_status=audit.get("clearance_status"),
        clearance_action=audit.get("clearance_action"),
        dhl_ticket=audit.get("dhl_ticket"),
        awb=audit.get("dhl_awb") or audit.get("awb"),
        updated_at=audit.get("clearance_updated_at"),
        found=True,
    )


def _translate_blocked_package(batch_id: str, pkg: dict) -> None:
    """Translate a function-internal guard block from
    generate_customs_description_package() into the matching HTTP 422 with
    row-level detail. No-op when the package was not blocked. Both generate
    routes call this so their behaviour matches the engine-internal guards that
    also protect the automation callers."""
    if not isinstance(pkg, dict) or not pkg.get("blocked"):
        return
    guard = pkg.get("guard")
    if guard == "descriptions_missing_for_customs":
        log.info("[%s] generation blocked (engine guard): %d line(s) lack an "
                 "approved customs description", batch_id, len(pkg.get("missing") or []))
        raise HTTPException(
            status_code=422,
            detail={
                "guard": guard,
                "error": "One or more product lines have no approved customs "
                         "description and would fall back to generic placeholder "
                         "text. Correct each line before generating.",
                "code":  guard,
                "rows":  pkg.get("missing") or [],
                "hint":  "Use the approved Product Master description or save a "
                         "shipment correction (action-proposals approve, "
                         "scope=shipment), then Recheck and regenerate.",
            },
        )
    if guard == "polish_desc_forbidden_tokens":
        log.warning("[%s] generation blocked (engine read-back): forbidden tokens %s; "
                    "PDF + SAD JSON unlinked, audit pointers NOT updated",
                    batch_id, pkg.get("tokens"))
        raise HTTPException(
            status_code=422,
            detail={
                "guard":  guard,
                "error":  "Generated customs document contained forbidden "
                          "placeholder text. Files were not saved.",
                "code":   guard,
                "tokens": pkg.get("tokens") or [],
                "hint":   "Correct the offending line(s) and regenerate.",
            },
        )
    raise HTTPException(
        status_code=422,
        detail={"guard": guard or "generation_blocked",
                "error": "Generation was blocked by a safety guard."},
    )


@router.post("/generate-description/{batch_id}", dependencies=[_auth, _op_auth])
async def generate_description(
    batch_id: str,
    awb: str = "",
    date_override: Optional[str] = None,
    force: bool = False,
    customs_view: str = "invoice_positions",
) -> Dict[str, Any]:
    """
    Manually trigger Polish customs description generation for a batch.
    Now calls generate_customs_description_package() which also generates the SAD-ready JSON.

    Parameters
    ----------
    batch_id      : batch ID to generate for
    awb           : AWB number (optional; taken from audit if not provided)
    date_override : date string YYYY-MM-DD (optional; defaults to today)
    force         : when True, evict cached audit["rows"] before rebuild.
                    Used by the operator's "force-regenerate" UI path to
                    purge stale aggregate rows from prior generations and
                    force the chain to re-read from packing_lines / DB.
                    The rebuild still goes through the full source chain
                    + reconciliation guard — no safety bypass.
    """
    from customs_description_engine import generate_customs_description_package

    audit = _load_audit(batch_id)
    if audit is None:
        raise HTTPException(status_code=404, detail=f"Batch {batch_id} not found.")

    # Force-regenerate:
    #  1. Evict cached audit rows so the source chain rebuilds
    #  2. For Global supplier — also re-reparse the packing PDF and
    #     refresh packing_lines so any rows persisted by an older
    #     parser version (empty metal, missing stone_type) are healed
    #     before the chain re-reads them
    #
    # Reconciliation guard is unchanged; this is a row-source refresh,
    # not a safety bypass.
    if force:
        had_rows = bool(audit.get("rows"))
        _purge_stale_audit_rows(audit)
        if had_rows:
            log.info("[%s] generate_description force=True: cleared "
                     "cached rows so chain rebuilds", batch_id)
        # Reparse packing for Global supplier so any stale packing_lines
        # rows (e.g. empty-metal rows from a pre-fix parser) are healed.
        try:
            if _detect_global_supplier_for_batch(batch_id):
                _force_reparse_global_packing(batch_id)
        except Exception as _exc:
            log.warning(
                "[%s] generate_description force=True: packing reparse "
                "failed (non-fatal — chain may still produce stale rows): %s",
                batch_id, _exc,
            )

    # ── Guard: DHL email check — RELAXED for external_agency_clearance ────────
    # For high-value shipments routed through Agencja Celna Spedycja, the
    # operational order is: generate description IMMEDIATELY → send agency
    # package → THEN handle DHL customs email when it arrives. Blocking the
    # description on dhl_email_received reverses the real workflow.
    _decision = audit.get("clearance_decision") or {}
    _is_agency_path = is_agency_clearance(_decision.get("clearance_path"))
    if not _is_agency_path:
        try:
            _dhl_advisory = guard_dhl_requires_email(audit)
            # advisory mode: returns advisory dict instead of raising
            if _dhl_advisory:
                log.info("[%s] DHL email guard advisory (advisory mode ON): %s",
                         batch_id, _dhl_advisory.get("code"))
                # Persist advisory as Inbox action_proposal (not just a log line)
                from ..pipelines.pz import _advisory_to_action_proposal, _write_advisory_proposal
                _adv_proposal = _advisory_to_action_proposal(
                    _dhl_advisory, batch_id, "dhl_clearance")
                _adv_audit_path = settings.storage_root / "outputs" / batch_id / "audit.json"
                _write_advisory_proposal(_adv_audit_path, _adv_proposal)
        except HTTPException as _ge:
            raise HTTPException(
                status_code=422,
                detail={
                    "guard":   "generate_description_requires_dhl_email",
                    "error":   "Polish description generation requires a DHL customs email or admin override.",
                    "code":    "dhl_email_not_received",
                    "hint":    "Use 'Scan DHL Inbox' first, or mark the email received via the DHL Pre-check panel.",
                },
            ) from _ge

    # ── Guard: CIF must be RESOLVED (positive) across the full authority ladder ──
    # Single platform authority: every customs/PZ/DHL action gates on the shared
    # require_resolved_cif() helper — never on a raw invoice CIF of 0. A CIF
    # resolved from AWB Custom Val or the OCR/AI vision fallback is sufficient to
    # proceed even when invoice parsing yielded 0. UNKNOWN blocks safely with an
    # extraction gap (code "cif_unresolved"); a genuine DECLARED_ZERO requires
    # explicit operator review (code "cif_declared_zero"). Downstream guards
    # (lines_missing_for_description, rows↔audit reconciliation) still protect
    # per-line PDF integrity; this gate only rejects an unresolved or declared
    # -zero customs value, so a valid carrier-declared value is never a false
    # block. See services/cif_authority.require_resolved_cif (wraps cif_resolver).
    from ..services.cif_authority import require_resolved_cif
    require_resolved_cif(audit, action="a Polish customs description")

    # Resolve AWB — check all locations where it may be stored
    resolved_awb = (
        awb or
        audit.get("dhl_awb") or
        audit.get("awb") or
        (audit.get("batch_meta") or {}).get("awb") or
        audit.get("tracking_no") or    # dashboard-upload shipments store AWB here
        ""
    )
    if not resolved_awb:
        raise HTTPException(
            status_code=422,
            detail="AWB not provided and not found in batch audit.json.",
        )

    # PR-206: Enrich audit with invoice rows.  Priority chain is
    # DB invoice_lines (this batch + shared-AWB union) → legacy XLSX
    # Rows sheet → (no rows → 422 below).  Synthetic per-piece averaging
    # in the engine is intentionally unreachable: when neither the DB
    # nor the XLSX has rows, the operator gets a clear manual-review
    # error instead of a misleading PDF.
    #
    # ``customs_view``: "invoice_positions" (default) aggregates the
    # per-row packing source into invoice-position rows (~8-10 positions,
    # 2-5 page PDF for customs use). "packing_rows" preserves the legacy
    # per-row authority (245 rows, 42 pages) for warehouse / audit /
    # detailed verification.
    _cv = (customs_view or "invoice_positions").lower()
    if _cv not in ("invoice_positions", "packing_rows"):
        _cv = "invoice_positions"
    audit = _inject_rows_from_sources(batch_id, audit, customs_view=_cv)

    # Global supplier-specific guard: packing file exists but extractor
    # produced 0 rows. Operator spec: NEVER silently fall through to the
    # aggregate synthesizer (which produces UNKNOWN / metal szlachetny /
    # grouped invoice aggregate). Block with a clear 422 instead so the
    # operator re-uploads or fixes the file.
    if audit.get("_global_packing_present_but_empty"):
        raise HTTPException(
            status_code=422,
            detail={
                "guard":  "global_packing_present_but_empty",
                "error":  "Global Jewellery packing file is on disk for this "
                          "batch but produced 0 parsed rows. Generating "
                          "from aggregate invoice totals would produce "
                          "'UNKNOWN' / 'metal szlachetny' placeholder rows "
                          "and lose per-line product authority.",
                "code":   "global_packing_present_but_empty",
                "hint":   "Re-run Reparse Packing against the Global packing "
                          "list, or re-upload the file. Aggregate fallback "
                          "is intentionally disabled for Global supplier "
                          "when a packing file is present.",
            },
        )

    if not audit.get("rows") and not audit.get("invoices"):
        raise HTTPException(
            status_code=422,
            detail={
                "guard":  "lines_missing_for_description",
                "error":  "No per-line invoice data found in DB invoice_lines, "
                          "XLSX Rows sheet, or audit.json. Generating from "
                          "aggregate totals would average per-piece values "
                          "and lose per-line HSN/karat/stone detail.",
                "code":   "lines_missing_for_description",
                "hint":   "Re-process the batch with valid invoice PDFs or "
                          "attach the PZ calculation XLSX.",
            },
        )

    # PR-206 / PR-547: Reconcile projected rows against aggregate audit totals.
    # Hard block on FOB drift > $1 or missing invoices (financial integrity).
    # Qty drift is advisory-only when FOB is exact — parser divergence between
    # invoice_intake_parser and pz_import_processor produces ±2-3 unit noise.
    _recon = _reconcile_rows_with_audit_totals(audit)
    if _recon.get("soft_warnings"):
        log.info("[%s] generate_description: qty advisory (non-blocking): %s",
                 batch_id, _recon["soft_warnings"])
    if not _recon["ok_hard"]:
        raise HTTPException(
            status_code=422,
            detail={
                "guard":    "rows_audit_reconciliation_failed",
                "error":    "Projected per-line rows do not reconcile with "
                            "the aggregate invoice_totals declared in "
                            "audit.json. Generating a customs document with "
                            "this mismatch would be unsafe.",
                "code":     "rows_audit_reconciliation_failed",
                "warnings": _recon["hard_warnings"],
                "details":  _recon["details"],
                "hint":     "Re-process the batch (Reparse all) or attach a "
                            "fresh PZ calculation XLSX, then retry.",
            },
        )

    _consignee_ov, _consignor_ov = (
        _resolve_customs_identities(batch_id)
        if settings.customs_identity_from_masters
        else (None, None)
    )

    # ── Apply operator-approved description corrections ───────────────────────
    # If the AI-validation layer detected a description mismatch and the
    # operator approved a correction via the Inbox, that correction is stored
    # in audit["description_corrections"][product_code].  Apply it to the
    # projected rows NOW, before passing ``audit`` to the engine, so the
    # engine renders the corrected material_pl rather than the placeholder.
    try:
        from ..services.customs_desc_checker import apply_description_corrections  # noqa: PLC0415
        apply_description_corrections(audit)
    except Exception as _corr_exc:
        log.warning("[%s] apply_description_corrections: non-fatal failure: %s",
                    batch_id, _corr_exc)

    # ── Guard: descriptions_missing_for_customs (pre-generation) ──────────────
    # Every projected line must resolve to an APPROVED, non-generic customs
    # description via the single Product Description Authority resolver
    # (description_engine.resolve_product_description_for_customs). This is the
    # same resolver V1 and V2 rely on through this shared route. A line that
    # would fall back to generic placeholder text ("Wyrób jubilerski" /
    # "metal szlachetny") is blocked HERE, before the PDF is built, with a
    # row-level explanation so the operator sees exactly which line needs a
    # correction — instead of an opaque post-generation forbidden-token 422.
    # The post-generation forbidden-token read-back below remains as a backstop.
    # NOTE (authority): this route-level guard screens audit["rows"] (the DHL
    # projection) as an EARLY, operator-facing block. The AUTHORITATIVE pre-gen
    # guard is engine-internal (generate_customs_description_package Guard #1):
    # it stamps the exact render items (_extract_invoices), fails closed on any
    # resolver error, and also protects the automation/CLI callers that never
    # reach this route. This route guard is a UX convenience, not the sole gate.
    # Resolve + STAMP approved descriptions onto each row so downstream
    # generation (process_batch_items → SAD JSON → PDF) consumes the resolver's
    # authoritative value, not the classifier's own text. Returns row-level
    # detail for any line that cannot be approved (→ block below).
    from ..services.description_engine import resolve_and_stamp_customs_descriptions  # noqa: PLC0415
    _missing_desc = resolve_and_stamp_customs_descriptions(
        audit.get("rows") or [],
        audit.get("description_corrections") or {},
    )
    if _missing_desc:
        log.info("[%s] generate_description blocked: %d line(s) lack an approved "
                 "customs description: %s", batch_id, len(_missing_desc),
                 [m.get("product_code") for m in _missing_desc])
        raise HTTPException(
            status_code=422,
            detail={
                "guard":  "descriptions_missing_for_customs",
                "error":  "One or more product lines have no approved customs "
                          "description and would fall back to generic "
                          "placeholder text. Correct each line before "
                          "generating the Polish description.",
                "code":   "descriptions_missing_for_customs",
                "rows":   _missing_desc,
                "hint":   "For each line: use the approved Product Master "
                          "description, or save a shipment correction via the "
                          "Inbox (action-proposals approve, scope=shipment), "
                          "then Recheck and regenerate. No generic fallback is "
                          "permitted in a customs document.",
            },
        )

    try:
        pkg = generate_customs_description_package(
            batch          = audit,
            awb            = resolved_awb,
            output_dir     = str(_POLISH_DESC_DIR),
            date_override  = date_override,
            consignee_name = _consignee_ov,
            consignor_name = _consignor_ov,
            corrections    = audit.get("description_corrections") or {},
        )
    except Exception as exc:
        log.error("Customs description generation failed: %s", exc, exc_info=True)
        raise HTTPException(status_code=500, detail=str(exc)) from exc

    # Function-internal guards (also cover automation callers) returned a block.
    _translate_blocked_package(batch_id, pkg)

    pdf_result = pkg.get("pdf") or {}
    if not pdf_result.get("generated"):
        raise HTTPException(
            status_code=500,
            detail=f"PDF generation failed: {pdf_result.get('error', 'unknown error')}",
        )

    # ── Validate-then-rollback overwrite safety ──────────────────────────────
    # Operator spec: a generated PDF that contains any of the forbidden
    # placeholder strings (UNKNOWN / metal szlachetny / Wyrób jubilerski /
    # grouped invoice aggregate) MUST NOT be saved. We can't pre-validate
    # (the engine writes directly), so we read-back the generated file,
    # scan its text, and unlink + raise 422 if any forbidden token is
    # present. The audit pointers are NOT updated in that case — the
    # previous state remains, and the operator sees a clear error.
    _generated_path = Path(pdf_result.get("output_path") or "")
    if _generated_path.is_file():
        try:
            import pdfplumber as _pp_validate  # noqa: PLC0415
            with _pp_validate.open(str(_generated_path)) as _pdf_v:
                _pdf_text = "\n".join(
                    (p.extract_text() or "") for p in _pdf_v.pages
                )
            # Single source of truth for forbidden tokens (includes the U+25A0
            # BLACK SQUARE glyph-failure marker). This route-level read-back is a
            # defense-in-depth backstop; the primary read-back now lives inside
            # generate_customs_description_package so automation callers are
            # covered too.
            from ..services.description_engine import PDF_FORBIDDEN_TOKENS as _FORBIDDEN_TOKENS  # noqa: PLC0415
            _hits = [t for t in _FORBIDDEN_TOKENS if t in _pdf_text]
            if _hits:
                # Roll back: unlink the bad file, do NOT touch audit pointers.
                try:
                    _generated_path.unlink()
                except OSError:
                    pass
                log.warning(
                    "[%s] generated PDF rejected — forbidden tokens %s. "
                    "File unlinked, audit pointers NOT updated.",
                    batch_id, _hits,
                )
                raise HTTPException(
                    status_code=422,
                    detail={
                        "guard":  "polish_desc_forbidden_tokens",
                        "error":  "Generated Polish description contains "
                                  "forbidden placeholder text. File not saved.",
                        "code":   "polish_desc_forbidden_tokens",
                        "tokens": _hits,
                        "hint":   "Reparse packing list and verify rows "
                                  "produce real PL/EN descriptions before "
                                  "regenerating.",
                    },
                )
        except HTTPException:
            raise
        except Exception as _ve:
            # Validation infrastructure failed (e.g. pdfplumber unavailable)
            # — degrade gracefully, log, do NOT roll back the file.
            log.warning(
                "[%s] forbidden-token validation skipped (%s); generated "
                "file preserved", batch_id, _ve,
            )

    # Update audit
    audit["clearance_status"]      = "polish_description_generated"
    audit["polish_desc_filename"]  = pdf_result.get("filename")
    audit["polish_desc_path"]      = pdf_result.get("output_path")
    audit["polish_desc_generated_at"] = datetime.now(timezone.utc).isoformat()
    audit["polish_desc_file_exists"]  = True
    json_result = pkg.get("json") or {}
    if json_result.get("generated"):
        audit["sad_ready_filename"] = json_result.get("filename")
        audit["sad_ready_path"]     = json_result.get("output_path")
    _write_audit(batch_id, audit)

    # Log timeline event
    for sub in ("outputs", "working"):
        _ap = settings.storage_root / sub / batch_id / "audit.json"
        if _ap.exists():
            tl.log_event(_ap, tl.EV_DESCRIPTION_READY, "dashboard", "admin",
                         detail={"filename": pdf_result.get("filename", "")})
            break

    # ── Auto-trigger agency package build for external_agency_clearance ───────
    # When CIF > threshold (agency path), Polish description → agency package
    # is the standard, deterministic next step. Build immediately so the operator
    # doesn't have to click again. Skip if already built.
    auto_agency_built = False
    auto_agency_error = None
    if _is_agency_path and not (audit.get("agency_reply_package") or {}).get("status"):
        try:
            from ..services.agency_email_builder import build_agency_package
            from ..services.email_service       import queue_email
            _pkg = build_agency_package(audit, batch_id)
            _missing_files = [
                a["label"] for a in _pkg.get("attachments", [])
                if not Path(a.get("path", "")).exists()
            ] + (_pkg.get("missing") or [])
            if _missing_files:
                auto_agency_error = {"missing": _missing_files,
                                     "hint": "regenerate missing files before agency build"}
            elif not (_pkg.get("to") or "").strip():
                auto_agency_error = {"error": "no recipients"}
            else:
                _body_text = f"{_pkg['body_pl']}\n\n---\n\n{_pkg['body_en']}".strip()
                _body_html = ("<div style='font-family:sans-serif'>"
                              f"<pre style='white-space:pre-wrap'>{_pkg['body_pl']}</pre><hr/>"
                              f"<pre style='white-space:pre-wrap'>{_pkg['body_en']}</pre></div>")
                _email_id = queue_email(
                    to          = _pkg["to"],
                    subject     = _pkg["subject"],
                    body_html   = _body_html,
                    body_text   = _body_text,
                    batch_id    = batch_id,
                    cc          = _pkg.get("cc", ""),
                    from_address= _pkg.get("from_address", ""),
                    email_type  = _pkg.get("email_type", "agency"),
                    # Pass attachments at queue time — integrity guard must fire
                    # on the synchronous SMTP attempt before audit.json is written.
                    attachments = _pkg.get("attachments", []),
                )
                from ..utils.io import write_json_atomic
                from datetime import datetime as _dt, timezone as _tz
                audit["agency_reply_package"] = {
                    "to":          _pkg["to"],
                    "to_list":     _pkg.get("to_list", []),
                    "cc":          _pkg.get("cc", ""),
                    "cc_list":     _pkg.get("cc_list", []),
                    "subject":     _pkg["subject"],
                    "body_pl":     _pkg["body_pl"],
                    "body_en":     _pkg["body_en"],
                    "attachments": _pkg["attachments"],
                    "email_id":    _email_id,
                    "status":      "queued",
                    "queued_at":   _dt.now(_tz.utc).isoformat(),
                    "source":      "auto_after_polish_desc",
                }
                _write_audit(batch_id, audit)
                auto_agency_built = True
                # Log timeline event
                for sub in ("outputs", "working"):
                    _ap = settings.storage_root / sub / batch_id / "audit.json"
                    if _ap.exists():
                        tl.log_event(_ap, "agency_package_auto_built", "system", "auto",
                                     detail={"email_id": _email_id, "trigger": "polish_desc_generated"})
                        break
                log.info("[auto-agency] package built and queued for batch %s (email_id=%s)",
                         batch_id, _email_id)
        except Exception as exc:
            log.warning("[auto-agency] auto-build failed for batch %s: %s", batch_id, exc)
            auto_agency_error = {"error": str(exc)}

    fn = pdf_result.get("filename", "")
    return {
        "ok":              True,
        "batch_id":        batch_id,
        "action":          "polish_description_generated",
        "status":          "generated",
        "generated":       True,
        "filename":        fn,
        "download_url":    f"{settings.fastapi_public_url}/api/v1/dhl/download/{fn}" if fn else None,
        "items_described": pdf_result.get("items_described"),
        "sad_json_filename": json_result.get("filename"),
        "auto_agency_built": auto_agency_built,
        "auto_agency_error": auto_agency_error,
        "errors":          [],
        "warnings":        [],
    }


@router.get("/download/{filename}", dependencies=[_auth])
async def download_dhl_file(filename: str) -> FileResponse:
    """
    Download a DHL clearance-related file (PDF or SAD-ready JSON).

    Searches in:
    - DSK output directory
    - Polish descriptions directory
    - SAD-ready directory

    **Cache policy (operator-locked):** generated PDFs are regenerable
    artifacts that change per click — the response sets
    ``Cache-Control: no-store, no-cache, must-revalidate, max-age=0``
    so the browser ALWAYS fetches a fresh copy. Prior behaviour cached
    for four hours (FastAPI FileResponse default), which caused stale
    PDFs to persist in the browser cache even after regenerate/delete.
    """
    # Sanitize filename — no path traversal
    if "/" in filename or "\\" in filename or ".." in filename:
        raise HTTPException(status_code=400, detail="Invalid filename.")
    ext = Path(filename).suffix.lower()
    if ext not in (".pdf", ".json"):
        raise HTTPException(status_code=400, detail="Only PDF and JSON files are served here.")

    found = _find_generated_file(filename)
    if found is None:
        raise HTTPException(status_code=404, detail=f"File not found: {filename}")

    media_type = "application/pdf" if ext == ".pdf" else "application/json"
    # Defeat aggressive browser caching of generated artifacts.
    # POLISH_DESC_*, DSK_*, SAD_READY_* files share the same filename
    # across regenerations (date-stamped) so without explicit no-store
    # the browser serves a stale cached copy for max-age=14400 (4h).
    no_cache_headers = {
        "Cache-Control": "no-store, no-cache, must-revalidate, max-age=0",
        "Pragma":        "no-cache",
        "Expires":       "0",
    }
    return FileResponse(
        path=str(found),
        media_type=media_type,
        filename=filename,
        headers=no_cache_headers,
    )


# ── New endpoints: customs description package, SAD-ready JSON, approval ──────

@router.post("/generate-customs-package/{batch_id}", dependencies=[_auth, _op_auth])
async def generate_customs_package(
    batch_id: str,
    body: GenerateCustomsPackageRequest,
) -> Dict[str, Any]:
    """
    Generate the full customs description package (Polish PDF + SAD-ready JSON) for a batch.

    Body: {"awb": "...", "dhl_email_id": "...", "date_override": "YYYY-MM-DD"}
    Returns combined result with download URLs for both files.
    """
    from customs_description_engine import generate_customs_description_package

    audit = _load_audit(batch_id)
    if audit is None:
        raise HTTPException(status_code=404, detail=f"Batch {batch_id} not found.")

    awb = (
        body.awb or
        audit.get("dhl_awb") or
        audit.get("awb") or
        (audit.get("batch_meta") or {}).get("awb") or
        audit.get("tracking_no") or
        ""
    )
    if not awb:
        raise HTTPException(status_code=422, detail="AWB not provided and not found in audit.json.")

    # ── Guard: CIF must be RESOLVED across the full authority ladder ──
    # The customs package (Polish PDF + SAD-ready JSON) declares the customs
    # value; it must gate on the SAME single CIF authority as the Polish
    # description and DSK — never on a raw invoice CIF of 0. UNKNOWN blocks with
    # an extraction gap (cif_unresolved); a genuine DECLARED_ZERO requires
    # operator review (cif_declared_zero). A value resolved from AWB Custom Val
    # or the OCR/AI fallback is sufficient to proceed.
    from ..services.cif_authority import require_resolved_cif
    require_resolved_cif(audit, action="a customs description package")

    # PR-206: DB-first row injection + lines-missing guard + reconciliation.
    audit = _inject_rows_from_sources(batch_id, audit)

    if not audit.get("rows") and not audit.get("invoices"):
        raise HTTPException(
            status_code=422,
            detail={
                "guard":  "lines_missing_for_description",
                "error":  "No per-line invoice data found in DB invoice_lines, "
                          "XLSX Rows sheet, or audit.json.",
                "code":   "lines_missing_for_description",
                "hint":   "Re-process the batch with valid invoice PDFs or "
                          "attach the PZ calculation XLSX.",
            },
        )

    _recon = _reconcile_rows_with_audit_totals(audit)
    if _recon.get("soft_warnings"):
        log.info("[%s] generate_customs_package: qty advisory (non-blocking): %s",
                 batch_id, _recon["soft_warnings"])
    if not _recon["ok_hard"]:
        raise HTTPException(
            status_code=422,
            detail={
                "guard":    "rows_audit_reconciliation_failed",
                "error":    "Projected rows do not reconcile with aggregate "
                            "invoice_totals declared in audit.json.",
                "code":     "rows_audit_reconciliation_failed",
                "warnings": _recon["hard_warnings"],
                "details":  _recon["details"],
                "hint":     "Re-process the batch or attach a fresh PZ XLSX.",
            },
        )

    # Parity with generate_description: apply operator-approved corrections to
    # the raw row fields before resolving/stamping.
    try:
        from ..services.customs_desc_checker import apply_description_corrections  # noqa: PLC0415
        apply_description_corrections(audit)
    except Exception as _corr_exc2:
        log.warning("[%s] apply_description_corrections (pkg): non-fatal failure: %s",
                    batch_id, _corr_exc2)

    # ── Guard: descriptions_missing_for_customs (pre-generation) ──────────────
    # Same single-authority resolver as generate_description — no bypass path.
    # NOTE (authority): like generate_description, this route-level guard screens
    # audit["rows"] as an EARLY, operator-facing block; the AUTHORITATIVE pre-gen
    # guard is engine-internal (generate_customs_description_package Guard #1),
    # which stamps the exact render items and fails closed on any resolver error.
    # Resolve + STAMP approved descriptions onto each row so downstream
    # generation (process_batch_items → SAD JSON → PDF) consumes the resolver's
    # authoritative value, not the classifier's own text. Returns row-level
    # detail for any line that cannot be approved (→ block below).
    from ..services.description_engine import resolve_and_stamp_customs_descriptions  # noqa: PLC0415
    _missing_desc = resolve_and_stamp_customs_descriptions(
        audit.get("rows") or [],
        audit.get("description_corrections") or {},
    )
    if _missing_desc:
        log.info("[%s] generate_customs_package blocked: %d line(s) lack an "
                 "approved customs description: %s", batch_id, len(_missing_desc),
                 [m.get("product_code") for m in _missing_desc])
        raise HTTPException(
            status_code=422,
            detail={
                "guard":  "descriptions_missing_for_customs",
                "error":  "One or more product lines have no approved customs "
                          "description and would fall back to generic "
                          "placeholder text.",
                "code":   "descriptions_missing_for_customs",
                "rows":   _missing_desc,
                "hint":   "Correct each line (approved Product Master "
                          "description or a shipment correction), then Recheck "
                          "and regenerate.",
            },
        )

    _consignee_ov2, _consignor_ov2 = (
        _resolve_customs_identities(batch_id)
        if settings.customs_identity_from_masters
        else (None, None)
    )

    try:
        pkg = generate_customs_description_package(
            batch          = audit,
            awb            = awb,
            output_dir     = str(_POLISH_DESC_DIR),
            dhl_email_id   = body.dhl_email_id or "",
            date_override  = body.date_override,
            consignee_name = _consignee_ov2,
            consignor_name = _consignor_ov2,
            corrections    = audit.get("description_corrections") or {},
        )
    except Exception as exc:
        log.error("generate_customs_package error: %s", exc, exc_info=True)
        raise HTTPException(status_code=500, detail=str(exc)) from exc

    # Engine-internal guards (missing-description / forbidden-token read-back)
    # returned a block — translate to the same 422 as generate_description.
    _translate_blocked_package(batch_id, pkg)

    pdf_result  = pkg.get("pdf") or {}
    json_result = pkg.get("json") or {}

    if not pdf_result.get("generated"):
        raise HTTPException(
            status_code=500,
            detail=f"PDF generation failed: {pdf_result.get('error', 'unknown')}",
        )
    if not json_result.get("generated"):
        raise HTTPException(
            status_code=500,
            detail=f"SAD JSON generation failed: {json_result.get('error', 'unknown')}",
        )

    # Update audit.json
    audit["clearance_status"]            = "polish_description_generated"
    audit["polish_desc_filename"]        = pdf_result.get("filename")
    audit["polish_desc_path"]            = pdf_result.get("output_path")
    audit["sad_ready_filename"]          = json_result.get("filename")
    audit["sad_ready_path"]              = json_result.get("output_path")
    audit["customs_package_generated_at"] = pkg.get("audit", {}).get("generated_at")
    audit["customs_pdf_hash"]            = pdf_result.get("pdf_hash")
    audit["customs_json_hash"]           = json_result.get("json_hash")
    _write_audit(batch_id, audit)

    pdf_fn  = pdf_result.get("filename", "")
    json_fn = json_result.get("filename", "")

    # Read classification_summary and error_flags from the generated SAD JSON
    classification_summary: Optional[dict] = None
    error_flags: Optional[dict] = None
    sad_output_path = json_result.get("output_path") or audit.get("sad_ready_path")
    if sad_output_path:
        try:
            sad_data = json.loads(Path(sad_output_path).read_text(encoding="utf-8"))
            classification_summary = sad_data.get("classification_summary")
            error_flags            = sad_data.get("error_flags")
        except Exception:
            pass

    return {
        "generated":             True,
        "batch_id":              batch_id,
        "awb":                   re.sub(r"\s+", "", awb),
        "classification_summary": classification_summary,
        "error_flags":           error_flags,
        "pdf": {
            "filename":        pdf_fn,
            "items_described": pdf_result.get("items_described"),
            "pdf_hash":        pdf_result.get("pdf_hash"),
            "download_url":    f"{settings.fastapi_public_url}/api/v1/dhl/download/{pdf_fn}" if pdf_fn else None,
        },
        "json": {
            "filename":    json_fn,
            "total_lines": json_result.get("total_lines"),
            "json_hash":   json_result.get("json_hash"),
            "download_url": f"{settings.fastapi_public_url}/api/v1/dhl/download/{json_fn}" if json_fn else None,
        },
        "audit": pkg.get("audit"),
    }


@router.get("/sad-ready/{batch_id}", dependencies=[_auth])
async def get_sad_ready(batch_id: str) -> Dict[str, Any]:
    """
    Return the parsed SAD-ready JSON data for a batch.

    This returns the structured data, not a raw file download.
    Use /download/{filename} for the raw JSON file.
    """
    audit = _load_audit(batch_id)
    if audit is None:
        raise HTTPException(status_code=404, detail=f"Batch {batch_id} not found.")

    sad_path = audit.get("sad_ready_path")
    awb      = audit.get("dhl_awb") or audit.get("awb") or ""

    # Try the stored path first
    if sad_path:
        p = Path(sad_path)
        if p.exists():
            try:
                return json.loads(p.read_text(encoding="utf-8"))
            except Exception as exc:
                raise HTTPException(status_code=500, detail=f"Could not parse SAD JSON: {exc}") from exc

    # Fallback: search by AWB
    found = _find_sad_json(batch_id, awb=awb)
    if found:
        try:
            return json.loads(found.read_text(encoding="utf-8"))
        except Exception as exc:
            raise HTTPException(status_code=500, detail=f"Could not parse SAD JSON: {exc}") from exc

    raise HTTPException(
        status_code=404,
        detail=f"SAD-ready JSON not found for batch {batch_id}. "
               "Run POST /api/v1/dhl/generate-customs-package/{batch_id} first.",
    )


def _update_batch_reply_delivery(batch_id: str, delivery_status: str, error: Optional[str] = None) -> bool:
    """
    Propagate email delivery confirmation back to the batch audit.json.
    Called by mark_email_sent (admin endpoint) after MCP confirms send/fail.

    delivery_status: "sent" | "failed"
    Returns True if the audit was found and updated.
    """
    audit = _load_audit(batch_id)
    if audit is None:
        log.warning("_update_batch_reply_delivery: batch %s not found", batch_id)
        return False

    now = datetime.now(timezone.utc).isoformat()
    audit["dhl_reply_status"] = delivery_status          # sent | failed
    audit["email_status"]     = delivery_status
    if delivery_status == "sent":
        audit["clearance_status"]    = "reply_sent"
        audit["dhl_reply_sent_at"]   = now
    else:
        audit["clearance_status"]    = "reply_failed"
        audit["dhl_reply_failed_at"] = now
        audit["dhl_reply_error"]     = error or "unknown"

    audit["clearance_updated_at"] = now
    _write_audit(batch_id, audit)
    log.info("DHL reply delivery propagated: batch=%s status=%s", batch_id, delivery_status)
    return True


@router.get("/reply-status/{batch_id}", dependencies=[_auth])
async def get_reply_status(batch_id: str) -> Dict[str, Any]:
    """
    Return current DHL reply delivery status for a batch.
    Cross-references audit.json (batch state) with email_queue.json (queue state).
    """
    if "/" in batch_id or ".." in batch_id:
        raise HTTPException(status_code=400, detail="Invalid batch_id.")

    from ..services import email_service as _email_svc

    audit = _load_audit(batch_id)
    if audit is None:
        raise HTTPException(status_code=404, detail=f"Batch {batch_id} not found.")

    email_id = audit.get("dhl_reply_email_id")

    # Cross-reference the email queue entry
    queue_entry = None
    if email_id:
        for e in _email_svc.get_all_emails(limit=200):
            if e.get("id") == email_id:
                queue_entry = e
                break

    # Determine live status — queue entry is authoritative over audit if available
    live_status = audit.get("dhl_reply_status") or "not_queued"
    if queue_entry:
        q_status = queue_entry.get("status", "pending")
        # Propagate to audit if queue shows a terminal state we haven't recorded yet
        if q_status == "sent" and live_status == "queued":
            _update_batch_reply_delivery(batch_id, "sent")
            live_status = "sent"
        elif q_status == "failed" and live_status not in ("failed",):
            _update_batch_reply_delivery(batch_id, "failed", queue_entry.get("error"))
            live_status = "failed"

    return {
        "batch_id":          batch_id,
        "clearance_status":  audit.get("clearance_status"),
        "dhl_reply_status":  live_status,
        "email_status":      audit.get("email_status"),
        "email_id":          email_id,
        "queue_status":      queue_entry.get("status") if queue_entry else None,
        "queued_at":         audit.get("dhl_reply_queued_at"),
        "sent_at":           queue_entry.get("sent_at") if queue_entry else audit.get("dhl_reply_sent_at"),
        "failed_at":         audit.get("dhl_reply_failed_at"),
        "error":             queue_entry.get("error") if queue_entry else audit.get("dhl_reply_error"),
        "to":                audit.get("dhl_reply_to"),
        "subject":           audit.get("dhl_reply_subject"),
    }


@router.post("/send-reply/{batch_id}", dependencies=[_auth, _op_auth])
async def send_dhl_reply(batch_id: str) -> Dict[str, Any]:
    """
    Queue the prepared DHL reply email for sending.

    Reads the reply package built by /email-package or /match-and-handle from audit.json,
    queues it via email_service, updates clearance_status → "reply_queued",
    and logs an EV_REPLY_APPROVED timeline event.

    Does NOT send immediately — the email goes to the JSON queue for MCP pickup.
    Admin clicks this button manually after reviewing the reply package.
    """
    from ..services import email_service

    audit = _load_audit(batch_id)
    if audit is None:
        raise HTTPException(status_code=404, detail=f"Batch {batch_id} not found.")

    # Guard: DHL email must have been received first
    clr_status = audit.get("clearance_status") or ""
    if not clr_status or clr_status in ("awaiting_dhl_customs_email", "shipment_created", ""):
        raise HTTPException(
            status_code=422,
            detail={
                "guard": "send_reply_requires_email",
                "error": "Cannot send reply — no DHL customs email has been received yet.",
                "code":  "dhl_email_not_received",
            },
        )

    # Resolve reply package from audit
    reply_pkg = audit.get("reply_package") or {}
    if not reply_pkg or not reply_pkg.get("to"):
        raise HTTPException(
            status_code=422,
            detail="Reply package not found. Run 'Build Reply Package' first.",
        )

    # ── Clearance routing guard (Step A: assert reply package valid) ────────────
    # Order: (1) reply_package validity → (2) DSK exists for high-value → (3) proceed
    try:
        from ..services.clearance_decision import assert_valid_dhl_reply, resolve_dhl_action
        assert_valid_dhl_reply(audit, reply_pkg)
        _dhl_action = resolve_dhl_action(audit)
        _is_dsk_transfer = _dhl_action["action"] == "dsk_transfer"
    except ValueError as _guard_exc:
        raise HTTPException(
            status_code=422,
            detail={
                "guard":   "clearance_routing",
                "error":   str(_guard_exc),
                "code":    "invalid_dhl_flow_high_value",
                "hint":    "Generate DSK first, then use 'Build DHL Reply Package' to include it.",
            },
        ) from _guard_exc
    except ImportError:
        _is_dsk_transfer = False   # clearance_decision module unavailable — allow legacy flow

    # ── Agency flow lock: DSK file must exist on disk before DHL transfer ────
    if _is_dsk_transfer:
        _dsk_file = audit.get("dsk_filename") or ""
        _dsk_on_disk = False
        if _dsk_file:
            for _sub in ("dsk_outputs", "outputs"):
                _candidate = settings.storage_root / _sub / _dsk_file
                if not _candidate.exists():
                    _candidate = settings.storage_root / _sub / batch_id / _dsk_file
                if _candidate.exists():
                    _dsk_on_disk = True
                    break
        if not _dsk_on_disk:
            raise HTTPException(
                status_code=422,
                detail={
                    "guard":  "dsk_file_required",
                    "error":  "DSK file must be generated before sending DHL transfer notification.",
                    "code":   "dsk_missing_high_value",
                    "hint":   "Click 'Generate DSK', verify the file appears, then rebuild the reply package.",
                },
            )

    # ── Polish description physical file check (carrier path) ────────────────
    # For both paths the description is an attachment; confirm file exists on disk.
    _pd_file = audit.get("polish_desc_filename") or audit.get("polish_desc_path") or ""
    if _pd_file:
        _pd_on_disk = False
        for _pd_sub in ("polish_descriptions", "outputs", "dsk_outputs"):
            _pd_candidate = settings.storage_root / _pd_sub / _pd_file
            if not _pd_candidate.exists():
                _pd_candidate = settings.storage_root / _pd_sub / batch_id / _pd_file
            if _pd_candidate.exists():
                _pd_on_disk = True
                break
        if not _pd_on_disk:
            raise HTTPException(
                status_code=422,
                detail={
                    "guard":  "polish_desc_missing",
                    "error":  f"Polish description file not found on disk: {_pd_file}",
                    "code":   "polish_desc_missing",
                    "hint":   "Re-generate Polish Description before building the reply package.",
                },
            )

    # ── Resolve TO/CC from centralized config ────────────────────────────────
    # Always send to DHL_TO from config (overrides anything stored in reply_package).
    # Keep subject from reply_package to preserve thread continuity.
    to_addr   = format_to(DHL_TO)   # "odprawacelna@dhl.com" (or override from config)
    cc_addr   = format_cc(INTERNAL_CC)  # internal always in CC
    subject   = reply_pkg.get("subject",  "")
    body_pl   = reply_pkg.get("body_pl",  "")
    body_en   = reply_pkg.get("body_en",  "")
    thread_id = reply_pkg.get("thread_id", "")

    # ── Validation guard: TO must be present ─────────────────────────────────
    if not to_addr:
        raise HTTPException(
            status_code=422,
            detail={"guard": "missing_recipients", "error": "DHL reply has no recipients (TO is empty)."},
        )

    # Build email body (Polish first, then EN)
    full_body_text = f"{body_pl}\n\n---\n\n{body_en}".strip()
    full_body_html = (
        f"<div style='font-family:sans-serif'>"
        f"<pre style='white-space:pre-wrap'>{body_pl}</pre>"
        f"<hr/>"
        f"<pre style='white-space:pre-wrap'>{body_en}</pre>"
        f"</div>"
    )

    # Queue the email
    # Pass attachment metadata directly into the queue entry so the integrity
    # guard can fire on the synchronous SMTP attempt inside queue_email(),
    # which fires before audit.json is updated.
    _reply_attachments = list(reply_pkg.get("attachments") or [])
    try:
        email_id = email_service.queue_email(
            to          = to_addr,
            subject     = subject,
            body_html   = full_body_html,
            body_text   = full_body_text,
            batch_id    = batch_id,   # links queue entry → batch for delivery propagation
            cc          = cc_addr,
            email_type  = "dhl_reply",
            attachments = _reply_attachments,
        )
    except Exception as exc:
        log.error("send_dhl_reply: queue_email failed: %s", exc, exc_info=True)
        raise HTTPException(status_code=500, detail=f"Failed to queue email: {exc}") from exc

    # Update audit
    sent_at = datetime.now(timezone.utc).isoformat()
    audit["clearance_status"]      = "reply_queued"   # NOT reply_sent — email is queued, not delivered
    audit["clearance_updated_at"]  = sent_at
    audit["dhl_reply_email_id"]    = email_id
    audit["dhl_reply_queued_at"]   = sent_at
    audit["dhl_reply_status"]      = "queued"         # explicit delivery state: queued / sent / failed
    audit["email_status"]          = "queued"         # MCP pickup pending
    audit["dhl_reply_to"]          = to_addr
    audit["dhl_reply_cc"]          = cc_addr
    audit["dhl_reply_subject"]     = subject
    _write_audit(batch_id, audit)

    # ── Mirror into email evidence store ─────────────────────────────────────
    _awb_reply = str(audit.get("dhl_awb") or audit.get("awb") or audit.get("tracking_no") or "")
    if _awb_reply:
        try:
            from ..services import email_evidence_store as evs
            evs.link_batch(_awb_reply, batch_id)
            evs.save_message(_awb_reply, {
                "message_id":      f"op_dhl_reply:{batch_id}",
                "thread_id":       f"op_dhl_reply:{batch_id}",
                "direction":       "outgoing",
                "sender":          "import@estrellajewels.eu",
                "to":              [to_addr] if to_addr else [],
                "cc":              [cc_addr] if cc_addr else [],
                "subject":         subject,
                "body_text":       f"DHL reply queued for batch {batch_id}.",
                "timestamp":       sent_at,
                "event_type":      "our_dhl_reply",
                "delivery_status": "queued",
                "matched_identifiers": {"awb": True},
                "attachments":     [],
                "source":          "operator_send",
            }, source="operator_send")
        except Exception as _evs_exc:
            log.warning("[send_dhl_reply] evidence store write failed (non-fatal): %s", _evs_exc)

    # Timeline event — distinguish DSK transfer from standard description reply
    _tl_event = tl.EV_DSK_TRANSFER_SENT if _is_dsk_transfer else tl.EV_REPLY_APPROVED
    for sub in ("outputs", "working"):
        audit_path = settings.storage_root / sub / batch_id / "audit.json"
        if audit_path.exists():
            tl.log_event(
                audit_path,
                _tl_event,
                trigger_source="dashboard",
                actor="admin",
                detail={
                    "email_id":    email_id,
                    "to":          to_addr,
                    "cc":          cc_addr,
                    "subject":     subject,
                    "reply_type":  "dsk_transfer" if _is_dsk_transfer else "description_reply",
                },
            )
            break

    log.info("DHL reply queued: batch=%s email_id=%s to=%s cc=%s type=%s",
             batch_id, email_id, to_addr, cc_addr, "dsk_transfer" if _is_dsk_transfer else "description")

    return {
        "ok":           True,
        "batch_id":     batch_id,
        "action":       "dsk_transfer" if _is_dsk_transfer else "description_reply",
        "status":       "reply_queued",
        "queued":       True,
        "email_id":     email_id,
        "to":           to_addr,
        "cc":           cc_addr,
        "subject":      subject,
        "thread_id":    thread_id,
        "queued_at":    sent_at,
        "clearance_status": "reply_queued",
        "errors":       [],
        "warnings":     [],
        "message": (
            "Reply queued for sending. Open the email queue (Admin → Email Queue) "
            "to send via ZohoMail MCP."
        ),
    }


@router.post("/approve/{batch_id}", dependencies=[_auth, _op_auth])
async def approve_description(
    batch_id: str,
    body: ApproveDescriptionRequest,
) -> Dict[str, Any]:
    """
    Approve the customs description package for a batch.

    Stores approved_by (name) and approved_at (ISO8601 timestamp) in both
    the SAD-ready JSON and the batch audit.json.

    Body: {"approved_by": "amit.gupta"}
    """
    if not body.approved_by or not body.approved_by.strip():
        raise HTTPException(status_code=422, detail="approved_by must not be empty.")

    audit = _load_audit(batch_id)
    if audit is None:
        raise HTTPException(status_code=404, detail=f"Batch {batch_id} not found.")

    approved_at  = datetime.now(timezone.utc).isoformat()
    approved_by  = body.approved_by.strip()

    # Update audit.json
    audit["customs_approved_by"] = approved_by
    audit["customs_approved_at"] = approved_at
    _write_audit(batch_id, audit)

    # Update SAD-ready JSON if it exists
    sad_path = audit.get("sad_ready_path")
    sad_updated = False
    if sad_path:
        p = Path(sad_path)
        if p.exists():
            try:
                sad_data = json.loads(p.read_text(encoding="utf-8"))
                sad_data["audit"]["approved_by"] = approved_by
                sad_data["audit"]["approved_at"] = approved_at
                tmp = p.with_suffix(".tmp")
                tmp.write_text(json.dumps(sad_data, indent=2, ensure_ascii=False), encoding="utf-8")
                tmp.replace(p)
                sad_updated = True
            except Exception as exc:
                log.warning("Could not update SAD JSON approval: %s", exc)

    return {
        "approved":     True,
        "batch_id":     batch_id,
        "approved_by":  approved_by,
        "approved_at":  approved_at,
        "sad_updated":  sad_updated,
    }


@router.post("/mark-email-received/{batch_id}", dependencies=[_auth, _op_auth])
async def mark_email_received(
    batch_id: str,
    body: Optional[MarkEmailReceivedRequest] = Body(default=None),
) -> Dict[str, Any]:
    """
    Admin manual override: mark a DHL customs email as received for this batch.

    This satisfies guard_dhl_requires_email so that Generate Description,
    Generate DSK, and Build Reply Package can proceed without waiting for
    the automatic email scan.

    Body fields are all optional — admin fills in what they know.
    """
    if body is None:
        raise HTTPException(
            status_code=422,
            detail="Request body required. Use dashboard form or send JSON fields: "
                   "{sender, subject, ticket, request_type, received_at, note} (all optional).",
        )

    if "/" in batch_id or ".." in batch_id:
        raise HTTPException(status_code=400, detail="Invalid batch_id.")

    audit = _load_audit(batch_id)
    if audit is None:
        raise HTTPException(status_code=404, detail=f"Batch {batch_id} not found.")

    received_at = body.received_at or datetime.now(timezone.utc).isoformat()
    _sender = body.sender or "odprawacelna@dhl.com"

    dhl_email = {
        "received":      True,
        "source":        "manual_admin",
        "sender":        _sender,
        "subject":       body.subject or "",
        "ticket":        body.ticket or "",
        "request_type":  body.request_type or "unknown",
        "received_at":   received_at,
        "note":          body.note or "",
    }
    if body.ticket:
        audit["dhl_ticket"] = body.ticket

    # ── DSK source tracking ───────────────────────────────────────────────────
    # If the email came from administracja_centralna@dhl.com (DHL_DSK_SOURCE),
    # it carries the broker (DSK) notification — record in audit.
    if is_dsk_source(_sender):
        audit["dsk_received"]    = True
        audit["dsk_source"]      = _sender
        audit["dsk_received_at"] = received_at
        log.info("DSK source email detected: batch=%s sender=%s", batch_id, _sender)

    audit["dhl_email"]       = dhl_email
    audit["clearance_status"] = "dhl_email_received"
    audit["clearance_updated_at"] = datetime.now(timezone.utc).isoformat()
    _write_audit(batch_id, audit)

    # ── Mirror into email evidence store ─────────────────────────────────────
    _awb = str(audit.get("dhl_awb") or audit.get("awb") or audit.get("tracking_no") or "")
    if _awb:
        try:
            from ..services import email_evidence_store as evs
            evs.link_batch(_awb, batch_id)
            evs.save_message(_awb, {
                "message_id":  f"op_dhl_request:{batch_id}",
                "thread_id":   f"op_dhl_request:{batch_id}",
                "direction":   "incoming",
                "sender":      _sender,
                "to":          [],
                "cc":          [],
                "subject":     body.subject or "",
                "body_text":   f"DHL customs email manually marked as received for batch {batch_id}.",
                "timestamp":   received_at,
                "event_type":  "dhl_request",
                "matched_identifiers": {"awb": True},
                "attachments": [],
                "source":      "manual_admin",
            }, source="manual_admin")
        except Exception as _evs_exc:
            log.warning("[mark_email_received] evidence store write failed (non-fatal): %s", _evs_exc)

    # Log timeline event
    for sub in ("outputs", "working"):
        audit_path = settings.storage_root / sub / batch_id / "audit.json"
        if audit_path.exists():
            tl.log_event(
                audit_path,
                tl.EV_DHL_EMAIL_RECEIVED,
                trigger_source="dashboard",
                actor="admin",
                detail={
                    "source":       "manual_admin",
                    "ticket":       body.ticket or "",
                    "request_type": body.request_type or "unknown",
                },
            )
            break

    log.info("DHL email manually marked received: batch=%s ticket=%s", batch_id, body.ticket or "—")

    return {
        "ok":               True,
        "batch_id":         batch_id,
        "action":           "dhl_email_marked_received",
        "status":           "dhl_email_received",
        "marked":           True,
        "clearance_status": "dhl_email_received",
        "dhl_email":        dhl_email,
        "errors":           [],
        "warnings":         [],
        "message":          "DHL email marked as received. Generate Description and DSK are now unlocked.",
    }


# ── Serialization helpers ─────────────────────────────────────────────────────

def _safe_dsk(dsk: Optional[dict]) -> Optional[dict]:
    """Return a serialization-safe subset of the DSK result."""
    if not dsk:
        return None
    return {
        "generated":        dsk.get("generated"),
        "filename":         dsk.get("filename"),
        "awb_clean":        dsk.get("awb_clean"),
        "awb_formatted":    dsk.get("awb_formatted"),
        "date":             dsk.get("date"),
        "skip_reason":      dsk.get("skip_reason"),
        "error":            dsk.get("error"),
        "version":          dsk.get("version"),
        "file_hash_sha256": dsk.get("file_hash_sha256"),
    }


def _safe_polish(pol: Optional[dict]) -> Optional[dict]:
    """Return a serialization-safe subset of the Polish description result."""
    if not pol:
        return None
    return {
        "generated":       pol.get("generated"),
        "filename":        pol.get("filename"),
        "items_described": pol.get("items_described"),
        "error":           pol.get("error"),
    }


# ── P2 Slice A: Proactive DHL customs dispatch ──────────────────────────────
# First-contact endpoint that creates an action proposal for proactive
# customs dispatch on a low-value DHL self-clearance shipment. Uses the
# existing routes_action_proposals approve/queue pipeline; no new execute
# endpoint. Locked under proposal_write_lock(batch_id) so concurrent POSTs
# for the same batch dedupe to a single proposal.

class ProactiveDispatchRequest(BaseModel):
    """Request body — operator_id ONLY. NO recipient/CC fields."""
    operator_id: str

    class Config:
        # Pydantic v1 — reject unknown fields so a malicious caller cannot
        # smuggle a `to:` or `cc:` override through the schema.
        extra = "forbid"


def _resolve_proactive_awb(audit: Dict[str, Any]) -> str:
    return (
        audit.get("dhl_awb")
        or audit.get("awb")
        or (audit.get("batch_meta") or {}).get("awb")
        or audit.get("tracking_no")
        or ""
    )


def _check_proactive_preconditions(audit: Dict[str, Any]) -> Optional[Dict[str, str]]:
    """
    Run G-PC1..G-PC8 against the live audit.

    Returns ``None`` on success, or a dict ``{"guard", "error", "code"}``
    that the caller raises as HTTP 422.
    """
    # G-PC7 — carrier must be DHL
    carrier = (audit.get("carrier") or "").upper()
    if carrier != "DHL":
        return {
            "guard": "carrier_not_dhl",
            "error": f"Carrier is {carrier!r}; proactive dispatch requires DHL.",
            "code":  "carrier_not_dhl",
        }

    # G-PC1 — must be carrier_self_clearance path
    cd = audit.get("clearance_decision") or {}
    clearance_path = (cd.get("clearance_path") or "").strip()
    if is_agency_clearance(clearance_path):
        return {
            "guard": "agency_path_active",
            "error": "Clearance path is agency clearance — proactive "
                     "dispatch not applicable.",
            "code":  "agency_path_active",
        }
    if clearance_path and not is_dhl_self_clearance(clearance_path):
        return {
            "guard": "not_self_clearance_path",
            "error": f"Clearance path is {clearance_path!r}; expected "
                     "carrier_self_clearance.",
            "code":  "not_self_clearance_path",
        }

    # G-PC2 — customs package must have been generated
    if not audit.get("customs_package_generated_at"):
        return {
            "guard": "customs_package_not_generated",
            "error": "Run /generate-customs-package before proactive dispatch.",
            "code":  "customs_package_not_generated",
        }

    # G-PC3 — must not have been dispatched
    if audit.get("proactive_dispatch_sent_at"):
        return {
            "guard": "already_dispatched",
            "error": "Proactive dispatch already sent for this batch.",
            "code":  "already_dispatched",
        }

    # G-PC5 — no agency
    if audit.get("agency_name") or (audit.get("agency_reply_package") or {}).get("status"):
        return {
            "guard": "agency_path_active",
            "error": "Agency forwarding active — proactive dispatch blocked.",
            "code":  "agency_path_active",
        }

    # G-PC6 — no DSK
    if audit.get("dsk_filename") or audit.get("dsk_reference"):
        return {
            "guard": "dsk_already_created",
            "error": "DSK has been generated — proactive dispatch blocked.",
            "code":  "dsk_already_created",
        }

    return None


@router.post("/proactive-dispatch/{batch_id}", dependencies=[_auth, _op_auth])
def request_proactive_dispatch(
    batch_id: str,
    body: ProactiveDispatchRequest,
) -> Dict[str, Any]:
    """
    Create an action proposal for proactive DHL customs dispatch.

    Step 1 of a two-step flow: this endpoint creates the proposal only.
    Approval and queueing go through the existing
    /api/v1/action-proposals/{proposal_id}/approve and /queue endpoints.

    No email is queued by this endpoint. No clearance_status is mutated.
    """
    from ..utils.proposal_lock import proposal_write_lock
    from ..api.routes_action_proposals import (
        create_proposal,
        _save_audit as _proposal_save_audit,
        _audit_path as _proposal_audit_path,
    )
    from ..services.dhl_proactive_dispatch_builder import (
        build_dhl_proactive_dispatch,
    )

    operator_id = (body.operator_id or "").strip()
    if not operator_id:
        raise HTTPException(
            status_code=422,
            detail={"guard": "missing_operator_id",
                    "error": "operator_id is required."},
        )

    # Phase 2.3.1 (Finding 1.2): reject operator_id values that match the
    # auto-actor sentinel space. Without this guard, an operator could mint
    # a self-approving proposal by spoofing the system actor name and bypass
    # the G9 self-approval block via the auto-actor exemption.
    from .routes_action_proposals import _is_auto_actor
    if _is_auto_actor(operator_id):
        raise HTTPException(
            status_code=422,
            detail={
                "code":    "auto_actor_sentinel_reserved",
                "guard":   "auto_actor_sentinel_reserved",
                "error":   f"operator_id is reserved for system actors and cannot be set by request.",
            },
        )

    # Acquire per-batch lock — concurrent POSTs serialise here.
    with proposal_write_lock(batch_id):
        audit = _load_audit(batch_id)
        if audit is None:
            raise HTTPException(
                status_code=404,
                detail={"guard": "batch_not_found",
                        "error": f"Batch {batch_id!r} not found."},
            )

        # G-PC8 — batch exists (just validated above)
        # Run remaining preconditions G-PC1..G-PC7
        gate_failure = _check_proactive_preconditions(audit)
        if gate_failure is not None:
            raise HTTPException(status_code=422, detail=gate_failure)

        # G-PC4 — idempotent dedup. create_proposal already returns
        # existing active proposal of the same type (lines 220-224 of
        # routes_action_proposals.py).
        proposal = create_proposal(
            audit         = audit,
            batch_id      = batch_id,
            proposal_type = "dhl_proactive_dispatch",
            reason        = "operator_initiated_proactive_dispatch",
            confidence    = "high",
        )

        # If we just created a NEW proposal (status pending_review and no
        # approved_by yet AND no created_by yet), stamp created_by + the
        # batch-level requested_at. If create_proposal returned an
        # existing active proposal, we must not stamp anything fresh.
        proposal_id = proposal["proposal_id"]
        is_new_proposal = (
            not proposal.get("created_by")
            and proposal.get("status") == "pending_review"
        )

        if is_new_proposal:
            # Re-verify attachments exist on disk via the builder's missing
            # list. Abort with 503 if anything is missing — do NOT persist
            # a partial proposal.
            draft = proposal.get("draft") or {}
            missing = draft.get("missing") or []
            if missing:
                # Roll back the in-memory append (not yet persisted)
                proposals = audit.get("action_proposals") or []
                audit["action_proposals"] = [
                    p for p in proposals
                    if p.get("proposal_id") != proposal_id
                ]
                raise HTTPException(
                    status_code=503,
                    detail={
                        "guard":   "attachment_missing",
                        "error":   "Required attachments missing on disk.",
                        "missing": missing,
                    },
                )

            # Stamp creator + batch-level timestamp
            proposal["created_by"] = operator_id
            audit["proactive_dispatch_requested_at"] = (
                datetime.now(timezone.utc).isoformat()
            )
            audit["proactive_dispatch_proposal_id"] = proposal_id

            _proposal_save_audit(batch_id, audit)

            tl.log_event(
                _proposal_audit_path(batch_id),
                tl.EV_DHL_PROACTIVE_DISPATCH_REQUESTED,
                "admin",
                actor=operator_id,
                detail={
                    "batch_id":         batch_id,
                    "proposal_id":      proposal_id,
                    "awb":              _resolve_proactive_awb(audit),
                    "operator_id":      operator_id,
                    "attachment_count": len(draft.get("attachments") or []),
                    "recipient":        draft.get("to") or "",
                },
            )
            log.info(
                "[proactive-dispatch] proposal created batch=%s proposal=%s by=%s",
                batch_id, proposal_id, operator_id,
            )
        else:
            log.info(
                "[proactive-dispatch] returning existing proposal batch=%s proposal=%s",
                batch_id, proposal_id,
            )

    return {
        "ok":           True,
        "batch_id":     batch_id,
        "proposal_id":  proposal_id,
        "status":       proposal.get("status", "pending_review"),
        "is_new":       bool(is_new_proposal),
    }


# ── W-5 P2: read-only self-clearance state for Mac dashboard state pill ──────
# Per Windows Atlas memory rule: no operator-facing write controls on Mac
# dashboard. This GET is the ONLY new surface from P2 on the Mac side.

@router.get("/selfclearance/state/{batch_id}", dependencies=[_auth])
def get_selfclearance_state(batch_id: str):
    """
    Return the current Path A self-clearance state for the batch.

    Response shape:
        {
          "batch_id": "...",
          "in_scope": true | false,
          "state": "awaiting_preemptive_send" | ... | "n/a",
          "p2_dispatch": {
            "shadow": <bool>,
            "message_id": <str|null>,
            "sent_at": <iso8601|null>,
          },
        }

    Returns `state: "n/a"` for non-Path-A shipments — the Mac dashboard
    renders the pill greyed-out in that case.
    """
    from ..services.dhl_clearance_coordinator import DhlClearanceCoordinator
    from ..services import dhl_clearance_manifest as _manifest

    # Load audit; if not present, return n/a (do not 500)
    audit_path = settings.storage_root / "outputs" / batch_id / "audit.json"
    if not audit_path.is_file():
        return {
            "batch_id":     batch_id,
            "in_scope":     False,
            "state":        "n/a",
            "p2_dispatch":  {"shadow": None, "message_id": None, "sent_at": None},
        }

    try:
        audit = json.loads(audit_path.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError):
        return {
            "batch_id":     batch_id,
            "in_scope":     False,
            "state":        "n/a",
            "p2_dispatch":  {"shadow": None, "message_id": None, "sent_at": None},
        }

    in_scope = DhlClearanceCoordinator.is_in_scope(audit)
    if not in_scope:
        return {
            "batch_id":     batch_id,
            "in_scope":     False,
            "state":        "n/a",
            "p2_dispatch":  {"shadow": None, "message_id": None, "sent_at": None},
        }

    block = audit.get(_manifest.MANIFEST_KEY) or {}
    p2 = block.get("p2_dispatch") or {}
    return {
        "batch_id":     batch_id,
        "in_scope":     True,
        "state":        block.get("state", "awaiting_preemptive_send"),
        "p2_dispatch":  {
            "shadow":     p2.get("shadow"),
            "message_id": p2.get("message_id"),
            "sent_at":    p2.get("sent_at"),
        },
    }
