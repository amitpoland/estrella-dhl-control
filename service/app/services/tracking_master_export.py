"""
tracking_master_export.py — Writes SHIPMENT_TRACKING_MASTER.xlsx from the DB.

One sheet: all events, 19 columns, newest first.
Regenerated in full on every export call.
"""
from __future__ import annotations

from datetime import datetime, timezone
from pathlib import Path
from typing import List, Dict, Any

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    _HAS_OPENPYXL = True
except ImportError:
    _HAS_OPENPYXL = False

from .tracking_db import get_all_events

_COLUMNS = [
    "id", "batch_id", "awb", "carrier", "stage", "status",
    "event_time", "captured_at", "source", "source_ref",
    "email_message_id", "raw_subject", "raw_sender",
    "location", "description", "normalized_stage",
    "confidence", "requires_manual_review", "created_at",
]

_HEADER_LABELS = {
    "id":                     "Event ID",
    "batch_id":               "Batch ID",
    "awb":                    "AWB",
    "carrier":                "Carrier",
    "stage":                  "Stage (Workflow)",
    "status":                 "Status",
    "event_time":             "Event Time (UTC)",
    "captured_at":            "Captured At (UTC)",
    "source":                 "Source",
    "source_ref":             "Source Ref",
    "email_message_id":       "Email Message ID",
    "raw_subject":            "Raw Subject",
    "raw_sender":             "Raw Sender",
    "location":               "Location",
    "description":            "Description",
    "normalized_stage":       "Normalized Stage (Movement)",
    "confidence":             "Confidence",
    "requires_manual_review": "Requires Manual Review",
    "created_at":             "Created At (UTC)",
}

_COL_WIDTHS = {
    "id":                     38,
    "batch_id":               18,
    "awb":                    16,
    "carrier":                10,
    "stage":                  32,
    "status":                 14,
    "event_time":             24,
    "captured_at":            24,
    "source":                 18,
    "source_ref":             22,
    "email_message_id":       36,
    "raw_subject":            40,
    "raw_sender":             28,
    "location":               20,
    "description":            40,
    "normalized_stage":       30,
    "confidence":             12,
    "requires_manual_review": 22,
    "created_at":             24,
}


def export_master_xlsx(output_path: Path, events: List[Dict[str, Any]] | None = None) -> Path:
    if not _HAS_OPENPYXL:
        raise RuntimeError("openpyxl is required for XLSX export")

    if events is None:
        events = get_all_events(limit=50000)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Tracking Events"

    header_font  = Font(bold=True, color="FFFFFF")
    header_fill  = PatternFill("solid", start_color="1F4E79")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    center       = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 32

    for col_idx, key in enumerate(_COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=_HEADER_LABELS.get(key, key))
        cell.font  = header_font
        cell.fill  = header_fill
        cell.alignment = header_align
        ws.column_dimensions[
            openpyxl.utils.get_column_letter(col_idx)
        ].width = _COL_WIDTHS.get(key, 18)

    alt_fill = PatternFill("solid", start_color="EBF3FB")
    for row_idx, ev in enumerate(events, start=2):
        fill = alt_fill if row_idx % 2 == 0 else None
        for col_idx, key in enumerate(_COLUMNS, start=1):
            val = ev.get(key, "")
            if key == "requires_manual_review":
                val = "Yes" if val else "No"
            elif key == "confidence" and isinstance(val, float):
                val = round(val, 3)
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.alignment = center
            if fill:
                cell.fill = fill

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(output_path))
    return output_path


def get_master_xlsx_path(storage_root: Path) -> Path:
    return storage_root / "SHIPMENT_TRACKING_MASTER.xlsx"
