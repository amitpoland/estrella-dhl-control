"""
bank_statement_import.py — adapter architecture for Treasury statement ingest.

Flow (never silent-commit):
  Upload → Parse → Preview → Validate → Reconcile → Confirm

Formats: CSV / XLSX initially. PDF only when a real bank format requires it.
Does not mutate wFirma. Writes only through treasury_db after Confirm.
"""
from __future__ import annotations

import csv
import io
import json
import uuid
from dataclasses import dataclass, asdict
from datetime import datetime, timezone
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any, Dict, List, Optional, Sequence, Tuple

from .treasury_db import (
    BalanceSnapshot,
    init_db as init_treasury_db,
    insert_balance_snapshot,
    _connect,
)


SUPPORTED_FORMATS = ("CSV", "XLSX")


@dataclass(frozen=True)
class ParsedBalanceRow:
    effective_date: str
    account_location: str
    currency: str
    closing_balance: Decimal
    reference_note: str = ""


@dataclass
class ImportPreview:
    batch_id: str
    filename: str
    format: str
    rows: List[ParsedBalanceRow]
    errors: List[str]
    warnings: List[str]

    @property
    def valid(self) -> bool:
        return not self.errors

    def to_dict(self) -> Dict[str, Any]:
        return {
            "batch_id": self.batch_id,
            "filename": self.filename,
            "format": self.format,
            "row_count": len(self.rows),
            "rows": [asdict(r) | {"closing_balance": format(r.closing_balance, "f")} for r in self.rows],
            "errors": list(self.errors),
            "warnings": list(self.warnings),
            "valid": self.valid,
        }


def _now() -> str:
    return datetime.now(timezone.utc).replace(microsecond=0).isoformat()


def _parse_decimal(raw: str) -> Decimal:
    s = (raw or "").strip().replace(" ", "").replace(",", ".")
    return Decimal(s)


def parse_csv_balances(
    content: bytes,
    *,
    filename: str = "upload.csv",
    default_account: str = "",
) -> ImportPreview:
    """Generic CSV: effective_date, account_location, currency, closing_balance[, note]."""
    batch_id = uuid.uuid4().hex
    errors: List[str] = []
    warnings: List[str] = []
    rows: List[ParsedBalanceRow] = []
    try:
        text = content.decode("utf-8-sig")
    except UnicodeDecodeError:
        text = content.decode("latin-1")
    reader = csv.DictReader(io.StringIO(text))
    if not reader.fieldnames:
        return ImportPreview(batch_id, filename, "CSV", [], ["CSV has no header row"], [])
    # normalize headers
    field_map = { (h or "").strip().lower(): h for h in reader.fieldnames }
    required = ("effective_date", "currency", "closing_balance")
    for req in required:
        if req not in field_map:
            errors.append(f"missing required column {req!r}")
    if errors:
        return ImportPreview(batch_id, filename, "CSV", [], errors, warnings)

    for i, raw in enumerate(reader, start=2):
        try:
            date = (raw.get(field_map["effective_date"]) or "").strip()
            ccy = (raw.get(field_map["currency"]) or "").strip().upper()
            bal = _parse_decimal(raw.get(field_map["closing_balance"]) or "")
            acct = ""
            if "account_location" in field_map:
                acct = (raw.get(field_map["account_location"]) or "").strip()
            if not acct:
                acct = default_account
            if not date or not ccy or not acct:
                errors.append(f"row {i}: effective_date, currency, account_location required")
                continue
            note = ""
            if "note" in field_map:
                note = (raw.get(field_map["note"]) or "").strip()
            elif "reference_note" in field_map:
                note = (raw.get(field_map["reference_note"]) or "").strip()
            rows.append(
                ParsedBalanceRow(
                    effective_date=date,
                    account_location=acct,
                    currency=ccy,
                    closing_balance=bal,
                    reference_note=note,
                )
            )
        except (InvalidOperation, KeyError, TypeError) as exc:
            errors.append(f"row {i}: {exc}")
    return ImportPreview(batch_id, filename, "CSV", rows, errors, warnings)


def parse_xlsx_balances(
    content: bytes,
    *,
    filename: str = "upload.xlsx",
    default_account: str = "",
) -> ImportPreview:
    """XLSX via openpyxl if available; otherwise clear unsupported error."""
    batch_id = uuid.uuid4().hex
    try:
        import openpyxl  # type: ignore
    except ImportError:
        return ImportPreview(
            batch_id,
            filename,
            "XLSX",
            [],
            ["openpyxl not installed — XLSX import unavailable on this host"],
            [],
        )
    try:
        wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        ws = wb.active
        headers = [str(c.value or "").strip().lower() for c in next(ws.iter_rows(min_row=1, max_row=1))]
        idx = {h: i for i, h in enumerate(headers) if h}
        for req in ("effective_date", "currency", "closing_balance"):
            if req not in idx:
                return ImportPreview(batch_id, filename, "XLSX", [], [f"missing column {req!r}"], [])
        rows: List[ParsedBalanceRow] = []
        errors: List[str] = []
        for rnum, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not row or all(v is None or str(v).strip() == "" for v in row):
                continue
            try:
                date = str(row[idx["effective_date"]] or "").strip()[:10]
                ccy = str(row[idx["currency"]] or "").strip().upper()
                bal = _parse_decimal(str(row[idx["closing_balance"]] or ""))
                acct = ""
                if "account_location" in idx:
                    acct = str(row[idx["account_location"]] or "").strip()
                if not acct:
                    acct = default_account
                if not date or not ccy or not acct:
                    errors.append(f"row {rnum}: effective_date, currency, account_location required")
                    continue
                note = ""
                if "note" in idx:
                    note = str(row[idx["note"]] or "").strip()
                rows.append(
                    ParsedBalanceRow(
                        effective_date=date,
                        account_location=acct,
                        currency=ccy,
                        closing_balance=bal,
                        reference_note=note,
                    )
                )
            except (InvalidOperation, IndexError, TypeError) as exc:
                errors.append(f"row {rnum}: {exc}")
        return ImportPreview(batch_id, filename, "XLSX", rows, errors, [])
    except Exception as exc:
        return ImportPreview(batch_id, filename, "XLSX", [], [f"XLSX parse failed: {exc}"], [])


def save_preview_batch(db_path: Path, preview: ImportPreview, uploaded_by: str = "") -> None:
    """Persist preview JSON in treasury_bank_import_batches (status=PREVIEW)."""
    init_treasury_db(db_path)
    with _connect(db_path) as conn:
        conn.execute(
            """
            INSERT OR REPLACE INTO treasury_bank_import_batches (
                id, filename, format, uploaded_at, uploaded_by, status,
                row_count, preview_json, validation_json, confirmed_at, detail
            ) VALUES (?,?,?,?,?,?,?,?,?,?,?)
            """,
            (
                preview.batch_id,
                preview.filename,
                preview.format,
                _now(),
                uploaded_by or None,
                "PREVIEW",
                len(preview.rows),
                json.dumps(preview.to_dict()),
                json.dumps({"errors": preview.errors, "warnings": preview.warnings}),
                None,
                None,
            ),
        )
        conn.commit()


def confirm_import_batch(
    db_path: Path,
    batch_id: str,
    *,
    operator: str = "",
) -> Dict[str, Any]:
    """Commit preview rows as BANK_IMPORT snapshots. Refuses if preview had errors.

    Atomic: all snapshot inserts + batch CONFIRMED update share one transaction.
    Concurrent confirm is rejected via conditional UPDATE (status must be PREVIEW).
    """
    init_treasury_db(db_path)
    with _connect(db_path) as conn:
        row = conn.execute(
            "SELECT * FROM treasury_bank_import_batches WHERE id = ?",
            (batch_id,),
        ).fetchone()
        if row is None:
            raise ValueError(f"import batch {batch_id!r} not found")
        if (row["status"] or "") == "CONFIRMED":
            raise ValueError(f"import batch {batch_id!r} already confirmed")
        if (row["status"] or "") != "PREVIEW":
            raise ValueError(
                f"import batch {batch_id!r} is not confirmable "
                f"(status={row['status']!r})"
            )
        preview = json.loads(row["preview_json"] or "{}")
        if preview.get("errors"):
            raise ValueError("refusing confirm: preview has validation errors")
        inserted = 0
        for r in preview.get("rows") or []:
            insert_balance_snapshot(
                db_path,
                BalanceSnapshot(
                    effective_date=r["effective_date"],
                    account_location=r["account_location"],
                    currency=r["currency"],
                    closing_balance=Decimal(str(r["closing_balance"])),
                    source="BANK_IMPORT",
                    operator=operator or None,
                    reference_note=r.get("reference_note") or None,
                    import_batch_id=batch_id,
                ),
                conn=conn,
            )
            inserted += 1
        cur = conn.execute(
            """
            UPDATE treasury_bank_import_batches
               SET status = 'CONFIRMED', confirmed_at = ?, detail = ?
             WHERE id = ? AND status = 'PREVIEW'
            """,
            (_now(), f"inserted={inserted}", batch_id),
        )
        if cur.rowcount != 1:
            raise ValueError(
                f"import batch {batch_id!r} confirm race — already confirmed "
                "or status changed"
            )
        conn.commit()
        return {"batch_id": batch_id, "status": "CONFIRMED", "inserted": inserted}
