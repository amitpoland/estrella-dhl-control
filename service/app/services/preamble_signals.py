"""
preamble_signals.py — Read-only identity-signal extractors for packing-list preambles
=====================================================================================

Pure functions that observe identity signals (VAT numbers, candidate heading rows)
from the top rows of a packing-list XLSX. Used in two distinct modes:

  PR 1 (visibility):     called observationally inside the draft-birth skip-event
                         emitter to populate `resolver_signals_seen` so the audit
                         records WHAT identity evidence the file contained at the
                         moment of failure. Does not influence draft creation.

  PR 2+ (authority):     called by the resolver chain as Pass 3.5 (VAT → wFirma
                         contractor lookup). Heading-candidate only trusted when
                         corroborated by VAT match (Lesson I authority chain).

Design rules (Lesson I):
  - Pure functions, zero side effects, no DB writes, no audit writes.
  - Read-only file open (openpyxl read_only=True).
  - Return None on any failure (missing file, unreadable, no signal found).
  - Never raise.

VAT regex coverage: EU two-letter prefix + 7..12 digits — covers SK, CZ, HU, PL,
DE, FR, IT, ES, NL, BE, AT, SE, DK, FI, PT, GR, IE, LU, MT, CY, SI, LV, LT, EE,
RO, BG, HR. Also catches non-EU patterns that share the [A-Z]{2}\\d{7,12} shape;
authoritative validation happens at wFirma lookup time (PR 2).
"""
from __future__ import annotations

import logging
import re
from pathlib import Path
from typing import Optional

log = logging.getLogger(__name__)

# ── VAT extraction ───────────────────────────────────────────────────────────
# Match: 2 uppercase letters + 7..12 digits. Allow inline separators ([-\s])
# inside the digit span because real files write "SK 1070 95376" or
# "SK-107095376". Strip separators before returning.
_VAT_RE = re.compile(r"\b([A-Z]{2})[\s-]?(\d{1,4}[\s-]?\d{1,5}[\s-]?\d{0,5})\b")
_DIGITS_RE = re.compile(r"\d")


def _normalize_vat(prefix: str, digits_raw: str) -> Optional[str]:
    """Return canonical VAT (PREFIX + digits, no separators) or None if length wrong."""
    digits = "".join(_DIGITS_RE.findall(digits_raw))
    if 7 <= len(digits) <= 12:
        return f"{prefix.upper()}{digits}"
    return None


def extract_vat_from_preamble(
    xlsx_path: str | Path,
    max_rows: int = 15,
) -> Optional[str]:
    """
    Scan the top `max_rows` rows of the active sheet for a VAT-shaped token.

    Returns the first hit as normalized "PREFIX+digits" (e.g. "SK107095376"),
    or None if no VAT is present, the file is missing, or any error occurs.

    Pure function — never raises, never writes.
    """
    if not xlsx_path:
        return None
    p = Path(xlsx_path)
    if not p.exists():
        return None
    try:
        import openpyxl as _opx  # type: ignore
    except ImportError:
        log.warning("preamble_signals: openpyxl not available; VAT extraction skipped")
        return None
    try:
        wb = _opx.load_workbook(str(p), read_only=True, data_only=True)
        try:
            ws = wb.active
            for row in ws.iter_rows(min_row=1, max_row=max_rows, values_only=True):
                for cell in row:
                    if cell is None:
                        continue
                    text = str(cell).strip()
                    if not text:
                        continue
                    for m in _VAT_RE.finditer(text.upper()):
                        prefix, digits_raw = m.group(1), m.group(2)
                        vat = _normalize_vat(prefix, digits_raw)
                        if vat:
                            return vat
        finally:
            wb.close()
    except Exception as exc:
        log.warning("preamble_signals.extract_vat_from_preamble failed (non-fatal): %s", exc)
    return None


# ── Heading candidate extraction ──────────────────────────────────────────────
# A heading candidate is a free-standing cell in the top rows that plausibly
# names the customer. PR 1 does NOT trust this in isolation; it only records it.
# PR 2/3 require VAT corroboration before binding.
#
# Filters (a heading candidate must NOT be):
#  - The sheet title "SHIPMENT PACKING LIST" or similar
#  - A labelled key:value line (contains ':' or '#')
#  - A starts-with-known-prefix line (Vat, Phone, Cont, Address, postal-code)
#  - Too short (<6) or too long (>80) to be a plausible name

_HEADING_TITLE_DENYLIST = {
    "shipment packing list",
    "packing list",
    "invoice",
    "shipment",
    "packing slip",
}

# Skip cells that begin with these tokens — they're metadata, not names.
_HEADING_PREFIX_DENYLIST = (
    "vat",
    "vat ",
    "vat-",
    "phone",
    "phone ",
    "phone-",
    "cont ",
    "cont-",
    "tel",
    "tel ",
    "tel-",
    "email",
    "e-mail",
    "fax",
    "address",
    "addr",
    "dated",
    "date",
    "invoice",
    "inv ",
    "inv#",
    "inv-",
    "p.o.",
    "po ",
    "po-",
)


def _looks_like_postal_or_phone(text: str) -> bool:
    """True if cell is mostly digits (postal code, phone, ID number)."""
    if not text:
        return True
    digit_ratio = sum(1 for c in text if c.isdigit()) / len(text)
    return digit_ratio >= 0.5


def _is_heading_candidate(text: str) -> bool:
    """Test whether a cell could plausibly be a customer-name heading."""
    if not text:
        return False
    t = text.strip()
    if len(t) < 6 or len(t) > 80:
        return False
    if ":" in t or "#" in t:
        return False
    lo = t.lower()
    if lo in _HEADING_TITLE_DENYLIST:
        return False
    for prefix in _HEADING_PREFIX_DENYLIST:
        if lo.startswith(prefix):
            return False
    if _looks_like_postal_or_phone(t):
        return False
    return True


def extract_heading_candidate(
    xlsx_path: str | Path,
    max_rows: int = 6,
) -> Optional[str]:
    """
    Return the first plausible standalone customer-name heading cell from the
    top `max_rows` rows, or None.

    Observation only. PR 1 records the candidate; PR 2/3 require VAT
    corroboration before treating it as authoritative.

    Pure function — never raises, never writes.
    """
    if not xlsx_path:
        return None
    p = Path(xlsx_path)
    if not p.exists():
        return None
    try:
        import openpyxl as _opx  # type: ignore
    except ImportError:
        log.warning("preamble_signals: openpyxl not available; heading extraction skipped")
        return None
    try:
        wb = _opx.load_workbook(str(p), read_only=True, data_only=True)
        try:
            ws = wb.active
            for row in ws.iter_rows(min_row=1, max_row=max_rows, values_only=True):
                for cell in row:
                    if cell is None:
                        continue
                    text = str(cell).strip()
                    if _is_heading_candidate(text):
                        return text
        finally:
            wb.close()
    except Exception as exc:
        log.warning("preamble_signals.extract_heading_candidate failed (non-fatal): %s", exc)
    return None


def extract_all_signals(xlsx_path: str | Path) -> dict:
    """
    Convenience wrapper returning both signals as a plain dict, ready to embed
    in an audit-event `detail.resolver_signals_seen` block.

    Always returns a dict with both keys (values may be None).
    """
    return {
        "vat": extract_vat_from_preamble(xlsx_path),
        "heading_candidate": extract_heading_candidate(xlsx_path),
    }
