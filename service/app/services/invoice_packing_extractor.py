"""
invoice_packing_extractor.py — Extract invoice lines and packing list rows.

Invoice lines come from pz_rows.json (engine output — not re-parsed).
Packing list rows are extracted from an uploaded PDF or XLSX file.

product_code convention:
  invoice_no + "-" + str(invoice_line_position)
  e.g. EJL/26-27/100-1, EJL/26-27/100-2

Matching strategy (in priority order):
  1. Direct: packing row has explicit invoice_no + line_position
  2. Fuzzy: same invoice_no + item_type + metal/karat + quantity match
  3. No match: product_code=None, requires_manual_review=True
"""
from __future__ import annotations

import hashlib
import logging
import re
import sys
from collections import defaultdict
from pathlib import Path
from typing import Any, Dict, FrozenSet, List, Optional, Tuple

from .excel_reader import read_excel_rows as _read_excel_rows

# Single item-type normalisation authority (lives next to ITEM_TYPE_PL/EN).
sys.path.insert(0, str(Path(__file__).resolve().parents[3]))
from description_grammar import (  # noqa: E402
    canonical_item_type_fuzzy as _canonical_grammar_item_type,
)

log = logging.getLogger(__name__)

_PARSER_NAME    = "invoice_packing_extractor"
_PARSER_VERSION = "1.0"


# ── Invoice lines: DB-first with pz_rows.json fallback ───────────────────────

def load_invoice_lines(
    batch_output_dir: Path,
    batch_id:         Optional[str] = None,
) -> List[Dict[str, Any]]:
    """
    Resolve invoice lines using the priority chain:
      1. document_db.invoice_lines    (canonical — source = "db_invoice_lines")
      2. pz_rows.json                 (legacy fallback — source = "legacy_pz_rows")
      3. []                           (no invoice context)

    The DB path is the modern source: invoice_lines are populated at intake time,
    long before PZ is generated. pz_rows.json is retained only for legacy batches
    that pre-date the intake pipeline.

    Returns list of dicts with:
      invoice_no, invoice_line_position, product_code,
      description_en, item_type, quantity, unit,
      unit_netto_pln, line_netto_pln, _source
    """
    # ── Priority 1: DB ────────────────────────────────────────────────────────
    if batch_id:
        try:
            from . import document_db as ddb
            db_rows = ddb.get_invoice_lines_for_batch(batch_id)
            if db_rows:
                result: List[Dict[str, Any]] = []
                for r in db_rows:
                    desc = r.get("description", "")
                    result.append({
                        "invoice_no":            r.get("invoice_no", ""),
                        "invoice_line_position": int(r.get("line_position") or 0),
                        "product_code":          r.get("product_code", ""),
                        "description":           desc,           # canonical
                        "description_en":        desc,           # legacy alias
                        # item_type isn't a separate column in invoice_lines;
                        # the matcher derives it from description if blank.
                        "item_type":             "",
                        "quantity":              float(r.get("quantity", 0) or 0),
                        "unit":                  "",
                        "rate_usd":              float(r.get("rate_usd",   r.get("unit_price",  0)) or 0),
                        "amount_usd":            float(r.get("amount_usd", r.get("total_value", 0)) or 0),
                        "unit_price":            float(r.get("rate_usd",   r.get("unit_price",  0)) or 0),
                        "total_value":           float(r.get("amount_usd", r.get("total_value", 0)) or 0),
                        "unit_netto_pln":        float(r.get("rate_usd",   r.get("unit_price",  0)) or 0),
                        "line_netto_pln":        float(r.get("amount_usd", r.get("total_value", 0)) or 0),
                        "gross_weight":          float(r.get("gross_weight", 0) or 0),
                        "net_weight":            float(r.get("net_weight",   0) or 0),
                        "hsn_code":              r.get("hsn_code", "") or r.get("hs_code", ""),
                        "_source":               "db_invoice_lines",
                    })
                log.info("[%s] invoice_lines loaded from DB (%d rows)",
                         batch_id, len(result))
                return result
        except Exception as exc:
            log.warning("[%s] DB invoice_lines lookup failed (falling back to pz_rows): %s",
                        batch_id, exc)

    # ── Priority 2: pz_rows.json (legacy) ─────────────────────────────────────
    pz_rows_path = batch_output_dir / "pz_rows.json"
    if not pz_rows_path.exists():
        log.warning("Neither DB invoice_lines nor pz_rows.json found for %s",
                    batch_id or batch_output_dir)
        return []

    import json
    raw: List[Dict[str, Any]] = json.loads(pz_rows_path.read_text(encoding="utf-8"))

    by_invoice: Dict[str, List[Dict[str, Any]]] = defaultdict(list)
    for row in raw:
        inv = row.get("invoice_no", "")
        by_invoice[inv].append(row)

    result: List[Dict[str, Any]] = []
    for inv_no, rows in sorted(by_invoice.items()):
        for pos, row in enumerate(rows, start=1):
            product_code = f"{inv_no}-{pos}"
            result.append({
                "invoice_no":            inv_no,
                "invoice_line_position": pos,
                "product_code":          product_code,
                "description_en":        row.get("description_en", ""),
                "item_type":             row.get("item_type", "").upper(),
                "quantity":              float(row.get("quantity", 0) or 0),
                "unit":                  row.get("unit", ""),
                "unit_netto_pln":        float(row.get("unit_netto_pln", 0) or 0),
                "line_netto_pln":        float(row.get("line_netto_pln", 0) or 0),
                "_source":               "legacy_pz_rows",
            })
    log.info("[%s] invoice_lines loaded from pz_rows.json (%d rows)",
             batch_id or "?", len(result))
    return result


# ── Packing list extraction ───────────────────────────────────────────────────

def file_sha256(path: Path) -> str:
    h = hashlib.sha256()
    with open(path, "rb") as f:
        for chunk in iter(lambda: f.read(65536), b""):
            h.update(chunk)
    return h.hexdigest()


def _new_diagnostic(file_type: str) -> Dict[str, Any]:
    """Build an empty parser_diagnostic skeleton."""
    return {
        "parser_name":          _PARSER_NAME,
        "parser_version":       _PARSER_VERSION,
        "file_type":            file_type,
        "workbook_sheet_names": [],
        "sheet_count":          0,
        "sheets_scanned":       [],
        "candidate_header_rows": [],
        "chosen_header":        None,
        "mapped_columns":       [],
        "unmatched_columns":    [],
        "alias_hits":           0,
        # Full per-column mapping audit: list of {col_index, original_header,
        # normalised, canonical_field, method, confidence, reason}.
        # Populated by _extract_packing_excel (primary) or _collect_excel_diagnostic
        # (fallback observability pass). Always present — never None.
        "column_mapping_audit": [],
        "row_count":            0,
        "failure_reason":       None,
        "exception_class":      None,
        "exception_message":    None,
        # C13B — client identity resolution (injected by routes_packing after extraction)
        # method: "filename" | "preamble" | "none"
        "client_name_resolution": None,
    }


def _collect_excel_diagnostic(path: Path, engine: str, diag: Dict[str, Any]) -> None:
    """Populate the diagnostic dict by re-reading the workbook for
    observability (sheet names, candidate header rows, chosen header,
    column mapping). Non-fatal — exceptions are swallowed and the
    diagnostic stays as far as it got.

    This NEVER changes parsing behaviour — it only inspects.
    """
    try:
        if engine == "openpyxl":
            import openpyxl as _opx
            wb = _opx.load_workbook(str(path), data_only=True, read_only=True)
            diag["workbook_sheet_names"] = list(wb.sheetnames)
            diag["sheet_count"] = len(wb.sheetnames)
            ws = wb.active
            diag["sheets_scanned"] = [ws.title] if ws is not None else []
            sheet_name = ws.title if ws is not None else "<active>"
        elif engine == "xlrd":
            import xlrd as _xlrd
            wb = _xlrd.open_workbook(str(path))
            diag["workbook_sheet_names"] = list(wb.sheet_names())
            diag["sheet_count"] = wb.nsheets
            diag["sheets_scanned"] = [wb.sheet_names()[0]] if wb.nsheets else []
            sheet_name = wb.sheet_names()[0] if wb.nsheets else "<sheet0>"
        else:
            return
    except Exception as exc:
        # Workbook unreadable in diagnostic pass — caller's failure_reason
        # already set; just leave the diag thin.
        log.debug("packing diagnostic workbook open failed: %s", exc)
        return

    # Re-read rows for header scoring (we don't trust the closed workbook
    # state; use the same path the parser uses).
    try:
        rows_raw = _read_excel_rows(path, engine)
    except Exception as exc:
        log.debug("packing diagnostic _read_excel_rows failed: %s", exc)
        return

    # Score each of the top 25 rows for alias hits.
    candidates: List[Dict[str, Any]] = []
    best_idx, best_hits = -1, 0
    best_raw: List[str] = []
    best_col_map: Dict[int, str] = {}
    for idx, row in enumerate(rows_raw[:25]):
        raw_cells = [str(c) if c is not None else "" for c in row]
        if not any(c.strip() for c in raw_cells):
            continue
        col_map = _map_headers(raw_cells)
        hits = len(col_map)
        if hits > 0:
            candidates.append({
                "sheet":            sheet_name,
                "row_index":        idx,
                "raw_cells_sample": raw_cells[:20],
                "alias_hits":       hits,
            })
        # _find_header_row uses a stricter rule (qty AND design); we record
        # the row that production parser would pick — re-derive via
        # _find_header_row over the same data.
        if hits > best_hits:
            best_hits, best_idx, best_raw, best_col_map = hits, idx, raw_cells, col_map

    diag["candidate_header_rows"] = candidates
    diag["alias_hits"] = best_hits

    hdr_idx = _find_header_row(rows_raw)
    if hdr_idx >= 0:
        hdr_cells = [str(c) if c is not None else "" for c in rows_raw[hdr_idx]]
        col_map = _map_headers(hdr_cells)
        diag["chosen_header"] = {
            "sheet":     sheet_name,
            "row_index": hdr_idx,
            "raw_cells": hdr_cells[:40],
        }
        diag["mapped_columns"] = [
            {"raw": hdr_cells[i] if i < len(hdr_cells) else "",
             "normalised": _normalise_header(hdr_cells[i]) if i < len(hdr_cells) else "",
             "canonical_field": fld}
            for i, fld in sorted(col_map.items())
        ]
        diag["unmatched_columns"] = [
            hdr_cells[i] for i in range(len(hdr_cells))
            if i not in col_map and hdr_cells[i].strip()
        ]
        # Populate column_mapping_audit if _extract_packing_excel didn't set it
        # (e.g. exception before line 783, early return, or exception-path call).
        if not diag.get("column_mapping_audit"):
            try:
                import dataclasses as _dc
                _, _audit = _map_headers_with_audit(hdr_cells)
                diag["column_mapping_audit"] = [_dc.asdict(m) for m in _audit]
            except Exception as _exc:
                log.debug("column_mapping_audit fallback failed: %s", _exc)
    elif best_idx >= 0:
        # Best-effort: surface the best-scoring candidate even when the
        # stricter _find_header_row rejected it. Helps the operator see
        # which row the parser ALMOST picked.
        diag["mapped_columns"] = [
            {"raw": best_raw[i] if i < len(best_raw) else "",
             "normalised": _normalise_header(best_raw[i]) if i < len(best_raw) else "",
             "canonical_field": fld}
            for i, fld in sorted(best_col_map.items())
        ]
        diag["unmatched_columns"] = [
            best_raw[i] for i in range(len(best_raw))
            if i not in best_col_map and best_raw[i].strip()
        ]
        # Populate column_mapping_audit from best-effort candidate row.
        if not diag.get("column_mapping_audit"):
            try:
                import dataclasses as _dc
                _, _audit = _map_headers_with_audit(best_raw)
                diag["column_mapping_audit"] = [_dc.asdict(m) for m in _audit]
            except Exception as _exc:
                log.debug("column_mapping_audit best-effort fallback failed: %s", _exc)


def _collect_pdf_diagnostic(path: Path, diag: Dict[str, Any]) -> None:
    """Best-effort observability for PDF packing lists. Captures page count
    and the first table-like row when present."""
    try:
        import pdfplumber as _pp  # type: ignore
        with _pp.open(str(path)) as pdf:
            page_count = len(pdf.pages)
            diag["sheet_count"] = page_count
            diag["workbook_sheet_names"] = [f"page_{i+1}" for i in range(page_count)]
            diag["sheets_scanned"] = diag["workbook_sheet_names"][:1]
            # Collect all table rows (mirrors _extract_packing_pdf) so the
            # diagnostic sees the same header the extractor sees.
            all_rows: List[List[Any]] = []
            for page in pdf.pages:
                for table in (page.extract_tables() or []):
                    if table:
                        for row in table:
                            if row is not None:
                                all_rows.append(row)
            if all_rows:
                first = [str(c) if c is not None else "" for c in all_rows[0]]
                diag["candidate_header_rows"] = [{
                    "sheet":            "page_1",
                    "row_index":        0,
                    "raw_cells_sample": first[:20],
                    "alias_hits":       len(_map_headers(first)),
                }]
                # Populate chosen_header when a real header exists so the
                # post-pass classifier can tell "header_not_detected" (no header)
                # from "empty_sheet" (header found, no usable data rows) — parity
                # with the Excel diagnostic.
                _hdr_idx = _find_header_row(all_rows)
                if _hdr_idx >= 0:
                    _hdr = [str(c) if c is not None else "" for c in all_rows[_hdr_idx]]
                    diag["chosen_header"] = {
                        "sheet":     "page_1",
                        "row_index": _hdr_idx,
                        "raw_cells": _hdr[:40],
                    }
    except Exception as exc:
        log.debug("packing diagnostic pdf open failed: %s", exc)


def extract_packing(
    path: Path,
    *,
    llm_fallback: bool = False,
    supplier_id: Optional[int] = None,
) -> Tuple[List[Dict[str, Any]], str, str, Dict[str, Any]]:
    """
    Dispatch to PDF or Excel extractor based on file extension.

    Returns (rows, parser_name, parser_version, parser_diagnostic).
    Each row is a raw dict of all extracted fields. The diagnostic dict
    carries observability data — see _new_diagnostic for the schema and
    docs/packing_diagnostics.md for the canonical contract. The
    diagnostic is ALWAYS returned (never None) so callers can record
    parser state regardless of success or failure.

    Supports:
      - .xlsx via openpyxl
      - .xls  via xlrd (legacy binary Excel)
      - .pdf  via pdfplumber

    Parser LOGIC is unchanged from prior versions — this wrapper only
    adds observability. The internal helpers (_extract_packing_excel,
    _extract_packing_pdf, _find_header_row, _map_headers) are untouched.

    Supplier routing: if the file is identified as a Global Jewellery
    document (via ``_detect_supplier_from_file``), the call is forwarded
    to ``global_packing_parser.parse_global_packing_excel`` which returns
    the same four-tuple.  The EJL path is completely unchanged.
    """
    # ── Supplier routing (additive gate — EJL path unchanged) ────────────
    try:
        _supplier = _detect_supplier_from_file(path)
    except Exception:
        _supplier = None

    if _supplier == "global_jewellery":
        # Route on file extension: PDF → text-positioned parser,
        # Excel → table parser. The Global supplier ships both formats
        # depending on the export they generate; both must work.
        suffix = path.suffix.lower()
        if suffix == ".pdf":
            from .global_packing_parser import parse_global_packing_pdf
            return parse_global_packing_pdf(path)
        from .global_packing_parser import parse_global_packing_excel
        return parse_global_packing_excel(path)

    # ── EJL / default path (original code below, unchanged) ──────────────
    suffix = path.suffix.lower()
    diag = _new_diagnostic(file_type=suffix)
    rows: List[Dict[str, Any]] = []

    try:
        if suffix == ".xlsx":
            rows = _extract_packing_excel(
                path, engine="openpyxl", _audit_dict=diag,
                llm_fallback=llm_fallback, supplier_id=supplier_id,
            )
            _collect_excel_diagnostic(path, "openpyxl", diag)
        elif suffix == ".xls":
            rows = _extract_packing_excel(
                path, engine="xlrd", _audit_dict=diag,
                llm_fallback=llm_fallback, supplier_id=supplier_id,
            )
            _collect_excel_diagnostic(path, "xlrd", diag)
        elif suffix == ".pdf":
            rows = _extract_packing_pdf(path)
            _collect_pdf_diagnostic(path, diag)
        else:
            diag["failure_reason"] = "unsupported_extension"
            diag["file_type"] = suffix or "<none>"
            raise ValueError(f"Unsupported packing list format: {suffix}")
    except ValueError:
        # Re-raise the explicit "unsupported_extension" case so existing
        # callers continue to receive ValueError.
        if diag["failure_reason"] == "unsupported_extension":
            raise
        diag["failure_reason"] = "parser_exception"
        diag["exception_class"] = "ValueError"
    except Exception as exc:
        diag["failure_reason"] = "parser_exception"
        diag["exception_class"] = type(exc).__name__
        diag["exception_message"] = str(exc)[:500]
        log.warning("extract_packing exception on %s: %s", path.name, exc)
        # Still attempt diagnostic collection (workbook may be partially
        # readable) for the file types where it makes sense.
        try:
            if suffix == ".xlsx":
                _collect_excel_diagnostic(path, "openpyxl", diag)
            elif suffix == ".xls":
                _collect_excel_diagnostic(path, "xlrd", diag)
            elif suffix == ".pdf":
                _collect_pdf_diagnostic(path, diag)
        except Exception:
            pass

    # ── Post-pass classification ─────────────────────────────────────────
    diag["row_count"] = len(rows)
    if diag["failure_reason"] is None:
        if not rows:
            # Distinguish header_not_detected from empty_sheet using the
            # diagnostic state.
            if diag["chosen_header"] is None and diag["candidate_header_rows"]:
                diag["failure_reason"] = "header_not_detected"
            elif diag["sheet_count"] == 0 and diag["file_type"] in (".xlsx", ".xls"):
                # File couldn't be opened as a workbook OR has zero sheets.
                diag["failure_reason"] = "file_corrupt"
            elif diag["chosen_header"] is None:
                diag["failure_reason"] = "header_not_detected"
            else:
                diag["failure_reason"] = "empty_sheet"

    return rows, _PARSER_NAME, _PARSER_VERSION, diag


def _normalise_header(h: str) -> str:
    return re.sub(r"[^a-z0-9]", "_", (h or "").strip().lower()).strip("_")


_FIELD_ALIASES: Dict[str, str] = {
    # design / product
    "design_no": "design_no",  "design": "design_no",  "style": "design_no",
    "model":     "design_no",  "designno":  "design_no",
    # batch
    "batch_no": "batch_no",   "batch": "batch_no",    "lot": "batch_no",
    "lot_no":   "batch_no",
    # bag / tray
    "bag_id":  "bag_id",   "bag":  "bag_id",   "bag_no":  "bag_id",
    "tray_id": "tray_id",  "tray": "tray_id",  "tray_no": "tray_id",
    # item type / category
    "item_type":  "item_type",  "type":     "item_type",  "category": "item_type",
    "description":"item_type",  "ctg":      "item_type",
    "pksr":       "line_position",   # EJL "PkSr" packing serial
    # uom
    "uom": "uom",  "unit": "uom",
    # quantities / weights
    "quantity":     "quantity",      "qty":      "quantity",   "pcs": "quantity",
    "pcs_qty":      "quantity",      "qty_pcs":  "quantity",
    "gross_weight": "gross_weight",  "gross_wt": "gross_weight","gw":  "gross_weight",
    "net_weight":   "net_weight",    "net_wt":   "net_weight",  "nw":  "net_weight",
    # Diamond weight — expanded supplier variants
    "dia_wt":          "diamond_weight",  "diamond_wt":     "diamond_weight",
    "diamond_weight":  "diamond_weight",  "dia_weight":     "diamond_weight",
    "diam_wt":         "diamond_weight",  "diamondwt":      "diamond_weight",
    "d_wt":            "diamond_weight",  "diamt":          "diamond_weight",
    # Color / colour weight — expanded supplier variants
    "col_wt":          "color_weight",    "color_weight":   "color_weight",
    "colour_weight":   "color_weight",    "color_wt":       "color_weight",
    "colour_wt":       "color_weight",    "col_weight":     "color_weight",
    "c_wt":            "color_weight",    "colwt":          "color_weight",
    "colourwt":        "color_weight",
    "qualtity":     "quality_string",  # observed typo in Pkg3806Rpt template
    # metal / karat / colour
    # EJL packing variants: "Kt/Color" combined OR "Kt"+"Col" separate cells.
    "metal":  "metal",  "material": "metal",  "kt_color": "metal",  "kt": "metal",
    "kt_col": "metal",   # observed: "Kt/Col" abbreviated header
    "col":    "metal_color",         # secondary (W/Y/RG indicator)
    "color":  "metal_color",
    "karat":  "karat",  "purity": "karat",  "fineness": "karat",
    # quality / stone — canonical field is quality_string
    "quality":    "quality_string",
    "stone_type": "stone_type", "stone": "stone_type",  "gemstone": "stone_type",
    # value (EJL packing list — used for fuzzy matching to invoice rate/amount)
    "value":       "unit_price",
    "price":       "unit_price",
    "unit_price":  "unit_price",
    "rate_usd":    "unit_price",
    "total_value": "total_value",
    "size":        "size",  "ring_size":   "size",  "sz":          "size",
    "size_mm":     "size",  "finger_size": "size",
    "client_po":   "client_po",      # sales-side PO reference
    "order_no":    "client_po",
    # invoice link (purchase + export references)
    "invoice_no":   "invoice_no",  "invoice": "invoice_no",  "inv_no": "invoice_no",
    "invoice_":     "invoice_no",
    "export_no":    "invoice_no",     # EJL sales packing uses "Export No"
    "invoice_line": "invoice_line_position",
    "line_position":"invoice_line_position",
    "line_no":      "invoice_line_position",
    "sr":           "line_position",  # generic Serial column
    # misc
    "remarks": "remarks",  "notes": "remarks",  "comment": "remarks",
    # Per-line currency (sales packing lists may carry one)
    "currency": "currency",  "ccy": "currency",
    # ── Global Jewellery column aliases (additive — do not rename existing) ──
    "style_no":   "design_no",   "styleno":   "design_no",
    "gross_wt":   "gross_weight","grosswt":   "gross_weight",
    "net_wt":     "net_weight",  "netwt":     "net_weight",
    "fob_value":  "unit_price",  "fobvalue":  "unit_price",
    "fob":        "unit_price",
    "srno":       "line_position",
}


# Currency markers stripped before numeric coercion so a legitimate
# currency-formatted value (e.g. "$ 993", "$ 1,554", "€ 538", "USD 993")
# normalises to its number instead of collapsing to 0.0. EJL packing lists
# render the Value/Total Value columns as "$ N" / "$ N,NNN". Bare symbols plus
# word-boundaried ISO codes; the Polish "zł" suffix is included explicitly.
_CURRENCY_MARKER_RE = re.compile(
    r"[$€£¥₹]|zł|\b(?:USD|EUR|PLN|GBP|CHF|JPY|INR)\b",
    re.IGNORECASE,
)


# Grouped-thousands form: 1,554 / 1,234,567 — every comma group exactly 3 digits.
# Anything else with a comma is read as a decimal comma (1234,56 -> 1234.56).
_THOUSANDS_COMMA_RE = re.compile(r"^[+-]?\d{1,3}(?:,\d{3})+$")


def _safe_float(val: Any) -> float:
    """Convert any value to float without raising.

    Canonical numeric-normalisation authority for the packing pipeline (upload
    AND reprocess persistence). Tolerates the string forms real packing lists
    carry: currency-prefixed values ("$ 993", "$ 1,554"), thousands separators
    (comma or space), decimal commas ("1234,56", "1.234,56"), and stray
    whitespace. Genuinely non-numeric strings ("ite 1", "Total") and any other
    error return 0.0 — never raises.

    Comma disambiguation (a bare comma is ambiguous across locales):
      - both "," and "." present -> the RIGHTMOST one is the decimal separator
      - comma only, and the whole value is grouped thousands ("1,554") -> separator
      - comma only, anything else ("1234,56", "1,5") -> decimal comma
    "1,554" therefore stays 1554.0, preserving the EJL packing-list contract.
    """
    if val is None:
        return 0.0
    if isinstance(val, (int, float)):
        return float(val)
    # Strip currency symbols / ISO codes, then any whitespace used as a thousands
    # separator. Bare split() covers Unicode spaces (NBSP, narrow NBSP) that
    # Excel exports emit, so no literal invisible characters live in this file.
    s = _CURRENCY_MARKER_RE.sub("", str(val))
    s = "".join(s.split())
    if "," in s:
        if "." in s:
            s = (s.replace(".", "").replace(",", ".")
                 if s.rfind(",") > s.rfind(".")
                 else s.replace(",", ""))
        elif _THOUSANDS_COMMA_RE.match(s):
            s = s.replace(",", "")
        else:
            s = s.replace(",", ".")
    try:
        return float(s)
    except (ValueError, TypeError):
        return 0.0


def _detect_supplier_from_file(path: Path) -> Optional[str]:
    """Read a short preview of the file and detect the supplier.

    For Excel files: scans the first 20 rows of the active sheet.
    For PDF files:   reads the first 500 characters of text.

    Returns the canonical supplier code (e.g. ``"global_jewellery"``) or
    ``None`` for unknown / EJL suppliers.
    """
    from .supplier_detect import detect_supplier

    try:
        suffix = path.suffix.lower()
        if suffix in (".xlsx", ".xls"):
            import openpyxl
            wb = openpyxl.load_workbook(str(path), data_only=True, read_only=True)
            ws = wb.active
            fragments = []
            for i, row in enumerate(ws.iter_rows(values_only=True)):
                if i >= 20:
                    break
                for cell in row:
                    if cell is not None:
                        fragments.append(str(cell))
            wb.close()
            return detect_supplier(" ".join(fragments))
        elif suffix == ".pdf":
            try:
                import pdfplumber
                with pdfplumber.open(str(path)) as pdf:
                    first_page_text = (pdf.pages[0].extract_text() or "") if pdf.pages else ""
                return detect_supplier(first_page_text[:500])
            except Exception:
                return None
    except Exception as exc:
        log.debug("_detect_supplier_from_file: %s — %s", path.name, exc)
    return None


_CURRENCY_TOKEN_RE = re.compile(
    r"\b(EUR|USD|PLN|GBP|CHF|JPY)\b", re.IGNORECASE,
)

# Symbol → ISO. Match longest-first so multi-char tokens win over '$'.
# 'zł' is the only common Polish symbol; CHF/JPY rarely appear as symbols.
_CURRENCY_SYMBOL_TO_ISO: List[tuple] = [
    ("zł",  "PLN"),
    ("CHF", "CHF"),
    ("EUR", "EUR"),
    ("USD", "USD"),
    ("PLN", "PLN"),
    ("GBP", "GBP"),
    ("JPY", "JPY"),
    ("€",   "EUR"),
    ("£",   "GBP"),
    ("¥",   "JPY"),
    ("$",   "USD"),
]


def _currency_from_format_string(fmt: str) -> str:
    """Parse an Excel ``cell.number_format`` string and return an ISO
    currency code, or "" if no recognised marker is present.

    Examples
    --------
    ``'[$-10409]"€"\\ 0;\\("€"\\ 0\\)'``    → ``"EUR"``
    ``'[$-10409]"$"\\ 0;\\("$"\\ 0\\)'``    → ``"USD"``
    ``'#,##0.00 "zł"'``                       → ``"PLN"``
    ``'General'``                              → ``""``
    """
    if not fmt:
        return ""
    s = str(fmt)
    # Excel locale prefixes like "[$-10409]" must NOT be confused with the
    # literal "$" symbol — strip them out before scanning.
    s_clean = re.sub(r"\[\$-[0-9A-Fa-f]+\]", "", s)
    for sym, iso in _CURRENCY_SYMBOL_TO_ISO:
        if sym in s_clean:
            return iso
    return ""


def _detect_packing_currency(rows: list, header_idx: int, headers: list) -> str:
    """
    Sales packing lists rarely carry a per-row currency cell. Detect from:
      1. Header text on Value/Total Value columns: "Value (EUR)", "Total USD"
      2. Sheet preamble cells (top 12 rows): "Currency: EUR", "USD"
    Returns the upper-case ISO code, or "" if nothing matched.
    """
    # 1. Header-level
    for h in headers or []:
        s = str(h or "")
        m = _CURRENCY_TOKEN_RE.search(s)
        if m:
            return m.group(1).upper()
    # 2. Preamble scan (rows above the header row)
    for r in (rows or [])[:max(header_idx, 0)]:
        for cell in r or []:
            s = str(cell or "")
            if not s:
                continue
            low = s.lower()
            if "currency" in low or "ccy" in low:
                m = _CURRENCY_TOKEN_RE.search(s)
                if m:
                    return m.group(1).upper()
            # Bare currency token in preamble (e.g. "All values in EUR")
            m = _CURRENCY_TOKEN_RE.search(s)
            if m and ("value" in low or "total" in low or "all " in low
                      or low.strip() in ("eur", "usd", "pln", "gbp",
                                           "chf", "jpy")):
                return m.group(1).upper()
    return ""


def _map_headers(raw_headers: List[str]) -> Dict[int, str]:
    """Return {col_index: canonical_field_name} for known headers.

    Tolerant of currency-annotated headers like ``Value (EUR)`` /
    ``Total USD``: a parenthetical or trailing currency token is stripped
    before alias lookup so ``value_eur`` resolves to the same canonical
    field as ``value``.
    """
    mapping: Dict[int, str] = {}
    for i, h in enumerate(raw_headers):
        raw = (h or "").strip()
        # Strip "(EUR)" / "(USD)" annotations from header text.
        cleaned = re.sub(r"\s*\(\s*(?:EUR|USD|PLN|GBP|CHF|JPY)\s*\)\s*",
                          "", raw, flags=re.IGNORECASE)
        # Strip trailing bare currency tokens: "Total USD" → "Total"
        cleaned = re.sub(r"\b(?:EUR|USD|PLN|GBP|CHF|JPY)\b\s*$",
                          "", cleaned, flags=re.IGNORECASE).strip()
        key = _normalise_header(cleaned)
        if key in _FIELD_ALIASES:
            mapping[i] = _FIELD_ALIASES[key]
    return mapping


def _map_headers_with_audit(
    raw_headers: List[str],
    *,
    llm_fallback: bool = False,
    supplier_id: Optional[int] = None,
) -> Tuple[Dict[int, str], List]:
    """Four-tier header mapping with per-column audit trail.

    Returns (col_map, column_mapping_audit) where col_map includes Tier-0
    (supplier_template), Tier-1 (alias), and accepted Tier-2 (fuzzy >= 0.90),
    and column_mapping_audit is a List[ColumnMapping] for the parser diagnostic.

    _map_headers() is left unchanged for backward-compatibility; this
    function delegates to excel_column_mapper.map_all_headers.
    """
    from .excel_column_mapper import build_col_map, map_all_headers
    mappings = map_all_headers(
        raw_headers, _FIELD_ALIASES,
        llm_fallback=llm_fallback,
        supplier_id=supplier_id,
    )
    col_map  = build_col_map(mappings)
    return col_map, mappings


def _derive_karat_from_metal(metal: Any) -> str:
    """Return the karat/purity token from a (possibly combined) metal string.

    The purchase packing templates alias the "Kt" / "Kt/Color" column into
    ``metal`` (e.g. ``"18KT/Y"``), so the standalone ``karat`` field that the
    variant signature reads stays empty unless the source literally has a
    "Karat" column. This lifts the purity token (the segment before any
    ``"/color"`` suffix) so ``karat`` is populated WITHOUT altering ``metal``.
    Only returns a token that carries a digit (a real karat/purity such as
    18KT, 14KT, 585, 950, PT950) so a colour-only value is never mistaken for a
    karat. Returns ``""`` when nothing usable is present.
    """
    m = str(metal or "").strip()
    if not m:
        return ""
    tok = m.split("/", 1)[0].strip()
    return tok if any(ch.isdigit() for ch in tok) else ""


def _row_to_dict(cells: List[Any], col_map: Dict[int, str]) -> Dict[str, Any]:
    row: Dict[str, Any] = {}
    for idx, field in col_map.items():
        if idx < len(cells):
            val = cells[idx]
            if val is None:
                val = ""
            row[field] = val
    return row




def _find_header_row(rows: List[List[Any]]) -> int:
    """
    Locate the header row by scanning for a row that contains BOTH a quantity
    cell (qty/quantity/pcs/pcs_qty) AND a design/category cell. Tolerant of
    minor template variations: header text may contain extra words.
    Returns -1 if no header row is found.

    Known limitation (shared by the Excel and PDF paths): only the top 25 rows
    are scanned, so an unusually long document preamble can push the real
    header out of range and yield -1. The PDF caller logs the collected row
    count on this path so such a failure is diagnosable rather than silent.
    """
    def _is_qty_header(h: str) -> bool:
        return h in ("qty", "quantity", "pcs") or "qty" in h or "pcs_qty" == h
    def _is_design_header(h: str) -> bool:
        return h in (
            "designno", "design", "design_no", "style", "ctg", "category",
            "style_no", "styleno", "sku", "item",
        ) or "design" in h or "style" in h

    for idx, row in enumerate(rows[:25]):  # only scan top 25 rows
        norm = [_normalise_header(str(c) if c is not None else "") for c in row]
        if any(_is_qty_header(h) for h in norm) and any(_is_design_header(h) for h in norm):
            return idx
    return -1


def _is_numeric_cell(v: Any) -> bool:
    """Return True if v parses as a finite number."""
    if v is None or v == "":
        return False
    if isinstance(v, (int, float)):
        return True
    s = str(v).strip().replace(",", ".")
    try:
        float(s)
        return True
    except ValueError:
        return False


def _is_subtotal_row(cells: List[Any], col_map: Dict[int, str]) -> bool:
    """Detect Total / Grand Total / sub-total rows so they don't pollute output."""
    for i, v in enumerate(cells):
        s = str(v or "").strip().lower()
        if not s:
            continue
        if "total" in s and i in col_map and col_map[i] == "design_no":
            return True
        if s.startswith(("grand total", "subtotal", "sub total", "total ")):
            return True
    return False


# ── Shared invoice-preamble scan (Excel + PDF paths derive invoice identity
#    identically, so the same logical document keys the same way regardless of
#    upload format). Two forms observed in EJL templates:
#      A. label + value in adjacent cells:   ["Invoice #", "EJL/26-27/013"]
#      B. label + value concatenated:        ["Export No : EJL/26-27/015"]
_PREAMBLE_LABELS_A = ("invoice", "invoice #", "invoice no", "export no", "export no.")
_PREAMBLE_LABEL_PATTERN_B = re.compile(
    r"^(?:invoice\s*(?:no|#)?|export\s*no\.?)\s*[:#]?\s*(.+)$",
    re.IGNORECASE,
)
_PREAMBLE_IS_INVOICE_LIKE = re.compile(r"^\s*(EJL|PROF|INV)[\s/\-]", re.IGNORECASE)


def _invoice_no_from_preamble(rows: List[List[Any]]) -> str:
    """Scan the first 12 preamble rows for an invoice / export reference.

    Returns the reference string, or "" if none is found. Shared by the Excel
    and PDF extraction paths. Behaviour mirrors the original inline Excel scan
    exactly: a Form-A label with no adjacent value does not stop the scan — the
    outer loop continues to the next row.
    """
    for r in (rows or [])[:12]:
        found = ""
        for i, cell in enumerate(r):
            raw = str(cell or "").strip()
            if not raw:
                continue
            # Form A — label only
            tag = raw.lower().rstrip("#:.").strip()
            if tag in _PREAMBLE_LABELS_A:
                for v in r[i + 1:]:
                    sv = str(v or "").strip()
                    if sv:
                        found = sv
                        break
                break
            # Form B — label and value in same cell
            m = _PREAMBLE_LABEL_PATTERN_B.match(raw)
            if m:
                cand = m.group(1).strip()
                if _PREAMBLE_IS_INVOICE_LIKE.match(cand):
                    found = cand
                    break
        if found:
            return found
    return ""


def _validate_and_normalise_row(
    d: Dict[str, Any], invoice_no_hint: str = "",
) -> Optional[Dict[str, Any]]:
    """Canonical packing-row validity + normalisation contract shared by the
    Excel and PDF extraction paths.

    Returns the normalised row dict, or ``None`` when the row is not a real
    product line and MUST be dropped (never emitted, never counted). This is
    the single guard that stops title/subtotal/footer/address rows and empty
    mappings from being treated as extracted business rows.

    Currency stamping is intentionally NOT here — it is format-specific (Excel
    ``number_format``) and remains in the Excel path.
    """
    # Real data rows ALWAYS have a numeric quantity — the strongest signal.
    # Drops subtotals, "frt"/"insu" footer rows, address rows, and — because an
    # empty mapping has no quantity — empty/degenerate dicts.
    if not _is_numeric_cell(d.get("quantity")):
        return None
    # And must carry a design / category identity cell.
    design_present = bool(str(d.get("design_no", "") or "").strip())
    item_type      = bool(str(d.get("item_type", "") or "").strip())
    if not design_present and not item_type:
        return None

    # ── Metal / colour split (Variants A/B/C) ────────────────────────────────
    # Goal: `metal` holds the full combined string (e.g. "14KT/W") AND
    # `metal_color` holds the standalone colour code (e.g. "W").
    if d.get("metal_color") and d.get("metal"):
        # Variant A / C: separate cells — merge karat + colour into metal.
        d["metal"] = f"{d['metal']}/{d['metal_color']}"
    elif not d.get("metal_color") and d.get("metal"):
        # Variant B: combined cell like "14KT/Y" — extract the colour suffix.
        combined = str(d["metal"]).strip()
        if "/" in combined:
            _, _, color_part = combined.partition("/")
            color_part = color_part.strip().rstrip("-").strip()
            if color_part and len(color_part) <= 4:
                d["metal_color"] = color_part

    # Derive standalone `karat` from the (possibly combined) metal.
    if not str(d.get("karat") or "").strip():
        _kt = _derive_karat_from_metal(d.get("metal"))
        if _kt:
            d["karat"] = _kt

    # Stamp invoice_no from the preamble hint when the row itself has none.
    if invoice_no_hint and not d.get("invoice_no"):
        d["invoice_no"] = invoice_no_hint

    return d


def _extract_packing_excel(
    path: Path,
    engine: str = "openpyxl",
    _audit_dict: Optional[Dict] = None,
    *,
    llm_fallback: bool = False,
    supplier_id: Optional[int] = None,
) -> List[Dict[str, Any]]:
    """
    Read an EJL packing list (XLSX via openpyxl, XLS via xlrd).

    The EJL template puts:
      - "Invoice #" / "EJL/26-27/013" in a stray cell pair near rows 3-5
      - The actual line table headers on row 9 (PkSr, Ctg, DesignNo, Kt/Color,
        Quality, Dia Wt, Col Wt, Qty, Value, Total Value, Size)
      - Real data rows starting at row 10
      - Sub-total rows ("Total PND-18KT", "Grand Total") interleaved

    This function locates the header row dynamically and drops sub-totals.
    """
    rows = _read_excel_rows(path, engine)
    if not rows:
        return []

    # ── Pull invoice_no from the sheet preamble (shared with the PDF path) ────
    invoice_no_from_sheet = _invoice_no_from_preamble(rows)

    # ── Locate the header row ─────────────────────────────────────────────
    hdr_idx = _find_header_row(rows)
    if hdr_idx < 0:
        log.warning("No recognisable column headers found in %s", path.name)
        return []

    headers = [str(c) if c is not None else "" for c in rows[hdr_idx]]
    if _audit_dict is not None:
        import dataclasses as _dc
        col_map, _mapping_audit = _map_headers_with_audit(
            headers, llm_fallback=llm_fallback, supplier_id=supplier_id,
        )
        _audit_dict["column_mapping_audit"] = [_dc.asdict(m) for m in _mapping_audit]
    else:
        col_map = _map_headers(headers)
    if not col_map:
        log.warning("Header row %d had no aliasable cells: %s", hdr_idx, headers)
        return []

    # Sheet-level currency (header text or preamble). Stamped onto every
    # row that doesn't already carry a per-row currency cell.
    sheet_currency = _detect_packing_currency(rows, hdr_idx, headers)

    # ── Cell-format currency (highest-priority source) ──────────────────
    # Excel encodes currency on the Value / Total Value cells via
    # ``number_format`` (e.g. ``[$-10409]"€" 0``). This is the AUTHORITATIVE
    # signal — it survives copy/paste and renders the symbol in the source
    # file the operator sees. Read once per workbook (engine=openpyxl only;
    # xlrd does not expose per-cell number_format reliably).
    cell_format_currencies: List[str] = []
    if engine == "openpyxl":
        try:
            import openpyxl as _opx  # type: ignore
            wb_fmt = _opx.load_workbook(str(path), data_only=False)
            ws_fmt = wb_fmt.active
            # Value-bearing columns: any column mapped to unit_price /
            # total_value. We scan ALL data rows (not just header) so a
            # mid-sheet currency switch is detectable.
            value_cols = [i + 1 for i, fld in col_map.items()
                          if fld in ("unit_price", "total_value")]
            for ri in range(hdr_idx + 2, ws_fmt.max_row + 1):
                for ci in value_cols:
                    cell = ws_fmt.cell(row=ri, column=ci)
                    iso = _currency_from_format_string(cell.number_format or "")
                    if iso:
                        cell_format_currencies.append(iso)
        except Exception as _exc:
            log.warning("number_format currency scan failed for %s: %s",
                         path.name, _exc)
    # Sheet-level cell-format currency: use the most common one. Mixed
    # → keep the most common but record the conflict for the caller via a
    # special marker on the parsed rows so intake can warn.
    sheet_cell_format_currency = ""
    sheet_cell_format_conflict = False
    if cell_format_currencies:
        from collections import Counter as _Counter
        counts = _Counter(cell_format_currencies)
        sheet_cell_format_currency = counts.most_common(1)[0][0]
        sheet_cell_format_conflict = len(counts) > 1

    result: List[Dict[str, Any]] = []
    for raw in rows[hdr_idx + 1:]:
        cells = list(raw)
        # Skip fully empty rows
        if all(c is None or str(c).strip() == "" for c in cells):
            continue
        if _is_subtotal_row(cells, col_map):
            continue
        d = _row_to_dict(cells, col_map)
        # Shared canonical validity + normalisation contract (numeric-qty +
        # design/category gates, metal/colour split, karat derivation, invoice
        # stamp). None => not a real product line; drop it.
        d = _validate_and_normalise_row(d, invoice_no_from_sheet)
        if d is None:
            continue

        # Currency priority on the row:
        #   1. cell number_format on Value/Total columns (authoritative)
        #   2. per-row currency cell (if present)
        #   3. sheet header / preamble token
        if not str(d.get("currency", "") or "").strip():
            if sheet_cell_format_currency:
                d["currency"]        = sheet_cell_format_currency
                d["currency_source"] = "excel_symbol"
            elif sheet_currency:
                d["currency"]        = sheet_currency
                d["currency_source"] = "excel_token"
        else:
            d.setdefault("currency_source", "excel_row")

        # Surface the multi-currency conflict so intake can warn.
        if sheet_cell_format_conflict:
            d["currency_conflict"] = True

        result.append(d)
    log.info("Packing extracted %d rows from %s (engine=%s, header_row=%d)",
             len(result), path.name, engine, hdr_idx)
    return result


# Backward-compatible alias (retained so any external caller still works)
def _extract_packing_xlsx(path: Path) -> List[Dict[str, Any]]:
    return _extract_packing_excel(path, engine="openpyxl")


def _extract_packing_pdf(path: Path) -> List[Dict[str, Any]]:
    """Read an EJL packing list PDF into the SAME canonical packing-row contract
    as the Excel path.

    EJL PDFs open with title / address / preamble rows (e.g.
    "SHIPMENT PACKING LIST") before the real column header, so the first table
    row is NOT the header. All table rows across every page are collected in
    document order, the real header is located with the shared
    ``_find_header_row`` scanner (skipping title/preamble rows), and every data
    row passes the shared ``_validate_and_normalise_row`` contract — so
    subtotal, footer, title and empty rows can never be emitted as products, and
    an empty mapping is never counted as an extracted row. The same canonical
    fields as the Excel path are emitted (notably ``line_position`` → ``pack_sr``
    and ``invoice_no``) so cross-format persistence keys align.
    """
    try:
        import pdfplumber
    except ImportError:
        raise RuntimeError("pdfplumber is required for PDF packing list extraction")

    # Collect every table row across all pages, in document order.
    all_rows: List[List[Any]] = []
    with pdfplumber.open(str(path)) as pdf:
        for page in pdf.pages:
            for table in (page.extract_tables() or []):
                if not table:
                    continue
                for row in table:
                    if row is not None:
                        all_rows.append(row)

    if not all_rows:
        log.warning("No table rows found in PDF %s", path.name)
        return []

    # Locate the real header (skips title / preamble rows) via the shared scanner.
    hdr_idx = _find_header_row(all_rows)
    if hdr_idx < 0:
        # _find_header_row scans only the top 25 rows; surface the collected row
        # count so a header pushed past that window by a long preamble is a
        # diagnosable failure, not a silent zero-row extraction.
        log.warning(
            "No recognisable column headers in the top %d of %d collected table "
            "rows for PDF %s (header may be beyond the 25-row scan window).",
            min(25, len(all_rows)), len(all_rows), path.name,
        )
        return []

    headers = [str(c) if c is not None else "" for c in all_rows[hdr_idx]]
    col_map = _map_headers(headers)
    if not col_map:
        log.warning("Header row %d had no aliasable cells in PDF %s: %s",
                    hdr_idx, path.name, headers)
        return []

    # Invoice reference from the preamble rows above the header (shared logic).
    invoice_no_from_sheet = _invoice_no_from_preamble(all_rows[:hdr_idx])

    result: List[Dict[str, Any]] = []
    for raw in all_rows[hdr_idx + 1:]:
        cells = [str(c).strip() if c is not None else "" for c in raw]
        # Skip fully empty rows and interleaved subtotal rows.
        if all(c == "" for c in cells):
            continue
        if _is_subtotal_row(cells, col_map):
            continue
        d = _row_to_dict(cells, col_map)
        d = _validate_and_normalise_row(d, invoice_no_from_sheet)
        if d is None:
            continue
        result.append(d)

    log.info("Packing PDF extracted %d rows from %s (header_row=%d)",
             len(result), path.name, hdr_idx)
    return result


# ── Matching ──────────────────────────────────────────────────────────────────

def _normalise_item_type(s: str) -> str:
    return re.sub(r"[^a-z]", "", s.lower())


def _canonical_item_type(s: str) -> str:
    """
    Normalise either an EJL packing 'Ctg' code (PND/RNG/ERG…) or an invoice
    description (containing PENDANT/RING/EARRING…) to a single canonical key.

    The alias map lives in description_grammar (the single item-type
    normalisation authority, next to ITEM_TYPE_PL / ITEM_TYPE_EN).  This wrapper
    only restores the lowercase contract this module's matching code relies on,
    and the "unrecognised → squashed input" fallback used as a grouping key.
    """
    return (_canonical_grammar_item_type(s) or _normalise_item_type(s)).lower()


# ── Metal / material code normalisation ──────────────────────────────────────
# Recognises the metal token written in either the invoice description
# ("18KT Gold", "PT950 Platinum") or the packing list "Kt/Color" column
# ("18KT/W", "PT950/-", "925/-", "14KT/Y").
#
# Returns canonical token: "18KT", "14KT", "22KT", "24KT", "PT950", "PT900",
# "PT850", "925", "999", or "" if unrecognised.
_METAL_PATTERNS = [
    (re.compile(r"\b(\d{3})\s*PT\b",      re.IGNORECASE),              lambda m: "PT" + m.group(1)),  # "950 PT"
    (re.compile(r"\bPT\s*(\d{3})\b",      re.IGNORECASE),              lambda m: "PT" + m.group(1)),  # "PT950"
    (re.compile(r"\b(\d{2})\s*KT?\b",     re.IGNORECASE),              lambda m: m.group(1) + "KT"),  # "18KT", "14K"
    # Silver: 925 / 999 with optional letter prefix (S, SL, SS, SLV, SILVER).
    # Negative lookbehind for digits prevents false hits like "1925" (year).
    # Negative lookahead prevents false hits like "9255".
    (re.compile(r"(?<![0-9])(925|999)(?![0-9])"),                       lambda m: m.group(1)),
]


def _canonical_metal(*texts: str) -> str:
    """First metal token found across one or more text fragments."""
    for raw in texts:
        if not raw:
            continue
        s = str(raw)
        for pat, fmt in _METAL_PATTERNS:
            m = pat.search(s)
            if m:
                return fmt(m).upper()
    return ""


def _metal_tokens(*texts: str) -> FrozenSet[str]:
    """EVERY metal token present across the given fragments, not just the first.

    A single canonical token cannot describe a two-tone piece.  An invoice
    description like "18KT Gold, LGD Stud PT950 Com Jewell RING" names two real
    metals; _canonical_metal() returns whichever pattern happens to be tried
    first (PT before KT), so comparing single tokens would report PT950 vs an
    18KT packing row as a *disagreement* when the two sides in fact agree on
    18KT.  Matching must compare sets: see _cmp_metal().
    """
    found = set()
    for raw in texts:
        if not raw:
            continue
        s = str(raw)
        for pat, fmt in _METAL_PATTERNS:
            for m in pat.finditer(s):
                found.add(fmt(m).upper())
    return frozenset(found)


def _approx(a: float, b: float, tol_abs: float = 0.01, tol_rel: float = 0.02) -> bool:
    """True if two floats agree within tol_abs OR tol_rel of the larger value."""
    if abs(a - b) <= tol_abs:
        return True
    base = max(abs(a), abs(b))
    return base > 0 and abs(a - b) / base <= tol_rel


# ── Match confidence ladder ──────────────────────────────────────────────────
# Confidence records HOW MUCH corroborating evidence a pair had.  It is never a
# licence to contradict evidence that WAS available — see _score_pair().
_CONF_DIRECT         = 1.0    # explicit (invoice_no, invoice_line_position)
_CONF_TYPE_QTY_RATE  = 0.95   # item type + quantity + unit rate agree
_CONF_TYPE_QTY_METAL = 0.85   # item type + quantity + metal agree (no rate)
_CONF_TYPE_QTY       = 0.80   # item type + quantity only
_CONF_QTY_RATE       = 0.70   # quantity + rate; packing list has no Ctg column
_CONF_AGGREGATE      = 0.60   # N:1 — row belongs to an aggregated invoice line

# Tiers that consume an invoice line exclusively (1:1), strongest first.
_EXCLUSIVE_TIERS = (
    _CONF_TYPE_QTY_RATE, _CONF_TYPE_QTY_METAL, _CONF_TYPE_QTY, _CONF_QTY_RATE,
)

# A row assigned below this floor carries requires_manual_review=True.
# _CONF_TYPE_QTY (item type + quantity, with neither rate nor metal available
# to corroborate) sits below the floor by design: that is the evidence level at
# which a 14KT pendant was assigned to a 925 invoice line.
_REVIEW_CONFIDENCE_FLOOR = 0.85

# Discriminator comparison outcomes.
_AGREE, _UNKNOWN, _CONFLICT = 1, 0, -1


def _cmp_token(a: str, b: str) -> int:
    """Compare two canonical tokens; absent on either side → _UNKNOWN."""
    if not a or not b:
        return _UNKNOWN
    return _AGREE if a == b else _CONFLICT


def _cmp_metal(a: FrozenSet[str], b: FrozenSet[str]) -> int:
    """Compare two metal token SETS; absent on either side → _UNKNOWN.

    Agreement is a non-empty intersection, not equality: a two-tone invoice
    line ({18KT, PT950}) and a packing row that records only its primary metal
    ({18KT}) describe the same piece.  Only genuinely disjoint sets — both
    sides declaring metals, with nothing in common — are a conflict.
    """
    if not a or not b:
        return _UNKNOWN
    return _AGREE if a & b else _CONFLICT


def _cmp_num(a: float, b: float, tol_abs: float = 0.01, tol_rel: float = 0.02) -> int:
    """Compare two positive numbers; missing on either side → _UNKNOWN."""
    if a <= 0 or b <= 0:
        return _UNKNOWN
    return _AGREE if _approx(a, b, tol_abs=tol_abs, tol_rel=tol_rel) else _CONFLICT


def _line_rate(il: Dict[str, Any]) -> float:
    return _safe_float(
        il.get("rate_usd") or il.get("unit_netto_pln") or il.get("unit_price") or 0
    )


def _line_total(il: Dict[str, Any], qty: float, rate: float) -> float:
    """Invoice line total — explicit column when present, else qty × rate."""
    for key in ("total_value", "amount_usd", "total_usd"):
        val = _safe_float(il.get(key))
        if val > 0:
            return val
    return qty * rate


def _invoice_line_facts(il: Dict[str, Any]) -> Dict[str, Any]:
    qty  = _safe_float(il.get("quantity"))
    rate = _line_rate(il)
    return {
        "line":   il,
        "inv_no": str(il.get("invoice_no", "") or "").strip(),
        "item":   _canonical_item_type(
            str(il.get("item_type", "") or il.get("description", "") or "")),
        "qty":    qty,
        "rate":   rate,
        "total":  _line_total(il, qty, rate),
        "metal":  _metal_tokens(il.get("description", ""), il.get("item_type", "")),
        "pos":    il.get("invoice_line_position"),
    }


def _packing_row_facts(r: Dict[str, Any], idx: int) -> Dict[str, Any]:
    return {
        "idx":     idx,
        "inv_no":  str(r.get("invoice_no", "") or "").strip(),
        "item":    _canonical_item_type(str(r.get("item_type", "") or "")),
        "qty":     _safe_float(r.get("quantity")),
        "rate":    _safe_float(r.get("unit_price")),
        "metal":   _metal_tokens(r.get("metal", ""), r.get("karat", "")),
        "sr":      _safe_float(r.get("pack_sr")),
        "raw_pos": r.get("invoice_line_position"),
    }


def _score_pair(rf: Dict[str, Any], lf: Dict[str, Any]) -> Optional[Tuple[float, str]]:
    """
    Strongest tier at which a packing row may be paired with an invoice line,
    or None if it may not be paired with it at all.

    INVARIANT — the rule this function exists to enforce:

        A discriminator that is present on BOTH sides and DISAGREES vetoes the
        pair at EVERY tier.  A weaker tier may relax a discriminator that is
        *unavailable*; it may never relax one that is available and contradicts.

    Without this invariant a low tier can claim an invoice line that a higher
    tier already refused on the evidence — e.g. a 14KT pendant taking the 925
    pendant's line on "item type + quantity" alone after the metal check had
    correctly rejected it, leaving the genuine 925 row with no line at all.
    """
    if not rf["inv_no"] or rf["inv_no"] != lf["inv_no"]:
        return None

    item  = _cmp_token(rf["item"],  lf["item"])
    metal = _cmp_metal(rf["metal"], lf["metal"])
    qty   = _cmp_num(rf["qty"],  lf["qty"])
    rate  = _cmp_num(rf["rate"], lf["rate"], tol_abs=0.05, tol_rel=0.05)

    if _CONFLICT in (item, metal, qty, rate):
        return None

    if item == _AGREE and qty == _AGREE and rate == _AGREE:
        return (_CONF_TYPE_QTY_RATE,
                "type+qty+rate+metal" if metal == _AGREE else "type+qty+rate")
    if item == _AGREE and qty == _AGREE and metal == _AGREE:
        return (_CONF_TYPE_QTY_METAL, "type+qty+metal")
    if item == _AGREE and qty == _AGREE:
        return (_CONF_TYPE_QTY, "type+qty")
    if qty == _AGREE and rate == _AGREE:
        return (_CONF_QTY_RATE, "qty+rate")
    return None


def _rate_distance(rf: Dict[str, Any], lf: Dict[str, Any]) -> float:
    """Relative rate gap, used only to break ties deterministically."""
    if rf["rate"] <= 0 or lf["rate"] <= 0:
        return 0.0
    base = max(rf["rate"], lf["rate"])
    return abs(rf["rate"] - lf["rate"]) / base if base > 0 else 0.0


def match_packing_to_invoice(
    packing_rows: List[Dict[str, Any]],
    invoice_lines: List[Dict[str, Any]],
) -> List[Dict[str, Any]]:
    """
    Match packing rows → invoice lines.

    Product codes are minted on INVOICE lines (`invoice_no + "-" + position`).
    This function copies them onto packing rows; a packing row that ends with
    `product_code = None` is a row this function failed to place, never a row
    whose code "was not generated".

    Assignment is evidence-ordered, not row-ordered:

      1. Direct      1.00 — explicit (invoice_no, invoice_line_position)
      2. type+qty+rate       0.95  ┐
      3. type+qty+metal      0.85  │ exclusive (1:1) tiers, strongest first,
      4. type+qty            0.80  │ mutually-unique pairs committed first
      5. qty+rate            0.70  ┘
      6. aggregate           0.60 — N:1, capacity-bounded
      7. unmatched           0.00 — requires_manual_review=True

    Two invariants distinguish this from a greedy single pass:

      * **No tier may contradict available evidence.** A discriminator present
        on both sides that disagrees vetoes the pair at every tier — a weaker
        tier may relax a discriminator that is *missing*, never one that is
        present and conflicting. See `_score_pair()`.
      * **Order independence.** Every (row, line) pair is scored before
        anything is committed, tiers are resolved strongest-first taking only
        mutually-unique pairs, and residual ties break on a deterministic key.
        Shuffling `packing_rows` cannot change the outcome.

    Rows assigned below `_REVIEW_CONFIDENCE_FLOOR`, rows placed by a tie-break,
    and unmatched rows all carry `requires_manual_review=True`.
    """
    rows: List[Dict[str, Any]] = [dict(r) for r in packing_rows]
    row_facts  = [_packing_row_facts(r, i) for i, r in enumerate(rows)]
    line_facts = [_invoice_line_facts(il) for il in invoice_lines]

    # row index → (line index, confidence, strategy, ambiguous)
    assigned: Dict[int, Tuple[int, float, str, bool]] = {}
    claimed: set = set()          # invoice line indexes consumed by a 1:1 tier

    # ── Tier 1 — direct (invoice_no, invoice_line_position) ───────────────────
    direct_map: Dict[Tuple[str, int], int] = {}
    for li, lf in enumerate(line_facts):
        if lf["inv_no"] and lf["pos"] is not None:
            try:
                direct_map.setdefault((lf["inv_no"], int(lf["pos"])), li)
            except (ValueError, TypeError):
                pass
    for rf in row_facts:
        if not rf["inv_no"] or rf["raw_pos"] is None:
            continue
        try:
            li = direct_map.get((rf["inv_no"], int(rf["raw_pos"])))
        except (ValueError, TypeError):
            continue
        if li is not None and li not in claimed:
            claimed.add(li)
            assigned[rf["idx"]] = (li, _CONF_DIRECT, "direct", False)

    # ── Score every remaining (row, line) pair once ───────────────────────────
    candidates: Dict[int, Dict[int, Tuple[float, str]]] = {}
    for rf in row_facts:
        if rf["idx"] in assigned:
            continue
        scored: Dict[int, Tuple[float, str]] = {}
        for li, lf in enumerate(line_facts):
            if li in claimed:
                continue
            hit = _score_pair(rf, lf)
            if hit is not None:
                scored[li] = hit
        candidates[rf["idx"]] = scored

    # ── Exclusive tiers, strongest first, unique pairs before contested ones ──
    for tier in _EXCLUSIVE_TIERS:
        while True:
            open_at_tier: Dict[int, List[int]] = {}
            for ridx, scored in candidates.items():
                if ridx in assigned:
                    continue
                lines = sorted(
                    li for li, (conf, _s) in scored.items()
                    if conf == tier and li not in claimed
                )
                if lines:
                    open_at_tier[ridx] = lines
            if not open_at_tier:
                break

            claimants: Dict[int, List[int]] = defaultdict(list)
            for ridx, lines in open_at_tier.items():
                for li in lines:
                    claimants[li].append(ridx)

            # Commit only pairs where the row has one option AND the line has
            # one suitor — then re-derive, because each commit can make other
            # pairs unique.  Fixed point, so the result cannot depend on order.
            progressed = False
            for ridx in sorted(open_at_tier):
                lines = [li for li in open_at_tier[ridx] if li not in claimed]
                if len(lines) != 1:
                    continue
                li = lines[0]
                if len(claimants[li]) != 1:
                    continue
                conf, strat = candidates[ridx][li]
                assigned[ridx] = (li, conf, strat, False)
                claimed.add(li)
                progressed = True
            if progressed:
                continue

            # Everything left at this tier is genuinely contested.  Break the
            # tie on a stable key — closest rate, then row order, then line
            # order — and mark the row for review.
            options = [
                (_rate_distance(row_facts[ridx], line_facts[li]), ridx, li)
                for ridx, lines in open_at_tier.items()
                for li in lines
            ]
            options.sort()
            _dist, ridx, li = options[0]
            conf, strat = candidates[ridx][li]
            assigned[ridx] = (li, conf, strat, True)
            claimed.add(li)

    # ── Tier 6 — N:1 aggregate, bounded by the invoice line's own capacity ────
    # An invoice line may summarise many packing rows ("21 RINGs @ $371.10").
    # Capacity is what stops one line absorbing rows that belong to another:
    # quantity is a hard bound, value a preference.
    rem_qty = [lf["qty"] for lf in line_facts]
    rem_val = [lf["total"] for lf in line_facts]
    for ridx, (li, _c, _s, _a) in assigned.items():
        rem_qty[li] -= row_facts[ridx]["qty"]
        rem_val[li] -= row_facts[ridx]["qty"] * row_facts[ridx]["rate"]

    # Capacity is consumed as rows are placed, so the ORDER rows are considered
    # in changes the outcome.  Order by the rows' own content — largest value
    # first (those have the fewest viable homes), then source serial — never by
    # the order the caller happened to supply them in.
    aggregate_queue = sorted(
        (rf for rf in row_facts if rf["idx"] not in assigned),
        key=lambda rf: (-(rf["qty"] * rf["rate"]), rf["sr"], rf["idx"]),
    )
    for rf in aggregate_queue:
        if rf["qty"] <= 0 or not rf["inv_no"] or not rf["item"]:
            continue
        row_val = rf["qty"] * rf["rate"]
        eligible = [
            li for li, lf in enumerate(line_facts)
            if lf["inv_no"] == rf["inv_no"]
            and _cmp_token(rf["item"], lf["item"]) == _AGREE
            and _cmp_metal(rf["metal"], lf["metal"]) != _CONFLICT
            and rem_qty[li] + 1e-9 >= rf["qty"]
        ]
        if not eligible:
            continue
        fits = [li for li in eligible if row_val <= rem_val[li] * 1.02 + 0.01]
        if fits:
            # Tightest value fit first — keeps rows off lines they would blow.
            li = min(fits, key=lambda i: (rem_val[i] - row_val, i))
        else:
            # Nothing fits by value; the largest remainder is the least-bad
            # home.  Quantity still reconciles; residual value error is
            # surfaced as an advisory, not silently absorbed.
            li = min(eligible, key=lambda i: (-rem_val[i], i))
        rem_qty[li] -= rf["qty"]
        rem_val[li] -= row_val
        assigned[rf["idx"]] = (li, _CONF_AGGREGATE, "type+metal_aggregate", False)

    # ── Emit, preserving the caller's row order ──────────────────────────────
    matched: List[Dict[str, Any]] = []
    for rf in row_facts:
        r = rows[rf["idx"]]
        hit = assigned.get(rf["idx"])
        if hit is None:
            r["invoice_line_position"]  = None
            r["product_code"]           = None
            r["requires_manual_review"] = True
            r["extracted_confidence"]   = 0.0
            r["match_strategy"]         = "unmatched"
            r["match_ambiguous"]        = False
        else:
            li, conf, strat, ambiguous = hit
            line = invoice_lines[li]
            r["invoice_line_position"]  = line.get("invoice_line_position")
            r["product_code"]           = line.get("product_code")
            r["invoice_no"]             = line.get("invoice_no") or rf["inv_no"]
            r["requires_manual_review"] = bool(
                ambiguous or conf < _REVIEW_CONFIDENCE_FLOOR
            )
            r["extracted_confidence"]   = conf
            r["match_strategy"]         = strat
            r["match_ambiguous"]        = ambiguous
        matched.append(r)

    return matched


# ── Global Jewellery pipeline ─────────────────────────────────────────────────

def _process_global_packing(
    batch_id: str,
    batch_output_dir: Path,
    packing_file_path: Path,
    raw_rows: List[Dict[str, Any]],
    parser_name: str,
    parser_version: str,
    parser_diagnostic: Dict[str, Any],
) -> Dict[str, Any]:
    """Pipeline for Global Jewellery packing lists.

    Global packing lists are the authority for item rows — there are no
    per-item invoice lines to match against.  Product codes are generated
    from the invoice_no (read from the batch's existing commercial invoice
    record) + the serial number from each packing row.

    Returns the same dict shape as ``process_packing_upload`` so the
    calling route layer (routes_packing.py / routes_intake.py) can handle
    it identically.  The ``"supplier"`` key in the returned dict signals
    the route layer to trigger description generation.
    """
    from . import document_db as ddb

    # Resolve invoice_no: prefer diag (parser already read preamble), else DB
    invoice_no: str = parser_diagnostic.get("invoice_no", "")
    if not invoice_no:
        # Look for an aggregate invoice_line with the 088/... pattern
        import re as _re
        _inv_re = _re.compile(r"\d{3}/\d{4}-\d{4}")
        for il in (ddb.get_invoice_lines_for_batch(batch_id) or []):
            cand = il.get("invoice_no", "")
            if _inv_re.search(str(cand)):
                invoice_no = cand
                break

    # Stamp product_code on rows that don't already have one
    enriched: List[Dict[str, Any]] = []
    for row in raw_rows:
        r = dict(row)
        if not r.get("product_code"):
            serial = r.get("serial_no") or r.get("invoice_line_position", 0)
            inv = invoice_no or "GLOBAL"
            r["product_code"]          = f"{inv}-{serial}"
            r["invoice_no"]            = inv
            r["requires_manual_review"] = False
        enriched.append(r)

    doc_invoice_no = invoice_no or (enriched[0].get("invoice_no") if enriched else "")

    return {
        "invoice_lines":        [],
        "invoice_lines_source": "global_packing_authority",
        "packing_rows":         enriched,
        "parser_diagnostic":    parser_diagnostic,
        "supplier":             "global_jewellery",
        "document": {
            "batch_id":          batch_id,
            "invoice_no":        doc_invoice_no,
            "source_file_path":  str(packing_file_path),
            "source_file_hash":  file_sha256(packing_file_path),
            "parser_name":       parser_name,
            "parser_version":    parser_version,
            "extraction_status": "complete" if enriched else "empty",
            "parser_diagnostic": parser_diagnostic,
        },
        "matched_count":   len(enriched),
        "unmatched_count": 0,
        "total_rows":      len(enriched),
    }


# ── Full pipeline ─────────────────────────────────────────────────────────────

def process_packing_upload(
    batch_id: str,
    batch_output_dir: Path,
    packing_file_path: Path,
    force_reextract: bool = False,
    supplier_id: Optional[int] = None,
) -> Dict[str, Any]:
    """
    Full pipeline: extract invoice lines from pz_rows.json, extract packing rows
    from the uploaded file, match them, and return structured result ready for DB insert.

    Returns:
      {
        "invoice_lines": [...],
        "packing_rows": [...],          # enriched with product_code
        "document": {...},              # packing_document fields
        "matched_count": int,           # rows that received a product_code
        "unmatched_count": int,         # rows with no invoice line at all
        "review_count": int,            # matched-but-uncertain + unmatched
        "total_rows": int,
      }
    """
    raw_rows, parser_name, parser_version, parser_diagnostic = extract_packing(
        packing_file_path, supplier_id=supplier_id,
    )

    # Auto-run LLM advisory tier when extraction yields 0 rows (XLSX only).
    # Rows from the second call are DISCARDED — only column_mapping_audit is merged.
    # build_col_map() already excludes method="llm", so no business records are affected.
    if not raw_rows and packing_file_path.suffix.lower() in (".xlsx", ".xls"):
        try:
            _, _, _, _llm_diag = extract_packing(packing_file_path, llm_fallback=True)
            parser_diagnostic["column_mapping_audit"] = (
                _llm_diag.get("column_mapping_audit") or []
            )
            parser_diagnostic["llm_auto_triggered"] = True
            parser_diagnostic["llm_auto_trigger_reason"] = (
                parser_diagnostic.get("failure_reason") or "zero_rows"
            )
        except Exception as _e:
            parser_diagnostic["llm_auto_triggered"] = False
            parser_diagnostic["llm_auto_trigger_error"] = str(_e)[:200]

    # ── Supplier-specific pipeline (Global Jewellery) ─────────────────────
    if parser_diagnostic.get("supplier") == "global_jewellery":
        return _process_global_packing(
            batch_id, batch_output_dir, packing_file_path,
            raw_rows, parser_name, parser_version, parser_diagnostic,
        )

    # ── EJL / default path (unchanged) ────────────────────────────────────
    invoice_lines = load_invoice_lines(batch_output_dir, batch_id=batch_id)
    inv_source = invoice_lines[0].get("_source", "unknown") if invoice_lines else "none"

    enriched = match_packing_to_invoice(raw_rows, invoice_lines)

    # Detect invoice_no from packing rows (majority vote)
    inv_nos = [r.get("invoice_no", "") for r in enriched if r.get("invoice_no")]
    if inv_nos:
        from collections import Counter
        doc_invoice_no = Counter(inv_nos).most_common(1)[0][0]
    else:
        doc_invoice_no = ""

    # "Matched" means the row was placed on an invoice line, i.e. it carries a
    # product_code.  It deliberately does NOT mean "clean": a row can be matched
    # and still need review (low confidence or a tie-break), which is what
    # review_count reports.  Conflating the two is how low-confidence
    # assignments used to pass as verified.
    matched   = sum(1 for r in enriched if r.get("product_code"))
    unmatched = len(enriched) - matched
    review    = sum(1 for r in enriched if r.get("requires_manual_review"))

    return {
        "invoice_lines":         invoice_lines,
        "invoice_lines_source":  inv_source,
        "packing_rows":          enriched,
        "parser_diagnostic":     parser_diagnostic,
        "document": {
            "batch_id":         batch_id,
            "invoice_no":       doc_invoice_no,
            "source_file_path": str(packing_file_path),
            "source_file_hash": file_sha256(packing_file_path),
            "parser_name":      parser_name,
            "parser_version":   parser_version,
            "extraction_status": "complete" if enriched else "empty",
            "parser_diagnostic": parser_diagnostic,
            "supplier_id":       supplier_id,
        },
        "matched_count":   matched,
        "unmatched_count": unmatched,
        "review_count":    review,
        "total_rows":      len(enriched),
    }
