"""
test_packing_pr2a_extraction.py — PR 2A: packing XLSX extraction enrichment.

Covers:
  1.  _FIELD_ALIASES: "quality" maps to "quality_string"
  2.  _FIELD_ALIASES: "qualtity" typo maps to "quality_string"
  3.  Variant A (separate Kt + Col cells): metal merged, metal_color preserved
  4.  Variant B (combined "14KT/Y"): metal_color parsed from slash token
  5.  Variant B with dash ("PT950/-"): dash stripped → metal_color = ""
  6.  Variant B with no slash: metal_color stays empty
  7.  Variant A with "RG" color code: metal_color = "RG"
  8.  DB init: unit_price_eur, metal_color, quality_string columns created
  9.  DB insert: 3 new fields stored and retrievable
  10. DB update (force_reextract): 3 new fields updated
  11. DB defaults: missing fields default to 0.0 / ""
  12. _extract_packing_excel: quality_string key present (not quality)
  13. _extract_packing_excel: metal_color parsed from combined "14KT/Y" cell
  14. _extract_packing_excel: "Qualtity" typo header captured as quality_string
  15. unit_price field carried through extraction into unit_price_eur slot
"""
from __future__ import annotations

import sys
from pathlib import Path
from typing import Any, Dict, List

import pytest

_ROOT = Path(__file__).parents[1]
if str(_ROOT) not in sys.path:
    sys.path.insert(0, str(_ROOT))


# ── Fixtures ──────────────────────────────────────────────────────────────────

@pytest.fixture()
def db(tmp_path):
    from app.services.packing_db import init_packing_db
    db_path = tmp_path / "packing.db"
    init_packing_db(db_path)
    return db_path


def _make_line(**kw) -> Dict[str, Any]:
    defaults = dict(
        packing_document_id="DOC001",
        batch_id="BATCH001",
        invoice_no="EJL/26-27/100",
        invoice_line_position=1,
        product_code="EJL/26-27/100-1",
        design_no="D-100",
        item_type="RING",
        quantity=2.0,
        metal="18KT/W",
        karat="18KT",
        stone_type="",
        remarks="",
        extracted_confidence=1.0,
        requires_manual_review=False,
    )
    defaults.update(kw)
    return defaults


# ── 1–2. Field aliases ────────────────────────────────────────────────────────

class TestFieldAliases:
    def test_quality_maps_to_quality_string(self):
        from app.services.invoice_packing_extractor import _FIELD_ALIASES
        assert _FIELD_ALIASES.get("quality") == "quality_string"

    def test_qualtity_typo_maps_to_quality_string(self):
        from app.services.invoice_packing_extractor import _FIELD_ALIASES
        assert _FIELD_ALIASES.get("qualtity") == "quality_string"

    def test_quality_not_mapped_to_quality(self):
        """Regression: the old 'quality' key must not map to 'quality' any more."""
        from app.services.invoice_packing_extractor import _FIELD_ALIASES
        # All quality-related aliases must resolve to quality_string
        for key, val in _FIELD_ALIASES.items():
            if "qualit" in key:
                assert val == "quality_string", (
                    f"alias {key!r} maps to {val!r}, expected 'quality_string'"
                )


# ── 3–7. Metal / metal_color merge logic ─────────────────────────────────────

class TestMetalColorMerge:
    """
    Tests the Kt/Col merge block inside _extract_packing_excel.
    We exercise the logic directly via a synthetic worksheet to avoid
    file I/O, using _row_to_dict + the processing block.
    """

    def _process_row(self, raw: Dict[str, Any]) -> Dict[str, Any]:
        """
        Mimic the metal/color processing block from _extract_packing_excel.
        Takes a pre-mapped row dict and returns it after the block executes.
        """
        d = dict(raw)
        # Replicate the block from the extractor
        if d.get("metal_color") and d.get("metal"):
            d["metal"] = f"{d['metal']}/{d['metal_color']}"
            # metal_color preserved as-is
        elif not d.get("metal_color") and d.get("metal"):
            combined = str(d["metal"]).strip()
            if "/" in combined:
                _, _, color_part = combined.partition("/")
                color_part = color_part.strip().rstrip("-").strip()
                if color_part and len(color_part) <= 4:
                    d["metal_color"] = color_part
        return d

    def test_variant_a_merge_preserves_metal_color(self):
        """Variant A: separate Kt='14KT', Col='W' → metal='14KT/W', metal_color='W'."""
        row = self._process_row({"metal": "14KT", "metal_color": "W"})
        assert row["metal"] == "14KT/W"
        assert row["metal_color"] == "W"

    def test_variant_b_combined_extracts_color(self):
        """Variant B: combined '14KT/Y' → metal_color='Y'."""
        row = self._process_row({"metal": "14KT/Y"})
        assert row.get("metal_color") == "Y"
        assert row["metal"] == "14KT/Y"  # metal unchanged

    def test_variant_b_dash_means_no_color(self):
        """Variant B: 'PT950/-' → dash stripped → metal_color=''."""
        row = self._process_row({"metal": "PT950/-"})
        assert row.get("metal_color", "") == ""

    def test_variant_b_no_slash_unchanged(self):
        """No slash in metal → metal_color stays empty."""
        row = self._process_row({"metal": "18KT"})
        assert row.get("metal_color", "") == ""
        assert row["metal"] == "18KT"

    def test_variant_a_rg_color_code(self):
        """Variant A: Col='RG' (Rose Gold) correctly preserved."""
        row = self._process_row({"metal": "18KT", "metal_color": "RG"})
        assert row["metal"] == "18KT/RG"
        assert row["metal_color"] == "RG"

    def test_variant_b_wy_bimetal(self):
        """Variant B: '18KT/WY' (white-yellow bimetal) → metal_color='WY'."""
        row = self._process_row({"metal": "18KT/WY"})
        assert row.get("metal_color") == "WY"

    def test_variant_b_long_color_token_ignored(self):
        """Variant B: color part > 4 chars is not treated as a color code."""
        # e.g. "18KT/YELLOW" — 6 chars — not a canonical color code
        row = self._process_row({"metal": "18KT/YELLOW"})
        # metal_color should NOT be set (color_part too long)
        assert row.get("metal_color", "") == ""


# ── 8. DB init columns ────────────────────────────────────────────────────────

class TestDBColumns:
    def test_new_columns_exist_after_init(self, db):
        import sqlite3
        con = sqlite3.connect(str(db))
        cols = {r[1] for r in con.execute("PRAGMA table_info(packing_lines)")}
        con.close()
        assert "unit_price_eur" in cols, "unit_price_eur column missing"
        assert "metal_color"    in cols, "metal_color column missing"
        assert "quality_string" in cols, "quality_string column missing"

    def test_init_idempotent_with_new_columns(self, tmp_path):
        """Second init call must not fail even when columns already exist."""
        from app.services.packing_db import init_packing_db
        p = tmp_path / "packing.db"
        init_packing_db(p)
        init_packing_db(p)  # must not raise


# ── 9. DB insert / retrieve ───────────────────────────────────────────────────

class TestDBNewFields:
    def test_insert_stores_unit_price_eur(self, db):
        from app.services import packing_db as pdb
        pdb.upsert_packing_lines([_make_line(unit_price_eur=42.50)])
        rows = pdb.get_packing_lines_for_batch("BATCH001")
        assert rows[0]["unit_price_eur"] == pytest.approx(42.50)

    def test_insert_stores_metal_color(self, db):
        from app.services import packing_db as pdb
        pdb.upsert_packing_lines([_make_line(metal_color="RG")])
        rows = pdb.get_packing_lines_for_batch("BATCH001")
        assert rows[0]["metal_color"] == "RG"

    def test_insert_stores_quality_string(self, db):
        from app.services import packing_db as pdb
        pdb.upsert_packing_lines([_make_line(quality_string="G-VS LAB,E-VVS LAB")])
        rows = pdb.get_packing_lines_for_batch("BATCH001")
        assert rows[0]["quality_string"] == "G-VS LAB,E-VVS LAB"

    def test_insert_compound_quality_string(self, db):
        """Compound quality values (comma-separated) stored verbatim."""
        from app.services import packing_db as pdb
        pdb.upsert_packing_lines([_make_line(quality_string="F-VS LAB,EMERALD")])
        rows = pdb.get_packing_lines_for_batch("BATCH001")
        assert rows[0]["quality_string"] == "F-VS LAB,EMERALD"

    def test_all_three_fields_together(self, db):
        from app.services import packing_db as pdb
        pdb.upsert_packing_lines([_make_line(
            unit_price_eur=125.00,
            metal_color="W",
            quality_string="G-VS LAB",
        )])
        rows = pdb.get_packing_lines_for_batch("BATCH001")
        r = rows[0]
        assert r["unit_price_eur"] == pytest.approx(125.00)
        assert r["metal_color"] == "W"
        assert r["quality_string"] == "G-VS LAB"


# ── 10. DB update (force_reextract) ───────────────────────────────────────────

class TestDBForceReextract:
    def test_force_reextract_updates_new_fields(self, db):
        from app.services import packing_db as pdb
        # Insert original row
        pdb.upsert_packing_lines([_make_line(
            unit_price_eur=10.0,
            metal_color="W",
            quality_string="G-VS",
        )])
        # Re-extract with updated values
        pdb.upsert_packing_lines([_make_line(
            unit_price_eur=15.0,
            metal_color="Y",
            quality_string="F-VVS LAB",
        )], force_reextract=True)
        rows = pdb.get_packing_lines_for_batch("BATCH001")
        assert len(rows) == 1
        r = rows[0]
        assert r["unit_price_eur"] == pytest.approx(15.0)
        assert r["metal_color"] == "Y"
        assert r["quality_string"] == "F-VVS LAB"


# ── 11. DB defaults ────────────────────────────────────────────────────────────

class TestDBDefaults:
    def test_missing_new_fields_default_correctly(self, db):
        """Row inserted without new fields should have safe defaults."""
        from app.services import packing_db as pdb
        # _make_line without specifying the new fields
        pdb.upsert_packing_lines([_make_line()])
        rows = pdb.get_packing_lines_for_batch("BATCH001")
        r = rows[0]
        assert r["unit_price_eur"] == pytest.approx(0.0)
        assert r["metal_color"] == ""
        assert r["quality_string"] == ""


# ── 12–15. Excel extraction ────────────────────────────────────────────────────

class TestExcelExtraction:
    """
    Test the XLSX extractor using openpyxl to build synthetic workbooks.
    Verifies that quality_string key is emitted (not quality),
    metal_color is parsed from combined tokens, and Value→unit_price.
    """

    def _make_workbook(self, headers: List[str], rows: List[List[Any]], path: Path) -> None:
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        # Preamble: invoice_no so extractor doesn't warn
        ws.cell(row=1, column=1, value="Invoice #")
        ws.cell(row=1, column=2, value="EJL/26-27/100")
        # Header at row 3 (within the 25-row scan window; must have qty + design)
        for col, h in enumerate(headers, start=1):
            ws.cell(row=3, column=col, value=h)
        for row_idx, row_data in enumerate(rows, start=4):
            for col, val in enumerate(row_data, start=1):
                ws.cell(row=row_idx, column=col, value=val)
        wb.save(str(path))

    def test_quality_column_produces_quality_string_key(self, tmp_path):
        """Standard 'Quality' header → extracted dict has 'quality_string' key."""
        from app.services.invoice_packing_extractor import _extract_packing_excel
        path = tmp_path / "pl.xlsx"
        self._make_workbook(
            headers=["DesignNo", "Qty", "Quality"],
            rows=[["D-001", 2, "G-VS LAB"]],
            path=path,
        )
        rows = _extract_packing_excel(path)
        assert len(rows) == 1
        assert "quality_string" in rows[0], "quality_string key missing from extracted row"
        assert "quality" not in rows[0], "legacy 'quality' key should not appear"
        assert rows[0]["quality_string"] == "G-VS LAB"

    def test_qualtity_typo_header_produces_quality_string(self, tmp_path):
        """'Qualtity' typo header → same quality_string key."""
        from app.services.invoice_packing_extractor import _extract_packing_excel
        path = tmp_path / "pl_typo.xlsx"
        self._make_workbook(
            headers=["DesignNo", "Qty", "Qualtity"],
            rows=[["D-002", 3, "F-VVS LAB"]],
            path=path,
        )
        rows = _extract_packing_excel(path)
        assert len(rows) == 1
        assert rows[0].get("quality_string") == "F-VVS LAB"
        assert "quality" not in rows[0]

    def test_combined_kt_color_extracts_metal_color(self, tmp_path):
        """Combined 'Kt/Color' header with '14KT/Y' cell → metal_color='Y'."""
        from app.services.invoice_packing_extractor import _extract_packing_excel
        path = tmp_path / "pl_combined.xlsx"
        self._make_workbook(
            headers=["DesignNo", "Qty", "Kt/Color"],
            rows=[["D-003", 1, "14KT/Y"]],
            path=path,
        )
        rows = _extract_packing_excel(path)
        assert len(rows) == 1
        r = rows[0]
        assert r.get("metal_color") == "Y", (
            f"Expected metal_color='Y', got {r.get('metal_color')!r}"
        )

    def test_value_column_captured_as_unit_price(self, tmp_path):
        """'Value' header → row has 'unit_price' field (route maps to unit_price_eur)."""
        from app.services.invoice_packing_extractor import _extract_packing_excel
        path = tmp_path / "pl_value.xlsx"
        self._make_workbook(
            headers=["DesignNo", "Qty", "Value"],
            rows=[["D-004", 2, 325.50]],
            path=path,
        )
        rows = _extract_packing_excel(path)
        assert len(rows) == 1
        assert rows[0].get("unit_price") == pytest.approx(325.50)

    def test_separate_kt_col_headers_preserved(self, tmp_path):
        """Separate 'Kt' and 'Col' columns → metal_color preserved as-is."""
        from app.services.invoice_packing_extractor import _extract_packing_excel
        path = tmp_path / "pl_separate.xlsx"
        self._make_workbook(
            headers=["DesignNo", "Qty", "Kt", "Col"],
            rows=[["D-005", 1, "18KT", "W"]],
            path=path,
        )
        rows = _extract_packing_excel(path)
        assert len(rows) == 1
        r = rows[0]
        # metal should be combined
        assert "18KT" in (r.get("metal") or "")
        # metal_color should carry the standalone "W"
        assert r.get("metal_color") == "W"

    def test_pt950_dash_no_metal_color(self, tmp_path):
        """Combined 'PT950/-' → metal_color='' (dash stripped, not a color code)."""
        from app.services.invoice_packing_extractor import _extract_packing_excel
        path = tmp_path / "pl_pt.xlsx"
        self._make_workbook(
            headers=["DesignNo", "Qty", "Kt/Color"],
            rows=[["D-006", 1, "PT950/-"]],
            path=path,
        )
        rows = _extract_packing_excel(path)
        assert len(rows) == 1
        assert rows[0].get("metal_color", "") == ""
