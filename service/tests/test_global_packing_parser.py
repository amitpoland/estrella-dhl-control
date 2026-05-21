"""
test_global_packing_parser.py — Unit tests for global_packing_parser.py

Known fixture values from Global Jewellery Invoice 088/2026-2027:
  Row 1: Bracelet JBR00377  metal 9   qty 1  gross 3.420  net 3.178  FOB 232.00
  Row 2: Bracelet JBR00368-3.00  metal 9  qty 1  gross 6.440  net 5.804  FOB 372.00
  Row 3: Pendant  J3806P00322  metal 925SL  qty 1  gross 1.470  net 0.942  FOB 5.00
  Total: 245 rows, FOB USD 3,172.00, Net 453.212g, Gross 505.103g
"""
from __future__ import annotations

import io
import sys
from pathlib import Path
from typing import Any, Dict, List

import pytest

_ROOT = Path(__file__).parents[1]
if str(_ROOT) not in sys.path:
    sys.path.insert(0, str(_ROOT))


# ── XLSX builder helper ───────────────────────────────────────────────────────

def _build_global_xlsx(
    rows: List[Dict[str, Any]],
    *,
    invoice_no: str = "088/2026-2027",
    tmp_path: Path,
) -> Path:
    """Build a minimal Global Jewellery packing list XLSX in tmp_path."""
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active

    # Preamble rows (rows 1-3)
    ws.append(["Global Jewellery Pvt. Ltd."])
    ws.append([f"Invoice No.: {invoice_no}"])
    ws.append([])

    # Header row (row 4)
    ws.append(["Sr", "Type", "Style No.", "Metal", "Qty", "Gross Wt", "Net Wt", "FOB Value"])

    # Data rows
    for r in rows:
        ws.append([
            r.get("sr", ""),
            r.get("type", ""),
            r.get("style_no", ""),
            r.get("metal", ""),
            r.get("qty", ""),
            r.get("gross_wt", ""),
            r.get("net_wt", ""),
            r.get("fob", ""),
        ])

    path = tmp_path / "packing_global.xlsx"
    wb.save(str(path))
    return path


def _make_row(n: int, item_type: str = "Bracelet", style: str = "BR001",
              metal: str = "18KT", qty: float = 1, gross: float = 5.0,
              net: float = 4.5, fob: float = 100.0) -> Dict[str, Any]:
    return dict(sr=n, type=item_type, style_no=style, metal=metal,
                qty=qty, gross_wt=gross, net_wt=net, fob=fob)


# ── Tests ─────────────────────────────────────────────────────────────────────

class TestSafeFloat:
    def test_numeric_string(self):
        from app.services.global_packing_parser import _safe_float
        assert _safe_float("3.420") == pytest.approx(3.420)

    def test_comma_formatted(self):
        from app.services.global_packing_parser import _safe_float
        assert _safe_float("3,172.00") == pytest.approx(3172.0)

    def test_ite_1_returns_zero(self):
        from app.services.global_packing_parser import _safe_float
        assert _safe_float("ite 1") == 0.0

    def test_none_returns_zero(self):
        from app.services.global_packing_parser import _safe_float
        assert _safe_float(None) == 0.0

    def test_integer_passthrough(self):
        from app.services.global_packing_parser import _safe_float
        assert _safe_float(5) == 5.0

    def test_empty_string_returns_zero(self):
        from app.services.global_packing_parser import _safe_float
        assert _safe_float("") == 0.0

    def test_total_string_returns_zero(self):
        from app.services.global_packing_parser import _safe_float
        assert _safe_float("Total") == 0.0


class TestHeaderDetection:
    def test_finds_correct_header_row(self, tmp_path):
        data = [_make_row(i) for i in range(1, 4)]
        path = _build_global_xlsx(data, tmp_path=tmp_path)
        from app.services.global_packing_parser import parse_global_packing_excel
        rows, name, version, diag = parse_global_packing_excel(path, invoice_no="088/2026-2027")
        assert diag["header_row_idx"] == 3  # row index 3 (0-based), row 4 in Excel

    def test_no_header_returns_empty(self, tmp_path):
        """A file with no recognisable header → failure_reason = header_not_detected."""
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Junk", "Data", "Row"])
        path = tmp_path / "junk.xlsx"
        wb.save(str(path))
        from app.services.global_packing_parser import parse_global_packing_excel
        rows, name, version, diag = parse_global_packing_excel(path)
        assert rows == []
        assert diag["failure_reason"] == "header_not_detected"


class TestFirstThreeRows:
    def test_row1_bracelet_jbr00377(self, tmp_path):
        data = [
            dict(sr=1, type="Bracelet", style_no="JBR00377", metal="9",
                 qty=1, gross_wt=3.420, net_wt=3.178, fob=232.00),
            dict(sr=2, type="Bracelet", style_no="JBR00368-3.00", metal="9",
                 qty=1, gross_wt=6.440, net_wt=5.804, fob=372.00),
            dict(sr=3, type="Pendant", style_no="J3806P00322", metal="925SL",
                 qty=1, gross_wt=1.470, net_wt=0.942, fob=5.00),
        ]
        path = _build_global_xlsx(data, tmp_path=tmp_path)
        from app.services.global_packing_parser import parse_global_packing_excel
        rows, _, _, _ = parse_global_packing_excel(path, invoice_no="088/2026-2027")

        r1 = rows[0]
        assert r1["item_type"] == "Bracelet"
        assert r1["design_no"] == "JBR00377"
        assert r1["quantity"] == pytest.approx(1.0)
        assert r1["gross_weight"] == pytest.approx(3.420)
        assert r1["net_weight"] == pytest.approx(3.178)
        assert r1["unit_price"] == pytest.approx(232.00)

    def test_row2(self, tmp_path):
        data = [
            dict(sr=1, type="Bracelet", style_no="JBR00377", metal="9",
                 qty=1, gross_wt=3.420, net_wt=3.178, fob=232.00),
            dict(sr=2, type="Bracelet", style_no="JBR00368-3.00", metal="9",
                 qty=1, gross_wt=6.440, net_wt=5.804, fob=372.00),
        ]
        path = _build_global_xlsx(data, tmp_path=tmp_path)
        from app.services.global_packing_parser import parse_global_packing_excel
        rows, _, _, _ = parse_global_packing_excel(path, invoice_no="088/2026-2027")
        assert rows[1]["design_no"] == "JBR00368-3.00"
        assert rows[1]["unit_price"] == pytest.approx(372.00)

    def test_row3_pendant(self, tmp_path):
        data = [
            dict(sr=1, type="Bracelet", style_no="JBR00377", metal="9",
                 qty=1, gross_wt=3.420, net_wt=3.178, fob=232.00),
            dict(sr=2, type="Bracelet", style_no="JBR00368-3.00", metal="9",
                 qty=1, gross_wt=6.440, net_wt=5.804, fob=372.00),
            dict(sr=3, type="Pendant", style_no="J3806P00322", metal="925SL",
                 qty=1, gross_wt=1.470, net_wt=0.942, fob=5.00),
        ]
        path = _build_global_xlsx(data, tmp_path=tmp_path)
        from app.services.global_packing_parser import parse_global_packing_excel
        rows, _, _, _ = parse_global_packing_excel(path, invoice_no="088/2026-2027")
        assert rows[2]["item_type"] == "Pendant"
        assert rows[2]["design_no"] == "J3806P00322"
        assert rows[2]["net_weight"] == pytest.approx(0.942)


class TestProductCodeFormat:
    def test_row1_product_code(self, tmp_path):
        data = [_make_row(1), _make_row(2)]
        path = _build_global_xlsx(data, tmp_path=tmp_path)
        from app.services.global_packing_parser import parse_global_packing_excel
        rows, _, _, _ = parse_global_packing_excel(path, invoice_no="088/2026-2027")
        assert rows[0]["product_code"] == "088/2026-2027-1"

    def test_row2_product_code(self, tmp_path):
        data = [_make_row(1), _make_row(2)]
        path = _build_global_xlsx(data, tmp_path=tmp_path)
        from app.services.global_packing_parser import parse_global_packing_excel
        rows, _, _, _ = parse_global_packing_excel(path, invoice_no="088/2026-2027")
        assert rows[1]["product_code"] == "088/2026-2027-2"


class TestNoiseRowSkipping:
    def test_ite_1_serial_skipped(self, tmp_path):
        """Row with serial 'ite 1' must be skipped — no crash."""
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Global Jewellery Pvt. Ltd."])
        ws.append(["088/2026-2027"])
        ws.append([])
        ws.append(["Sr", "Type", "Style No.", "Metal", "Qty", "Gross Wt", "Net Wt", "FOB Value"])
        ws.append(["ite 1", "Bracelet", "JBR00377", "9", 1, 3.420, 3.178, 232.00])
        ws.append([1, "Bracelet", "JBR00377", "9", 1, 3.420, 3.178, 232.00])
        path = tmp_path / "packing_ite.xlsx"
        wb.save(str(path))

        from app.services.global_packing_parser import parse_global_packing_excel
        rows, _, _, diag = parse_global_packing_excel(path, invoice_no="088/2026-2027")
        # "ite 1" row should be skipped; clean row should be kept
        assert len(rows) == 1
        assert diag["rows_skipped"] == 1

    def test_total_row_skipped(self, tmp_path):
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Global Jewellery Pvt. Ltd."])
        ws.append([])
        ws.append(["Sr", "Type", "Style No.", "Metal", "Qty", "Gross Wt", "Net Wt", "FOB Value"])
        ws.append([1, "Bracelet", "JBR00377", "9", 1, 3.420, 3.178, 232.00])
        ws.append(["Total", "Grand Total", "", "", 1, 3.420, 3.178, 232.00])
        path = tmp_path / "packing_total.xlsx"
        wb.save(str(path))

        from app.services.global_packing_parser import parse_global_packing_excel
        rows, _, _, diag = parse_global_packing_excel(path, invoice_no="088/2026-2027")
        assert len(rows) == 1

    def test_blank_rows_skipped(self, tmp_path):
        data = [_make_row(1)]
        path = _build_global_xlsx(data, tmp_path=tmp_path)
        # Verify the blank preamble rows don't get included
        from app.services.global_packing_parser import parse_global_packing_excel
        rows, _, _, _ = parse_global_packing_excel(path, invoice_no="088/2026-2027")
        # Only the data row should be returned
        assert all(r["item_type"] for r in rows)


class TestTotals:
    def test_3_rows_extracted(self, tmp_path):
        data = [_make_row(i) for i in range(1, 4)]
        path = _build_global_xlsx(data, tmp_path=tmp_path)
        from app.services.global_packing_parser import parse_global_packing_excel
        rows, _, _, diag = parse_global_packing_excel(path, invoice_no="088/2026-2027")
        assert len(rows) == 3
        assert diag["rows_extracted"] == 3

    def test_fob_total_accumulates(self, tmp_path):
        data = [
            _make_row(1, fob=232.00),
            _make_row(2, fob=372.00),
            _make_row(3, fob=5.00),
        ]
        path = _build_global_xlsx(data, tmp_path=tmp_path)
        from app.services.global_packing_parser import parse_global_packing_excel
        _, _, _, diag = parse_global_packing_excel(path, invoice_no="088/2026-2027")
        assert diag["total_fob_usd"] == pytest.approx(609.00)

    def test_parser_name_and_version(self, tmp_path):
        data = [_make_row(1)]
        path = _build_global_xlsx(data, tmp_path=tmp_path)
        from app.services.global_packing_parser import parse_global_packing_excel
        _, name, version, _ = parse_global_packing_excel(path, invoice_no="088/2026-2027")
        assert name == "global_packing_v1"
        assert version == "1.0"


class TestInvoiceNoPreamble:
    def test_invoice_no_read_from_preamble(self, tmp_path):
        """If invoice_no not passed, parser reads it from preamble."""
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Global Jewellery Pvt. Ltd."])
        ws.append(["Invoice No.: 088/2026-2027"])
        ws.append([])
        ws.append(["Sr", "Type", "Style No.", "Metal", "Qty", "Gross Wt", "Net Wt", "FOB Value"])
        ws.append([1, "Bracelet", "JBR00377", "9", 1, 3.420, 3.178, 232.00])
        path = tmp_path / "packing_preamble.xlsx"
        wb.save(str(path))

        from app.services.global_packing_parser import parse_global_packing_excel
        rows, _, _, diag = parse_global_packing_excel(path)  # no invoice_no arg
        assert diag["invoice_no"] == "088/2026-2027"
        assert rows[0]["product_code"] == "088/2026-2027-1"

    def test_fallback_product_code_without_invoice_no(self, tmp_path):
        """Without invoice_no, product_code uses 'GLOBAL' prefix."""
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Some Supplier"])  # no invoice number in preamble
        ws.append([])
        ws.append(["Sr", "Type", "Style No.", "Metal", "Qty", "Gross Wt", "Net Wt", "FOB Value"])
        ws.append([1, "Bracelet", "JBR00377", "9", 1, 3.420, 3.178, 232.00])
        path = tmp_path / "packing_noinv.xlsx"
        wb.save(str(path))

        from app.services.global_packing_parser import parse_global_packing_excel
        rows, _, _, _ = parse_global_packing_excel(path)
        assert rows[0]["product_code"].startswith("GLOBAL-")
