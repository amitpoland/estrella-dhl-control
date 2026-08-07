"""POST /api/v1/packing/{batch_id}/rematch — governed re-extraction endpoint.

End-to-end against a real temp DB and a real xlsx: persist rows under a
deliberately WRONG assignment, then verify the dry run proposes the correction
without writing, and that apply is triple-gated (admin session, confirmation
token, non-blocking plan) before the same computation persists.

All fixtures synthetic — the repository is public.
"""
from __future__ import annotations

import json
from pathlib import Path

import pytest
from fastapi.testclient import TestClient

INV = "TEST/00-00/001"
BID = "B-REMATCH-1"


# ── Fixtures ─────────────────────────────────────────────────────────────────

def _build_purchase_xlsx(path: Path) -> None:
    """Two-row EJL-style purchase packing list.

    Row 1: silver pendant  $5   → belongs to invoice line 1 (925, $5)
    Row 2: gold pendant  $106   → belongs to invoice line 2 (14KT, $106)
    """
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    # Preamble: the parser reads the invoice reference from here and stamps it
    # onto every row — without it every matcher tier vetoes on invoice_no.
    ws.append(["Invoice #", INV])
    ws.append([])
    ws.append(["PkSr", "Ctg", "DesignNo", "Kt/Color", "Quality",
               "Dia Wt", "Col Wt", "Qty", "Value", "Total Value", "Size"])
    ws.append([1, "PND", "SYN-001", "925/W", "PLAIN", 0, 0, 1, 5.0, 5.0, ""])
    ws.append([2, "PND", "SYN-002", "14KT/Y", "G-VS", 0.5, 0, 1, 106.0, 106.0, ""])
    wb.save(str(path))


@pytest.fixture()
def env(tmp_path, monkeypatch):
    monkeypatch.setenv("STORAGE_ROOT", str(tmp_path))
    monkeypatch.setenv("PZ_STORAGE_ROOT", str(tmp_path))
    from app.core.config import settings
    monkeypatch.setattr(settings, "storage_root", tmp_path, raising=False)
    monkeypatch.setattr(settings, "api_key", "", raising=False)
    from app.main import app
    from app.auth.dependencies import get_current_user, require_admin
    # Admin session by default (Lesson O: session override, popped in teardown).
    app.dependency_overrides[get_current_user] = lambda: {
        "id": "t", "email": "t@l", "role": "admin"}
    app.dependency_overrides[require_admin] = lambda: {
        "id": "t", "email": "t@l", "role": "admin"}

    from app.services import document_db as ddb
    from app.services import packing_db as pdb
    ddb.init_document_db(tmp_path / "documents.db")
    pdb.init_packing_db(tmp_path / "packing.db")
    try:
        yield TestClient(app), tmp_path, ddb, pdb
    finally:
        app.dependency_overrides.clear()


def _seed(tmp_path, ddb, pdb, *, wrong: bool = True) -> Path:
    """Batch with invoice authority + a packing file + persisted rows.

    ``wrong=True`` persists the historical mis-assignment: both rows on line 2,
    which over-fills a qty-1 line and leaves line 1 starved.
    """
    out = tmp_path / "outputs" / BID
    (out / "source" / "packing").mkdir(parents=True, exist_ok=True)
    (out / "audit.json").write_text(json.dumps(
        {"batch_id": BID, "tracking_no": BID, "awb": BID,
         "carrier": "DHL", "timeline": []}), encoding="utf-8")

    # Purchase authority: 2 one-piece lines the two rows discriminate on metal.
    inv_doc = ddb.register_document(
        batch_id=BID, document_type="purchase_invoice",
        file_name="invoice.pdf", file_path=str(out / "invoice.pdf"),
        file_hash="synthetic-invoice-hash", source="intake") or ""
    ddb.store_invoice_lines(inv_doc, BID, [
        {"invoice_no": INV, "line_position": 1, "product_code": f"{INV}-1",
         "description": "PCS, SL925 SILVER Plain Jewellery PENDANT",
         "quantity": 1, "unit_price": 5.0, "total_value": 5.0,
         "rate_usd": 5.0, "amount_usd": 5.0},
        {"invoice_no": INV, "line_position": 2, "product_code": f"{INV}-2",
         "description": "PCS, 14KT Gold Studded PENDANT",
         "quantity": 1, "unit_price": 106.0, "total_value": 106.0,
         "rate_usd": 106.0, "amount_usd": 106.0},
    ])

    # The stored packing source file + its content-addressed document record.
    pf = out / "source" / "packing" / "purchase_syn.xlsx"
    _build_purchase_xlsx(pf)
    from app.services.invoice_packing_extractor import file_sha256
    doc_id = pdb.upsert_packing_document(
        batch_id=BID, invoice_no=INV,
        source_file_path=str(pf), source_file_hash=file_sha256(pf),
        parser_name="test", parser_version="1", extraction_status="complete")

    # Persisted rows: sr1 (the silver $5 pendant) sits on the WRONG line when
    # wrong=True — the exact incident shape this endpoint exists to repair.
    pdb.upsert_packing_lines([
        {"batch_id": BID, "packing_document_id": doc_id, "invoice_no": INV,
         "pack_sr": 1, "invoice_line_position": 2 if wrong else 1,
         "product_code": f"{INV}-2" if wrong else f"{INV}-1",
         "design_no": "SYN-001", "item_type": "PENDANT", "quantity": 1,
         "unit_price": 5.0, "total_value": 5.0, "metal": "925",
         "match_strategy": "type+qty", "extracted_confidence": 0.80,
         "requires_manual_review": False},
        {"batch_id": BID, "packing_document_id": doc_id, "invoice_no": INV,
         "pack_sr": 2, "invoice_line_position": 2,
         "product_code": f"{INV}-2",
         "design_no": "SYN-002", "item_type": "PENDANT", "quantity": 1,
         "unit_price": 106.0, "total_value": 106.0, "metal": "14KT",
         "match_strategy": "type+qty+rate+metal", "extracted_confidence": 0.95,
         "requires_manual_review": False},
    ])
    return pf


def _rows_by_sr(pdb):
    return {r["pack_sr"]: r for r in pdb.get_packing_lines_for_batch(BID)}


def _timeline(tmp) -> list:
    audit = json.loads((tmp / "outputs" / BID / "audit.json").read_text(encoding="utf-8"))
    return audit.get("timeline", [])


# ── Dry run ──────────────────────────────────────────────────────────────────

def test_dry_run_proposes_the_correction_and_writes_nothing(env):
    cli, tmp, ddb, pdb = env
    _seed(tmp, ddb, pdb)
    before = _rows_by_sr(pdb)

    r = cli.post(f"/api/v1/packing/{BID}/rematch")
    assert r.status_code == 200, r.text
    body = r.json()
    assert body["dry_run"] is True and body["applied"] is False

    plan = body["plan"]
    assert plan["blocking"] is False
    changed = {c["pack_sr"]: c for c in plan["row_changes"]}
    # The mis-placed silver pendant moves to its metal-matching line.
    assert changed[1]["old"]["product_code"] == f"{INV}-2"
    assert changed[1]["new"]["product_code"] == f"{INV}-1"
    # The line reconciliation shows the over/starve pair resolving.
    recon = {l["product_code"]: l for l in plan["line_reconciliation"]}
    assert recon[f"{INV}-2"]["before"]["qty_status"] == "over"
    assert recon[f"{INV}-2"]["after"]["qty_status"] == "ok"
    assert recon[f"{INV}-1"]["before"]["qty_status"] == "short"
    assert recon[f"{INV}-1"]["after"]["qty_status"] == "ok"

    # And nothing was written.
    assert _rows_by_sr(pdb) == before


def test_dry_run_on_batch_with_no_stored_rows_is_404_not_an_empty_write(env):
    cli, tmp, ddb, pdb = env
    out = tmp / "outputs" / "B-NOROWS"
    (out / "source" / "packing").mkdir(parents=True, exist_ok=True)
    (out / "audit.json").write_text(json.dumps(
        {"batch_id": "B-NOROWS", "timeline": []}), encoding="utf-8")
    _build_purchase_xlsx(out / "source" / "packing" / "p.xlsx")
    r = cli.post("/api/v1/packing/B-NOROWS/rematch")
    assert r.status_code == 404


def test_unregistered_source_file_blocks_instead_of_guessing_a_document(env):
    """Content hash is the authority: an unregistered (or edited) file must not
    resolve by filename and must render the plan blocking."""
    cli, tmp, ddb, pdb = env
    pf = _seed(tmp, ddb, pdb)
    # Edit the file after registration → its hash no longer matches the record.
    _build_purchase_xlsx(pf)  # rebuild alone keeps content identical…
    import openpyxl
    wb = openpyxl.load_workbook(pf)
    wb.active.append([3, "RNG", "SYN-003", "14KT/Y", "G-VS", 0, 0, 1, 50.0, 50.0, ""])
    wb.save(str(pf))

    r = cli.post(f"/api/v1/packing/{BID}/rematch")
    assert r.status_code == 200, r.text
    plan = r.json()["plan"]
    assert plan["blocking"] is True
    assert "no_unique_document_for_file" in {b["code"] for b in plan["blockers"]}


# ── Apply gating ─────────────────────────────────────────────────────────────

def test_apply_without_confirmation_token_refuses_and_writes_nothing(env):
    cli, tmp, ddb, pdb = env
    _seed(tmp, ddb, pdb)
    before = _rows_by_sr(pdb)

    for confirm in ("", "yes", "true", BID.lower() + "x"):
        r = cli.post(f"/api/v1/packing/{BID}/rematch",
                     params={"apply": "true", "confirm": confirm})
        assert r.status_code == 200, r.text
        body = r.json()
        assert body["applied"] is False
        assert body["refused"] == "confirmation_token_missing_or_mismatched"

    assert _rows_by_sr(pdb) == before


def test_apply_with_blocking_plan_refuses_and_writes_nothing(env):
    cli, tmp, ddb, pdb = env
    pf = _seed(tmp, ddb, pdb)
    import openpyxl
    wb = openpyxl.load_workbook(pf)
    wb.active.append([3, "RNG", "SYN-003", "14KT/Y", "G-VS", 0, 0, 1, 50.0, 50.0, ""])
    wb.save(str(pf))  # hash mismatch → blocking plan
    before = _rows_by_sr(pdb)

    r = cli.post(f"/api/v1/packing/{BID}/rematch",
                 params={"apply": "true", "confirm": BID})
    assert r.status_code == 200, r.text
    body = r.json()
    assert body["applied"] is False
    assert body["refused"] == "plan_is_blocking"
    assert _rows_by_sr(pdb) == before


def test_apply_with_confirmation_persists_exactly_the_dry_run_plan(env):
    cli, tmp, ddb, pdb = env
    _seed(tmp, ddb, pdb)

    dry = cli.post(f"/api/v1/packing/{BID}/rematch").json()["plan"]
    r = cli.post(f"/api/v1/packing/{BID}/rematch",
                 params={"apply": "true", "confirm": BID})
    assert r.status_code == 200, r.text
    body = r.json()
    assert body["applied"] is True
    assert body["rows_written"] == len(dry["row_changes"])

    after = _rows_by_sr(pdb)
    # sr1 landed on its metal-matching line; sr2 untouched.
    assert after[1]["product_code"] == f"{INV}-1"
    assert after[1]["invoice_line_position"] == 1
    assert after[2]["product_code"] == f"{INV}-2"
    assert after[2]["invoice_line_position"] == 2
    # After the write, a fresh plan projects zero changes.
    assert body["plan_after"]["rows_changed"] == 0


def test_apply_is_idempotent(env):
    cli, tmp, ddb, pdb = env
    _seed(tmp, ddb, pdb)
    cli.post(f"/api/v1/packing/{BID}/rematch",
             params={"apply": "true", "confirm": BID})
    snap = _rows_by_sr(pdb)

    r = cli.post(f"/api/v1/packing/{BID}/rematch",
                 params={"apply": "true", "confirm": BID})
    body = r.json()
    assert body["applied"] is False or body.get("rows_written", 0) == 0
    assert _rows_by_sr(pdb) == snap


def test_operator_confirmed_row_survives_an_apply(env):
    """The human decision survives; its damage is advisory, not a batch veto.

    Historical semantics (pre gate-refinement): the confirmed pin keeping
    line 2 over authority made the WHOLE plan blocking, so this scenario was
    pinned as `blocking is True` + refused apply. That gate conflated "the
    write would raise a line over authority" with "an over line exists in the
    batch": here the write adds nothing to line 2 (before == after — the only
    proposed change, sr1's correction, is preserved and never written), so the
    pre-existing violation is now surfaced as a `line_over_authority_preexisting`
    ADVISORY and the plan is non-blocking. The apply then finds no writable
    change and writes nothing — and the confirmed row survives byte for byte,
    which is this test's actual subject.
    """
    cli, tmp, ddb, pdb = env
    _seed(tmp, ddb, pdb)
    # Operator confirms the (wrong) mapping on sr1. The machine may disagree;
    # it must not overwrite a human decision.
    rows = _rows_by_sr(pdb)
    import sqlite3
    with sqlite3.connect(tmp / "packing.db") as con:
        con.execute(
            "UPDATE packing_lines SET operator_review_status='confirmed' WHERE id=?",
            (rows[1]["id"],))
        con.commit()

    dry = cli.post(f"/api/v1/packing/{BID}/rematch").json()["plan"]
    preserved = {e["pack_sr"] for e in dry["operator_confirmed_preserved"]}
    assert preserved == {1}
    # The pre-existing over line the write never touches is an advisory now,
    # not a blocker — one wrongly-confirmed pin must not veto the batch.
    assert dry["blocking"] is False
    assert dry["blockers"] == []
    adv = [a for a in dry["advisories"]
           if a["code"] == "line_over_authority_preexisting"]
    assert adv and adv[0]["product_code"] == f"{INV}-2"
    assert adv[0]["assigned_qty_before"] == adv[0]["assigned_qty_after"] == 2.0

    before = _rows_by_sr(pdb)
    r = cli.post(f"/api/v1/packing/{BID}/rematch",
                 params={"apply": "true", "confirm": BID})
    body = r.json()
    # sr1's correction is preserved (not writable) and nothing else changes,
    # so the apply has no row_changes and writes nothing.
    assert body["applied"] is False
    assert body.get("refused") is None
    assert _rows_by_sr(pdb) == before
    assert _rows_by_sr(pdb)[1]["operator_review_status"] == "confirmed"


def test_apply_lands_corrections_while_a_preexisting_advisory_remains(env):
    """The live-incident shape, end to end: advisory + write in ONE apply.

    Line 2 (authority 1) is pinned over authority by TWO operator-confirmed
    rows — releasable only by an operator ruling, and the write never adds to
    it. The unconfirmed silver row sr1, wrongly stored on line 2, is the one
    correction the plan can make (line 2 drains 3 → 2, still over). The apply
    must land exactly that correction, leave both confirmed rows byte for
    byte, and record the advisory in the response and the audit timeline.
    """
    cli, tmp, ddb, pdb = env
    out = tmp / "outputs" / BID
    (out / "source" / "packing").mkdir(parents=True, exist_ok=True)
    (out / "audit.json").write_text(json.dumps(
        {"batch_id": BID, "tracking_no": BID, "awb": BID,
         "carrier": "DHL", "timeline": []}), encoding="utf-8")

    inv_doc = ddb.register_document(
        batch_id=BID, document_type="purchase_invoice",
        file_name="invoice.pdf", file_path=str(out / "invoice.pdf"),
        file_hash="synthetic-invoice-hash-adv", source="intake") or ""
    ddb.store_invoice_lines(inv_doc, BID, [
        {"invoice_no": INV, "line_position": 1, "product_code": f"{INV}-1",
         "description": "PCS, SL925 SILVER Plain Jewellery PENDANT",
         "quantity": 1, "unit_price": 5.0, "total_value": 5.0,
         "rate_usd": 5.0, "amount_usd": 5.0},
        {"invoice_no": INV, "line_position": 2, "product_code": f"{INV}-2",
         "description": "PCS, 14KT Gold Studded PENDANT",
         "quantity": 1, "unit_price": 106.0, "total_value": 106.0,
         "rate_usd": 106.0, "amount_usd": 106.0},
    ])

    pf = out / "source" / "packing" / "purchase_syn_adv.xlsx"
    _build_three_row_xlsx(pf)
    from app.services.invoice_packing_extractor import file_sha256
    doc_id = pdb.upsert_packing_document(
        batch_id=BID, invoice_no=INV,
        source_file_path=str(pf), source_file_hash=file_sha256(pf),
        parser_name="test", parser_version="1", extraction_status="complete")

    common = {"batch_id": BID, "packing_document_id": doc_id, "invoice_no": INV,
              "item_type": "PENDANT", "quantity": 1,
              "requires_manual_review": False}
    pdb.upsert_packing_lines([
        # sr1: the silver piece, wrongly persisted on the gold line —
        # unconfirmed, so its correction is writable.
        {**common, "pack_sr": 1, "invoice_line_position": 2,
         "product_code": f"{INV}-2", "design_no": "SYN-001", "metal": "925",
         "unit_price": 5.0, "total_value": 5.0,
         "match_strategy": "type+qty", "extracted_confidence": 0.70},
        # sr2 + sr3: two gold pieces both pinned to the 1-pc gold line by the
        # operator — the pre-existing over-authority the write cannot fix.
        {**common, "pack_sr": 2, "invoice_line_position": 2,
         "product_code": f"{INV}-2", "design_no": "SYN-002", "metal": "14KT",
         "unit_price": 106.0, "total_value": 106.0,
         "match_strategy": "operator", "extracted_confidence": 1.0},
        {**common, "pack_sr": 3, "invoice_line_position": 2,
         "product_code": f"{INV}-2", "design_no": "SYN-003", "metal": "14KT",
         "unit_price": 106.0, "total_value": 106.0,
         "match_strategy": "operator", "extracted_confidence": 1.0},
    ])
    rows = _rows_by_sr(pdb)
    import sqlite3
    with sqlite3.connect(tmp / "packing.db") as con:
        for sr in (2, 3):
            con.execute(
                "UPDATE packing_lines SET operator_review_status='confirmed' WHERE id=?",
                (rows[sr]["id"],))
        con.commit()

    dry = cli.post(f"/api/v1/packing/{BID}/rematch").json()["plan"]
    assert dry["blocking"] is False
    adv = [a for a in dry["advisories"]
           if a["code"] == "line_over_authority_preexisting"]
    assert adv and adv[0]["product_code"] == f"{INV}-2"
    assert adv[0]["assigned_qty_before"] == 3.0
    assert adv[0]["assigned_qty_after"] == 2.0
    # The one writable correction: sr1 leaves the gold line.
    assert {c["pack_sr"] for c in dry["row_changes"]} == {1}

    before = _rows_by_sr(pdb)
    r = cli.post(f"/api/v1/packing/{BID}/rematch",
                 params={"apply": "true", "confirm": BID})
    assert r.status_code == 200, r.text
    body = r.json()
    assert body["applied"] is True
    assert body["rows_written"] == 1

    after = _rows_by_sr(pdb)
    # The correction landed.
    assert after[1]["invoice_line_position"] == 1
    assert after[1]["product_code"] == f"{INV}-1"
    # Both confirmed rows survive byte for byte — the write never saw them.
    assert after[2] == before[2]
    assert after[3] == before[3]
    assert after[2]["operator_review_status"] == "confirmed"
    # The audit event records that an advisory remained at apply time.
    events = [e for e in _timeline(tmp) if e["event"] == "packing_rematch_applied"]
    assert len(events) == 1
    assert events[0]["detail"]["advisories"] == 1


# ── Re-parse failure (the reparse_failed blocker branch) ─────────────────────

def test_reparse_failure_is_a_blocker_and_apply_refuses(env, monkeypatch):
    """An exception inside the re-parse pipeline must block, never skip.

    The extractor itself is fail-soft by design (a corrupt workbook returns
    zero rows plus a diagnostic, it does not raise), so the only way an
    exception reaches this branch is an unexpected pipeline failure. Simulate
    one at the route's import boundary and verify the route converts it into a
    ``reparse_failed`` blocker — not a silent skip that would leave the file's
    stored rows unaccounted for — and that the apply path then refuses.
    """
    cli, tmp, ddb, pdb = env
    pf = _seed(tmp, ddb, pdb)
    before = _rows_by_sr(pdb)

    from app.api import routes_packing

    def _boom(**kwargs):
        raise RuntimeError("synthetic: workbook exploded mid-parse")

    monkeypatch.setattr(routes_packing, "process_packing_upload", _boom)

    r = cli.post(f"/api/v1/packing/{BID}/rematch")
    assert r.status_code == 200, r.text
    plan = r.json()["plan"]
    assert plan["blocking"] is True
    blockers = {b["code"]: b for b in plan["blockers"]}
    assert "reparse_failed" in blockers
    assert blockers["reparse_failed"]["file"] == pf.name
    assert "workbook exploded" in blockers["reparse_failed"]["detail"]
    # The per-file report carries the error too, so the operator sees WHICH
    # file failed, not just that something did.
    files = {f["file"]: f for f in plan["files"]}
    assert files[pf.name]["rows_extracted"] == 0
    assert "workbook exploded" in files[pf.name]["error"]

    # Even a fully-confirmed apply request must refuse against this plan.
    r = cli.post(f"/api/v1/packing/{BID}/rematch",
                 params={"apply": "true", "confirm": BID})
    body = r.json()
    assert body["applied"] is False
    assert body["refused"] == "plan_is_blocking"
    assert _rows_by_sr(pdb) == before


# ── Sales impact (advisory — reported end-to-end, never a gate) ──────────────

def test_sales_impact_reports_over_bill_resolution_without_gating(env):
    """With sales rows present, the endpoint response carries the downstream
    position per product_code — and the over-bill it reports is advisory: it
    must not block the purchase-side apply (Lesson N / Lesson R).

    Seeded position: both packing rows sit on line 2, so code -1 has zero
    packing backing while sales has sold one piece of it (over-billed), and
    code -2 has two rows against authority one (capped, so sales of one piece
    is still covered). The correction moves sr1 back to line 1, resolving the
    code -1 over-bill.
    """
    cli, tmp, ddb, pdb = env
    _seed(tmp, ddb, pdb)
    ddb.store_sales_packing_lines("SD-REMATCH-1", BID, [
        {"client_name": "SYNTH CLIENT", "product_code": f"{INV}-1",
         "quantity": 1, "unit_price": 9.0, "total_value": 9.0},
        {"client_name": "SYNTH CLIENT", "product_code": f"{INV}-2",
         "quantity": 1, "unit_price": 150.0, "total_value": 150.0},
    ])

    r = cli.post(f"/api/v1/packing/{BID}/rematch")
    assert r.status_code == 200, r.text
    plan = r.json()["plan"]

    impact = {s["product_code"]: s for s in plan["sales_impact"]}
    starved = impact[f"{INV}-1"]
    assert starved["sales_qty"] == 1
    assert starved["available_before"] == 0
    assert starved["available_after"] == 1
    assert starved["over_billed_before"] is True
    assert starved["over_billed_after"] is False
    assert starved["verdict"] == "over_bill_resolved"
    # Availability is bounded by invoice authority (1), not by the two packing
    # rows that happened to carry the code.
    covered = impact[f"{INV}-2"]
    assert covered["available_before"] == 1
    assert covered["verdict"] == "ok"
    assert plan["counts"]["over_bills_resolved"] == 1

    # The over-bill is advisory: the plan is NOT blocking and the apply lands.
    assert plan["blocking"] is False
    r = cli.post(f"/api/v1/packing/{BID}/rematch",
                 params={"apply": "true", "confirm": BID})
    assert r.json()["applied"] is True


# ── Partial apply: confirmed row preserved, the rest still lands ─────────────

def _build_three_row_xlsx(path: Path) -> None:
    """Three-row purchase packing list: one silver piece, two gold pieces."""
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(["Invoice #", INV])
    ws.append([])
    ws.append(["PkSr", "Ctg", "DesignNo", "Kt/Color", "Quality",
               "Dia Wt", "Col Wt", "Qty", "Value", "Total Value", "Size"])
    ws.append([1, "PND", "SYN-001", "925/W", "PLAIN", 0, 0, 1, 5.0, 5.0, ""])
    ws.append([2, "PND", "SYN-002", "14KT/Y", "G-VS", 0.5, 0, 1, 106.0, 106.0, ""])
    ws.append([3, "PND", "SYN-003", "14KT/Y", "G-VS", 0.5, 0, 1, 106.0, 106.0, ""])
    wb.save(str(path))


def test_partial_apply_preserves_confirmed_row_and_lands_the_rest(env):
    """A confirmed row must not veto the corrections around it.

    The existing preserved-row test pins the confirmed row where preserving it
    keeps a line over authority, so the whole plan blocks. Here the pinned
    line has slack (authority qty 3), so preserving the human decision leaves
    a NON-blocking plan — and the apply must land the machine's remaining
    correction while leaving the confirmed row untouched. Partial repair,
    not all-or-nothing.

    Position: line 1 (925, qty 1) / line 2 (14KT, qty 3). Stored: sr1 (silver,
    operator-confirmed) and sr2 (gold) on line 2, sr3 (gold) wrongly on
    line 1. Matcher: sr1→line 1 (shown, preserved), sr3→line 2 (lands).
    After: line 2 carries sr1+sr2+sr3 = 3 ≤ 3, so nothing is over.
    """
    cli, tmp, ddb, pdb = env
    out = tmp / "outputs" / BID
    (out / "source" / "packing").mkdir(parents=True, exist_ok=True)
    (out / "audit.json").write_text(json.dumps(
        {"batch_id": BID, "tracking_no": BID, "awb": BID,
         "carrier": "DHL", "timeline": []}), encoding="utf-8")

    inv_doc = ddb.register_document(
        batch_id=BID, document_type="purchase_invoice",
        file_name="invoice.pdf", file_path=str(out / "invoice.pdf"),
        file_hash="synthetic-invoice-hash-3", source="intake") or ""
    ddb.store_invoice_lines(inv_doc, BID, [
        {"invoice_no": INV, "line_position": 1, "product_code": f"{INV}-1",
         "description": "PCS, SL925 SILVER Plain Jewellery PENDANT",
         "quantity": 1, "unit_price": 5.0, "total_value": 5.0,
         "rate_usd": 5.0, "amount_usd": 5.0},
        {"invoice_no": INV, "line_position": 2, "product_code": f"{INV}-2",
         "description": "PCS, 14KT Gold Studded PENDANT",
         "quantity": 3, "unit_price": 106.0, "total_value": 318.0,
         "rate_usd": 106.0, "amount_usd": 318.0},
    ])

    pf = out / "source" / "packing" / "purchase_syn3.xlsx"
    _build_three_row_xlsx(pf)
    from app.services.invoice_packing_extractor import file_sha256
    doc_id = pdb.upsert_packing_document(
        batch_id=BID, invoice_no=INV,
        source_file_path=str(pf), source_file_hash=file_sha256(pf),
        parser_name="test", parser_version="1", extraction_status="complete")

    common = {"batch_id": BID, "packing_document_id": doc_id, "invoice_no": INV,
              "item_type": "PENDANT", "quantity": 1,
              "requires_manual_review": False}
    pdb.upsert_packing_lines([
        # sr1: silver piece pinned (wrongly) to the gold line by the operator.
        {**common, "pack_sr": 1, "invoice_line_position": 2,
         "product_code": f"{INV}-2", "design_no": "SYN-001", "metal": "925",
         "unit_price": 5.0, "total_value": 5.0,
         "match_strategy": "operator", "extracted_confidence": 1.0},
        {**common, "pack_sr": 2, "invoice_line_position": 2,
         "product_code": f"{INV}-2", "design_no": "SYN-002", "metal": "14KT",
         "unit_price": 106.0, "total_value": 106.0,
         "match_strategy": "type+qty+rate+metal", "extracted_confidence": 0.95},
        # sr3: gold piece persisted on the silver line — the historical bug.
        {**common, "pack_sr": 3, "invoice_line_position": 1,
         "product_code": f"{INV}-1", "design_no": "SYN-003", "metal": "14KT",
         "unit_price": 106.0, "total_value": 106.0,
         "match_strategy": "type+qty", "extracted_confidence": 0.70},
    ])
    rows = _rows_by_sr(pdb)
    import sqlite3
    with sqlite3.connect(tmp / "packing.db") as con:
        con.execute(
            "UPDATE packing_lines SET operator_review_status='confirmed' WHERE id=?",
            (rows[1]["id"],))
        con.commit()

    dry = cli.post(f"/api/v1/packing/{BID}/rematch").json()["plan"]
    # The machine disagrees with the confirmed row and SHOWS it — but as
    # preserved, not as a change the write would make.
    preserved = {e["pack_sr"]: e for e in dry["operator_confirmed_preserved"]}
    assert set(preserved) == {1}
    assert preserved[1]["new"]["invoice_line_position"] == 1
    assert 1 not in {c["pack_sr"] for c in dry["row_changes"]}
    # sr3's correction is in the plan, and preserving sr1 leaves line 2 at
    # exactly its authority — so nothing blocks.
    changes = {c["pack_sr"]: c for c in dry["row_changes"]}
    assert changes[3]["new"]["invoice_line_position"] == 2
    assert changes[3]["new"]["product_code"] == f"{INV}-2"
    assert dry["blocking"] is False

    r = cli.post(f"/api/v1/packing/{BID}/rematch",
                 params={"apply": "true", "confirm": BID})
    assert r.status_code == 200, r.text
    body = r.json()
    assert body["applied"] is True
    assert body["rows_written"] == len(dry["row_changes"])

    after = _rows_by_sr(pdb)
    # The human decision survived, byte for byte.
    assert after[1]["invoice_line_position"] == 2
    assert after[1]["product_code"] == f"{INV}-2"
    assert after[1]["operator_review_status"] == "confirmed"
    # The machine's correction landed around it.
    assert after[3]["invoice_line_position"] == 2
    assert after[3]["product_code"] == f"{INV}-2"


# ── Audit timeline attribution ───────────────────────────────────────────────

def test_apply_lands_a_rematch_event_in_the_batch_timeline(env):
    """A confirm-gated rewrite of purchase-authority assignments must be
    attributable in the operator-visible batch timeline, not only in the app
    log (GATE-4 follow-up from the PR #1099 gate, persistence FLAG 2)."""
    cli, tmp, ddb, pdb = env
    _seed(tmp, ddb, pdb)

    r = cli.post(f"/api/v1/packing/{BID}/rematch",
                 params={"apply": "true", "confirm": BID})
    body = r.json()
    assert body["applied"] is True

    events = [e for e in _timeline(tmp) if e["event"] == "packing_rematch_applied"]
    assert len(events) == 1
    ev = events[0]
    assert ev["trigger_source"] == "packing_rematch"
    # The actor is the authenticated admin's identity, not a role literal —
    # a multi-admin install must be able to say WHO applied the rewrite.
    assert ev["actor"] == "t@l"
    assert ev["detail"]["batch_id"] == BID
    assert ev["detail"]["rows_written"] == body["rows_written"]
    # rows_written and proposed_changes are recorded separately by design (an
    # upsert may legitimately land fewer rows than planned); the fixture has
    # exactly one mis-placed row, so both are 1 here.
    assert ev["detail"]["rows_written"] == 1
    assert ev["detail"]["proposed_changes"] == 1


def test_dry_run_and_refused_apply_emit_no_rematch_event(env):
    """Only a write that actually landed may claim one in the timeline."""
    cli, tmp, ddb, pdb = env
    _seed(tmp, ddb, pdb)

    assert cli.post(f"/api/v1/packing/{BID}/rematch").json()["dry_run"] is True
    refused = cli.post(f"/api/v1/packing/{BID}/rematch",
                       params={"apply": "true", "confirm": "wrong"}).json()
    assert refused["applied"] is False

    assert [e for e in _timeline(tmp) if e["event"] == "packing_rematch_applied"] == []


def test_blocking_plan_apply_emits_no_rematch_event(env):
    cli, tmp, ddb, pdb = env
    pf = _seed(tmp, ddb, pdb)
    import openpyxl
    wb = openpyxl.load_workbook(pf)
    wb.active.append([3, "RNG", "SYN-003", "14KT/Y", "G-VS", 0, 0, 1, 50.0, 50.0, ""])
    wb.save(str(pf))  # hash mismatch → blocking plan

    refused = cli.post(f"/api/v1/packing/{BID}/rematch",
                       params={"apply": "true", "confirm": BID}).json()
    assert refused["refused"] == "plan_is_blocking"
    assert [e for e in _timeline(tmp) if e["event"] == "packing_rematch_applied"] == []


def test_noop_second_apply_emits_no_second_rematch_event(env):
    """After a successful apply, re-applying finds zero row_changes and must
    not write — and must not claim a second rewrite in the timeline."""
    cli, tmp, ddb, pdb = env
    _seed(tmp, ddb, pdb)

    first = cli.post(f"/api/v1/packing/{BID}/rematch",
                     params={"apply": "true", "confirm": BID}).json()
    assert first["applied"] is True

    second = cli.post(f"/api/v1/packing/{BID}/rematch",
                      params={"apply": "true", "confirm": BID}).json()
    assert second["applied"] is False or second.get("rows_written", 0) == 0

    events = [e for e in _timeline(tmp) if e["event"] == "packing_rematch_applied"]
    assert len(events) == 1


# ── Auth (Lesson O: session-guarded route, tests migrate with it) ────────────

def test_rematch_requires_admin_not_just_a_session(env):
    cli, tmp, ddb, pdb = env
    _seed(tmp, ddb, pdb)
    from app.main import app
    from app.auth.dependencies import get_current_user, require_admin
    # Non-admin session: restore the real require_admin so the role check runs.
    app.dependency_overrides[get_current_user] = lambda: {
        "id": "u", "email": "u@l", "role": "viewer"}
    app.dependency_overrides.pop(require_admin, None)
    try:
        r = cli.post(f"/api/v1/packing/{BID}/rematch")
        assert r.status_code == 403
    finally:
        app.dependency_overrides[get_current_user] = lambda: {
            "id": "t", "email": "t@l", "role": "admin"}
        app.dependency_overrides[require_admin] = lambda: {
            "id": "t", "email": "t@l", "role": "admin"}


def test_rematch_rejects_unauthenticated(env):
    cli, tmp, ddb, pdb = env
    _seed(tmp, ddb, pdb)
    from app.main import app
    from app.auth.dependencies import get_current_user, require_admin
    app.dependency_overrides.pop(get_current_user, None)
    app.dependency_overrides.pop(require_admin, None)
    try:
        r = cli.post(f"/api/v1/packing/{BID}/rematch")
        assert r.status_code == 401
    finally:
        app.dependency_overrides[get_current_user] = lambda: {
            "id": "t", "email": "t@l", "role": "admin"}
        app.dependency_overrides[require_admin] = lambda: {
            "id": "t", "email": "t@l", "role": "admin"}


# ── Invoice-scoped apply (one blocked invoice must not hold the batch hostage) ─

INV_B = "TEST/00-00/002"


def _build_invoice_xlsx(path, inv, rows):
    """EJL-style packing list for one invoice; rows = (sr, design, ktcolor, val)."""
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(["Invoice #", inv])
    ws.append([])
    ws.append(["PkSr", "Ctg", "DesignNo", "Kt/Color", "Quality",
               "Dia Wt", "Col Wt", "Qty", "Value", "Total Value", "Size"])
    for sr, design, ktcolor, val in rows:
        ws.append([sr, "PND", design, ktcolor, "PLAIN", 0, 0, 1, val, val, ""])
    wb.save(str(path))


def _seed_two_invoices(tmp_path, ddb, pdb, *, b_confirmed_over=True):
    """One batch, two invoices in separate source files — the prod layout.

    Invoice A (TEST/00-00/001): the SAFE invoice — the standard wrong-assignment
    scenario whose correction is clean (mirrors prod invoice 491).
    Invoice B (TEST/00-00/002): when ``b_confirmed_over`` — two operator-CONFIRMED
    rows on a qty-1 line, so preserving them keeps the line over authority and
    the plan blocks on B alone (mirrors prod invoice 485).
    """
    out = tmp_path / "outputs" / BID
    (out / "source" / "packing").mkdir(parents=True, exist_ok=True)
    (out / "audit.json").write_text(json.dumps(
        {"batch_id": BID, "tracking_no": BID, "awb": BID,
         "carrier": "DHL", "timeline": []}), encoding="utf-8")

    inv_doc = ddb.register_document(
        batch_id=BID, document_type="purchase_invoice",
        file_name="invoice.pdf", file_path=str(out / "invoice.pdf"),
        file_hash="synthetic-invoice-hash-2inv", source="intake") or ""
    ddb.store_invoice_lines(inv_doc, BID, [
        {"invoice_no": INV, "line_position": 1, "product_code": f"{INV}-1",
         "description": "PCS, SL925 SILVER Plain Jewellery PENDANT",
         "quantity": 1, "unit_price": 5.0, "total_value": 5.0,
         "rate_usd": 5.0, "amount_usd": 5.0},
        {"invoice_no": INV, "line_position": 2, "product_code": f"{INV}-2",
         "description": "PCS, 14KT Gold Studded PENDANT",
         "quantity": 1, "unit_price": 106.0, "total_value": 106.0,
         "rate_usd": 106.0, "amount_usd": 106.0},
        {"invoice_no": INV_B, "line_position": 1, "product_code": f"{INV_B}-1",
         "description": "PCS, SL925 SILVER Plain Jewellery PENDANT",
         "quantity": 1, "unit_price": 7.0, "total_value": 7.0,
         "rate_usd": 7.0, "amount_usd": 7.0},
    ])

    from app.services.invoice_packing_extractor import file_sha256
    pf_a = out / "source" / "packing" / "purchase_A.xlsx"
    _build_invoice_xlsx(pf_a, INV, [(1, "SYN-001", "925/W", 5.0),
                                    (2, "SYN-002", "14KT/Y", 106.0)])
    doc_a = pdb.upsert_packing_document(
        batch_id=BID, invoice_no=INV,
        source_file_path=str(pf_a), source_file_hash=file_sha256(pf_a),
        parser_name="test", parser_version="1", extraction_status="complete")

    pf_b = out / "source" / "packing" / "purchase_B.xlsx"
    _build_invoice_xlsx(pf_b, INV_B, [(1, "SYN-B01", "925/W", 7.0),
                                      (2, "SYN-B02", "925/W", 7.0)])
    doc_b = pdb.upsert_packing_document(
        batch_id=BID, invoice_no=INV_B,
        source_file_path=str(pf_b), source_file_hash=file_sha256(pf_b),
        parser_name="test", parser_version="1", extraction_status="complete")

    pdb.upsert_packing_lines([
        # Invoice A stored WRONG (sr1 on line 2), correctable — same as _seed().
        {"batch_id": BID, "packing_document_id": doc_a, "invoice_no": INV,
         "pack_sr": 1, "invoice_line_position": 2, "product_code": f"{INV}-2",
         "design_no": "SYN-001", "item_type": "PENDANT", "quantity": 1,
         "unit_price": 5.0, "total_value": 5.0, "metal": "925",
         "match_strategy": "type+qty", "extracted_confidence": 0.80,
         "requires_manual_review": False},
        {"batch_id": BID, "packing_document_id": doc_a, "invoice_no": INV,
         "pack_sr": 2, "invoice_line_position": 2, "product_code": f"{INV}-2",
         "design_no": "SYN-002", "item_type": "PENDANT", "quantity": 1,
         "unit_price": 106.0, "total_value": 106.0, "metal": "14KT",
         "match_strategy": "type+qty+rate+metal", "extracted_confidence": 0.95,
         "requires_manual_review": False},
        # Invoice B: two pieces on a qty-1 line — the historical over-assignment.
        {"batch_id": BID, "packing_document_id": doc_b, "invoice_no": INV_B,
         "pack_sr": 1, "invoice_line_position": 1, "product_code": f"{INV_B}-1",
         "design_no": "SYN-B01", "item_type": "PENDANT", "quantity": 1,
         "unit_price": 7.0, "total_value": 7.0, "metal": "925",
         "match_strategy": "type+qty", "extracted_confidence": 0.80,
         "requires_manual_review": False},
        {"batch_id": BID, "packing_document_id": doc_b, "invoice_no": INV_B,
         "pack_sr": 2, "invoice_line_position": 1, "product_code": f"{INV_B}-1",
         "design_no": "SYN-B02", "item_type": "PENDANT", "quantity": 1,
         "unit_price": 7.0, "total_value": 7.0, "metal": "925",
         "match_strategy": "type+qty", "extracted_confidence": 0.80,
         "requires_manual_review": False},
    ])

    if b_confirmed_over:
        # The operator confirmed BOTH B rows — preserving them keeps line B-1
        # at 2 assigned vs authority 1, so the plan blocks on B alone.
        import sqlite3
        rows = {(r["invoice_no"], r["pack_sr"]): r
                for r in pdb.get_packing_lines_for_batch(BID)}
        with sqlite3.connect(tmp_path / "packing.db") as con:
            for sr in (1, 2):
                con.execute(
                    "UPDATE packing_lines SET operator_review_status='confirmed' "
                    "WHERE id=?", (rows[(INV_B, sr)]["id"],))
            con.commit()
    return pf_a, pf_b


def _rows_full(pdb, inv=None):
    rows = pdb.get_packing_lines_for_batch(BID)
    if inv is not None:
        rows = [r for r in rows if r.get("invoice_no") == inv]
    return {r["pack_sr"]: r for r in rows}


def test_unscoped_blocked_batch_still_refuses_exactly_as_before(env):
    """Byte-for-behavior compatibility: without a scope, one blocked invoice
    still refuses the whole batch with the same refusal code."""
    cli, tmp, ddb, pdb = env
    _seed_two_invoices(tmp, ddb, pdb)
    before = _rows_full(pdb)

    dry = cli.post(f"/api/v1/packing/{BID}/rematch").json()
    assert dry["plan"]["blocking"] is True
    assert "scope_invoice" not in dry  # no scope fields on unscoped calls

    r = cli.post(f"/api/v1/packing/{BID}/rematch",
                 params={"apply": "true", "confirm": BID})
    body = r.json()
    assert body["applied"] is False
    assert body["refused"] == "plan_is_blocking"
    assert _rows_full(pdb) == before


def test_scoped_apply_lands_safe_invoice_while_other_stays_blocked(env):
    """THE incident shape: blockers live only in invoice B; scoping to invoice A
    applies A's correction while B — confirmed rows included — stays untouched."""
    cli, tmp, ddb, pdb = env
    _seed_two_invoices(tmp, ddb, pdb)
    b_before = _rows_full(pdb, INV_B)

    dry = cli.post(f"/api/v1/packing/{BID}/rematch",
                   params={"invoice_no": INV}).json()
    assert dry["dry_run"] is True
    assert dry["batch_blocking"] is True
    assert dry["scope_blocking"] is False
    assert dry["scope_invoice"] == INV
    # B's blocker is attributed to B alone: visible batch-wide, not gating A.
    assert {b["code"] for b in dry["plan"]["blockers"]} == {"line_over_authority_after"}
    assert all(b["scope_invoices"] == [INV_B] for b in dry["plan"]["blockers"])

    r = cli.post(f"/api/v1/packing/{BID}/rematch",
                 params={"apply": "true", "confirm": BID, "invoice_no": INV})
    assert r.status_code == 200, r.text
    body = r.json()
    assert body["applied"] is True
    assert body["batch_blocking"] is True
    assert body["scope_blocking"] is False
    assert body["rows_written"] >= 1

    # A corrected: the misplaced silver pendant moved home.
    a_after = _rows_full(pdb, INV)
    assert a_after[1]["product_code"] == f"{INV}-1"
    assert a_after[1]["invoice_line_position"] == 1
    # B untouched — every field, including the confirmations.
    assert _rows_full(pdb, INV_B) == b_before
    assert all(r["operator_review_status"] == "confirmed"
               for r in b_before.values())

    # The audit event names the scope and both counts.
    ev = [e for e in _timeline(tmp) if e.get("event") == "packing_rematch_applied"]
    assert len(ev) == 1
    d = ev[0]["detail"]
    assert d["invoice_scope"] == INV
    assert d["scoped_proposed_changes"] == d["rows_written"] == body["rows_written"]
    assert d["batch_proposed_changes"] >= d["scoped_proposed_changes"]


def test_scope_with_its_own_blocker_refuses(env):
    """Scoping to the BLOCKED invoice must refuse — the scope is not a bypass."""
    cli, tmp, ddb, pdb = env
    _seed_two_invoices(tmp, ddb, pdb)
    before = _rows_full(pdb)

    r = cli.post(f"/api/v1/packing/{BID}/rematch",
                 params={"apply": "true", "confirm": BID, "invoice_no": INV_B})
    body = r.json()
    assert body["applied"] is False
    assert body["refused"] == "scope_is_blocking"
    assert body["scope_blocking"] is True
    assert _rows_full(pdb) == before


def test_global_blocker_vetoes_every_scope(env):
    """A blocker that cannot be attributed (edited file, hash unresolved) is
    GLOBAL: it must veto even a scope whose own invoice is clean."""
    cli, tmp, ddb, pdb = env
    pf_a, pf_b = _seed_two_invoices(tmp, ddb, pdb, b_confirmed_over=False)
    # Edit B's file after registration: its hash no longer resolves, and the
    # resulting blocker has empty scope_invoices.
    import openpyxl
    wb = openpyxl.load_workbook(pf_b)
    wb.active.append([9, "PND", "SYN-B99", "925/W", "PLAIN", 0, 0, 1, 1.0, 1.0, ""])
    wb.save(str(pf_b))
    before = _rows_full(pdb)

    r = cli.post(f"/api/v1/packing/{BID}/rematch",
                 params={"apply": "true", "confirm": BID, "invoice_no": INV})
    body = r.json()
    assert body["applied"] is False
    assert body["refused"] == "scope_is_blocking"
    assert any(b["code"] == "no_unique_document_for_file"
               and b["scope_invoices"] == [] for b in body["scope_blockers"])
    assert _rows_full(pdb) == before


def test_unknown_invoice_scope_refuses(env):
    cli, tmp, ddb, pdb = env
    _seed_two_invoices(tmp, ddb, pdb)
    before = _rows_full(pdb)
    r = cli.post(f"/api/v1/packing/{BID}/rematch",
                 params={"apply": "true", "confirm": BID,
                         "invoice_no": "TEST/00-00/999"})
    body = r.json()
    assert body["ok"] is False
    assert body["applied"] is False
    assert body["refused"] == "invoice_scope_unknown"
    assert _rows_full(pdb) == before


def test_hash_failure_before_document_resolution_is_a_graceful_global_blocker(env, monkeypatch):
    """If file_sha256 itself raises, no document ever resolved — the route must
    still answer 200 with a reparse_failed blocker whose scope is GLOBAL (empty),
    not crash on an unbound doc_ids. Gate finding from the #1118 backend review."""
    cli, tmp, ddb, pdb = env
    _seed_two_invoices(tmp, ddb, pdb, b_confirmed_over=False)

    from app.api import routes_packing

    def _boom(_pf):
        raise OSError("synthetic: file vanished before hashing")

    monkeypatch.setattr(routes_packing, "file_sha256", _boom)
    r = cli.post(f"/api/v1/packing/{BID}/rematch",
                 params={"apply": "true", "confirm": BID, "invoice_no": INV})
    assert r.status_code == 200, r.text
    body = r.json()
    assert body["applied"] is False
    assert body["refused"] == "scope_is_blocking"
    assert any(b["code"] == "reparse_failed" and b["scope_invoices"] == []
               for b in body["scope_blockers"])


def test_second_scoped_apply_is_noop(env):
    cli, tmp, ddb, pdb = env
    _seed_two_invoices(tmp, ddb, pdb)
    cli.post(f"/api/v1/packing/{BID}/rematch",
             params={"apply": "true", "confirm": BID, "invoice_no": INV})
    snap = _rows_full(pdb)
    events_after_first = len(_timeline(tmp))

    r = cli.post(f"/api/v1/packing/{BID}/rematch",
                 params={"apply": "true", "confirm": BID, "invoice_no": INV})
    body = r.json()
    assert body["applied"] is False or body.get("rows_written", 0) == 0
    assert _rows_full(pdb) == snap
    assert len(_timeline(tmp)) == events_after_first
