"""
test_tracking_db.py — Tests for the SQLite tracking events store.

Covers:
  1. init_tracking_db creates table
  2. record_event inserts a row
  3. dedup: same (batch_id, awb, stage, event_time, source_ref, email_message_id) skipped
  4. dedup: different stage → inserted
  5. dedup: different event_time → inserted
  6. dedup: different source_ref → inserted
  7. get_events_for_batch returns only that batch, ordered by event_time
  8. get_events_for_awb returns only that AWB
  9. get_all_events returns all rows, newest first
  10. record_events_batch inserts multiple, skips dupes, returns count
  11. get_latest_stage_for_batch returns most recent stage
  12. DHL email workflow events recorded via dhl pipeline hook
  13. XLSX export writes a file with correct headers
  14. XLSX export includes all events
"""
from __future__ import annotations

import json
import sys
from datetime import datetime, timezone
from pathlib import Path

import pytest

_ROOT = Path(__file__).parents[1]
if str(_ROOT) not in sys.path:
    sys.path.insert(0, str(_ROOT))


# ── Fixtures ──────────────────────────────────────────────────────────────────

@pytest.fixture()
def db(tmp_path):
    from app.services.tracking_db import init_tracking_db
    db_path = tmp_path / "tracking_events.db"
    init_tracking_db(db_path)
    return db_path


def _ev(**kw):
    defaults = dict(
        batch_id="BATCH001",
        awb="1234567890",
        stage="DHL_FIRST_EMAIL_RECEIVED",
        event_time="2026-05-01T10:00:00+00:00",
        source="dhl_monitor",
        source_ref="",
        email_message_id="",
    )
    defaults.update(kw)
    return defaults


# ── Table creation ────────────────────────────────────────────────────────────

class TestInit:
    def test_table_created(self, db):
        import sqlite3
        con = sqlite3.connect(str(db))
        tables = [r[0] for r in con.execute("SELECT name FROM sqlite_master WHERE type='table'")]
        assert "shipment_tracking_events" in tables
        con.close()

    def test_init_idempotent(self, tmp_path):
        from app.services.tracking_db import init_tracking_db
        db_path = tmp_path / "tracking_events.db"
        init_tracking_db(db_path)
        init_tracking_db(db_path)  # should not raise


# ── record_event ──────────────────────────────────────────────────────────────

class TestRecordEvent:
    def test_insert_returns_true(self, db):
        from app.services import tracking_db as tdb
        ok = tdb.record_event(**_ev())
        assert ok is True

    def test_row_persisted(self, db):
        from app.services import tracking_db as tdb
        tdb.record_event(**_ev(description="First DHL email"))
        rows = tdb.get_events_for_batch("BATCH001")
        assert len(rows) == 1
        assert rows[0]["stage"] == "DHL_FIRST_EMAIL_RECEIVED"
        assert rows[0]["description"] == "First DHL email"

    def test_dedup_same_key_skipped(self, db):
        from app.services import tracking_db as tdb
        tdb.record_event(**_ev())
        ok = tdb.record_event(**_ev())  # exact same key
        assert ok is False
        assert len(tdb.get_events_for_batch("BATCH001")) == 1

    def test_dedup_different_stage_inserted(self, db):
        from app.services import tracking_db as tdb
        tdb.record_event(**_ev(stage="DHL_FIRST_EMAIL_RECEIVED"))
        ok = tdb.record_event(**_ev(stage="DHL_EMAIL_SCAN_ACTIVATED"))
        assert ok is True
        assert len(tdb.get_events_for_batch("BATCH001")) == 2

    def test_dedup_different_event_time_inserted(self, db):
        from app.services import tracking_db as tdb
        tdb.record_event(**_ev(event_time="2026-05-01T10:00:00+00:00"))
        ok = tdb.record_event(**_ev(event_time="2026-05-02T10:00:00+00:00"))
        assert ok is True
        assert len(tdb.get_events_for_batch("BATCH001")) == 2

    def test_dedup_different_source_ref_inserted(self, db):
        from app.services import tracking_db as tdb
        tdb.record_event(**_ev(source_ref="ticket-001"))
        ok = tdb.record_event(**_ev(source_ref="ticket-002"))
        assert ok is True

    def test_all_fields_stored(self, db):
        from app.services import tracking_db as tdb
        tdb.record_event(
            batch_id="B1", awb="AWB1", stage="CUSTOMS_PENDING",
            event_time="2026-05-01T12:00:00+00:00", source="dhl_api",
            carrier="DHL", status="customs",
            source_ref="ref-1", email_message_id="msg-1",
            raw_subject="Subject line", raw_sender="sender@dhl.com",
            location="WARSAW - PL", description="Customs pending",
            normalized_stage="CUSTOMS_PENDING", confidence=0.9,
            requires_manual_review=False,
        )
        rows = tdb.get_events_for_batch("B1")
        assert len(rows) == 1
        r = rows[0]
        assert r["carrier"] == "DHL"
        assert r["location"] == "WARSAW - PL"
        assert r["normalized_stage"] == "CUSTOMS_PENDING"
        assert abs(r["confidence"] - 0.9) < 0.001
        assert r["requires_manual_review"] == 0


# ── Queries ───────────────────────────────────────────────────────────────────

class TestQueries:
    def test_get_events_for_batch_isolates(self, db):
        from app.services import tracking_db as tdb
        tdb.record_event(**_ev(batch_id="BATCH001"))
        tdb.record_event(**_ev(batch_id="BATCH002", stage="PICKED_UP"))
        rows = tdb.get_events_for_batch("BATCH001")
        assert len(rows) == 1
        assert rows[0]["batch_id"] == "BATCH001"

    def test_get_events_for_batch_ordered_by_time(self, db):
        from app.services import tracking_db as tdb
        tdb.record_event(**_ev(event_time="2026-05-02T10:00:00+00:00", stage="PICKED_UP"))
        tdb.record_event(**_ev(event_time="2026-05-01T08:00:00+00:00", stage="LABEL_CREATED"))
        rows = tdb.get_events_for_batch("BATCH001")
        assert rows[0]["event_time"] < rows[1]["event_time"]

    def test_get_events_for_awb(self, db):
        from app.services import tracking_db as tdb
        tdb.record_event(**_ev(awb="AWB_X", batch_id="B1"))
        tdb.record_event(**_ev(awb="AWB_Y", batch_id="B2", stage="PICKED_UP"))
        rows = tdb.get_events_for_awb("AWB_X")
        assert len(rows) == 1
        assert rows[0]["awb"] == "AWB_X"

    def test_get_all_events_newest_first(self, db):
        from app.services import tracking_db as tdb
        tdb.record_event(**_ev(event_time="2026-05-01T08:00:00+00:00", stage="LABEL_CREATED"))
        tdb.record_event(**_ev(event_time="2026-05-02T14:00:00+00:00", stage="DELIVERED"))
        rows = tdb.get_all_events()
        assert rows[0]["event_time"] > rows[1]["event_time"]

    def test_get_latest_stage(self, db):
        from app.services import tracking_db as tdb
        tdb.record_event(**_ev(event_time="2026-05-01T08:00:00+00:00", stage="LABEL_CREATED"))
        tdb.record_event(**_ev(event_time="2026-05-02T14:00:00+00:00", stage="DELIVERED"))
        stage = tdb.get_latest_stage_for_batch("BATCH001")
        assert stage == "DELIVERED"

    def test_get_latest_stage_none_when_empty(self, db):
        from app.services import tracking_db as tdb
        assert tdb.get_latest_stage_for_batch("NONEXISTENT") is None


# ── Batch insert ──────────────────────────────────────────────────────────────

class TestRecordEventsBatch:
    def test_batch_inserts_all(self, db):
        from app.services import tracking_db as tdb
        events = [
            {"batch_id": "B1", "awb": "AWB1", "stage": "LABEL_CREATED",
             "event_time": "2026-05-01T08:00:00+00:00", "source": "dhl_api",
             "normalized_stage": "LABEL_CREATED", "confidence": 1.0},
            {"batch_id": "B1", "awb": "AWB1", "stage": "PICKED_UP",
             "event_time": "2026-05-01T10:00:00+00:00", "source": "dhl_api",
             "normalized_stage": "PICKED_UP", "confidence": 1.0},
        ]
        count = tdb.record_events_batch(events)
        assert count == 2

    def test_batch_skips_dupes(self, db):
        from app.services import tracking_db as tdb
        ev = {"batch_id": "B1", "awb": "AWB1", "stage": "LABEL_CREATED",
              "event_time": "2026-05-01T08:00:00+00:00", "source": "dhl_api"}
        tdb.record_events_batch([ev])
        count = tdb.record_events_batch([ev])
        assert count == 0

    def test_batch_uses_normalized_stage_as_stage_fallback(self, db):
        from app.services import tracking_db as tdb
        ev = {"batch_id": "B1", "awb": "AWB1",
              "normalized_stage": "IN_TRANSIT",
              "event_time": "2026-05-01T10:00:00+00:00",
              "source": "dhl_api", "raw_description": "Processed at facility"}
        count = tdb.record_events_batch([ev])
        assert count == 1
        rows = tdb.get_events_for_batch("B1")
        assert rows[0]["stage"] == "IN_TRANSIT"


# ── DHL pipeline hook ─────────────────────────────────────────────────────────

class TestDhlPipelineHook:
    @staticmethod
    def _run(coro):
        """Run a coroutine safely regardless of event loop state left by prior tests."""
        import asyncio
        loop = asyncio.new_event_loop()
        try:
            return loop.run_until_complete(coro)
        finally:
            loop.close()

    def test_receive_dhl_email_writes_two_events(self, tmp_path, db):
        import json
        from app.services.tracking_db import init_tracking_db

        batch_dir = tmp_path / "outputs" / "BATCH_DHL"
        batch_dir.mkdir(parents=True)
        audit_path = batch_dir / "audit.json"
        audit = {"batch_id": "BATCH_DHL", "awb": "9876543210", "timeline": []}
        audit_path.write_text(json.dumps(audit), encoding="utf-8")

        from app.pipelines.dhl import receive_dhl_email
        self._run(
            receive_dhl_email(audit, audit_path, ticket="TKT123", awb="9876543210")
        )

        from app.services import tracking_db as tdb
        events = tdb.get_events_for_batch("BATCH_DHL")
        stages = [e["stage"] for e in events]
        assert "DHL_FIRST_EMAIL_RECEIVED" in stages
        assert "DHL_EMAIL_SCAN_ACTIVATED" in stages

    def test_receive_dhl_email_sets_tracking_active(self, tmp_path, db):
        import json
        batch_dir = tmp_path / "outputs" / "BATCH_ACTIVE"
        batch_dir.mkdir(parents=True)
        audit_path = batch_dir / "audit.json"
        audit = {"batch_id": "BATCH_ACTIVE", "awb": "1111111111", "timeline": []}
        audit_path.write_text(json.dumps(audit), encoding="utf-8")

        from app.pipelines.dhl import receive_dhl_email
        self._run(
            receive_dhl_email(audit, audit_path, ticket="", awb="1111111111")
        )

        updated = json.loads(audit_path.read_text(encoding="utf-8"))
        assert updated.get("tracking_active") is True

    def test_receive_dhl_email_dedup_same_batch(self, tmp_path, db):
        """Calling receive_dhl_email twice for the same batch creates only 2 events."""
        import json

        batch_dir = tmp_path / "outputs" / "BATCH_DEDUP"
        batch_dir.mkdir(parents=True)
        audit_path = batch_dir / "audit.json"
        audit = {"batch_id": "BATCH_DEDUP", "awb": "5555555555", "timeline": []}
        audit_path.write_text(json.dumps(audit), encoding="utf-8")

        from app.pipelines.dhl import receive_dhl_email
        from app.services import tracking_db as tdb

        self._run(
            receive_dhl_email(audit, audit_path, ticket="TKT-A", awb="5555555555")
        )
        first_count = len(tdb.get_events_for_batch("BATCH_DEDUP"))

        self._run(
            receive_dhl_email(audit, audit_path, ticket="TKT-A", awb="5555555555")
        )
        second_count = len(tdb.get_events_for_batch("BATCH_DEDUP"))

        assert first_count == 2
        assert second_count == 2  # no duplicates added


# ── XLSX export ───────────────────────────────────────────────────────────────

class TestXlsxExport:
    def test_export_creates_file(self, tmp_path, db):
        from app.services import tracking_db as tdb
        from app.services.tracking_master_export import export_master_xlsx
        tdb.record_event(**_ev(description="Test event"))
        out = tmp_path / "SHIPMENT_TRACKING_MASTER.xlsx"
        result = export_master_xlsx(out)
        assert result.exists()
        assert result.stat().st_size > 0

    def test_export_headers(self, tmp_path, db):
        import openpyxl
        from app.services import tracking_db as tdb
        from app.services.tracking_master_export import export_master_xlsx, _COLUMNS, _HEADER_LABELS
        tdb.record_event(**_ev())
        out = tmp_path / "out.xlsx"
        export_master_xlsx(out)
        wb = openpyxl.load_workbook(str(out))
        ws = wb.active
        headers = [ws.cell(row=1, column=c).value for c in range(1, len(_COLUMNS) + 1)]
        for col in ("Event ID", "Batch ID", "AWB", "Stage (Workflow)", "Event Time (UTC)"):
            assert col in headers

    def test_export_all_events(self, tmp_path, db):
        import openpyxl
        from app.services import tracking_db as tdb
        from app.services.tracking_master_export import export_master_xlsx
        tdb.record_event(**_ev(stage="LABEL_CREATED", event_time="2026-05-01T08:00:00+00:00"))
        tdb.record_event(**_ev(stage="PICKED_UP",     event_time="2026-05-01T10:00:00+00:00"))
        out = tmp_path / "out.xlsx"
        export_master_xlsx(out)
        wb = openpyxl.load_workbook(str(out))
        ws = wb.active
        assert ws.max_row == 3  # 1 header + 2 data rows

    def test_export_empty_db(self, tmp_path, db):
        from app.services.tracking_master_export import export_master_xlsx
        out = tmp_path / "empty.xlsx"
        export_master_xlsx(out)
        assert out.exists()
