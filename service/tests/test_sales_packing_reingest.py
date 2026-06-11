"""
test_sales_packing_reingest.py — Idempotent backfill for sales_packing_lines.

Covers:
  - app.services.document_db.replace_sales_packing_lines (helper-level)
  - POST /api/v1/shipment/sales-packing/reingest (route-level)

Pins (each maps to a numbered scope rule):
  1. re-upload same file replaces rows, no duplicates
  2. unrelated client rows preserved (other batch / other sales_doc)
  3. sales prices/currency persisted on the new rows
  4. Clear-Diamonds USD persists from Excel symbol
  5. Anastazia EUR persists from Excel symbol
  6. PND tiebreak persists 5.13 USD → 123-3 and 51.30 USD → 123-2
  7. mixed-currency file blocks insert (no override)
  8. response includes counts + warnings
"""
from __future__ import annotations

import json
import sqlite3
import uuid
from pathlib import Path
from unittest.mock import patch

import openpyxl
import pytest
from fastapi.testclient import TestClient

from app.core.config import settings
from app.services import wfirma_client as _wc
from app.services import packing_db   as pdb
from app.services import warehouse_db as wdb
from app.services import document_db  as ddb
from app.services import wfirma_db    as wfdb
from app.services import proforma_service_charges_db as scdb


URL = "/api/v1/shipment/sales-packing/reingest"
BATCH = "BATCH_REINGEST"


# ── fixtures ────────────────────────────────────────────────────────────────

@pytest.fixture(autouse=True)
def _prime_vat():
    _wc._VAT_CODE_ID_CACHE["23"]  = "222"
    _wc._VAT_CODE_ID_CACHE["WDT"] = "228"
    _wc._VAT_CODE_ID_CACHE["EXP"] = "229"
    yield


@pytest.fixture()
def storage(tmp_path):
    pdb.init_packing_db(tmp_path / "packing.db")
    wdb.init_warehouse_db(tmp_path / "warehouse.db")
    ddb.init_document_db(tmp_path / "documents.db")
    wfdb.init_wfirma_db(tmp_path / "wfirma.db")
    scdb.init(tmp_path / "proforma_links.db")
    return tmp_path


@pytest.fixture()
def client(storage):
    from app.main import app
    with patch.object(settings, "storage_root", storage):
        with TestClient(app, raise_server_exceptions=True) as c:
            yield c


def _auth():
    return {"X-API-KEY": settings.api_key or "test-key"}


# ── helpers ─────────────────────────────────────────────────────────────────

def _make_xlsx(tmp_path: Path, *, name: str, currency_format: str = "",
                rows: list = None, invoice_no: str = "EJL/X-1") -> Path:
    """Build a minimal sales packing Excel with optional currency
    number_format on the Value/Total cells."""
    rows = rows or [{"design": "JE-001", "qty": 1, "value": 200.0}]
    wb = openpyxl.Workbook(); ws = wb.active
    ws.append([f"Export No : {invoice_no}"])
    ws.append([""])
    ws.append(["PkSr", "DesignNo", "Qty", "Value", "Total Value"])
    for i, r in enumerate(rows, start=4):
        ws.append([i - 3, r["design"], r["qty"], r["value"],
                    r["value"] * r["qty"]])
        if currency_format:
            for col in ("D", "E"):
                ws[f"{col}{i}"].number_format = currency_format
    p = tmp_path / name; wb.save(str(p))
    return p


def _seed_sales_doc(client_name: str, *, batch=BATCH, ref="REF",
                     doc_no="SO-RE") -> str:
    """Create a sales_documents row and return its id."""
    return ddb.store_sales_document(
        batch_id=batch, document_id=str(uuid.uuid4()),
        data={"client_name": client_name, "client_ref": ref,
              "sales_doc_no": doc_no},
    )


def _seed_initial_rows(sales_doc_id: str, client: str, n: int = 1,
                        *, batch=BATCH) -> int:
    rows = [{
        "client_name":  client, "client_ref": "REF",
        "product_code": f"OLD-{i}", "design_no": f"OLD-{i}",
        "bag_id": "", "quantity": 1.0, "remarks": "",
        "unit_price": 999.0, "currency": "PLN",  # legacy / wrong
        "total_value": 999.0, "price_source": "stale",
    } for i in range(n)]
    return ddb.store_sales_packing_lines(sales_doc_id, batch, rows)


def _row_count(storage, sales_doc_id: str) -> int:
    with sqlite3.connect(str(storage / "documents.db")) as con:
        return con.execute(
            "SELECT COUNT(*) FROM sales_packing_lines "
            "WHERE sales_document_id=?",
            (sales_doc_id,),
        ).fetchone()[0]


def _all_rows(storage, batch_id: str = BATCH):
    with sqlite3.connect(str(storage / "documents.db")) as con:
        con.row_factory = sqlite3.Row
        return [dict(r) for r in con.execute(
            "SELECT * FROM sales_packing_lines WHERE batch_id=? "
            "ORDER BY client_name, product_code",
            (batch_id,),
        )]


def _post(client_obj, *, files, sales_blocks, override_currency=""):
    data = {
        "batch_id":          BATCH,
        "metadata":          json.dumps({"sales_blocks": sales_blocks}),
        "override_currency": override_currency,
    }
    fobjs = []
    for path in files:
        fobjs.append(("files", (Path(path).name, open(path, "rb"),
                                  "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")))
    return client_obj.post(URL, headers=_auth(), files=fobjs, data=data)


# ── 1. helper-level: replace is idempotent ─────────────────────────────────

def test_replace_helper_is_idempotent(storage):
    sd = _seed_sales_doc("ACME")
    _seed_initial_rows(sd, "ACME", n=3)
    assert _row_count(storage, sd) == 3
    new_rows = [{
        "client_name": "ACME", "client_ref": "REF",
        "product_code": "PC-NEW", "design_no": "D-NEW",
        "bag_id": "", "quantity": 1.0, "remarks": "",
        "unit_price": 200.0, "currency": "EUR",
        "total_value": 200.0, "price_source": "packing_list",
    }]
    r1 = ddb.replace_sales_packing_lines(sd, BATCH, new_rows)
    assert r1 == {"deleted": 3, "inserted": 1}
    assert _row_count(storage, sd) == 1
    # Idempotent: same call again replaces identically.
    r2 = ddb.replace_sales_packing_lines(sd, BATCH, new_rows)
    assert r2 == {"deleted": 1, "inserted": 1}
    assert _row_count(storage, sd) == 1


def test_replace_does_not_touch_other_sales_documents(storage):
    sd1 = _seed_sales_doc("ACME-1")
    sd2 = _seed_sales_doc("ACME-2", doc_no="SO-2")
    _seed_initial_rows(sd1, "ACME-1", n=2)
    _seed_initial_rows(sd2, "ACME-2", n=4)
    ddb.replace_sales_packing_lines(sd1, BATCH, [])
    assert _row_count(storage, sd1) == 0
    assert _row_count(storage, sd2) == 4   # untouched


def test_replace_does_not_touch_other_batches(storage):
    sd_a = _seed_sales_doc("X", batch="BATCH_A")
    sd_b = _seed_sales_doc("X", batch="BATCH_B")
    _seed_initial_rows(sd_a, "X", n=2, batch="BATCH_A")
    _seed_initial_rows(sd_b, "X", n=3, batch="BATCH_B")
    ddb.replace_sales_packing_lines(sd_a, "BATCH_A", [])
    assert _row_count(storage, sd_a) == 0
    assert _row_count(storage, sd_b) == 3


# ── 1. & 2. route: re-upload same file replaces, no duplicates ──────────────

def test_reupload_replaces_no_duplicates(client, storage, tmp_path):
    sd = _seed_sales_doc("Anastazia Panakova")
    _seed_initial_rows(sd, "Anastazia Panakova", n=1)
    p = _make_xlsx(tmp_path, name="anast.xlsx",
                    currency_format='[$-10409]"€"\\ 0',
                    rows=[{"design": "CSTR07718", "qty": 1, "value": 306.0}])
    # First re-ingest.
    r1 = _post(client, files=[p],
                sales_blocks=[{"packing_index": 0,
                                "client_name": "Anastazia Panakova",
                                "client_ref":  "REF"}])
    assert r1.status_code == 200, r1.text
    body1 = r1.json()
    assert body1["files"][0]["deleted_count"]  == 1
    assert body1["files"][0]["inserted_count"] == 1
    assert _row_count(storage, sd) == 1

    # Second re-ingest of the same file → still exactly 1 row.
    r2 = _post(client, files=[p],
                sales_blocks=[{"packing_index": 0,
                                "client_name": "Anastazia Panakova",
                                "client_ref":  "REF"}])
    body2 = r2.json()
    assert body2["files"][0]["deleted_count"]  == 1
    assert body2["files"][0]["inserted_count"] == 1
    assert _row_count(storage, sd) == 1


def test_unrelated_client_rows_preserved(client, storage, tmp_path):
    """Re-ingest of one client must not touch another client's rows."""
    sd_a = _seed_sales_doc("Anastazia Panakova")
    sd_b = _seed_sales_doc("OTHER",          doc_no="SO-OTH")
    _seed_initial_rows(sd_a, "Anastazia Panakova", n=1)
    _seed_initial_rows(sd_b, "OTHER",                n=5)
    p = _make_xlsx(tmp_path, name="anast.xlsx",
                    currency_format='[$-10409]"€"\\ 0',
                    rows=[{"design": "CSTR07718", "qty": 1, "value": 306.0}])
    _post(client, files=[p],
           sales_blocks=[{"packing_index": 0,
                           "client_name": "Anastazia Panakova"}])
    assert _row_count(storage, sd_a) == 1
    assert _row_count(storage, sd_b) == 5  # untouched


# ── 3. sales prices/currency persisted ─────────────────────────────────────

def test_prices_and_currency_persisted(client, storage, tmp_path):
    sd = _seed_sales_doc("ACME")
    p = _make_xlsx(tmp_path, name="acme.xlsx",
                    currency_format='[$-10409]"€"\\ 0',
                    rows=[{"design": "JE-001", "qty": 2, "value": 150.0}])
    _post(client, files=[p],
           sales_blocks=[{"packing_index": 0, "client_name": "ACME"}])
    rows = _all_rows(storage)
    assert len(rows) == 1
    assert rows[0]["unit_price"]   == 150.0
    assert rows[0]["currency"]     == "EUR"
    assert rows[0]["total_value"]  == 300.0   # 150 × 2
    assert rows[0]["price_source"] == "packing_list"


# ── 4. Clear-Diamonds USD persists ─────────────────────────────────────────

def test_clear_diamonds_usd_persists(client, storage, tmp_path):
    sd = _seed_sales_doc("Clear-Diamonds")
    p = _make_xlsx(tmp_path, name="clear.xlsx",
                    currency_format='[$-10409]"$"\\ 0',
                    rows=[{"design": "JR05671", "qty": 1, "value": 251.37}])
    r = _post(client, files=[p],
               sales_blocks=[{"packing_index": 0,
                               "client_name": "Clear-Diamonds"}])
    body = r.json()
    assert body["files"][0]["currency"]        == "USD"
    assert body["files"][0]["currency_source"] == "excel_symbol"
    rows = _all_rows(storage)
    assert rows[0]["currency"] == "USD"


def test_clear_diamonds_default_eur_does_not_override_excel_usd(
    client, storage, tmp_path,
):
    """Even if customer default is EUR, an Excel ``$`` symbol wins."""
    wfdb.upsert_customer(client_name="Clear-Diamonds",
                          wfirma_customer_id="X", country="LV",
                          vat_id="LV12345", match_status="matched")
    wfdb.set_customer_default_currency(client_name="Clear-Diamonds",
                                        currency="EUR")
    sd = _seed_sales_doc("Clear-Diamonds")
    p = _make_xlsx(tmp_path, name="cd.xlsx",
                    currency_format='[$-10409]"$"\\ 0',
                    rows=[{"design": "JE-1", "qty": 1, "value": 50.0}])
    body = _post(client, files=[p],
                  sales_blocks=[{"packing_index": 0,
                                  "client_name": "Clear-Diamonds"}]).json()
    assert body["files"][0]["currency"]        == "USD"
    assert body["files"][0]["currency_source"] == "excel_symbol"


# ── 5. Anastazia EUR persists ──────────────────────────────────────────────

def test_anastazia_eur_persists(client, storage, tmp_path):
    sd = _seed_sales_doc("Anastazia Panakova")
    p = _make_xlsx(tmp_path, name="anast.xlsx",
                    currency_format='[$-10409]"€"\\ 0',
                    rows=[{"design": "CSTR07718", "qty": 1, "value": 306.0}])
    body = _post(client, files=[p],
                  sales_blocks=[{"packing_index": 0,
                                  "client_name": "Anastazia Panakova"}]).json()
    assert body["files"][0]["currency"]        == "EUR"
    assert body["files"][0]["currency_source"] == "excel_symbol"
    rows = _all_rows(storage)
    assert rows[0]["currency"]   == "EUR"
    assert rows[0]["unit_price"] == 306.0


# ── 6. PND tiebreak persists ────────────────────────────────────────────────

def test_pnd_tiebreak_persists(client, storage, tmp_path):
    """Two PND rows with prices 5.13 and 51.30 USD → matched against
    supplier-side EJL/26-27/123-2 ($36) and EJL/26-27/123-3 ($4)."""
    # Seed supplier-side packing for invoice EJL/26-27/123: two pendants.
    # NB: distinct pack_sr is required — packing_db's primary dedup key is
    # (batch_id, invoice_no, packing_document_id, pack_sr).
    pdb.upsert_packing_lines([
        {"batch_id": BATCH, "invoice_no": "EJL/26-27/123",
         "invoice_line_position": 2, "product_code": "EJL/26-27/123-2",
         "design_no": "", "bag_id": "", "tray_id": "",
         "item_type": "PND", "uom": "PCS", "quantity": 1.0,
         "gross_weight": 0, "net_weight": 0, "metal": "", "karat": "",
         "stone_type": "", "remarks": "", "extracted_confidence": 1.0,
         "requires_manual_review": False, "pack_sr": 1.0,
         "unit_price": 0, "total_value": 0},
        {"batch_id": BATCH, "invoice_no": "EJL/26-27/123",
         "invoice_line_position": 3, "product_code": "EJL/26-27/123-3",
         "design_no": "", "bag_id": "", "tray_id": "",
         "item_type": "PND", "uom": "PCS", "quantity": 1.0,
         "gross_weight": 0, "net_weight": 0, "metal": "", "karat": "",
         "stone_type": "", "remarks": "", "extracted_confidence": 1.0,
         "requires_manual_review": False, "pack_sr": 2.0,
         "unit_price": 0, "total_value": 0},
    ])
    # Seed import invoice prices so the disambiguator has supplier prices.
    ddb.store_invoice_lines("doc-x", BATCH, [
        {"invoice_no": "EJL/26-27/123", "line_position": 2,
         "product_code": "EJL/26-27/123-2", "description": "",
         "quantity": 1.0, "unit_price": 36.0, "total_value": 36.0,
         "currency": "USD", "rate_usd": 36.0, "amount_usd": 36.0},
        {"invoice_no": "EJL/26-27/123", "line_position": 3,
         "product_code": "EJL/26-27/123-3", "description": "",
         "quantity": 1.0, "unit_price": 4.0, "total_value": 4.0,
         "currency": "USD", "rate_usd": 4.0, "amount_usd": 4.0},
    ])
    sd = _seed_sales_doc("Clear-Diamonds")
    p = _make_xlsx(tmp_path, name="cd.xlsx",
                    currency_format='[$-10409]"$"\\ 0',
                    invoice_no="EJL/26-27/123",
                    rows=[
                        {"design": "PND", "qty": 1, "value": 51.30},
                        {"design": "PND", "qty": 1, "value":  5.13},
                    ])
    body = _post(client, files=[p],
                  sales_blocks=[{"packing_index": 0,
                                  "client_name": "Clear-Diamonds"}]).json()
    assert body["files"][0]["pnd_summary"]["applied"] is True
    rows = _all_rows(storage)
    by_price = {r["unit_price"]: r["product_code"] for r in rows}
    assert by_price[51.30] == "EJL/26-27/123-2"
    assert by_price[5.13]  == "EJL/26-27/123-3"
    # Currency persisted on both rows.
    assert {r["currency"] for r in rows} == {"USD"}


# ── 7. mixed currency blocks ────────────────────────────────────────────────

def test_mixed_currency_blocks_insert(client, storage, tmp_path):
    sd = _seed_sales_doc("ACME")
    _seed_initial_rows(sd, "ACME", n=2)   # baseline rows we expect untouched
    # Two rows with different currency symbols on Value/Total cells.
    wb = openpyxl.Workbook(); ws = wb.active
    ws.append(["Export No : EJL/X-1"]); ws.append([""])
    ws.append(["PkSr", "DesignNo", "Qty", "Value", "Total Value"])
    ws.append([1, "JE-1", 1, 100, 100])
    ws.append([2, "JE-2", 1, 200, 200])
    ws["D4"].number_format = '[$-10409]"$"\\ 0'
    ws["E4"].number_format = '[$-10409]"$"\\ 0'
    ws["D5"].number_format = '[$-10409]"€"\\ 0'
    ws["E5"].number_format = '[$-10409]"€"\\ 0'
    p = tmp_path / "mixed.xlsx"; wb.save(str(p))

    r = _post(client, files=[p],
               sales_blocks=[{"packing_index": 0, "client_name": "ACME"}])
    body = r.json()
    f = body["files"][0]
    assert f["currency_conflict"]   is True
    assert f["currency_source"]     == "mixed_excel_currencies_block"
    assert f["inserted_count"]      == 0
    assert f["deleted_count"]       == 0   # blocked → no replace happened
    # Pre-existing rows untouched.
    assert _row_count(storage, sd) == 2


def test_mixed_currency_with_override_proceeds(client, storage, tmp_path):
    """Operator can clear the conflict by supplying override_currency."""
    sd = _seed_sales_doc("ACME")
    wb = openpyxl.Workbook(); ws = wb.active
    ws.append(["Export No : EJL/X-1"]); ws.append([""])
    ws.append(["PkSr", "DesignNo", "Qty", "Value", "Total Value"])
    ws.append([1, "JE-1", 1, 100, 100])
    ws.append([2, "JE-2", 1, 200, 200])
    ws["D4"].number_format = '[$-10409]"$"\\ 0'
    ws["E4"].number_format = '[$-10409]"$"\\ 0'
    ws["D5"].number_format = '[$-10409]"€"\\ 0'
    ws["E5"].number_format = '[$-10409]"€"\\ 0'
    p = tmp_path / "mixed.xlsx"; wb.save(str(p))
    body = _post(client, files=[p],
                  sales_blocks=[{"packing_index": 0, "client_name": "ACME"}],
                  override_currency="USD").json()
    f = body["files"][0]
    assert f["inserted_count"]    == 2
    assert f["currency"]          == "USD"
    assert f["currency_source"]   == "operator_override"
    rows = _all_rows(storage)
    assert {r["currency"] for r in rows} == {"USD"}


# ── 8. response shape ───────────────────────────────────────────────────────

def test_response_includes_counts_and_warnings(client, storage, tmp_path):
    sd = _seed_sales_doc("ACME")
    _seed_initial_rows(sd, "ACME", n=2)
    p = _make_xlsx(tmp_path, name="x.xlsx",
                    currency_format='[$-10409]"€"\\ 0',
                    rows=[{"design": "JE-1", "qty": 1, "value": 50.0}])
    body = _post(client, files=[p],
                  sales_blocks=[{"packing_index": 0,
                                  "client_name": "ACME"}]).json()
    f = body["files"][0]
    for key in ("file", "client_name", "before_count", "deleted_count",
                "inserted_count", "currency", "currency_source",
                "currency_conflict", "warnings", "pnd_summary"):
        assert key in f, key
    assert f["before_count"]   == 2
    assert f["deleted_count"]  == 2
    assert f["inserted_count"] == 1


def test_unknown_client_warns_and_skips(client, storage, tmp_path):
    """No matching sales_document → file skipped with warning."""
    p = _make_xlsx(tmp_path, name="ghost.xlsx",
                    currency_format='[$-10409]"€"\\ 0',
                    rows=[{"design": "JE-1", "qty": 1, "value": 50.0}])
    body = _post(client, files=[p],
                  sales_blocks=[{"packing_index": 0,
                                  "client_name": "GHOST"}]).json()
    f = body["files"][0]
    assert f["inserted_count"] == 0
    assert any("no sales_document" in w for w in f["warnings"])


def test_missing_batch_id_rejected(client):
    r = client.post(URL, headers=_auth(),
                     data={"batch_id": "", "metadata": "{}"})
    # FastAPI may surface "" as 422 (missing) or our explicit 400 — both
    # are valid rejections for an unusable batch_id.
    assert r.status_code in (400, 422)


def test_missing_files_rejected(client):
    r = client.post(URL, headers=_auth(),
                     data={"batch_id": BATCH, "metadata": "{}"})
    assert r.status_code == 400


# ── Orphan-repair: sales_document_id bypass for empty-client rows ─────────────
# Covers the case where initial intake stored a sales_documents row with
# client_name='' because the operator omitted it from the upload metadata.
# The reingest endpoint now accepts sales_document_id in the metadata block
# as a bypass key, backfills client_name, and creates the proforma draft.

def test_orphan_repair_backfills_client_name(client, storage, tmp_path):
    """Reingest with sales_document_id fixes empty-client row and inserts lines."""
    # Seed a sales_documents row with empty client_name (orphaned)
    orphan_id = ddb.store_sales_document(
        batch_id=BATCH, document_id=str(uuid.uuid4()),
        data={"client_name": "", "client_ref": "",
              "sales_doc_no": "SO-ORPHAN"},
    )
    assert _row_count(storage, orphan_id) == 0

    p = _make_xlsx(tmp_path, name="orphan.xlsx",
                    rows=[{"design": "JE-999", "qty": 2, "value": 150.0}])
    r = _post(client, files=[p], sales_blocks=[{
        "packing_index":   0,
        "client_name":     "UAB Monodija",
        "client_ref":      "REF-258",
        "sales_document_id": orphan_id,
    }])
    assert r.status_code == 200, r.text
    body = r.json()
    f0 = body["files"][0]
    assert f0["inserted_count"] == 1
    assert f0.get("client_name_repaired") is True
    assert f0.get("repaired_from_doc_id") == orphan_id
    assert _row_count(storage, orphan_id) == 1

    # Verify client_name was backfilled in sales_documents
    with sqlite3.connect(str(storage / "documents.db")) as con:
        row = con.execute(
            "SELECT client_name FROM sales_documents WHERE id=?",
            (orphan_id,),
        ).fetchone()
    assert row is not None
    assert row[0] == "UAB Monodija"


def test_orphan_repair_unknown_doc_id_warns(client, storage, tmp_path):
    """Reingest with unknown sales_document_id returns a warning, not 500."""
    p = _make_xlsx(tmp_path, name="unknown.xlsx",
                    rows=[{"design": "JE-001", "qty": 1, "value": 100.0}])
    r = _post(client, files=[p], sales_blocks=[{
        "packing_index":     0,
        "client_name":       "Ghost Client",
        "sales_document_id": "aaaabbbb-0000-0000-0000-000000000000",
    }])
    assert r.status_code == 200, r.text
    f0 = r.json()["files"][0]
    assert any("not found in batch" in w for w in f0["warnings"])
    assert f0["inserted_count"] == 0


def test_orphan_repair_no_sales_document_id_still_warns(client, storage, tmp_path):
    """Without sales_document_id, the original 'no sales_document found' warning fires."""
    p = _make_xlsx(tmp_path, name="nowarn.xlsx",
                    rows=[{"design": "JE-002", "qty": 1, "value": 50.0}])
    r = _post(client, files=[p], sales_blocks=[{
        "packing_index": 0,
        "client_name":   "Nonexistent Client",
    }])
    assert r.status_code == 200, r.text
    f0 = r.json()["files"][0]
    assert any("no sales_document found" in w for w in f0["warnings"])
    assert f0["inserted_count"] == 0


def test_orphan_repair_normal_path_unaffected(client, storage, tmp_path):
    """Normal reingest (client_name present in DB) still works when orphan path is in place."""
    sd = _seed_sales_doc("Normal Client")
    _seed_initial_rows(sd, "Normal Client", n=1)
    p = _make_xlsx(tmp_path, name="normal.xlsx",
                    rows=[{"design": "JE-777", "qty": 3, "value": 75.0}])
    r = _post(client, files=[p], sales_blocks=[{
        "packing_index": 0,
        "client_name":   "Normal Client",
    }])
    assert r.status_code == 200, r.text
    f0 = r.json()["files"][0]
    assert f0["inserted_count"] == 1
    assert f0.get("client_name_repaired") is None
    assert _row_count(storage, sd) == 1
