"""
test_parser_currency_symbols.py — Currency symbol detection from Excel
``cell.number_format`` and the resolution priority across all sources.

Pins (each maps to a numbered scope rule):
  1. ``$`` symbol on Value/Total cells parses as USD.
  2. ``€`` parses as EUR.
  3. Excel symbol overrides operator/customer default in the intake ladder.
  4. Operator currency used only when Excel has no currency.
  5. Customer default used only when Excel/operator missing.
  6. Clear-Diamonds real source file dry-run returns USD.
  7. Multiple symbols in one file emits warning + blocks dominant fallback.

Plus thin guards on the helper:
  - locale prefix ``[$-10409]`` is NOT confused with the literal "$" symbol.
  - ``"zł"`` parses as PLN.
  - blank / "General" format → "".
"""
from __future__ import annotations

from pathlib import Path

import openpyxl
import pytest

from app.services.invoice_packing_extractor import (
    extract_packing,
    _currency_from_format_string,
)


# ── Helper-level: format-string parser ─────────────────────────────────────

@pytest.mark.parametrize("fmt, expected", [
    ('[$-10409]"€"\\ 0;\\("€"\\ 0\\)',          "EUR"),
    ('[$-10409]"$"\\ 0;\\("$"\\ 0\\)',          "USD"),
    ('#,##0.00 "zł"',                            "PLN"),
    ('"£"#,##0.00',                              "GBP"),
    ('CHF #,##0.00',                             "CHF"),
    ('General',                                   ""),
    ('',                                          ""),
    ('mm-dd-yy',                                  ""),
    ('[$-10409]0.00;\\(0.00\\)',                 ""),     # locale only, no symbol
    ('"USD" #,##0',                              "USD"),
])
def test_currency_from_format_string(fmt, expected):
    assert _currency_from_format_string(fmt) == expected


def test_locale_prefix_is_not_dollar_sign():
    """`[$-10409]` is Excel's English-US locale tag, NOT a USD marker.
    The parser must strip it before scanning for symbols."""
    # No actual currency symbol — only the locale prefix.
    assert _currency_from_format_string('[$-10409]0.00') == ""
    # Locale + Polish złoty — symbol still wins.
    assert _currency_from_format_string('[$-10415]#,##0.00 "zł"') == "PLN"


# ── Excel-level: build minimal sheets and parse ────────────────────────────

def _make_sheet(tmp_path: Path, *, fmt: str, name: str = "x.xlsx",
                value: float = 200.0, design: str = "JE-100",
                invoice_no: str = "EJL/X-1") -> Path:
    wb = openpyxl.Workbook(); ws = wb.active
    ws.append([f"Export No : {invoice_no}"])
    ws.append([""])
    ws.append(["PkSr", "DesignNo", "Qty", "Value", "Total Value"])
    ws.append([1, design, 1, value, value])
    # Stamp number_format on the Value (col D) and Total Value (col E) cells.
    for col_letter in ("D", "E"):
        ws[f"{col_letter}4"].number_format = fmt
    p = tmp_path / name; wb.save(str(p))
    return p


def test_dollar_symbol_format_parses_usd(tmp_path):
    p = _make_sheet(tmp_path, fmt='[$-10409]"$"\\ 0;\\("$"\\ 0\\)')
    rows, _, _ = extract_packing(p)
    assert len(rows) == 1
    assert rows[0]["currency"]        == "USD"
    assert rows[0]["currency_source"] == "excel_symbol"


def test_euro_symbol_format_parses_eur(tmp_path):
    p = _make_sheet(tmp_path, fmt='[$-10409]"€"\\ 0;\\("€"\\ 0\\)')
    rows, _, _ = extract_packing(p)
    assert rows[0]["currency"]        == "EUR"
    assert rows[0]["currency_source"] == "excel_symbol"


def test_zloty_format_parses_pln(tmp_path):
    p = _make_sheet(tmp_path, fmt='#,##0.00 "zł"')
    rows, _, _ = extract_packing(p)
    assert rows[0]["currency"]        == "PLN"
    assert rows[0]["currency_source"] == "excel_symbol"


def test_no_currency_in_format_falls_back_to_token_or_blank(tmp_path):
    """No symbol in number_format AND no header/preamble token → blank."""
    p = _make_sheet(tmp_path, fmt='[$-10409]0.00;\\(0.00\\)')
    rows, _, _ = extract_packing(p)
    assert (rows[0].get("currency") or "") == ""


# ── Multi-currency conflict in one file ────────────────────────────────────

def test_mixed_currency_symbols_flagged(tmp_path):
    """Two Value cells with different currency symbols → conflict flag."""
    wb = openpyxl.Workbook(); ws = wb.active
    ws.append(["Export No : EJL/X-1"])
    ws.append([""])
    ws.append(["PkSr", "DesignNo", "Qty", "Value", "Total Value"])
    ws.append([1, "JE-1", 1, 100, 100])
    ws.append([2, "JE-2", 1, 200, 200])
    ws["D4"].number_format = '[$-10409]"$"\\ 0'
    ws["E4"].number_format = '[$-10409]"$"\\ 0'
    ws["D5"].number_format = '[$-10409]"€"\\ 0'
    ws["E5"].number_format = '[$-10409]"€"\\ 0'
    p = tmp_path / "mixed.xlsx"; wb.save(str(p))
    rows, _, _ = extract_packing(p)
    # Each row is stamped with the dominant detected currency, AND each
    # row carries the conflict flag so intake can refuse to silently
    # commit the dominant value.
    assert any(r.get("currency_conflict") for r in rows)


# ── 6. Real Clear-Diamonds source file ─────────────────────────────────────

_REAL_FILES_ROOT = Path(
    "/Users/amitgupta/Library/Application Support/estrellajewels/storage/"
    "outputs/SHIPMENT_6049349806_2026-05_7409ac77/source/sales"
)


@pytest.mark.skipif(not _REAL_FILES_ROOT.exists(),
                     reason="real AWB 6049349806 source files not on disk")
def test_real_clear_diamonds_file_returns_usd():
    candidates = list(_REAL_FILES_ROOT.glob("*Clear-Diamonds*.xlsx"))
    assert candidates, "Clear-Diamonds source file missing"
    rows, _, _ = extract_packing(candidates[0])
    currencies = {r.get("currency") for r in rows if r.get("currency")}
    assert currencies == {"USD"}, currencies
    assert all(r.get("currency_source") == "excel_symbol"
               for r in rows if r.get("currency"))


@pytest.mark.skipif(not _REAL_FILES_ROOT.exists(),
                     reason="real AWB 6049349806 source files not on disk")
def test_real_eur_files_return_eur():
    """Anastazia / OMARA / Impact Gallery all carry the € symbol."""
    for pattern in ("*Anastazia*.xlsx", "*OMARA*.xlsx", "*Impact*.xlsx"):
        cands = list(_REAL_FILES_ROOT.glob(pattern))
        assert cands, pattern
        rows, _, _ = extract_packing(cands[0])
        currencies = {r.get("currency") for r in rows if r.get("currency")}
        assert currencies == {"EUR"}, (pattern, currencies)


# ── 3-5. Resolution ladder (route-level via direct simulation) ─────────────

def test_excel_symbol_overrides_customer_default(tmp_path):
    """Even if the customer default is EUR, an Excel ``$`` symbol wins."""
    p = _make_sheet(tmp_path, fmt='[$-10409]"$"\\ 0;\\("$"\\ 0\\)')
    rows, _, _ = extract_packing(p)
    # Simulate the intake ladder (Excel → operator → customer_default → blank).
    operator_currency = ""
    customer_default  = "EUR"
    first_excel_ccy   = next((r["currency"] for r in rows
                                if (r.get("currency") or "")), "")
    first_excel_src   = next((r.get("currency_source") for r in rows
                                if (r.get("currency") or "")), "")
    if first_excel_ccy:
        chosen, src = first_excel_ccy, first_excel_src
    elif operator_currency:
        chosen, src = operator_currency, "operator"
    elif customer_default:
        chosen, src = customer_default, "customer_default"
    else:
        chosen, src = "", "missing"
    assert chosen == "USD"
    assert src    == "excel_symbol"


def test_operator_currency_only_when_excel_missing(tmp_path):
    p = _make_sheet(tmp_path, fmt='[$-10409]0.00')   # no symbol
    rows, _, _ = extract_packing(p)
    operator_currency = "EUR"
    customer_default  = "PLN"
    first = next((r["currency"] for r in rows if r.get("currency")), "")
    if first:
        chosen, src = first, "excel"
    elif operator_currency:
        chosen, src = operator_currency, "operator"
    elif customer_default:
        chosen, src = customer_default, "customer_default"
    else:
        chosen, src = "", "missing"
    assert chosen == "EUR"
    assert src    == "operator"


def test_customer_default_only_when_excel_and_operator_missing(tmp_path):
    p = _make_sheet(tmp_path, fmt='General')         # no symbol
    rows, _, _ = extract_packing(p)
    operator_currency = ""
    customer_default  = "EUR"
    first = next((r["currency"] for r in rows if r.get("currency")), "")
    if first:
        chosen, src = first, "excel"
    elif operator_currency:
        chosen, src = operator_currency, "operator"
    elif customer_default:
        chosen, src = customer_default, "customer_default"
    else:
        chosen, src = "", "missing"
    assert chosen == "EUR"
    assert src    == "customer_default"
