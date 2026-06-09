"""
Verification suite for the Excel ingestion stack.
Run: python service/tests/test_excel_stack.py
"""
import importlib.metadata as meta
import os
import tempfile

PACKAGES = [
    ("pandas", "pandas"),
    ("openpyxl", "openpyxl"),
    ("pyxlsb", "pyxlsb"),
    ("rapidfuzz", "rapidfuzz"),
    ("xlsxwriter", "xlsxwriter"),
    ("python-dateutil", "dateutil"),
    ("chardet", "chardet"),
]


def test_imports():
    print("=== Import verification ===")
    all_ok = True
    for dist_name, import_name in PACKAGES:
        try:
            __import__(import_name)
            v = meta.version(dist_name)
            print(f"  OK  {dist_name}=={v}")
        except Exception as e:
            print(f"  FAIL {dist_name}: {e}")
            all_ok = False
    return all_ok


def test_functional():
    import pandas as pd
    import openpyxl
    from rapidfuzz import fuzz
    from dateutil import parser as du_parser

    print("\n=== Functional validation ===")

    # --- Write XLSX via xlsxwriter ---
    tmp = tempfile.mktemp(suffix=".xlsx")
    import xlsxwriter as xlsxw
    wb = xlsxw.Workbook(tmp)
    ws = wb.add_worksheet("Sheet1")
    headers = ["AWB", "Customer", "Qty", "Value EUR", "Date"]
    rows = [
        ["4789974092", "ACME Jewels", 10, 1250.00, "2026-01-15"],
        ["9938632830", "Star Fashion", 5, 875.50, "2026-02-20"],
        ["1122334455", "Gold Shop",   3, 520.00,  "2026-03-10"],
    ]
    for c, h in enumerate(headers):
        ws.write(0, c, h)
    for r, row in enumerate(rows, start=1):
        for c, val in enumerate(row):
            ws.write(r, c, val)
    wb.close()
    print(f"  xlsxwriter: wrote {tmp}")

    # --- Read XLSX via pandas/openpyxl engine ---
    df = pd.read_excel(tmp, engine="openpyxl")
    assert len(df) == 3, f"Expected 3 rows, got {len(df)}"
    assert list(df.columns) == headers, f"Column mismatch: {list(df.columns)}"
    print(f"  pandas/openpyxl: read {len(df)} rows, {len(df.columns)} cols — OK")

    # --- Read via openpyxl directly ---
    wb2 = openpyxl.load_workbook(tmp)
    ws2 = wb2.active
    row_count = ws2.max_row - 1  # exclude header
    assert row_count == 3, f"openpyxl row count mismatch: {row_count}"
    print(f"  openpyxl direct: {row_count} data rows — OK")

    # --- rapidfuzz column mapping simulation ---
    candidate_headers = ["awb number", "tracking_id", "customer name", "quantity", "value", "date"]
    target = "AWB"
    scores = [(h, fuzz.WRatio(target.lower(), h.lower())) for h in candidate_headers]
    best = max(scores, key=lambda x: x[1])
    print(f"  rapidfuzz: best match for '{target}' -> '{best[0]}' (score {best[1]}) - OK")

    # --- python-dateutil date parsing ---
    dates = ["15 Jan 2026", "2026-02-20", "20/03/2026", "March 10, 2026"]
    parsed = [str(du_parser.parse(d).date()) for d in dates]
    print(f"  dateutil: parsed {len(parsed)} date formats — OK")

    # --- chardet encoding detection ---
    import chardet
    sample = "Złoty naszyjnik z brylantami".encode("utf-8")
    detected = chardet.detect(sample)
    print(f"  chardet: detected encoding={detected['encoding']} confidence={detected['confidence']:.2f} — OK")

    # cleanup
    os.unlink(tmp)
    print(f"  Temp file removed.")
    return True


if __name__ == "__main__":
    ok1 = test_imports()
    ok2 = test_functional()
    if ok1 and ok2:
        print("\nExcel stack OK — all checks passed.")
    else:
        raise SystemExit(1)
